"""
Endpoints para gestión de etiquetas de envío flex.

Permite:
- Subir archivos .zip/.txt con etiquetas ZPL de MercadoEnvíos
- Escanear QR individuales con pistola de código de barras
- Listar etiquetas con datos de envío (destinatario, dirección, CP, cordón)
- Asignar logísticas a etiquetas individual o masivamente
- Cambiar fecha de envío (reprogramar)
- Estadísticas de distribución por cordón, logística y estado

Los archivos ZPL contienen JSONs embebidos en los QR codes con formato:
{"id":"46458064834","sender_id":413658225,"hash_code":"...","security_digit":"0"}

El campo "id" del QR = mlshippingid en tb_mercadolibre_orders_shipping.
"""

import logging
import re
import io
import json
import zipfile
from io import BytesIO
from datetime import date, datetime, UTC

from pathlib import Path as FilePath

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session, aliased
from sqlalchemy import func, cast, case, and_, or_, desc, Numeric, text
from typing import Any, List, Optional
from pydantic import BaseModel, ConfigDict, Field

from app.core.database import get_db
from app.api.deps import get_current_user
from app.models.usuario import Usuario
from app.models.etiqueta_envio import EtiquetaEnvio
from app.models.etiqueta_colecta import EtiquetaColecta
from app.models.logistica import Logistica
from app.models.transporte import Transporte
from app.models.mercadolibre_order_shipping import MercadoLibreOrderShipping
from app.models.codigo_postal_cordon import CodigoPostalCordon
from app.models.sale_order_header import SaleOrderHeader
from app.models.sale_order_status import SaleOrderStatus
from app.models.etiqueta_envio_audit import EtiquetaEnvioAudit
from app.models.operador import Operador
from app.models.operador_actividad import OperadorActividad
from app.models.logistica_costo_cordon import LogisticaCostoCordon
from app.services.etiqueta_enrichment_service import (
    enriquecer_etiquetas_sync,
    re_enriquecer_desde_db,
    re_enriquecer_por_http,
)
from app.services.ml_webhook_service import fetch_shipment_label_zpl
from app.services.geocoding_service import geocode_address
from app.models.sale_order_detail import SaleOrderDetail
from app.models.tb_item import TBItem
from app.models.mercadolibre_order_header import MercadoLibreOrderHeader
from app.models.mercadolibre_user_data import MercadoLibreUserData
from app.models.configuracion import Configuracion
from app.services.permisos_service import verificar_permiso

router = APIRouter()

# Regex para extraer JSONs del QR embebidos en ZPL
QR_JSON_REGEX = re.compile(r'\{"id":"[^}]+\}')


logger = logging.getLogger(__name__)


def _check_permiso(db: Session, user: Usuario, codigo: str) -> None:
    """Verifica permiso y lanza 403 si no lo tiene."""
    if not verificar_permiso(db, user, codigo):
        raise HTTPException(
            status_code=403,
            detail=f"No tenés permiso: {codigo}",
        )


def _get_lluvia_config(db: Session) -> tuple[str, float]:
    """Lee la configuración de offset por lluvia desde la tabla configuracion.

    Returns:
        (tipo, valor) — ej: ("fijo", 1800.0) o ("porcentaje", 50.0)
        Si no existe, devuelve ("fijo", 0.0) (sin offset).
    """
    tipo_row = db.query(Configuracion.valor).filter(Configuracion.clave == "lluvia_offset_tipo").first()
    valor_row = db.query(Configuracion.valor).filter(Configuracion.clave == "lluvia_offset_valor").first()
    tipo = tipo_row[0] if tipo_row else "fijo"
    try:
        valor = float(valor_row[0]) if valor_row else 0.0
    except (ValueError, TypeError):
        valor = 0.0
    return tipo, valor


def _build_costo_case(
    costo_turbo_col: Any,
    costo_normal_col: Any,
    lluvia_tipo: str,
    lluvia_valor: float,
) -> Any:
    """Construye la CASE expression de costo_envio con soporte lluvia.

    Lógica:
    1. turbo + lluvia → costo_turbo + offset (fijo o %)
    2. turbo           → costo_turbo (o fallback a normal si no hay turbo)
    3. else            → costo_normal

    El costo_override (manual) se aplica en el COALESCE externo.
    """
    costo_turbo_eff = func.coalesce(costo_turbo_col, costo_normal_col)

    if lluvia_valor > 0:
        if lluvia_tipo == "porcentaje":
            # costo_turbo * (1 + pct/100)
            costo_lluvia = cast(
                costo_turbo_eff * (1 + lluvia_valor / 100),
                Numeric(12, 2),
            )
        else:
            # costo_turbo + monto fijo
            costo_lluvia = cast(
                costo_turbo_eff + lluvia_valor,
                Numeric(12, 2),
            )
    else:
        # Sin offset → lluvia = turbo
        costo_lluvia = cast(costo_turbo_eff, Numeric(12, 2))

    return case(
        (
            and_(EtiquetaEnvio.es_turbo.is_(True), EtiquetaEnvio.es_lluvia.is_(True)),
            costo_lluvia,
        ),
        (
            EtiquetaEnvio.es_turbo.is_(True),
            cast(costo_turbo_eff, Numeric(12, 2)),
        ),
        else_=cast(costo_normal_col, Numeric(12, 2)),
    )


async def _geocode_envio_manual(
    shipping_id: str,
    street_name: str,
    street_number: str,
    city_name: str,
    transporte_id: Optional[int],
    zip_code: Optional[str] = None,
) -> None:
    """
    Background task: geocodifica un envío manual y guarda lat/lng.

    Lógica:
      1. Si tiene transporte asignado y el transporte YA tiene lat/lng → usar esos.
      2. Si tiene transporte con dirección pero sin lat/lng → geocodificar la
         dirección del transporte, guardar en AMBOS (transporte + etiqueta).
      3. Si no tiene transporte → geocodificar la dirección del cliente.
    """
    from app.core.database import SessionLocal

    db = SessionLocal()
    try:
        etiqueta = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id == shipping_id).first()
        if not etiqueta:
            return

        lat, lng = None, None

        # Caso 1 y 2: tiene transporte
        if transporte_id is not None:
            transporte = db.query(Transporte).filter(Transporte.id == transporte_id).first()
            if transporte:
                if transporte.latitud and transporte.longitud:
                    # Caso 1: transporte ya geocodificado
                    lat, lng = transporte.latitud, transporte.longitud
                elif transporte.direccion:
                    # Caso 2: geocodificar dirección del transporte
                    ciudad_transp = transporte.localidad or "Buenos Aires"
                    coords = await geocode_address(transporte.direccion, ciudad=ciudad_transp, db=db)
                    if coords:
                        lat, lng = coords
                        transporte.latitud = lat
                        transporte.longitud = lng

        # Caso 3: sin transporte o transporte sin dirección → geocodificar cliente
        if lat is None and street_name:
            direccion_cliente = f"{street_name} {street_number}".strip()
            ciudad = city_name or "Buenos Aires"
            coords = await geocode_address(direccion_cliente, ciudad=ciudad, zip_code=zip_code, db=db)
            if coords:
                lat, lng = coords

        if lat is not None and lng is not None:
            etiqueta.latitud = lat
            etiqueta.longitud = lng
            db.commit()
            logger.info("Geocoding OK para envío manual %s → (%.6f, %.6f)", shipping_id, lat, lng)
        else:
            logger.warning("Geocoding sin resultado para envío manual %s", shipping_id)

    except Exception:
        logger.exception("Error geocodificando envío manual %s", shipping_id)
        db.rollback()
    finally:
        db.close()


# ── Schemas ──────────────────────────────────────────────────────────


class UploadResultResponse(BaseModel):
    """Resultado de la carga de etiquetas desde archivo."""

    total: int
    nuevas: int
    duplicadas: int
    errores: int = 0
    detalle_errores: List[str] = []


class ManualScanRequest(BaseModel):
    """Payload del escaneo manual con pistola."""

    json_data: str = Field(
        description='JSON raw del QR, ej: {"id":"46458064834","sender_id":413658225,...}',
    )


class ManualScanResponse(BaseModel):
    """Resultado del escaneo manual."""

    duplicada: bool
    shipping_id: str
    mensaje: str


class EtiquetaEnvioResponse(BaseModel):
    """Etiqueta con datos enriquecidos de envío, dirección y estado."""

    shipping_id: str
    sender_id: Optional[int] = None
    nombre_archivo: Optional[str] = None
    fecha_envio: date
    logistica_id: Optional[int] = None
    logistica_nombre: Optional[str] = None
    logistica_color: Optional[str] = None

    # Order ID de MercadoLibre (para link a detalle de venta en ML)
    ml_order_id: Optional[str] = None

    # Buyer nickname de MercadoLibre (ej: GUSY2007)
    mluser_nickname: Optional[str] = None

    # Datos de ML shipping
    mlreceiver_name: Optional[str] = None
    mlstreet_name: Optional[str] = None
    mlstreet_number: Optional[str] = None
    mlzip_code: Optional[str] = None
    mlcity_name: Optional[str] = None
    mlstatus: Optional[str] = None

    # Cordón
    cordon: Optional[str] = None

    # Datos enriquecidos (ML webhook)
    latitud: Optional[float] = None
    longitud: Optional[float] = None
    direccion_completa: Optional[str] = None
    direccion_comentario: Optional[str] = None

    # Estado ERP
    ssos_id: Optional[int] = None
    ssos_name: Optional[str] = None
    ssos_color: Optional[str] = None

    # Pistoleado
    pistoleado_at: Optional[str] = None
    pistoleado_caja: Optional[str] = None
    pistoleado_operador_nombre: Optional[str] = None

    # Costo de envío (calculado desde logistica_costo_cordon, o override manual)
    costo_envio: Optional[float] = None
    costo_override: Optional[float] = None

    # Envío manual (sin ML)
    es_manual: bool = False
    manual_bra_id: Optional[int] = None
    manual_soh_id: Optional[int] = None
    manual_cust_id: Optional[int] = None
    manual_comment: Optional[str] = None
    manual_phone: Optional[str] = None

    # Outlet (título de item contiene "outlet")
    es_outlet: bool = False

    # Turbo (mlshipping_method_id == "515282")
    es_turbo: bool = False

    # Lluvia — recargo extra sobre turbo
    es_lluvia: bool = False

    # Flag de envío (mal pasado, cancelado, duplicado, otro)
    flag_envio: Optional[str] = None
    flag_envio_motivo: Optional[str] = None
    flag_envio_at: Optional[str] = None
    flag_envio_usuario_nombre: Optional[str] = None

    # Creado por usuario del sistema (cuando viene de Pedidos Pendientes)
    creado_por_usuario_nombre: Optional[str] = None

    # Transporte interprovincial
    transporte_id: Optional[int] = None
    transporte_nombre: Optional[str] = None
    transporte_color: Optional[str] = None
    transporte_direccion: Optional[str] = None
    transporte_cp: Optional[str] = None
    transporte_localidad: Optional[str] = None
    transporte_telefono: Optional[str] = None
    transporte_horario: Optional[str] = None

    model_config = ConfigDict(from_attributes=True)


class EstadisticasEnvioResponse(BaseModel):
    """Estadísticas de distribución de etiquetas."""

    total: int
    por_cordon: dict[str, int]
    sin_cordon: int
    por_logistica: dict[str, int]
    sin_logistica: int
    por_estado_ml: dict[str, int]
    por_estado_erp: dict[str, int]
    costo_total: float = 0.0
    costo_por_logistica: dict[str, float] = {}
    flagged: int = 0


class EstadisticaDiaItem(BaseModel):
    """Estadísticas de un día individual para la vista calendario."""

    fecha: date
    total: int = 0
    flex: int = 0
    manuales: int = 0
    por_cordon: dict[str, int] = {}
    sin_cordon: int = 0
    con_logistica: int = 0
    sin_logistica: int = 0
    enviados: int = 0
    no_entregados: int = 0


class EstadisticasPorDiaResponse(BaseModel):
    """Respuesta del endpoint de estadísticas agrupadas por día."""

    dias: List[EstadisticaDiaItem]


class AsignarLogisticaRequest(BaseModel):
    """Payload para asignar logística a una etiqueta."""

    logistica_id: Optional[int] = Field(
        None,
        description="ID de logística. None para desasignar.",
    )


class CambiarFechaRequest(BaseModel):
    """Payload para cambiar fecha de envío."""

    fecha_envio: date


class CostoOverrideRequest(BaseModel):
    """Payload para establecer o quitar un costo override en una etiqueta."""

    costo: Optional[float] = Field(
        None,
        ge=0,
        description="Costo manual. None para eliminar el override y volver al calculado.",
    )
    operador_id: int = Field(description="Operador autenticado con PIN")


class CrearEnvioManualRequest(BaseModel):
    """Payload para crear un envío manual (sin MercadoLibre)."""

    fecha_envio: date
    receiver_name: str = Field(max_length=500, description="Nombre del destinatario")
    street_name: str = Field(max_length=500, description="Calle")
    street_number: str = Field(max_length=50, description="Número")
    zip_code: str = Field(max_length=50, description="Código postal")
    city_name: str = Field(max_length=500, description="Ciudad / Localidad")
    status: str = Field(
        description="Estado del envío: ready_to_ship, shipped, delivered",
        pattern="^(ready_to_ship|shipped|delivered)$",
    )
    cust_id: Optional[int] = Field(
        None, description="ID de cliente del ERP. Se resuelve automáticamente si se envía soh_id + bra_id."
    )
    bra_id: Optional[int] = Field(None, description="Sucursal (tb_branch.bra_id)")
    soh_id: Optional[int] = Field(None, description="N° pedido ERP (soh_id de SaleOrderHeader)")
    logistica_id: Optional[int] = Field(None, description="Logística asignada")
    transporte_id: Optional[int] = Field(None, description="Transporte interprovincial asignado")
    comment: Optional[str] = Field(None, max_length=1000, description="Observaciones")
    phone: Optional[str] = Field(None, max_length=100, description="Teléfono del destinatario")
    operador_id: int = Field(description="Operador autenticado con PIN")


class CrearEnvioManualResponse(BaseModel):
    """Resultado de la creación de un envío manual."""

    ok: bool = True
    shipping_id: str
    cordon: Optional[str] = None
    mensaje: str


class CrearDesdePedidoRequest(BaseModel):
    """Payload para crear un envío flex desde la tab Pedidos Pendientes.

    Si se envían soh_id + bra_id, se resuelve el cust_id automáticamente.
    Si no se envían, se crea un envío manual puro (sin pedido asociado).
    """

    fecha_envio: date
    soh_id: Optional[int] = Field(None, description="N° pedido ERP (opcional)")
    bra_id: Optional[int] = Field(None, description="Sucursal (opcional)")
    receiver_name: str = Field(max_length=500, description="Nombre del destinatario")
    street_name: str = Field(max_length=500, description="Dirección completa")
    street_number: str = Field(max_length=50, default="S/N", description="Número")
    zip_code: str = Field(max_length=50, description="Código postal")
    city_name: str = Field(max_length=500, description="Ciudad / Localidad")
    comment: Optional[str] = Field(None, max_length=1000, description="Observaciones")
    phone: Optional[str] = Field(None, max_length=100, description="Teléfono del destinatario")
    logistica_id: Optional[int] = Field(None, description="Logística asignada")
    transporte_id: Optional[int] = Field(None, description="Transporte interprovincial asignado")
    cust_id: Optional[int] = Field(None, description="ID cliente ERP (si no hay pedido)")
    status: Optional[str] = Field(
        None,
        description="Estado del envío (default: ready_to_ship)",
        pattern="^(ready_to_ship|shipped|delivered)$",
    )


class AsignarMasivoRequest(BaseModel):
    """Payload para asignación masiva de logística."""

    shipping_ids: List[str] = Field(min_length=1)
    logistica_id: int


class AsignarTransporteMasivoRequest(BaseModel):
    """Payload para asignación masiva de transporte."""

    shipping_ids: List[str] = Field(min_length=1)
    transporte_id: Optional[int] = Field(None, description="ID del transporte (None para desasignar)")


class ShippingIdsRequest(BaseModel):
    """Payload genérico con lista de shipping_ids."""

    shipping_ids: List[str] = Field(min_length=1)


FLAG_ENVIO_VALIDOS = {"mal_pasado", "envio_cancelado", "duplicado", "otro"}


class FlagEnvioRequest(BaseModel):
    """Payload para flaggear un envío."""

    flag_envio: Optional[str] = Field(
        None,
        description="Tipo de flag: mal_pasado, envio_cancelado, duplicado, otro. None para quitar flag.",
    )
    motivo: Optional[str] = Field(
        None,
        max_length=500,
        description="Observación libre (se muestra como tooltip en el badge)",
    )


class FlagEnvioMasivoRequest(BaseModel):
    """Payload para flaggear múltiples envíos."""

    shipping_ids: List[str] = Field(min_length=1)
    flag_envio: Optional[str] = Field(
        None,
        description="Tipo de flag: mal_pasado, envio_cancelado, duplicado, otro. None para quitar flag.",
    )
    motivo: Optional[str] = Field(
        None,
        max_length=500,
        description="Observación libre",
    )


# ── Schemas Pistoleado ───────────────────────────────────────────────


class PistolearRequest(BaseModel):
    """Payload del pistoleado: escaneo de QR de etiqueta en depósito."""

    shipping_id: str = Field(description="shipping_id extraído del QR de la etiqueta")
    caja: str = Field(max_length=50, description="Contenedor activo (CAJA 1, SUELTOS 1, etc.)")
    logistica_id: int = Field(description="Logística que el operador está pistoleando")
    operador_id: int = Field(description="Operador autenticado con PIN")
    bulto: Optional[int] = Field(None, description="N° de bulto escaneado (del QR). None = bulto único.")
    total_bultos: Optional[int] = Field(None, description="Total bultos del envío (del QR)")
    forzar_asignacion: bool = Field(
        False, description="Si True, asigna la logística aunque no tenga pistoleado_asigna (doble escaneo)"
    )


class PistolearResponse(BaseModel):
    """Resultado exitoso de un pistoleado."""

    ok: bool = True
    shipping_id: str
    caja: str
    operador: str
    receiver_name: Optional[str] = None
    ciudad: Optional[str] = None
    cordon: Optional[str] = None
    pistoleado_at: str
    bulto: Optional[int] = None
    total_bultos: Optional[int] = None
    bultos_pistoleados: int = Field(0, description="Cantidad de bultos pistoleados hasta ahora")
    count: int = Field(description="Total pistoleadas en esta sesión (fecha + logística + operador)")
    estado_erp: Optional[str] = Field(None, description="Nombre del estado ERP del pedido (ssos_name)")
    logistica_asignada: bool = Field(
        False,
        description="True si la logística fue asignada por el pistoleado (modo pistoleado_asigna)",
    )

    model_config = ConfigDict(from_attributes=True)


class PistoleadoConflictResponse(BaseModel):
    """Respuesta cuando la etiqueta ya fue pistoleada."""

    detail: str = "Ya pistoleada"
    pistoleado_por: str
    pistoleado_at: str
    pistoleado_caja: str


class PistoleadoStatsResponse(BaseModel):
    """Estadísticas de pistoleado para una fecha/logística."""

    total_etiquetas: int
    pistoleadas: int
    pendientes: int
    porcentaje: float
    en_preparacion: int = Field(0, description="Pistoleadas cuyo pedido ERP está En Preparación")
    por_caja: dict[str, int]
    por_operador: dict[str, int]


# ── Helpers ──────────────────────────────────────────────────────────


def _parse_qr_json(raw: str) -> dict:
    """
    Parsea el JSON del QR de la etiqueta ZPL.
    Retorna dict con shipping_id, sender_id, hash_code.
    Raises ValueError si el JSON no tiene el campo 'id'.
    """
    data = json.loads(raw)
    shipping_id = data.get("id")
    if not shipping_id:
        raise ValueError("JSON del QR no tiene campo 'id'")

    return {
        "shipping_id": str(shipping_id),
        "sender_id": data.get("sender_id"),
        "hash_code": data.get("hash_code"),
    }


def _extraer_qrs_de_texto(text: str) -> List[str]:
    """Extrae todos los JSONs de QR de un texto ZPL."""
    return QR_JSON_REGEX.findall(text)


def _soh_status_subquery(db: Session, shipping_ids_sub=None):
    """
    Subquery deduplicada: estado ERP por shipping_id (envíos ML).

    Cruza orders_shipping (por mlshippingid) → sale_order_header (por mlo_id).
    Un mismo mlo_id puede tener múltiples filas en SaleOrderHeader
    (una por combinación comp_id/bra_id). Se toma el pedido más reciente
    (mayor soh_cd) con ROW_NUMBER OVER (PARTITION BY mlshippingid ORDER BY soh_cd DESC).

    Si shipping_ids_sub se proporciona, restringe a esos IDs para evitar
    escanear la tabla completa (performance).
    """
    base_q = (
        db.query(
            MercadoLibreOrderShipping.mlshippingid.label("shipping_id_str"),
            SaleOrderHeader.ssos_id.label("soh_ssos_id"),
            func.row_number()
            .over(
                partition_by=MercadoLibreOrderShipping.mlshippingid,
                order_by=desc(SaleOrderHeader.soh_cd),
            )
            .label("rn"),
        )
        .join(
            SaleOrderHeader,
            MercadoLibreOrderShipping.mlo_id == SaleOrderHeader.mlo_id,
        )
        .filter(
            MercadoLibreOrderShipping.mlshippingid.isnot(None),
            MercadoLibreOrderShipping.mlo_id.isnot(None),
            SaleOrderHeader.mlo_id.isnot(None),
        )
    )

    if shipping_ids_sub is not None:
        base_q = base_q.filter(MercadoLibreOrderShipping.mlshippingid.in_(shipping_ids_sub))

    ranked = base_q.subquery()

    return (
        db.query(
            ranked.c.shipping_id_str,
            ranked.c.soh_ssos_id,
        )
        .filter(ranked.c.rn == 1)
        .subquery()
    )


def _manual_soh_status_subquery(db: Session, shipping_ids_sub=None):
    """
    Subquery: estado ERP para envíos manuales (por manual_soh_id + manual_bra_id).

    Los envíos manuales creados desde pedidos tienen manual_soh_id y manual_bra_id
    que referencian directamente a SaleOrderHeader.

    Si shipping_ids_sub se proporciona, restringe a esos IDs para evitar
    escanear la tabla completa (performance).
    """
    base_q = (
        db.query(
            EtiquetaEnvio.shipping_id.label("shipping_id_str"),
            SaleOrderHeader.ssos_id.label("manual_ssos_id"),
            func.row_number()
            .over(
                partition_by=EtiquetaEnvio.shipping_id,
                order_by=desc(SaleOrderHeader.soh_cd),
            )
            .label("rn"),
        )
        .join(
            SaleOrderHeader,
            and_(
                EtiquetaEnvio.manual_soh_id == SaleOrderHeader.soh_id,
                EtiquetaEnvio.manual_bra_id == SaleOrderHeader.bra_id,
            ),
        )
        .filter(
            EtiquetaEnvio.es_manual.is_(True),
            EtiquetaEnvio.manual_soh_id.isnot(None),
            EtiquetaEnvio.manual_bra_id.isnot(None),
        )
    )

    if shipping_ids_sub is not None:
        base_q = base_q.filter(EtiquetaEnvio.shipping_id.in_(shipping_ids_sub))

    ranked = base_q.subquery()

    return (
        db.query(
            ranked.c.shipping_id_str,
            ranked.c.manual_ssos_id,
        )
        .filter(ranked.c.rn == 1)
        .subquery()
    )


def _shipping_dedup_subquery(db: Session, shipping_ids_sub=None):
    """
    Subquery deduplicada: una fila por mlshippingid de MercadoLibreOrderShipping.

    Un mismo mlshippingid puede tener múltiples filas en la tabla
    (una por item/order del envío, con distinto mlm_id).
    Se toma la fila más reciente (mayor mlm_id) con
    ROW_NUMBER OVER (PARTITION BY mlshippingid ORDER BY mlm_id DESC).

    Esto evita que los JOINs multipliquen filas en listado, export y estadísticas.

    Si shipping_ids_sub se proporciona, restringe a esos IDs para evitar
    escanear las 88k+ filas completas (performance: ~50 IDs vs tabla completa).
    """
    base_q = db.query(
        MercadoLibreOrderShipping.mlshippingid.label("mlshippingid"),
        MercadoLibreOrderShipping.mlo_id,
        MercadoLibreOrderShipping.mlreceiver_name,
        MercadoLibreOrderShipping.mlstreet_name,
        MercadoLibreOrderShipping.mlstreet_number,
        MercadoLibreOrderShipping.mlzip_code,
        MercadoLibreOrderShipping.mlcity_name,
        MercadoLibreOrderShipping.mlstatus,
        func.row_number()
        .over(
            partition_by=MercadoLibreOrderShipping.mlshippingid,
            order_by=desc(MercadoLibreOrderShipping.mlm_id),
        )
        .label("rn"),
    ).filter(MercadoLibreOrderShipping.mlshippingid.isnot(None))

    if shipping_ids_sub is not None:
        base_q = base_q.filter(MercadoLibreOrderShipping.mlshippingid.in_(shipping_ids_sub))

    ranked = base_q.subquery()

    return (
        db.query(
            ranked.c.mlshippingid,
            ranked.c.mlo_id,
            ranked.c.mlreceiver_name,
            ranked.c.mlstreet_name,
            ranked.c.mlstreet_number,
            ranked.c.mlzip_code,
            ranked.c.mlcity_name,
            ranked.c.mlstatus,
        )
        .filter(ranked.c.rn == 1)
        .subquery()
    )


def _insertar_etiqueta(
    db: Session,
    shipping_id: str,
    sender_id: Optional[int],
    hash_code: Optional[str],
    nombre_archivo: Optional[str],
    fecha_envio: date,
) -> bool:
    """
    Inserta una etiqueta si no existe.
    Retorna True si se insertó (nueva), False si ya existía (duplicada).
    """
    existente = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id == shipping_id).first()
    if existente:
        return False

    etiqueta = EtiquetaEnvio(
        shipping_id=shipping_id,
        sender_id=sender_id,
        hash_code=hash_code,
        nombre_archivo=nombre_archivo,
        fecha_envio=fecha_envio,
    )
    db.add(etiqueta)
    return True


# ── Endpoints ────────────────────────────────────────────────────────


@router.post(
    "/etiquetas-envio/upload",
    response_model=UploadResultResponse,
    summary="Subir archivo ZPL (.zip o .txt) con etiquetas",
)
def upload_etiquetas(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(..., description="Archivo .zip o .txt con etiquetas ZPL"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> UploadResultResponse:
    """
    Recibe un .zip (que contiene .txt) o directamente un .txt con etiquetas ZPL.
    Extrae los JSONs de los QR codes, parsea shipping_id/sender_id/hash_code
    e inserta en etiquetas_envio con ON CONFLICT(shipping_id) DO NOTHING.
    """
    _check_permiso(db, current_user, "envios_flex.subir_etiquetas")

    if not file.filename:
        raise HTTPException(400, "Archivo sin nombre")

    filename = file.filename.lower()
    if not filename.endswith((".zip", ".txt")):
        raise HTTPException(400, "Solo se aceptan archivos .zip o .txt")

    contents = file.file.read()
    nombre_archivo = file.filename
    text_content = ""

    if filename.endswith(".zip"):
        try:
            with zipfile.ZipFile(BytesIO(contents)) as zf:
                # Buscar el primer .txt dentro del zip
                txt_files = [
                    name for name in zf.namelist() if name.lower().endswith(".txt") and not name.startswith("__MACOSX")
                ]
                if not txt_files:
                    raise HTTPException(400, "El .zip no contiene archivos .txt")

                text_content = zf.read(txt_files[0]).decode("utf-8", errors="replace")
        except zipfile.BadZipFile:
            raise HTTPException(400, "Archivo .zip inválido o corrupto")
    else:
        text_content = contents.decode("utf-8", errors="replace")

    # Extraer QR JSONs
    qr_jsons = _extraer_qrs_de_texto(text_content)

    if not qr_jsons:
        raise HTTPException(400, "No se encontraron etiquetas QR en el archivo")

    total = len(qr_jsons)
    nuevas = 0
    duplicadas = 0
    errores = 0
    detalle_errores: List[str] = []
    nuevos_shipping_ids: List[str] = []
    hoy = date.today()

    for raw_json in qr_jsons:
        try:
            parsed = _parse_qr_json(raw_json)
            es_nueva = _insertar_etiqueta(
                db=db,
                shipping_id=parsed["shipping_id"],
                sender_id=parsed["sender_id"],
                hash_code=parsed["hash_code"],
                nombre_archivo=nombre_archivo,
                fecha_envio=hoy,
            )
            if es_nueva:
                nuevas += 1
                nuevos_shipping_ids.append(parsed["shipping_id"])
            else:
                duplicadas += 1
        except (json.JSONDecodeError, ValueError) as e:
            errores += 1
            detalle_errores.append(f"Error parseando QR: {str(e)[:100]}")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Error guardando en base de datos: {str(e)}")

    # Enriquecer etiquetas nuevas en background (coords, dirección, comentario)
    if nuevos_shipping_ids:
        background_tasks.add_task(enriquecer_etiquetas_sync, nuevos_shipping_ids)

    return UploadResultResponse(
        total=total,
        nuevas=nuevas,
        duplicadas=duplicadas,
        errores=errores,
        detalle_errores=detalle_errores[:20],
    )


@router.post(
    "/etiquetas-envio/manual",
    response_model=ManualScanResponse,
    summary="Registrar etiqueta individual (escaneo con pistola)",
)
def registrar_manual(
    payload: ManualScanRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> ManualScanResponse:
    """
    Recibe el JSON del QR escaneado con pistola de código de barras.
    Inserta la etiqueta si no existe.
    """
    _check_permiso(db, current_user, "envios_flex.subir_etiquetas")

    try:
        parsed = _parse_qr_json(payload.json_data)
    except (json.JSONDecodeError, ValueError) as e:
        raise HTTPException(400, f"JSON inválido: {str(e)}")

    es_nueva = _insertar_etiqueta(
        db=db,
        shipping_id=parsed["shipping_id"],
        sender_id=parsed["sender_id"],
        hash_code=parsed["hash_code"],
        nombre_archivo="escaneo_manual",
        fecha_envio=date.today(),
    )

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Error guardando: {str(e)}")

    shipping_id = parsed["shipping_id"]

    if es_nueva:
        # Enriquecer en background (coords, dirección, comentario)
        background_tasks.add_task(enriquecer_etiquetas_sync, [shipping_id])

        return ManualScanResponse(
            duplicada=False,
            shipping_id=shipping_id,
            mensaje=f"Cargado: {shipping_id}",
        )
    else:
        return ManualScanResponse(
            duplicada=True,
            shipping_id=shipping_id,
            mensaje=f"Ya existía: {shipping_id}",
        )


@router.get(
    "/etiquetas-envio",
    response_model=List[EtiquetaEnvioResponse],
    summary="Listar etiquetas con datos de envío",
)
def listar_etiquetas(
    fecha_envio: Optional[date] = Query(None, description="Filtrar por fecha de envío exacta"),
    fecha_desde: Optional[date] = Query(None, description="Filtrar desde fecha (inclusive)"),
    fecha_hasta: Optional[date] = Query(None, description="Filtrar hasta fecha (inclusive)"),
    cordon: Optional[str] = Query(None, description="Filtrar por cordón"),
    logistica_id: Optional[int] = Query(None, description="Filtrar por logística"),
    sin_logistica: bool = Query(False, description="Solo etiquetas sin logística asignada"),
    mlstatus: Optional[str] = Query(None, description="Filtrar por estado ML"),
    ssos_id: Optional[int] = Query(None, description="Filtrar por estado ERP"),
    solo_outlet: bool = Query(False, description="Solo etiquetas de productos outlet"),
    solo_turbo: bool = Query(False, description="Solo etiquetas de envíos turbo"),
    pistoleado: Optional[str] = Query(None, pattern="^(si|no)$", description="Filtrar por pistoleado: si/no"),
    sin_cordon: bool = Query(False, description="Solo etiquetas sin cordón asignado"),
    search: Optional[str] = Query(None, description="Buscar por shipping_id, destinatario o dirección"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> List[EtiquetaEnvioResponse]:
    """
    Lista etiquetas de envío con JOINs a:
    - tb_mercadolibre_orders_shipping (datos del envío)
    - cp_cordones (cordón del CP)
    - tb_sale_order_header (ssos_id del pedido ERP)
    - tb_sale_order_status (nombre y color del estado)
    - logisticas (nombre y color de la logística)
    """
    _check_permiso(db, current_user, "envios_flex.ver")

    # ── Pre-filtrar shipping_ids por fecha ───────────────────────
    # Obtener solo los IDs que coinciden con el rango de fechas ANTES de
    # armar las subqueries pesadas. Esto reduce el scan de 88k+ filas a
    # ~50-200 del día, haciendo que los JOINs sean instantáneos.
    ids_fecha_q = db.query(EtiquetaEnvio.shipping_id)
    if fecha_envio:
        ids_fecha_q = ids_fecha_q.filter(EtiquetaEnvio.fecha_envio == fecha_envio)
    elif fecha_desde or fecha_hasta:
        if fecha_desde:
            ids_fecha_q = ids_fecha_q.filter(EtiquetaEnvio.fecha_envio >= fecha_desde)
        if fecha_hasta:
            ids_fecha_q = ids_fecha_q.filter(EtiquetaEnvio.fecha_envio <= fecha_hasta)
    ids_fecha_sub = ids_fecha_q.subquery()

    soh_sub = _soh_status_subquery(db, shipping_ids_sub=ids_fecha_sub)
    manual_soh_sub = _manual_soh_status_subquery(db, shipping_ids_sub=ids_fecha_sub)
    ManualSaleOrderStatus = aliased(SaleOrderStatus)

    # Subquery: costo vigente por (logistica_id, cordon) donde vigente_desde <= hoy.
    # Usamos max(id) como criterio único — el registro más reciente (mayor id) es
    # siempre la última intención del usuario, incluso si hay duplicados por fecha.
    # cp_cordones usa tildes (Cordón) pero logistica_costo_cordon no (Cordon),
    # así que normalizamos con REPLACE en la condición del JOIN.
    hoy = date.today()
    max_costo_sub = (
        db.query(
            LogisticaCostoCordon.logistica_id.label("costo_logistica_id"),
            LogisticaCostoCordon.cordon.label("costo_cordon"),
            func.max(LogisticaCostoCordon.id).label("max_id"),
        )
        .filter(LogisticaCostoCordon.vigente_desde <= hoy)
        .group_by(
            LogisticaCostoCordon.logistica_id,
            LogisticaCostoCordon.cordon,
        )
        .subquery()
    )

    costo_sub = (
        db.query(
            LogisticaCostoCordon.logistica_id.label("costo_log_id"),
            LogisticaCostoCordon.cordon.label("costo_cordon_val"),
            LogisticaCostoCordon.costo.label("costo_valor"),
            LogisticaCostoCordon.costo_turbo.label("costo_turbo_valor"),
        )
        .join(
            max_costo_sub,
            LogisticaCostoCordon.id == max_costo_sub.c.max_id,
        )
        .subquery()
    )

    # Expresión para normalizar cordón: "Cordón 1" → "Cordon 1" (quitar tilde)
    cordon_normalizado = func.replace(CodigoPostalCordon.cordon, "ó", "o")

    # Subquery deduplicada: una fila por mlshippingid (evita duplicados por items)
    shipping_sub = _shipping_dedup_subquery(db, shipping_ids_sub=ids_fecha_sub)

    # Expresiones COALESCE: para envíos manuales priorizar manual_*, sino ML shipping
    eff_receiver = func.coalesce(EtiquetaEnvio.manual_receiver_name, shipping_sub.c.mlreceiver_name)
    eff_street = func.coalesce(EtiquetaEnvio.manual_street_name, shipping_sub.c.mlstreet_name)
    eff_street_num = func.coalesce(EtiquetaEnvio.manual_street_number, shipping_sub.c.mlstreet_number)
    eff_zip = func.coalesce(EtiquetaEnvio.manual_zip_code, shipping_sub.c.mlzip_code)
    eff_city = func.coalesce(EtiquetaEnvio.manual_city_name, shipping_sub.c.mlcity_name)
    eff_status = func.coalesce(EtiquetaEnvio.manual_status, shipping_sub.c.mlstatus)

    # CP efectivo para resolver cordón: si hay transporte con CP, usar ese;
    # sino, usar el CP del cliente (eff_zip).  La etiqueta sigue mostrando
    # el CP del cliente — esto solo afecta la resolución de cordón/costo.
    eff_zip_for_cordon = func.coalesce(Transporte.cp, eff_zip)

    # Lluvia offset config
    lluvia_tipo, lluvia_valor = _get_lluvia_config(db)

    # Alias para segundo join a Usuario (flag_envio_usuario)
    FlagUsuario = aliased(Usuario)

    query = (
        db.query(
            EtiquetaEnvio.shipping_id,
            EtiquetaEnvio.sender_id,
            EtiquetaEnvio.nombre_archivo,
            EtiquetaEnvio.fecha_envio,
            EtiquetaEnvio.logistica_id,
            EtiquetaEnvio.transporte_id,
            EtiquetaEnvio.latitud,
            EtiquetaEnvio.longitud,
            EtiquetaEnvio.direccion_completa,
            EtiquetaEnvio.direccion_comentario,
            EtiquetaEnvio.pistoleado_at,
            EtiquetaEnvio.pistoleado_caja,
            EtiquetaEnvio.es_manual,
            EtiquetaEnvio.manual_bra_id,
            EtiquetaEnvio.manual_soh_id,
            EtiquetaEnvio.manual_cust_id,
            EtiquetaEnvio.manual_comment,
            EtiquetaEnvio.manual_phone,
            EtiquetaEnvio.es_outlet,
            EtiquetaEnvio.es_turbo,
            EtiquetaEnvio.es_lluvia,
            EtiquetaEnvio.flag_envio,
            EtiquetaEnvio.flag_envio_motivo,
            EtiquetaEnvio.flag_envio_at,
            Operador.nombre.label("pistoleado_operador_nombre"),
            Logistica.nombre.label("logistica_nombre"),
            Logistica.color.label("logistica_color"),
            Transporte.nombre.label("transporte_nombre"),
            Transporte.color.label("transporte_color"),
            Transporte.direccion.label("transporte_direccion"),
            Transporte.cp.label("transporte_cp"),
            Transporte.localidad.label("transporte_localidad"),
            Transporte.telefono.label("transporte_telefono"),
            Transporte.horario.label("transporte_horario"),
            eff_receiver.label("mlreceiver_name"),
            eff_street.label("mlstreet_name"),
            eff_street_num.label("mlstreet_number"),
            eff_zip.label("mlzip_code"),
            eff_city.label("mlcity_name"),
            eff_status.label("mlstatus"),
            CodigoPostalCordon.cordon,
            func.coalesce(soh_sub.c.soh_ssos_id, manual_soh_sub.c.manual_ssos_id).label("ssos_id"),
            func.coalesce(SaleOrderStatus.ssos_name, ManualSaleOrderStatus.ssos_name).label("ssos_name"),
            func.coalesce(SaleOrderStatus.ssos_color, ManualSaleOrderStatus.ssos_color).label("ssos_color"),
            func.coalesce(
                cast(EtiquetaEnvio.costo_override, Numeric(12, 2)),
                _build_costo_case(
                    costo_sub.c.costo_turbo_valor,
                    costo_sub.c.costo_valor,
                    lluvia_tipo,
                    lluvia_valor,
                ),
            ).label("costo_envio"),
            EtiquetaEnvio.costo_override,
            MercadoLibreUserData.nickname.label("mluser_nickname"),
            MercadoLibreOrderHeader.mlorder_id.label("ml_order_id"),
            Usuario.nombre.label("creado_por_usuario_nombre"),
            FlagUsuario.nombre.label("flag_envio_usuario_nombre"),
        )
        .outerjoin(
            Logistica,
            EtiquetaEnvio.logistica_id == Logistica.id,
        )
        .outerjoin(
            Transporte,
            EtiquetaEnvio.transporte_id == Transporte.id,
        )
        .outerjoin(
            shipping_sub,
            EtiquetaEnvio.shipping_id == shipping_sub.c.mlshippingid,
        )
        .outerjoin(
            MercadoLibreOrderHeader,
            shipping_sub.c.mlo_id == MercadoLibreOrderHeader.mlo_id,
        )
        .outerjoin(
            MercadoLibreUserData,
            MercadoLibreOrderHeader.mluser_id == MercadoLibreUserData.mluser_id,
        )
        .outerjoin(
            CodigoPostalCordon,
            eff_zip_for_cordon == CodigoPostalCordon.codigo_postal,
        )
        .outerjoin(
            soh_sub,
            soh_sub.c.shipping_id_str == EtiquetaEnvio.shipping_id,
        )
        .outerjoin(
            SaleOrderStatus,
            soh_sub.c.soh_ssos_id == SaleOrderStatus.ssos_id,
        )
        .outerjoin(
            manual_soh_sub,
            manual_soh_sub.c.shipping_id_str == EtiquetaEnvio.shipping_id,
        )
        .outerjoin(
            ManualSaleOrderStatus,
            manual_soh_sub.c.manual_ssos_id == ManualSaleOrderStatus.ssos_id,
        )
        .outerjoin(
            Operador,
            EtiquetaEnvio.pistoleado_operador_id == Operador.id,
        )
        .outerjoin(
            Usuario,
            EtiquetaEnvio.creado_por_usuario_id == Usuario.id,
        )
        .outerjoin(
            FlagUsuario,
            EtiquetaEnvio.flag_envio_usuario_id == FlagUsuario.id,
        )
        .outerjoin(
            costo_sub,
            and_(
                costo_sub.c.costo_log_id == EtiquetaEnvio.logistica_id,
                costo_sub.c.costo_cordon_val == cordon_normalizado,
            ),
        )
    )

    # ── Filtros ──────────────────────────────────────────────────

    # fecha_envio = exacta (backward compatible), fecha_desde/fecha_hasta = rango
    if fecha_envio:
        query = query.filter(EtiquetaEnvio.fecha_envio == fecha_envio)
    else:
        if fecha_desde:
            query = query.filter(EtiquetaEnvio.fecha_envio >= fecha_desde)
        if fecha_hasta:
            query = query.filter(EtiquetaEnvio.fecha_envio <= fecha_hasta)

    if cordon:
        query = query.filter(CodigoPostalCordon.cordon == cordon)

    if logistica_id is not None:
        query = query.filter(EtiquetaEnvio.logistica_id == logistica_id)

    if sin_logistica:
        query = query.filter(EtiquetaEnvio.logistica_id.is_(None))

    if solo_outlet:
        query = query.filter(EtiquetaEnvio.es_outlet.is_(True))

    if solo_turbo:
        query = query.filter(EtiquetaEnvio.es_turbo.is_(True))

    if mlstatus:
        query = query.filter(eff_status == mlstatus)

    if ssos_id is not None:
        query = query.filter(or_(soh_sub.c.soh_ssos_id == ssos_id, manual_soh_sub.c.manual_ssos_id == ssos_id))

    if pistoleado == "si":
        query = query.filter(EtiquetaEnvio.pistoleado_at.isnot(None))
    elif pistoleado == "no":
        query = query.filter(EtiquetaEnvio.pistoleado_at.is_(None))

    if sin_cordon:
        query = query.filter(CodigoPostalCordon.cordon.is_(None))

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            (EtiquetaEnvio.shipping_id.ilike(search_term))
            | (eff_receiver.ilike(search_term))
            | (eff_street.ilike(search_term))
            | (eff_city.ilike(search_term))
            | (MercadoLibreUserData.nickname.ilike(search_term))
        )

    # Excluir etiquetas que existen en colecta (son de otro flujo)
    colecta_ids = db.query(EtiquetaColecta.shipping_id).subquery()
    query = query.filter(~EtiquetaEnvio.shipping_id.in_(db.query(colecta_ids.c.shipping_id)))

    # Ordenar por shipping_id desc (más recientes primero)
    query = query.order_by(EtiquetaEnvio.shipping_id.desc())

    rows = query.all()

    return [
        EtiquetaEnvioResponse(
            shipping_id=row.shipping_id,
            sender_id=row.sender_id,
            nombre_archivo=row.nombre_archivo,
            fecha_envio=row.fecha_envio,
            logistica_id=row.logistica_id,
            logistica_nombre=row.logistica_nombre,
            logistica_color=row.logistica_color,
            mluser_nickname=row.mluser_nickname,
            mlreceiver_name=row.mlreceiver_name,
            mlstreet_name=row.mlstreet_name,
            mlstreet_number=row.mlstreet_number,
            mlzip_code=row.mlzip_code,
            mlcity_name=row.mlcity_name,
            mlstatus=row.mlstatus,
            cordon=row.cordon,
            latitud=row.latitud,
            longitud=row.longitud,
            direccion_completa=row.direccion_completa,
            direccion_comentario=row.direccion_comentario,
            ssos_id=row.ssos_id,
            ssos_name=row.ssos_name,
            ssos_color=row.ssos_color,
            pistoleado_at=str(row.pistoleado_at) if row.pistoleado_at else None,
            pistoleado_caja=row.pistoleado_caja,
            pistoleado_operador_nombre=row.pistoleado_operador_nombre,
            costo_envio=float(row.costo_envio) if row.costo_envio is not None else None,
            costo_override=float(row.costo_override) if row.costo_override is not None else None,
            es_manual=row.es_manual,
            manual_bra_id=row.manual_bra_id,
            manual_soh_id=row.manual_soh_id,
            manual_cust_id=row.manual_cust_id,
            manual_comment=row.manual_comment,
            manual_phone=row.manual_phone,
            es_outlet=row.es_outlet,
            es_turbo=row.es_turbo,
            es_lluvia=row.es_lluvia,
            flag_envio=row.flag_envio,
            flag_envio_motivo=row.flag_envio_motivo,
            flag_envio_at=str(row.flag_envio_at) if row.flag_envio_at else None,
            flag_envio_usuario_nombre=row.flag_envio_usuario_nombre,
            creado_por_usuario_nombre=row.creado_por_usuario_nombre,
            transporte_id=row.transporte_id,
            transporte_nombre=row.transporte_nombre,
            transporte_color=row.transporte_color,
            transporte_direccion=row.transporte_direccion,
            transporte_cp=row.transporte_cp,
            transporte_localidad=row.transporte_localidad,
            transporte_telefono=row.transporte_telefono,
            transporte_horario=row.transporte_horario,
        )
        for row in rows
    ]


@router.get(
    "/etiquetas-envio/estadisticas",
    response_model=EstadisticasEnvioResponse,
    summary="Estadísticas de distribución de etiquetas",
)
def estadisticas_etiquetas(
    fecha_envio: Optional[date] = Query(None, description="Fecha de envío exacta (por defecto hoy)"),
    fecha_desde: Optional[date] = Query(None, description="Filtrar desde fecha (inclusive)"),
    fecha_hasta: Optional[date] = Query(None, description="Filtrar hasta fecha (inclusive)"),
    cordon: Optional[str] = Query(None, description="Filtrar por cordón"),
    logistica_id: Optional[int] = Query(None, description="Filtrar por logística"),
    sin_logistica: bool = Query(False, description="Solo etiquetas sin logística asignada"),
    mlstatus: Optional[str] = Query(None, description="Filtrar por estado ML"),
    ssos_id: Optional[int] = Query(None, description="Filtrar por estado ERP"),
    solo_outlet: bool = Query(False, description="Solo etiquetas de productos outlet"),
    solo_turbo: bool = Query(False, description="Solo etiquetas de envíos turbo"),
    pistoleado: Optional[str] = Query(None, pattern="^(si|no)$", description="Filtrar por pistoleado: si/no"),
    sin_cordon: bool = Query(False, description="Solo etiquetas sin cordón asignado"),
    search: Optional[str] = Query(None, description="Buscar por shipping_id, destinatario o dirección"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> EstadisticasEnvioResponse:
    """
    Distribución de etiquetas por cordón, logística y estado.

    Acepta los mismos filtros que el listado para que las stats sean
    el resumen exacto de lo que el usuario ve en la tabla.
    """
    _check_permiso(db, current_user, "envios_flex.ver")

    # Determinar filtro de fechas: exacta (backward compatible) o rango
    if fecha_envio:
        fecha_filter = EtiquetaEnvio.fecha_envio == fecha_envio
        fecha_costo = fecha_envio
    elif fecha_desde or fecha_hasta:
        conditions = []
        if fecha_desde:
            conditions.append(EtiquetaEnvio.fecha_envio >= fecha_desde)
        if fecha_hasta:
            conditions.append(EtiquetaEnvio.fecha_envio <= fecha_hasta)
        fecha_filter = and_(*conditions) if len(conditions) > 1 else conditions[0]
        fecha_costo = fecha_hasta or fecha_desde or date.today()
    else:
        fecha_filter = EtiquetaEnvio.fecha_envio == date.today()
        fecha_costo = date.today()

    # ── Pre-filtrar shipping_ids por fecha ────────────────────────
    # Obtener solo los IDs del rango de fechas ANTES de armar subqueries
    # pesadas. Reduce scan de 88k+ filas a ~50-200 (performance).
    ids_fecha_stats = db.query(EtiquetaEnvio.shipping_id).filter(fecha_filter).subquery()

    # ── Subquery de IDs filtrados ────────────────────────────────
    # Construye la misma lógica de filtros del listado como subquery
    # de shipping_ids. Todas las queries de stats la usan para limitar
    # al mismo set que ve el usuario en la tabla.
    shipping_stats = _shipping_dedup_subquery(db, shipping_ids_sub=ids_fecha_stats)
    soh_sub = _soh_status_subquery(db, shipping_ids_sub=ids_fecha_stats)
    manual_soh_sub = _manual_soh_status_subquery(db, shipping_ids_sub=ids_fecha_stats)

    stats_eff_zip = func.coalesce(
        EtiquetaEnvio.manual_zip_code,
        shipping_stats.c.mlzip_code,
    )
    stats_eff_status = func.coalesce(
        EtiquetaEnvio.manual_status,
        shipping_stats.c.mlstatus,
    )
    stats_eff_receiver = func.coalesce(
        EtiquetaEnvio.manual_receiver_name,
        shipping_stats.c.mlreceiver_name,
    )
    stats_eff_street = func.coalesce(
        EtiquetaEnvio.manual_street_name,
        shipping_stats.c.mlstreet_name,
    )
    stats_eff_city = func.coalesce(
        EtiquetaEnvio.manual_city_name,
        shipping_stats.c.mlcity_name,
    )

    # CP efectivo para resolver cordón en stats: si hay transporte con CP,
    # usar ese; sino, usar el CP del cliente.  Mismo patrón que listar_etiquetas.
    stats_eff_zip_for_cordon = func.coalesce(Transporte.cp, stats_eff_zip)

    filtered_ids_q = (
        db.query(EtiquetaEnvio.shipping_id)
        .outerjoin(
            shipping_stats,
            EtiquetaEnvio.shipping_id == shipping_stats.c.mlshippingid,
        )
        .outerjoin(
            Transporte,
            EtiquetaEnvio.transporte_id == Transporte.id,
        )
        .outerjoin(
            CodigoPostalCordon,
            stats_eff_zip_for_cordon == CodigoPostalCordon.codigo_postal,
        )
        .outerjoin(
            soh_sub,
            soh_sub.c.shipping_id_str == EtiquetaEnvio.shipping_id,
        )
        .outerjoin(
            manual_soh_sub,
            manual_soh_sub.c.shipping_id_str == EtiquetaEnvio.shipping_id,
        )
        .filter(fecha_filter)
    )

    # Aplicar los mismos filtros que el listado
    if cordon:
        filtered_ids_q = filtered_ids_q.filter(CodigoPostalCordon.cordon == cordon)
    if logistica_id:
        filtered_ids_q = filtered_ids_q.filter(EtiquetaEnvio.logistica_id == logistica_id)
    if sin_logistica:
        filtered_ids_q = filtered_ids_q.filter(EtiquetaEnvio.logistica_id.is_(None))
    if solo_outlet:
        filtered_ids_q = filtered_ids_q.filter(EtiquetaEnvio.es_outlet.is_(True))
    if solo_turbo:
        filtered_ids_q = filtered_ids_q.filter(EtiquetaEnvio.es_turbo.is_(True))
    if mlstatus:
        filtered_ids_q = filtered_ids_q.filter(stats_eff_status == mlstatus)
    if ssos_id is not None:
        filtered_ids_q = filtered_ids_q.filter(
            or_(soh_sub.c.soh_ssos_id == ssos_id, manual_soh_sub.c.manual_ssos_id == ssos_id)
        )
    if pistoleado == "si":
        filtered_ids_q = filtered_ids_q.filter(EtiquetaEnvio.pistoleado_at.isnot(None))
    elif pistoleado == "no":
        filtered_ids_q = filtered_ids_q.filter(EtiquetaEnvio.pistoleado_at.is_(None))
    if sin_cordon:
        filtered_ids_q = filtered_ids_q.filter(CodigoPostalCordon.cordon.is_(None))
    if search:
        search_term = f"%{search}%"
        filtered_ids_q = filtered_ids_q.filter(
            (EtiquetaEnvio.shipping_id.ilike(search_term))
            | (stats_eff_receiver.ilike(search_term))
            | (stats_eff_street.ilike(search_term))
            | (stats_eff_city.ilike(search_term))
        )

    # Excluir etiquetas que existen en colecta (son de otro flujo)
    colecta_ids_stats = db.query(EtiquetaColecta.shipping_id).subquery()
    filtered_ids_q = filtered_ids_q.filter(~EtiquetaEnvio.shipping_id.in_(db.query(colecta_ids_stats.c.shipping_id)))

    filtered_ids_sub = filtered_ids_q.subquery()

    # Filtro reutilizable: restringe a los shipping_ids filtrados
    ids_filter = EtiquetaEnvio.shipping_id.in_(db.query(filtered_ids_sub.c.shipping_id))

    # Base: total de etiquetas que coinciden con los filtros
    total = db.query(EtiquetaEnvio).filter(ids_filter).count()

    # Flaggeadas (mal pasado, cancelado, etc.)
    flagged = db.query(EtiquetaEnvio).filter(ids_filter, EtiquetaEnvio.flag_envio.isnot(None)).count()

    # Por cordón
    cordon_rows = (
        db.query(
            CodigoPostalCordon.cordon,
            func.count().label("cantidad"),
        )
        .select_from(EtiquetaEnvio)
        .outerjoin(
            shipping_stats,
            EtiquetaEnvio.shipping_id == shipping_stats.c.mlshippingid,
        )
        .outerjoin(
            Transporte,
            EtiquetaEnvio.transporte_id == Transporte.id,
        )
        .join(
            CodigoPostalCordon,
            stats_eff_zip_for_cordon == CodigoPostalCordon.codigo_postal,
        )
        .filter(
            ids_filter,
            CodigoPostalCordon.cordon.isnot(None),
        )
        .group_by(CodigoPostalCordon.cordon)
        .all()
    )
    por_cordon = {row.cordon: row.cantidad for row in cordon_rows}
    con_cordon = sum(por_cordon.values())

    # Por logística
    logistica_rows = (
        db.query(
            Logistica.nombre,
            func.count().label("cantidad"),
        )
        .join(EtiquetaEnvio, EtiquetaEnvio.logistica_id == Logistica.id)
        .filter(ids_filter)
        .group_by(Logistica.nombre)
        .all()
    )
    por_logistica = {row.nombre: row.cantidad for row in logistica_rows}
    con_logistica = sum(por_logistica.values())

    # Por estado ML
    ml_status_rows = (
        db.query(
            stats_eff_status.label("eff_mlstatus"),
            func.count().label("cantidad"),
        )
        .select_from(EtiquetaEnvio)
        .outerjoin(
            shipping_stats,
            EtiquetaEnvio.shipping_id == shipping_stats.c.mlshippingid,
        )
        .filter(
            ids_filter,
            stats_eff_status.isnot(None),
        )
        .group_by(stats_eff_status)
        .all()
    )
    por_estado_ml = {row.eff_mlstatus: row.cantidad for row in ml_status_rows}

    # Por estado ERP (ML + manuales)
    erp_status_rows = (
        db.query(
            SaleOrderStatus.ssos_name,
            func.count().label("cantidad"),
        )
        .join(soh_sub, soh_sub.c.soh_ssos_id == SaleOrderStatus.ssos_id)
        .join(
            EtiquetaEnvio,
            EtiquetaEnvio.shipping_id == soh_sub.c.shipping_id_str,
        )
        .filter(ids_filter)
        .group_by(SaleOrderStatus.ssos_name)
        .all()
    )
    por_estado_erp = {row.ssos_name: row.cantidad for row in erp_status_rows}

    # Agregar manuales con pedido ERP
    manual_erp_rows = (
        db.query(
            SaleOrderStatus.ssos_name,
            func.count().label("cantidad"),
        )
        .join(manual_soh_sub, manual_soh_sub.c.manual_ssos_id == SaleOrderStatus.ssos_id)
        .join(
            EtiquetaEnvio,
            EtiquetaEnvio.shipping_id == manual_soh_sub.c.shipping_id_str,
        )
        .filter(ids_filter)
        .group_by(SaleOrderStatus.ssos_name)
        .all()
    )
    for row in manual_erp_rows:
        por_estado_erp[row.ssos_name] = por_estado_erp.get(row.ssos_name, 0) + row.cantidad

    # ── Costos de envío ─────────────────────────────────────────
    # max(id) como criterio único — evita duplicados cuando hay múltiples registros
    # con la misma (logistica_id, cordon, vigente_desde).
    max_costo_stats = (
        db.query(
            LogisticaCostoCordon.logistica_id.label("costo_logistica_id"),
            LogisticaCostoCordon.cordon.label("costo_cordon"),
            func.max(LogisticaCostoCordon.id).label("max_id"),
        )
        .filter(LogisticaCostoCordon.vigente_desde <= fecha_costo)
        .group_by(
            LogisticaCostoCordon.logistica_id,
            LogisticaCostoCordon.cordon,
        )
        .subquery()
    )

    costo_stats = (
        db.query(
            LogisticaCostoCordon.logistica_id.label("costo_log_id"),
            LogisticaCostoCordon.cordon.label("costo_cordon_val"),
            LogisticaCostoCordon.costo.label("costo_valor"),
            LogisticaCostoCordon.costo_turbo.label("costo_turbo_valor"),
        )
        .join(
            max_costo_stats,
            LogisticaCostoCordon.id == max_costo_stats.c.max_id,
        )
        .subquery()
    )

    cordon_norm = func.replace(CodigoPostalCordon.cordon, "ó", "o")

    # Lluvia offset config
    lluvia_tipo_s, lluvia_valor_s = _get_lluvia_config(db)

    costo_efectivo = func.coalesce(
        cast(EtiquetaEnvio.costo_override, Numeric(12, 2)),
        _build_costo_case(
            costo_stats.c.costo_turbo_valor,
            costo_stats.c.costo_valor,
            lluvia_tipo_s,
            lluvia_valor_s,
        ),
    )

    costo_rows = (
        db.query(
            Logistica.nombre.label("log_nombre"),
            func.coalesce(func.sum(costo_efectivo), 0).label("costo_sum"),
        )
        .select_from(EtiquetaEnvio)
        .join(Logistica, EtiquetaEnvio.logistica_id == Logistica.id)
        .outerjoin(
            shipping_stats,
            EtiquetaEnvio.shipping_id == shipping_stats.c.mlshippingid,
        )
        .outerjoin(
            Transporte,
            EtiquetaEnvio.transporte_id == Transporte.id,
        )
        .join(
            CodigoPostalCordon,
            stats_eff_zip_for_cordon == CodigoPostalCordon.codigo_postal,
        )
        .outerjoin(
            costo_stats,
            and_(
                costo_stats.c.costo_log_id == EtiquetaEnvio.logistica_id,
                costo_stats.c.costo_cordon_val == cordon_norm,
            ),
        )
        .filter(
            ids_filter,
            CodigoPostalCordon.cordon.isnot(None),
        )
        .group_by(Logistica.nombre)
        .all()
    )

    costo_por_logistica = {row.log_nombre: float(row.costo_sum) for row in costo_rows}
    costo_total = sum(costo_por_logistica.values())

    # Sumar también etiquetas con costo_override que NO tienen logística
    costo_sin_logistica = (
        db.query(
            func.coalesce(func.sum(cast(EtiquetaEnvio.costo_override, Numeric(12, 2))), 0),
        )
        .filter(
            ids_filter,
            EtiquetaEnvio.logistica_id.is_(None),
            EtiquetaEnvio.costo_override.isnot(None),
        )
        .scalar()
    )
    costo_total += float(costo_sin_logistica or 0)

    return EstadisticasEnvioResponse(
        total=total,
        por_cordon=por_cordon,
        sin_cordon=max(0, total - con_cordon),
        por_logistica=por_logistica,
        sin_logistica=max(0, total - con_logistica),
        por_estado_ml=por_estado_ml,
        por_estado_erp=por_estado_erp,
        costo_total=costo_total,
        costo_por_logistica=costo_por_logistica,
        flagged=flagged,
    )


@router.get(
    "/etiquetas-envio/estadisticas-por-dia",
    response_model=EstadisticasPorDiaResponse,
    summary="Estadísticas agrupadas por día para vista calendario",
)
def estadisticas_por_dia(
    fecha_desde: date = Query(..., description="Fecha inicio (inclusive)"),
    fecha_hasta: date = Query(..., description="Fecha fin (inclusive)"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> EstadisticasPorDiaResponse:
    """
    Devuelve estadísticas agrupadas por fecha_envio para la vista calendario.

    Para cada día en el rango devuelve:
    - total, flex (no manual), manuales (es_manual=True)
    - distribución por cordón (CABA, Cordón 1, etc.)
    - con/sin logística asignada

    Solo incluye días que tienen al menos 1 etiqueta.
    """
    _check_permiso(db, current_user, "envios_flex.ver")

    # Pre-filtrar shipping_ids por rango de fechas (performance)
    ids_fecha_cal = (
        db.query(EtiquetaEnvio.shipping_id)
        .filter(
            EtiquetaEnvio.fecha_envio >= fecha_desde,
            EtiquetaEnvio.fecha_envio <= fecha_hasta,
        )
        .subquery()
    )

    shipping_stats = _shipping_dedup_subquery(db, shipping_ids_sub=ids_fecha_cal)

    stats_eff_zip = func.coalesce(
        EtiquetaEnvio.manual_zip_code,
        shipping_stats.c.mlzip_code,
    )
    stats_eff_zip_for_cordon = func.coalesce(Transporte.cp, stats_eff_zip)

    # Estado ML efectivo (manual_status overrides mlstatus de ML)
    stats_eff_status = func.coalesce(
        EtiquetaEnvio.manual_status,
        shipping_stats.c.mlstatus,
    )

    # Query base: una fila por (fecha_envio, es_manual, cordon, tiene_logistica, mlstatus)
    base_q = (
        db.query(
            EtiquetaEnvio.fecha_envio,
            EtiquetaEnvio.es_manual,
            CodigoPostalCordon.cordon,
            case(
                (EtiquetaEnvio.logistica_id.isnot(None), True),
                else_=False,
            ).label("tiene_logistica"),
            stats_eff_status.label("mlstatus"),
            func.count().label("cantidad"),
        )
        .outerjoin(
            shipping_stats,
            EtiquetaEnvio.shipping_id == shipping_stats.c.mlshippingid,
        )
        .outerjoin(
            Transporte,
            EtiquetaEnvio.transporte_id == Transporte.id,
        )
        .outerjoin(
            CodigoPostalCordon,
            stats_eff_zip_for_cordon == CodigoPostalCordon.codigo_postal,
        )
        .filter(
            EtiquetaEnvio.fecha_envio >= fecha_desde,
            EtiquetaEnvio.fecha_envio <= fecha_hasta,
            # Excluir etiquetas que existen en colecta
            ~EtiquetaEnvio.shipping_id.in_(db.query(EtiquetaColecta.shipping_id)),
        )
        .group_by(
            EtiquetaEnvio.fecha_envio,
            EtiquetaEnvio.es_manual,
            CodigoPostalCordon.cordon,
            "tiene_logistica",
            stats_eff_status,
        )
        .all()
    )

    # Agrupar resultados por fecha
    dias_map: dict[date, EstadisticaDiaItem] = {}

    for row in base_q:
        fecha = row.fecha_envio
        if fecha not in dias_map:
            dias_map[fecha] = EstadisticaDiaItem(fecha=fecha)

        dia = dias_map[fecha]
        cantidad = row.cantidad

        dia.total += cantidad

        if row.es_manual:
            dia.manuales += cantidad
        else:
            dia.flex += cantidad

        if row.cordon:
            dia.por_cordon[row.cordon] = dia.por_cordon.get(row.cordon, 0) + cantidad
        else:
            dia.sin_cordon += cantidad

        if row.tiene_logistica:
            dia.con_logistica += cantidad
        else:
            dia.sin_logistica += cantidad

        if row.mlstatus == "shipped":
            dia.enviados += cantidad
        elif row.mlstatus == "not_delivered":
            dia.no_entregados += cantidad

    # Ordenar por fecha
    dias = sorted(dias_map.values(), key=lambda d: d.fecha)

    return EstadisticasPorDiaResponse(dias=dias)


# ── Columnas disponibles para export ──────────────────────────────

EXPORT_COLUMNS = {
    "shipping_id": "Shipping ID",
    "fecha_envio": "Fecha Envío",
    "destinatario": "Destinatario",
    "direccion": "Dirección",
    "cp": "CP",
    "localidad": "Localidad",
    "cordon": "Cordón",
    "logistica": "Logística",
    "costo_envio": "Costo Envío",
    "estado_ml": "Estado ML",
    "estado_erp": "Estado ERP",
    "pistoleado": "Pistoleado",
    "caja": "Caja",
    "turbo": "Turbo",
    "lluvia": "Lluvia",
    "flag_envio": "Flag",
    "flag_envio_motivo": "Motivo Flag",
}


@router.get(
    "/etiquetas-envio/export",
    summary="Exportar etiquetas a Excel (XLSX)",
    response_class=StreamingResponse,
)
def exportar_etiquetas(
    fecha_desde: date = Query(..., description="Desde fecha (inclusive)"),
    fecha_hasta: date = Query(..., description="Hasta fecha (inclusive)"),
    columnas: Optional[str] = Query(
        None,
        description="Columnas a incluir (comma-separated). Si no se especifica, todas.",
    ),
    cordon: Optional[str] = Query(None, description="Filtrar por cordón"),
    logistica_id: Optional[int] = Query(None, description="Filtrar por logística"),
    sin_logistica: bool = Query(False, description="Solo sin logística asignada"),
    mlstatus: Optional[str] = Query(None, description="Filtrar por estado ML"),
    solo_outlet: bool = Query(False, description="Solo etiquetas de productos outlet"),
    solo_turbo: bool = Query(False, description="Solo etiquetas de envíos turbo"),
    search: Optional[str] = Query(None, description="Buscar por shipping_id, destinatario o dirección"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> StreamingResponse:
    """
    Exporta etiquetas filtradas a un archivo Excel (.xlsx).
    Soporta selección de columnas y todos los filtros de la vista.
    """
    _check_permiso(db, current_user, "envios_flex.exportar")

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    # Validar columnas solicitadas
    if columnas:
        cols_solicitadas = [c.strip() for c in columnas.split(",") if c.strip()]
        invalidas = [c for c in cols_solicitadas if c not in EXPORT_COLUMNS]
        if invalidas:
            raise HTTPException(400, f"Columnas inválidas: {', '.join(invalidas)}")
    else:
        cols_solicitadas = list(EXPORT_COLUMNS.keys())

    # Pre-filtrar shipping_ids por rango de fechas (performance)
    ids_fecha_exp = (
        db.query(EtiquetaEnvio.shipping_id)
        .filter(
            EtiquetaEnvio.fecha_envio >= fecha_desde,
            EtiquetaEnvio.fecha_envio <= fecha_hasta,
        )
        .subquery()
    )

    # Reusar subquery deduplicada de estado ERP
    soh_sub = _soh_status_subquery(db, shipping_ids_sub=ids_fecha_exp)
    manual_soh_sub = _manual_soh_status_subquery(db, shipping_ids_sub=ids_fecha_exp)
    ManualSaleOrderStatus = aliased(SaleOrderStatus)

    hoy_export = date.today()
    max_costo_exp = (
        db.query(
            LogisticaCostoCordon.logistica_id.label("costo_logistica_id"),
            LogisticaCostoCordon.cordon.label("costo_cordon"),
            func.max(LogisticaCostoCordon.id).label("max_id"),
        )
        .filter(LogisticaCostoCordon.vigente_desde <= hoy_export)
        .group_by(LogisticaCostoCordon.logistica_id, LogisticaCostoCordon.cordon)
        .subquery()
    )

    costo_exp = (
        db.query(
            LogisticaCostoCordon.logistica_id.label("costo_log_id"),
            LogisticaCostoCordon.cordon.label("costo_cordon_val"),
            LogisticaCostoCordon.costo.label("costo_valor"),
            LogisticaCostoCordon.costo_turbo.label("costo_turbo_valor"),
        )
        .join(
            max_costo_exp,
            LogisticaCostoCordon.id == max_costo_exp.c.max_id,
        )
        .subquery()
    )

    cordon_norm_exp = func.replace(CodigoPostalCordon.cordon, "ó", "o")

    # Lluvia offset config
    lluvia_tipo_e, lluvia_valor_e = _get_lluvia_config(db)

    # Subquery deduplicada: una fila por mlshippingid (evita duplicados por items)
    shipping_exp = _shipping_dedup_subquery(db, shipping_ids_sub=ids_fecha_exp)

    # COALESCE para envíos manuales en export
    exp_receiver = func.coalesce(EtiquetaEnvio.manual_receiver_name, shipping_exp.c.mlreceiver_name)
    exp_street = func.coalesce(EtiquetaEnvio.manual_street_name, shipping_exp.c.mlstreet_name)
    exp_street_num = func.coalesce(EtiquetaEnvio.manual_street_number, shipping_exp.c.mlstreet_number)
    exp_zip = func.coalesce(EtiquetaEnvio.manual_zip_code, shipping_exp.c.mlzip_code)
    exp_city = func.coalesce(EtiquetaEnvio.manual_city_name, shipping_exp.c.mlcity_name)
    exp_status = func.coalesce(EtiquetaEnvio.manual_status, shipping_exp.c.mlstatus)

    query = (
        db.query(
            EtiquetaEnvio.shipping_id,
            EtiquetaEnvio.fecha_envio,
            EtiquetaEnvio.pistoleado_at,
            EtiquetaEnvio.pistoleado_caja,
            Operador.nombre.label("pistoleado_operador_nombre"),
            Logistica.nombre.label("logistica_nombre"),
            exp_receiver.label("mlreceiver_name"),
            exp_street.label("mlstreet_name"),
            exp_street_num.label("mlstreet_number"),
            exp_zip.label("mlzip_code"),
            exp_city.label("mlcity_name"),
            exp_status.label("mlstatus"),
            CodigoPostalCordon.cordon,
            func.coalesce(SaleOrderStatus.ssos_name, ManualSaleOrderStatus.ssos_name).label("ssos_name"),
            func.coalesce(
                cast(EtiquetaEnvio.costo_override, Numeric(12, 2)),
                _build_costo_case(
                    costo_exp.c.costo_turbo_valor,
                    costo_exp.c.costo_valor,
                    lluvia_tipo_e,
                    lluvia_valor_e,
                ),
            ).label("costo_envio"),
            EtiquetaEnvio.es_turbo,
            EtiquetaEnvio.es_lluvia,
            EtiquetaEnvio.flag_envio,
            EtiquetaEnvio.flag_envio_motivo,
        )
        .outerjoin(Logistica, EtiquetaEnvio.logistica_id == Logistica.id)
        .outerjoin(
            shipping_exp,
            EtiquetaEnvio.shipping_id == shipping_exp.c.mlshippingid,
        )
        .outerjoin(
            CodigoPostalCordon,
            exp_zip == CodigoPostalCordon.codigo_postal,
        )
        .outerjoin(soh_sub, soh_sub.c.shipping_id_str == EtiquetaEnvio.shipping_id)
        .outerjoin(SaleOrderStatus, soh_sub.c.soh_ssos_id == SaleOrderStatus.ssos_id)
        .outerjoin(manual_soh_sub, manual_soh_sub.c.shipping_id_str == EtiquetaEnvio.shipping_id)
        .outerjoin(ManualSaleOrderStatus, manual_soh_sub.c.manual_ssos_id == ManualSaleOrderStatus.ssos_id)
        .outerjoin(Operador, EtiquetaEnvio.pistoleado_operador_id == Operador.id)
        .outerjoin(
            costo_exp,
            and_(
                costo_exp.c.costo_log_id == EtiquetaEnvio.logistica_id,
                costo_exp.c.costo_cordon_val == cordon_norm_exp,
            ),
        )
    )

    # Filtros
    query = query.filter(
        EtiquetaEnvio.fecha_envio >= fecha_desde,
        EtiquetaEnvio.fecha_envio <= fecha_hasta,
    )

    if cordon:
        query = query.filter(CodigoPostalCordon.cordon == cordon)
    if logistica_id is not None:
        query = query.filter(EtiquetaEnvio.logistica_id == logistica_id)
    if sin_logistica:
        query = query.filter(EtiquetaEnvio.logistica_id.is_(None))
    if solo_outlet:
        query = query.filter(EtiquetaEnvio.es_outlet.is_(True))
    if solo_turbo:
        query = query.filter(EtiquetaEnvio.es_turbo.is_(True))
    if mlstatus:
        query = query.filter(exp_status == mlstatus)
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            (EtiquetaEnvio.shipping_id.ilike(search_term))
            | (exp_receiver.ilike(search_term))
            | (exp_street.ilike(search_term))
            | (exp_city.ilike(search_term))
        )

    query = query.order_by(EtiquetaEnvio.fecha_envio, EtiquetaEnvio.shipping_id.desc())
    rows = query.all()

    # ── Generar XLSX ──────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Envíos Flex"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Headers
    headers = [EXPORT_COLUMNS[c] for c in cols_solicitadas]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Mapeo de columna → valor
    def get_cell_value(col_key: str, row: object) -> object:
        """Extrae el valor de una columna para una fila de resultados."""
        if col_key == "shipping_id":
            return row.shipping_id
        elif col_key == "fecha_envio":
            return row.fecha_envio
        elif col_key == "destinatario":
            return row.mlreceiver_name or ""
        elif col_key == "direccion":
            parts = [row.mlstreet_name or "", row.mlstreet_number or ""]
            return " ".join(p for p in parts if p).strip()
        elif col_key == "cp":
            return row.mlzip_code or ""
        elif col_key == "localidad":
            return row.mlcity_name or ""
        elif col_key == "cordon":
            return row.cordon or ""
        elif col_key == "logistica":
            return row.logistica_nombre or ""
        elif col_key == "costo_envio":
            return float(row.costo_envio) if row.costo_envio is not None else ""
        elif col_key == "estado_ml":
            return row.mlstatus or ""
        elif col_key == "estado_erp":
            return row.ssos_name or ""
        elif col_key == "pistoleado":
            if row.pistoleado_at:
                ts = str(row.pistoleado_at)[:16]  # YYYY-MM-DD HH:MM
                operador = row.pistoleado_operador_nombre or ""
                return f"{ts} — {operador}" if operador else ts
            return ""
        elif col_key == "caja":
            return row.pistoleado_caja or ""
        elif col_key == "turbo":
            return "Turbo" if row.es_turbo else ""
        elif col_key == "lluvia":
            return "Lluvia" if row.es_lluvia else ""
        elif col_key == "flag_envio":
            labels = {
                "mal_pasado": "Mal pasado",
                "envio_cancelado": "Cancelado",
                "duplicado": "Duplicado",
                "otro": "Otro",
            }
            return labels.get(row.flag_envio, row.flag_envio or "")
        elif col_key == "flag_envio_motivo":
            return row.flag_envio_motivo or ""
        return ""

    # Datos
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_key in enumerate(cols_solicitadas, start=1):
            ws.cell(row=row_idx, column=col_idx, value=get_cell_value(col_key, row))

    # Anchos automáticos (estimados)
    col_widths = {
        "shipping_id": 16,
        "fecha_envio": 14,
        "destinatario": 25,
        "direccion": 35,
        "cp": 8,
        "localidad": 20,
        "cordon": 12,
        "logistica": 18,
        "costo_envio": 14,
        "estado_ml": 18,
        "estado_erp": 18,
        "pistoleado": 22,
        "caja": 14,
        "turbo": 8,
        "lluvia": 8,
        "flag_envio": 14,
        "flag_envio_motivo": 30,
    }
    for col_idx, col_key in enumerate(cols_solicitadas, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = col_widths.get(col_key, 15)

    # Generar archivo en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"envios_flex_{fecha_desde}_{fecha_hasta}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.put(
    "/etiquetas-envio/{shipping_id}/logistica",
    response_model=dict,
    summary="Asignar logística a una etiqueta",
)
def asignar_logistica(
    shipping_id: str,
    payload: AsignarLogisticaRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """Asigna o desasigna la logística de una etiqueta."""
    _check_permiso(db, current_user, "envios_flex.asignar_logistica")

    etiqueta = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id == shipping_id).first()
    if not etiqueta:
        raise HTTPException(404, f"Etiqueta {shipping_id} no encontrada")

    if payload.logistica_id is not None:
        logistica = db.query(Logistica).filter(Logistica.id == payload.logistica_id, Logistica.activa.is_(True)).first()
        if not logistica:
            raise HTTPException(404, "Logística no encontrada o inactiva")

    etiqueta.logistica_id = payload.logistica_id
    db.commit()

    return {"ok": True, "shipping_id": shipping_id, "logistica_id": payload.logistica_id}


@router.put(
    "/etiquetas-envio/{shipping_id}/fecha",
    response_model=dict,
    summary="Cambiar fecha de envío (reprogramar)",
)
def cambiar_fecha(
    shipping_id: str,
    payload: CambiarFechaRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """Reprograma la fecha de envío de una etiqueta."""
    _check_permiso(db, current_user, "envios_flex.cambiar_fecha")

    etiqueta = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id == shipping_id).first()
    if not etiqueta:
        raise HTTPException(404, f"Etiqueta {shipping_id} no encontrada")

    etiqueta.fecha_envio = payload.fecha_envio
    db.commit()

    return {"ok": True, "shipping_id": shipping_id, "fecha_envio": str(payload.fecha_envio)}


@router.put(
    "/etiquetas-envio/{shipping_id}/costo",
    response_model=dict,
    summary="Establecer o quitar costo override de una etiqueta",
)
def set_costo_override(
    shipping_id: str,
    payload: CostoOverrideRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Establece un costo manual (override) para una etiqueta de envío.

    Si costo es None, elimina el override y vuelve al costo calculado
    automáticamente por logistica_costo_cordon.

    Requiere envios_flex.config y operador autenticado con PIN.
    Registra la acción en operador_actividad para auditoría.
    """
    _check_permiso(db, current_user, "envios_flex.config")

    etiqueta = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id == shipping_id).first()
    if not etiqueta:
        raise HTTPException(404, f"Etiqueta {shipping_id} no encontrada")

    # Validar operador activo
    operador = db.query(Operador).filter(Operador.id == payload.operador_id, Operador.activo.is_(True)).first()
    if not operador:
        raise HTTPException(404, "Operador no encontrado o inactivo")

    valor_anterior = float(etiqueta.costo_override) if etiqueta.costo_override is not None else None
    etiqueta.costo_override = payload.costo

    # Registrar actividad del operador
    actividad = OperadorActividad(
        operador_id=payload.operador_id,
        usuario_id=current_user.id,
        tab_key="envios-flex",
        accion="costo_override",
        detalle={
            "shipping_id": shipping_id,
            "costo_anterior": valor_anterior,
            "costo_nuevo": payload.costo,
        },
    )
    db.add(actividad)

    db.commit()

    return {
        "ok": True,
        "shipping_id": shipping_id,
        "costo_override": payload.costo,
    }


@router.put(
    "/etiquetas-envio/asignar-masivo",
    response_model=dict,
    summary="Asignar logística a múltiples etiquetas",
)
def asignar_masivo(
    payload: AsignarMasivoRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """Asigna la misma logística a un lote de etiquetas."""
    _check_permiso(db, current_user, "envios_flex.asignar_logistica")

    # Verificar que la logística existe
    logistica = db.query(Logistica).filter(Logistica.id == payload.logistica_id, Logistica.activa.is_(True)).first()
    if not logistica:
        raise HTTPException(404, "Logística no encontrada o inactiva")

    # Update masivo
    updated = (
        db.query(EtiquetaEnvio)
        .filter(EtiquetaEnvio.shipping_id.in_(payload.shipping_ids))
        .update(
            {EtiquetaEnvio.logistica_id: payload.logistica_id},
            synchronize_session="fetch",
        )
    )

    db.commit()

    return {
        "ok": True,
        "actualizadas": updated,
        "logistica_id": payload.logistica_id,
        "logistica_nombre": logistica.nombre,
    }


class BorrarEtiquetasRequest(BaseModel):
    """Payload para borrar etiquetas con auditoría."""

    shipping_ids: List[str] = Field(min_length=1)
    comment: Optional[str] = Field(None, max_length=500, description="Motivo del borrado")


@router.delete(
    "/etiquetas-envio",
    response_model=dict,
    summary="Borrar etiquetas de envío (con auditoría)",
)
def borrar_etiquetas(
    payload: BorrarEtiquetasRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Elimina etiquetas por shipping_id.
    Antes de borrar, copia cada etiqueta a etiquetas_envio_audit
    con el usuario que borró, timestamp y comentario opcional.
    """
    _check_permiso(db, current_user, "envios_flex.eliminar")

    # Buscar etiquetas a borrar
    etiquetas = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id.in_(payload.shipping_ids)).all()

    if not etiquetas:
        raise HTTPException(404, "No se encontraron etiquetas para borrar")

    # Copiar a audit antes de borrar
    for etiq in etiquetas:
        audit = EtiquetaEnvioAudit(
            shipping_id=etiq.shipping_id,
            sender_id=etiq.sender_id,
            hash_code=etiq.hash_code,
            nombre_archivo=etiq.nombre_archivo,
            fecha_envio=etiq.fecha_envio,
            logistica_id=etiq.logistica_id,
            latitud=etiq.latitud,
            longitud=etiq.longitud,
            direccion_completa=etiq.direccion_completa,
            direccion_comentario=etiq.direccion_comentario,
            pistoleado_at=etiq.pistoleado_at,
            pistoleado_caja=etiq.pistoleado_caja,
            pistoleado_operador_id=etiq.pistoleado_operador_id,
            original_created_at=etiq.created_at,
            original_updated_at=etiq.updated_at,
            deleted_by=current_user.id,
            delete_comment=payload.comment,
        )
        db.add(audit)

    # Desasociar items de RMA que referencian estas etiquetas
    if payload.shipping_ids:
        placeholders = ", ".join(f":id_{i}" for i in range(len(payload.shipping_ids)))
        params = {f"id_{i}": sid for i, sid in enumerate(payload.shipping_ids)}
        db.execute(
            text(f"UPDATE rma_caso_items SET shipping_id = NULL WHERE shipping_id IN ({placeholders})"),
            params,
        )

    # Borrar originales
    deleted = (
        db.query(EtiquetaEnvio)
        .filter(EtiquetaEnvio.shipping_id.in_(payload.shipping_ids))
        .delete(synchronize_session="fetch")
    )

    db.commit()

    return {"ok": True, "eliminadas": deleted}


# ── Envío manual (sin ML) ─────────────────────────────────────────────


@router.get(
    "/etiquetas-envio/lookup-pedido",
    response_model=dict,
    summary="Buscar pedido ERP por soh_id + bra_id → devuelve cust_id",
)
def lookup_pedido(
    soh_id: int = Query(..., description="N° pedido ERP"),
    bra_id: int = Query(..., description="Sucursal"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Busca un pedido en SaleOrderHeader por comp_id=1 + bra_id + soh_id
    y devuelve el cust_id asociado para autocompletar la dirección
    del cliente en el modal de envío manual.

    Acepta envios_flex.subir_etiquetas (tab Envíos Flex) O
    pedidos.crear_envio_flex (tab Pedidos Pendientes).
    """
    if not verificar_permiso(db, current_user, "envios_flex.subir_etiquetas") and not verificar_permiso(
        db, current_user, "pedidos.crear_envio_flex"
    ):
        raise HTTPException(
            status_code=403,
            detail="No tenés permiso: envios_flex.subir_etiquetas o pedidos.crear_envio_flex",
        )

    from app.models.tb_customer import TBCustomer

    soh = (
        db.query(
            SaleOrderHeader.cust_id,
            SaleOrderHeader.soh_id,
        )
        .filter(
            SaleOrderHeader.comp_id == 1,
            SaleOrderHeader.bra_id == bra_id,
            SaleOrderHeader.soh_id == soh_id,
        )
        .first()
    )
    if not soh:
        raise HTTPException(
            404,
            f"Pedido {soh_id} no encontrado en sucursal {bra_id}",
        )

    # Buscar datos del cliente para autocompletar
    cliente = (
        db.query(
            TBCustomer.cust_id,
            TBCustomer.cust_name,
            TBCustomer.cust_address,
            TBCustomer.cust_city,
            TBCustomer.cust_zip,
            TBCustomer.cust_phone1,
            TBCustomer.cust_cellphone,
        )
        .filter(
            TBCustomer.comp_id == 1,
            TBCustomer.cust_id == soh.cust_id,
        )
        .first()
    )

    return {
        "soh_id": soh_id,
        "bra_id": bra_id,
        "cust_id": soh.cust_id,
        "cust_name": cliente.cust_name if cliente else None,
        "cust_address": cliente.cust_address if cliente else None,
        "cust_city": cliente.cust_city if cliente else None,
        "cust_zip": cliente.cust_zip if cliente else None,
        "cust_phone1": cliente.cust_phone1 if cliente else None,
        "cust_cellphone": cliente.cust_cellphone if cliente else None,
    }


@router.post(
    "/etiquetas-envio/desde-pedido",
    response_model=CrearEnvioManualResponse,
    summary="Crear envío flex desde Pedidos Pendientes",
)
def crear_envio_desde_pedido(
    payload: CrearDesdePedidoRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> CrearEnvioManualResponse:
    """
    Crea un envío flex manual desde la tab Pedidos Pendientes.

    Similar a crear_envio_manual pero NO requiere operador (lo crea otro sector).
    Guarda el usuario del sistema que lo creó en creado_por_usuario_id para
    trazabilidad visual en la grilla de Envíos Flex.
    """
    _check_permiso(db, current_user, "pedidos.crear_envio_flex")

    # Resolver cust_id desde SaleOrderHeader (solo si se pasó soh_id + bra_id)
    resolved_cust_id: Optional[int] = payload.cust_id
    if payload.soh_id and payload.bra_id:
        soh = (
            db.query(SaleOrderHeader.cust_id)
            .filter(
                SaleOrderHeader.comp_id == 1,
                SaleOrderHeader.bra_id == payload.bra_id,
                SaleOrderHeader.soh_id == payload.soh_id,
            )
            .first()
        )
        if not soh:
            raise HTTPException(
                404,
                f"Pedido {payload.soh_id} no encontrado en sucursal {payload.bra_id}",
            )
        resolved_cust_id = soh.cust_id

    # Generar shipping_id único: MAN_{timestamp}_{seq}
    ahora = datetime.now(UTC)
    ts = ahora.strftime("%Y%m%d%H%M%S")
    prefix = f"MAN_{ts}_"
    count = (
        db.query(func.count()).select_from(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id.like(f"{prefix}%")).scalar()
        or 0
    )
    shipping_id = f"{prefix}{count + 1}"

    # Validar transporte si se envió
    if payload.transporte_id is not None:
        transporte = (
            db.query(Transporte).filter(Transporte.id == payload.transporte_id, Transporte.activa.is_(True)).first()
        )
        if not transporte:
            raise HTTPException(404, "Transporte no encontrado o inactivo")

    etiqueta = EtiquetaEnvio(
        shipping_id=shipping_id,
        fecha_envio=payload.fecha_envio,
        es_manual=True,
        manual_receiver_name=payload.receiver_name,
        manual_street_name=payload.street_name,
        manual_street_number=payload.street_number,
        manual_zip_code=payload.zip_code,
        manual_city_name=payload.city_name,
        manual_status=payload.status or "ready_to_ship",
        manual_cust_id=resolved_cust_id,
        manual_bra_id=payload.bra_id,
        manual_soh_id=payload.soh_id,
        manual_comment=payload.comment,
        manual_phone=payload.phone,
        logistica_id=payload.logistica_id,
        transporte_id=payload.transporte_id,
        nombre_archivo="desde_pedido",
        creado_por_usuario_id=current_user.id,
    )
    db.add(etiqueta)

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Error creando envío desde pedido: {str(e)}")

    # Resolver cordón: si hay transporte con CP, usar ese CP (zona de entrega
    # de la logística); sino, usar el CP del cliente.
    cp_for_cordon = payload.zip_code
    if payload.transporte_id is not None and transporte and transporte.cp:
        cp_for_cordon = transporte.cp

    cordon_val = None
    if cp_for_cordon:
        cordon_row = (
            db.query(CodigoPostalCordon.cordon).filter(CodigoPostalCordon.codigo_postal == cp_for_cordon).first()
        )
        cordon_val = cordon_row.cordon if cordon_row else None

    # Geocodificar en background (no bloquea la respuesta)
    background_tasks.add_task(
        _geocode_envio_manual,
        shipping_id=shipping_id,
        street_name=payload.street_name,
        street_number=payload.street_number,
        city_name=payload.city_name,
        transporte_id=payload.transporte_id,
        zip_code=payload.zip_code,
    )

    soh_label = f" desde pedido GBP:{payload.soh_id}" if payload.soh_id else ""
    return CrearEnvioManualResponse(
        ok=True,
        shipping_id=shipping_id,
        cordon=cordon_val,
        mensaje=f"Envío flex creado{soh_label}",
    )


@router.post(
    "/etiquetas-envio/manual-envio",
    response_model=CrearEnvioManualResponse,
    summary="Crear envío manual (sin MercadoLibre)",
)
def crear_envio_manual(
    payload: CrearEnvioManualRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> CrearEnvioManualResponse:
    """
    Crea una etiqueta de envío manual para envíos fuera de MercadoLibre.

    Genera un shipping_id con formato "MAN_{timestamp}_{seq}" para no
    colisionar con los IDs de ML.  Los datos de dirección se guardan
    en los campos manual_* de la etiqueta.

    Requiere envios_flex.subir_etiquetas y operador autenticado con PIN.
    Registra la acción en operador_actividad para auditoría.
    """
    _check_permiso(db, current_user, "envios_flex.subir_etiquetas")

    # Validar operador activo
    operador = db.query(Operador).filter(Operador.id == payload.operador_id, Operador.activo.is_(True)).first()
    if not operador:
        raise HTTPException(404, "Operador no encontrado o inactivo")

    # Validar logística si se envió
    if payload.logistica_id is not None:
        logistica = db.query(Logistica).filter(Logistica.id == payload.logistica_id, Logistica.activa.is_(True)).first()
        if not logistica:
            raise HTTPException(404, "Logística no encontrada o inactiva")

    # Validar transporte si se envió
    if payload.transporte_id is not None:
        transporte = (
            db.query(Transporte).filter(Transporte.id == payload.transporte_id, Transporte.activa.is_(True)).first()
        )
        if not transporte:
            raise HTTPException(404, "Transporte no encontrado o inactivo")

    # Si viene soh_id + bra_id, resolver cust_id desde SaleOrderHeader
    resolved_cust_id = payload.cust_id
    if payload.soh_id and payload.bra_id and not resolved_cust_id:
        soh = (
            db.query(SaleOrderHeader.cust_id)
            .filter(
                SaleOrderHeader.comp_id == 1,
                SaleOrderHeader.bra_id == payload.bra_id,
                SaleOrderHeader.soh_id == payload.soh_id,
            )
            .first()
        )
        if not soh:
            raise HTTPException(
                404,
                f"Pedido {payload.soh_id} no encontrado en sucursal {payload.bra_id}",
            )
        resolved_cust_id = soh.cust_id

    # Generar shipping_id único: MAN_{timestamp}_{seq}
    ahora = datetime.now(UTC)
    ts = ahora.strftime("%Y%m%d%H%M%S")

    # Secuencia: contar cuántos manuales hay con el mismo segundo
    prefix = f"MAN_{ts}_"
    count = (
        db.query(func.count()).select_from(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id.like(f"{prefix}%")).scalar()
        or 0
    )
    shipping_id = f"{prefix}{count + 1}"

    # Crear etiqueta manual
    etiqueta = EtiquetaEnvio(
        shipping_id=shipping_id,
        fecha_envio=payload.fecha_envio,
        logistica_id=payload.logistica_id,
        transporte_id=payload.transporte_id,
        es_manual=True,
        manual_receiver_name=payload.receiver_name,
        manual_street_name=payload.street_name,
        manual_street_number=payload.street_number,
        manual_zip_code=payload.zip_code,
        manual_city_name=payload.city_name,
        manual_status=payload.status,
        manual_cust_id=resolved_cust_id,
        manual_bra_id=payload.bra_id,
        manual_soh_id=payload.soh_id,
        manual_comment=payload.comment,
        manual_phone=payload.phone,
        nombre_archivo="envio_manual",
    )
    db.add(etiqueta)

    # Registrar actividad del operador
    actividad = OperadorActividad(
        operador_id=payload.operador_id,
        usuario_id=current_user.id,
        tab_key="envios-flex",
        accion="crear_envio_manual",
        detalle={
            "shipping_id": shipping_id,
            "receiver_name": payload.receiver_name,
            "street_name": payload.street_name,
            "street_number": payload.street_number,
            "zip_code": payload.zip_code,
            "city_name": payload.city_name,
            "status": payload.status,
            "cust_id": resolved_cust_id,
            "bra_id": payload.bra_id,
            "soh_id": payload.soh_id,
            "logistica_id": payload.logistica_id,
            "transporte_id": payload.transporte_id,
            "comment": payload.comment,
        },
    )
    db.add(actividad)

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Error guardando envío manual: {str(e)}")

    # Resolver cordón: si hay transporte con CP, usar ese CP (zona de entrega
    # de la logística); sino, usar el CP del cliente.
    cp_for_cordon = payload.zip_code
    if payload.transporte_id is not None and transporte and transporte.cp:
        cp_for_cordon = transporte.cp

    cordon_val = None
    if cp_for_cordon:
        cordon_row = (
            db.query(CodigoPostalCordon.cordon).filter(CodigoPostalCordon.codigo_postal == cp_for_cordon).first()
        )
        cordon_val = cordon_row.cordon if cordon_row else None

    # Geocodificar en background (no bloquea la respuesta)
    background_tasks.add_task(
        _geocode_envio_manual,
        shipping_id=shipping_id,
        street_name=payload.street_name,
        street_number=payload.street_number,
        city_name=payload.city_name,
        transporte_id=payload.transporte_id,
        zip_code=payload.zip_code,
    )

    return CrearEnvioManualResponse(
        ok=True,
        shipping_id=shipping_id,
        cordon=cordon_val,
        mensaje=f"Envío manual creado: {shipping_id}",
    )


@router.put(
    "/etiquetas-envio/manual-envio/{shipping_id}",
    response_model=dict,
    summary="Editar envío manual existente",
)
async def editar_envio_manual(
    shipping_id: str,
    payload: CrearEnvioManualRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Edita los datos de un envío manual existente (es_manual=True).

    Permite corregir destinatario, dirección, logística, estado, etc.
    sin tener que borrar y recrear el envío.
    Requiere envios_flex.subir_etiquetas y operador autenticado con PIN.
    """
    _check_permiso(db, current_user, "envios_flex.subir_etiquetas")

    etiqueta = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id == shipping_id).first()
    if not etiqueta:
        raise HTTPException(404, f"Etiqueta {shipping_id} no encontrada")

    if not etiqueta.es_manual:
        raise HTTPException(400, "Solo se pueden editar envíos manuales")

    # Validar operador activo
    operador = db.query(Operador).filter(Operador.id == payload.operador_id, Operador.activo.is_(True)).first()
    if not operador:
        raise HTTPException(404, "Operador no encontrado o inactivo")

    # Validar logística si se envió
    if payload.logistica_id is not None:
        logistica = db.query(Logistica).filter(Logistica.id == payload.logistica_id, Logistica.activa.is_(True)).first()
        if not logistica:
            raise HTTPException(404, "Logística no encontrada o inactiva")

    # Validar transporte si se envió
    if payload.transporte_id is not None:
        transporte_obj = (
            db.query(Transporte).filter(Transporte.id == payload.transporte_id, Transporte.activa.is_(True)).first()
        )
        if not transporte_obj:
            raise HTTPException(404, "Transporte no encontrado o inactivo")

    # Resolver cust_id desde pedido si corresponde
    resolved_cust_id = payload.cust_id
    if payload.soh_id and payload.bra_id and not resolved_cust_id:
        soh = (
            db.query(SaleOrderHeader.cust_id)
            .filter(
                SaleOrderHeader.comp_id == 1,
                SaleOrderHeader.bra_id == payload.bra_id,
                SaleOrderHeader.soh_id == payload.soh_id,
            )
            .first()
        )
        if not soh:
            raise HTTPException(
                404,
                f"Pedido {payload.soh_id} no encontrado en sucursal {payload.bra_id}",
            )
        resolved_cust_id = soh.cust_id

    # Detectar si la dirección cambió → necesita re-geocoding
    direccion_cambio = (
        (etiqueta.manual_street_name or "") != (payload.street_name or "")
        or (etiqueta.manual_street_number or "") != (payload.street_number or "")
        or (etiqueta.manual_city_name or "") != (payload.city_name or "")
        or (etiqueta.manual_zip_code or "") != (payload.zip_code or "")
        or (etiqueta.transporte_id or None) != (payload.transporte_id or None)
    )

    # Guardar estado anterior para auditoría
    estado_anterior = {
        "receiver_name": etiqueta.manual_receiver_name,
        "street_name": etiqueta.manual_street_name,
        "street_number": etiqueta.manual_street_number,
        "zip_code": etiqueta.manual_zip_code,
        "city_name": etiqueta.manual_city_name,
        "status": etiqueta.manual_status,
        "logistica_id": etiqueta.logistica_id,
        "transporte_id": etiqueta.transporte_id,
        "cust_id": etiqueta.manual_cust_id,
        "bra_id": etiqueta.manual_bra_id,
        "soh_id": etiqueta.manual_soh_id,
        "comment": etiqueta.manual_comment,
        "phone": etiqueta.manual_phone,
    }

    # Si la dirección cambió, limpiar coords viejas para forzar re-geocoding
    if direccion_cambio:
        etiqueta.latitud = None
        etiqueta.longitud = None

    # Actualizar campos
    etiqueta.fecha_envio = payload.fecha_envio
    etiqueta.manual_receiver_name = payload.receiver_name
    etiqueta.manual_street_name = payload.street_name
    etiqueta.manual_street_number = payload.street_number
    etiqueta.manual_zip_code = payload.zip_code
    etiqueta.manual_city_name = payload.city_name
    etiqueta.manual_status = payload.status
    etiqueta.manual_cust_id = resolved_cust_id
    etiqueta.manual_bra_id = payload.bra_id
    etiqueta.manual_soh_id = payload.soh_id
    etiqueta.manual_comment = payload.comment
    etiqueta.manual_phone = payload.phone
    etiqueta.logistica_id = payload.logistica_id
    etiqueta.transporte_id = payload.transporte_id

    # Registrar actividad
    actividad = OperadorActividad(
        operador_id=payload.operador_id,
        usuario_id=current_user.id,
        tab_key="envios-flex",
        accion="editar_envio_manual",
        detalle={
            "shipping_id": shipping_id,
            "anterior": estado_anterior,
            "nuevo": {
                "receiver_name": payload.receiver_name,
                "street_name": payload.street_name,
                "street_number": payload.street_number,
                "zip_code": payload.zip_code,
                "city_name": payload.city_name,
                "status": payload.status,
                "logistica_id": payload.logistica_id,
                "transporte_id": payload.transporte_id,
                "cust_id": resolved_cust_id,
                "bra_id": payload.bra_id,
                "soh_id": payload.soh_id,
                "comment": payload.comment,
                "phone": payload.phone,
            },
        },
    )
    db.add(actividad)

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Error guardando cambios: {str(e)}")

    # Si la dirección cambió, re-geocodificar en background
    if direccion_cambio:
        background_tasks.add_task(
            _geocode_envio_manual,
            shipping_id=shipping_id,
            street_name=payload.street_name,
            street_number=payload.street_number,
            city_name=payload.city_name,
            transporte_id=payload.transporte_id,
            zip_code=payload.zip_code,
        )
        logger.info(
            "Dirección cambió para %s → re-geocodificando en background",
            shipping_id,
        )

    # Resolver cordón: si hay transporte con CP, usar ese CP; sino, CP del cliente.
    cp_for_cordon = payload.zip_code
    if payload.transporte_id is not None and transporte_obj and transporte_obj.cp:
        cp_for_cordon = transporte_obj.cp

    cordon_val = None
    if cp_for_cordon:
        cordon_row = (
            db.query(CodigoPostalCordon.cordon).filter(CodigoPostalCordon.codigo_postal == cp_for_cordon).first()
        )
        cordon_val = cordon_row.cordon if cordon_row else None

    return {
        "ok": True,
        "shipping_id": shipping_id,
        "cordon": cordon_val,
        "mensaje": f"Envío {shipping_id} actualizado",
    }


@router.put(
    "/etiquetas-envio/{shipping_id}/estado-ml",
    response_model=dict,
    summary="Cambiar estado ML de un envío manual",
)
def cambiar_estado_ml(
    shipping_id: str,
    status: str = Query(
        ...,
        description="Nuevo estado: ready_to_ship, shipped, delivered",
        pattern="^(ready_to_ship|shipped|delivered)$",
    ),
    operador_id: int = Query(..., description="Operador autenticado con PIN"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Cambia el estado ML (manual_status) de un envío manual.

    Solo aplica a etiquetas con es_manual=True.
    Registra la acción en operador_actividad.
    """
    _check_permiso(db, current_user, "envios_flex.config")

    etiqueta = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id == shipping_id).first()
    if not etiqueta:
        raise HTTPException(404, f"Etiqueta {shipping_id} no encontrada")

    if not etiqueta.es_manual:
        raise HTTPException(
            400,
            "Solo se puede cambiar el estado ML de envíos manuales",
        )

    # Validar operador activo
    operador = db.query(Operador).filter(Operador.id == operador_id, Operador.activo.is_(True)).first()
    if not operador:
        raise HTTPException(404, "Operador no encontrado o inactivo")

    estado_anterior = etiqueta.manual_status
    etiqueta.manual_status = status

    # Registrar actividad
    actividad = OperadorActividad(
        operador_id=operador_id,
        usuario_id=current_user.id,
        tab_key="envios-flex",
        accion="cambiar_estado_manual",
        detalle={
            "shipping_id": shipping_id,
            "estado_anterior": estado_anterior,
            "estado_nuevo": status,
        },
    )
    db.add(actividad)

    db.commit()

    return {
        "ok": True,
        "shipping_id": shipping_id,
        "estado_anterior": estado_anterior,
        "estado_nuevo": status,
    }


@router.put(
    "/etiquetas-envio/{shipping_id}/turbo",
    response_model=dict,
    summary="Marcar o desmarcar envío como turbo",
)
def toggle_turbo(
    shipping_id: str,
    es_turbo: bool = Query(..., description="True para marcar como turbo, False para desmarcar"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Marca o desmarca una etiqueta como turbo (mlshipping_method_id = '515282').

    Requiere permiso envios_flex.config.
    """
    _check_permiso(db, current_user, "envios_flex.config")

    etiqueta = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id == shipping_id).first()
    if not etiqueta:
        raise HTTPException(404, f"Etiqueta {shipping_id} no encontrada")

    valor_anterior = etiqueta.es_turbo
    etiqueta.es_turbo = es_turbo
    db.commit()

    return {
        "ok": True,
        "shipping_id": shipping_id,
        "es_turbo": es_turbo,
        "valor_anterior": valor_anterior,
    }


@router.put(
    "/etiquetas-envio/turbo-masivo",
    response_model=dict,
    summary="Marcar o desmarcar turbo en múltiples etiquetas",
)
def toggle_turbo_masivo(
    payload: ShippingIdsRequest,
    es_turbo: bool = Query(..., description="True para marcar como turbo, False para desmarcar"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Marca o desmarca múltiples etiquetas como turbo en una sola operación.

    Requiere permiso envios_flex.config.
    """
    _check_permiso(db, current_user, "envios_flex.config")

    updated = (
        db.query(EtiquetaEnvio)
        .filter(EtiquetaEnvio.shipping_id.in_(payload.shipping_ids))
        .update({EtiquetaEnvio.es_turbo: es_turbo}, synchronize_session="fetch")
    )
    db.commit()

    return {
        "ok": True,
        "es_turbo": es_turbo,
        "actualizados": updated,
        "solicitados": len(payload.shipping_ids),
    }


@router.put(
    "/etiquetas-envio/{shipping_id}/lluvia",
    response_model=dict,
    summary="Marcar o desmarcar envío como lluvia",
)
def toggle_lluvia(
    shipping_id: str,
    es_lluvia: bool = Query(..., description="True para marcar como lluvia, False para desmarcar"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Marca o desmarca una etiqueta como lluvia (offset sobre costo turbo).

    Solo aplica cuando la etiqueta también es turbo.
    Requiere permiso envios_flex.config.
    """
    _check_permiso(db, current_user, "envios_flex.config")

    etiqueta = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id == shipping_id).first()
    if not etiqueta:
        raise HTTPException(404, f"Etiqueta {shipping_id} no encontrada")

    valor_anterior = etiqueta.es_lluvia
    etiqueta.es_lluvia = es_lluvia
    db.commit()

    return {
        "ok": True,
        "shipping_id": shipping_id,
        "es_lluvia": es_lluvia,
        "valor_anterior": valor_anterior,
    }


@router.put(
    "/etiquetas-envio/lluvia-masivo",
    response_model=dict,
    summary="Marcar o desmarcar lluvia en múltiples etiquetas",
)
def toggle_lluvia_masivo(
    payload: ShippingIdsRequest,
    es_lluvia: bool = Query(..., description="True para marcar como lluvia, False para desmarcar"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Marca o desmarca múltiples etiquetas como lluvia en una sola operación.

    Requiere permiso envios_flex.config.
    """
    _check_permiso(db, current_user, "envios_flex.config")

    updated = (
        db.query(EtiquetaEnvio)
        .filter(EtiquetaEnvio.shipping_id.in_(payload.shipping_ids))
        .update({EtiquetaEnvio.es_lluvia: es_lluvia}, synchronize_session="fetch")
    )
    db.commit()

    return {
        "ok": True,
        "es_lluvia": es_lluvia,
        "actualizados": updated,
        "solicitados": len(payload.shipping_ids),
    }


# ── Flag de envío (mal pasado, cancelado, duplicado, otro) ───────


@router.put(
    "/etiquetas-envio/{shipping_id}/flag",
    response_model=dict,
    summary="Flaggear o desflaggear un envío",
)
def toggle_flag_envio(
    shipping_id: str,
    payload: FlagEnvioRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Marca un envío con un flag (mal_pasado, envio_cancelado, duplicado, otro)
    o lo quita (flag_envio=None).

    Los envíos flaggeados se muestran con badge visual en TabEnviosFlex
    y se contabilizan separados en las estadísticas.

    Requiere permiso envios_flex.config.
    """
    _check_permiso(db, current_user, "envios_flex.config")

    if payload.flag_envio and payload.flag_envio not in FLAG_ENVIO_VALIDOS:
        raise HTTPException(
            422,
            f"Flag inválido. Valores permitidos: {', '.join(sorted(FLAG_ENVIO_VALIDOS))}",
        )

    etiqueta = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id == shipping_id).first()
    if not etiqueta:
        raise HTTPException(404, f"Etiqueta {shipping_id} no encontrada")

    valor_anterior = etiqueta.flag_envio

    if payload.flag_envio:
        etiqueta.flag_envio = payload.flag_envio
        etiqueta.flag_envio_motivo = payload.motivo
        etiqueta.flag_envio_at = datetime.now(UTC)
        etiqueta.flag_envio_usuario_id = current_user.id
    else:
        # Quitar flag
        etiqueta.flag_envio = None
        etiqueta.flag_envio_motivo = None
        etiqueta.flag_envio_at = None
        etiqueta.flag_envio_usuario_id = None

    db.commit()

    return {
        "ok": True,
        "shipping_id": shipping_id,
        "flag_envio": etiqueta.flag_envio,
        "valor_anterior": valor_anterior,
    }


@router.put(
    "/etiquetas-envio/flag-masivo",
    response_model=dict,
    summary="Flaggear o desflaggear múltiples envíos",
)
def toggle_flag_envio_masivo(
    payload: FlagEnvioMasivoRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Marca o desmarca múltiples etiquetas con un flag en una sola operación.

    Requiere permiso envios_flex.config.
    """
    _check_permiso(db, current_user, "envios_flex.config")

    if payload.flag_envio and payload.flag_envio not in FLAG_ENVIO_VALIDOS:
        raise HTTPException(
            422,
            f"Flag inválido. Valores permitidos: {', '.join(sorted(FLAG_ENVIO_VALIDOS))}",
        )

    update_values: dict = {}
    if payload.flag_envio:
        update_values = {
            EtiquetaEnvio.flag_envio: payload.flag_envio,
            EtiquetaEnvio.flag_envio_motivo: payload.motivo,
            EtiquetaEnvio.flag_envio_at: datetime.now(UTC),
            EtiquetaEnvio.flag_envio_usuario_id: current_user.id,
        }
    else:
        update_values = {
            EtiquetaEnvio.flag_envio: None,
            EtiquetaEnvio.flag_envio_motivo: None,
            EtiquetaEnvio.flag_envio_at: None,
            EtiquetaEnvio.flag_envio_usuario_id: None,
        }

    updated = (
        db.query(EtiquetaEnvio)
        .filter(EtiquetaEnvio.shipping_id.in_(payload.shipping_ids))
        .update(update_values, synchronize_session="fetch")
    )
    db.commit()

    return {
        "ok": True,
        "flag_envio": payload.flag_envio,
        "actualizados": updated,
        "solicitados": len(payload.shipping_ids),
    }


@router.put(
    "/etiquetas-envio/transporte-masivo",
    response_model=dict,
    summary="Asignar o quitar transporte a múltiples etiquetas",
)
def asignar_transporte_masivo(
    payload: AsignarTransporteMasivoRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Asigna (o desasigna) el mismo transporte a un lote de etiquetas.

    Si transporte_id es None, desasigna el transporte de todas las
    etiquetas indicadas.  Requiere permiso envios_flex.asignar_logistica.
    """
    _check_permiso(db, current_user, "envios_flex.asignar_logistica")

    transporte_nombre: Optional[str] = None
    if payload.transporte_id is not None:
        transporte = (
            db.query(Transporte).filter(Transporte.id == payload.transporte_id, Transporte.activa.is_(True)).first()
        )
        if not transporte:
            raise HTTPException(404, "Transporte no encontrado o inactivo")
        transporte_nombre = transporte.nombre

    updated = (
        db.query(EtiquetaEnvio)
        .filter(EtiquetaEnvio.shipping_id.in_(payload.shipping_ids))
        .update(
            {EtiquetaEnvio.transporte_id: payload.transporte_id},
            synchronize_session="fetch",
        )
    )
    db.commit()

    return {
        "ok": True,
        "actualizadas": updated,
        "transporte_id": payload.transporte_id,
        "transporte_nombre": transporte_nombre,
    }


# ── Pistoleado ───────────────────────────────────────────────────────


@router.post(
    "/etiquetas-envio/pistolear",
    response_model=PistolearResponse,
    summary="Pistolear etiqueta (escaneo de paquete en depósito)",
    responses={
        409: {"model": PistoleadoConflictResponse, "description": "Ya pistoleada"},
        422: {"description": "Logística no coincide"},
    },
)
def pistolear_etiqueta(
    payload: PistolearRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> PistolearResponse:
    """
    Registra el pistoleado de una etiqueta de envío.

    El operador escanea el QR de la etiqueta con pistola de barras.

    Dos modos de operación:
    - **Bulto único / ML**: si `bulto` es None → comportamiento clásico.
      Duplicado completo → 409.
    - **Multi-bulto (manuales)**: si `bulto` y `total_bultos` están presentes
      y total_bultos > 1 → tracking per-bulto en `pistoleado_bultos` (JSON array).
      Duplicado de bulto específico → 409. `pistoleado_at` se setea solo cuando
      TODOS los bultos fueron escaneados.

    Validaciones:
    1. La etiqueta debe existir en el sistema (→ 404).
    2. La logística asignada debe coincidir (→ 422).
    3. No debe haber sido pistoleada antes — por envío completo o por bulto (→ 409).

    Side effects:
    - Graba pistoleado_at, pistoleado_caja, pistoleado_operador_id.
    - En multi-bulto: actualiza pistoleado_bultos JSON array y total_bultos.
    - Registra actividad en operador_actividad.
    """
    _check_permiso(db, current_user, "envios_flex.pistoleado")

    # Validar operador activo
    operador = (
        db.query(Operador)
        .filter(
            Operador.id == payload.operador_id,
            Operador.activo.is_(True),
        )
        .first()
    )
    if not operador:
        raise HTTPException(404, "Operador no encontrado o inactivo")

    # Buscar etiqueta
    etiqueta = (
        db.query(EtiquetaEnvio)
        .filter(
            EtiquetaEnvio.shipping_id == payload.shipping_id,
        )
        .first()
    )
    if not etiqueta:
        raise HTTPException(404, f"Etiqueta {payload.shipping_id} no encontrada en el sistema")

    # Validar logística coincide — o asignar si pistoleado_asigna está activo
    logistica_pistoleando = db.query(Logistica).filter(Logistica.id == payload.logistica_id).first()
    if not logistica_pistoleando:
        raise HTTPException(404, "Logística de pistoleado no encontrada")

    fue_asignada = False
    if etiqueta.logistica_id is not None and etiqueta.logistica_id != payload.logistica_id:
        if logistica_pistoleando.pistoleado_asigna:
            # Modo asignación: reasignar la logística de la etiqueta
            etiqueta.logistica_id = payload.logistica_id
            fue_asignada = True
        else:
            # Modo estricto: rechazar si no coincide
            logistica_etiq = db.query(Logistica).filter(Logistica.id == etiqueta.logistica_id).first()
            raise HTTPException(
                422,
                detail={
                    "detail": "Logística no coincide",
                    "etiqueta_logistica": logistica_etiq.nombre if logistica_etiq else "Desconocida",
                    "etiqueta_logistica_id": etiqueta.logistica_id,
                    "pistoleando_logistica": logistica_pistoleando.nombre,
                    "pistoleando_logistica_id": payload.logistica_id,
                },
            )
    elif etiqueta.logistica_id is None and logistica_pistoleando.pistoleado_asigna:
        # Sin logística asignada + modo asignación → asignar
        etiqueta.logistica_id = payload.logistica_id
        fue_asignada = True
    elif etiqueta.logistica_id is None and not logistica_pistoleando.pistoleado_asigna:
        if payload.forzar_asignacion:
            # Doble escaneo: el operador confirmó asignar esta logística
            etiqueta.logistica_id = payload.logistica_id
            fue_asignada = True
        else:
            # Sin logística asignada + modo estricto → rechazar (primer escaneo)
            raise HTTPException(
                422,
                detail={
                    "detail": "Etiqueta sin logística asignada",
                    "etiqueta_logistica": "Sin asignar",
                    "etiqueta_logistica_id": None,
                    "pistoleando_logistica": logistica_pistoleando.nombre,
                    "pistoleando_logistica_id": payload.logistica_id,
                },
            )

    ahora = datetime.now(UTC)

    # --- Per-bulto tracking (solo envíos manuales multi-bulto) ---
    is_multi_bulto = payload.bulto is not None and payload.total_bultos is not None and payload.total_bultos > 1
    bultos_pistoleados_count = 0

    if is_multi_bulto:
        # Parsear array existente de bultos pistoleados
        bultos_arr: list[dict] = []
        if etiqueta.pistoleado_bultos:
            try:
                bultos_arr = json.loads(etiqueta.pistoleado_bultos)
            except (json.JSONDecodeError, TypeError):
                bultos_arr = []

        # Check duplicado de ESTE bulto específico
        bultos_ya_escaneados = {b["bulto"] for b in bultos_arr if "bulto" in b}
        if payload.bulto in bultos_ya_escaneados:
            # Buscar quién lo pistoleó
            entry_previo = next((b for b in bultos_arr if b.get("bulto") == payload.bulto), None)
            op_previo_id = entry_previo.get("operador_id") if entry_previo else None
            op_previo = db.query(Operador).filter(Operador.id == op_previo_id).first() if op_previo_id else None
            nombre_previo = op_previo.nombre if op_previo else "Desconocido"
            raise HTTPException(
                409,
                detail={
                    "detail": f"Bulto {payload.bulto}/{payload.total_bultos} ya pistoleado",
                    "pistoleado_por": nombre_previo,
                    "pistoleado_at": entry_previo.get("at", "") if entry_previo else "",
                    "pistoleado_caja": entry_previo.get("caja", "") if entry_previo else "",
                },
            )

        # Appendear nuevo bulto
        bultos_arr.append(
            {
                "bulto": payload.bulto,
                "at": ahora.isoformat(),
                "caja": payload.caja,
                "operador_id": payload.operador_id,
            }
        )
        etiqueta.pistoleado_bultos = json.dumps(bultos_arr, separators=(",", ":"))
        bultos_pistoleados_count = len(bultos_arr)

        # Guardar total_bultos en la etiqueta si no estaba
        if etiqueta.total_bultos is None:
            etiqueta.total_bultos = payload.total_bultos

        # Cuando TODOS los bultos fueron escaneados → marcar pistoleado_at (backward compat)
        if bultos_pistoleados_count >= payload.total_bultos:
            etiqueta.pistoleado_at = ahora
            etiqueta.pistoleado_caja = payload.caja
            etiqueta.pistoleado_operador_id = payload.operador_id
    else:
        # --- Comportamiento original: bulto único / ML etiquetas ---
        if etiqueta.pistoleado_at is not None:
            op_previo = db.query(Operador).filter(Operador.id == etiqueta.pistoleado_operador_id).first()
            nombre_previo = op_previo.nombre if op_previo else "Desconocido"
            raise HTTPException(
                409,
                detail={
                    "detail": "Ya pistoleada",
                    "pistoleado_por": nombre_previo,
                    "pistoleado_at": str(etiqueta.pistoleado_at),
                    "pistoleado_caja": etiqueta.pistoleado_caja or "",
                },
            )

        etiqueta.pistoleado_at = ahora
        etiqueta.pistoleado_caja = payload.caja
        etiqueta.pistoleado_operador_id = payload.operador_id
        bultos_pistoleados_count = 1

    # Registrar actividad
    detalle_actividad: dict = {
        "shipping_id": payload.shipping_id,
        "caja": payload.caja,
        "logistica_id": payload.logistica_id,
        "fecha_envio": str(etiqueta.fecha_envio) if etiqueta.fecha_envio else None,
    }
    if is_multi_bulto:
        detalle_actividad["bulto"] = payload.bulto
        detalle_actividad["total_bultos"] = payload.total_bultos
        detalle_actividad["bultos_pistoleados"] = bultos_pistoleados_count

    actividad = OperadorActividad(
        operador_id=payload.operador_id,
        usuario_id=current_user.id,
        tab_key="pistoleado",
        accion="pistoleado",
        detalle=detalle_actividad,
    )
    db.add(actividad)

    db.commit()

    # Obtener datos de ML shipping para el feedback
    ml_shipping = (
        db.query(MercadoLibreOrderShipping)
        .filter(
            MercadoLibreOrderShipping.mlshippingid == payload.shipping_id,
        )
        .first()
    )

    # Obtener cordón
    cordon_val = None
    if ml_shipping and ml_shipping.mlzip_code:
        cordon_row = (
            db.query(CodigoPostalCordon.cordon)
            .filter(
                CodigoPostalCordon.codigo_postal == ml_shipping.mlzip_code,
            )
            .first()
        )
        cordon_val = cordon_row.cordon if cordon_row else None

    # Obtener estado ERP del pedido (via soh_sub → SaleOrderStatus)
    estado_erp_name = None
    if ml_shipping and ml_shipping.mlo_id:
        soh_row = (
            db.query(SaleOrderHeader.ssos_id)
            .filter(SaleOrderHeader.mlo_id == ml_shipping.mlo_id)
            .order_by(desc(SaleOrderHeader.soh_cd))
            .first()
        )
        if soh_row and soh_row.ssos_id:
            ssos_row = db.query(SaleOrderStatus.ssos_name).filter(SaleOrderStatus.ssos_id == soh_row.ssos_id).first()
            estado_erp_name = ssos_row.ssos_name if ssos_row else None

    # Contar pistoleadas de este operador + logística + fecha (para TTS counter)
    count = (
        db.query(func.count())
        .select_from(EtiquetaEnvio)
        .filter(
            EtiquetaEnvio.pistoleado_operador_id == payload.operador_id,
            EtiquetaEnvio.logistica_id == payload.logistica_id,
            EtiquetaEnvio.pistoleado_at.isnot(None),
            func.date(EtiquetaEnvio.pistoleado_at) == date.today(),
        )
        .scalar()
        or 0
    )

    return PistolearResponse(
        ok=True,
        shipping_id=payload.shipping_id,
        caja=payload.caja,
        operador=operador.nombre,
        receiver_name=ml_shipping.mlreceiver_name if ml_shipping else None,
        ciudad=ml_shipping.mlcity_name if ml_shipping else None,
        cordon=cordon_val,
        pistoleado_at=str(ahora),
        bulto=payload.bulto,
        total_bultos=payload.total_bultos,
        bultos_pistoleados=bultos_pistoleados_count,
        count=count,
        estado_erp=estado_erp_name,
        logistica_asignada=fue_asignada,
    )


@router.get(
    "/etiquetas-envio/pistoleado/stats",
    response_model=PistoleadoStatsResponse,
    summary="Estadísticas de pistoleado por fecha y logística",
)
def stats_pistoleado(
    fecha: Optional[date] = Query(None, description="Fecha de envío (default: hoy)"),
    logistica_id: Optional[int] = Query(None, description="Filtrar por logística"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> PistoleadoStatsResponse:
    """
    Estadísticas de pistoleado: total, pistoleadas, pendientes, porcentaje,
    desglose por caja y por operador.
    """
    _check_permiso(db, current_user, "envios_flex.pistoleado")

    fecha_filtro = fecha or date.today()

    # Base query: etiquetas de la fecha
    base = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.fecha_envio == fecha_filtro)

    if logistica_id is not None:
        base = base.filter(EtiquetaEnvio.logistica_id == logistica_id)

    total = base.count()
    pistoleadas = base.filter(EtiquetaEnvio.pistoleado_at.isnot(None)).count()
    pendientes = total - pistoleadas
    porcentaje = round((pistoleadas / total * 100), 1) if total > 0 else 0.0

    # Por caja
    caja_rows = db.query(
        EtiquetaEnvio.pistoleado_caja,
        func.count().label("cantidad"),
    ).filter(
        EtiquetaEnvio.fecha_envio == fecha_filtro,
        EtiquetaEnvio.pistoleado_at.isnot(None),
        EtiquetaEnvio.pistoleado_caja.isnot(None),
    )
    if logistica_id is not None:
        caja_rows = caja_rows.filter(EtiquetaEnvio.logistica_id == logistica_id)
    caja_rows = caja_rows.group_by(EtiquetaEnvio.pistoleado_caja).all()
    por_caja = {row.pistoleado_caja: row.cantidad for row in caja_rows}

    # Por operador
    op_rows = (
        db.query(
            Operador.nombre,
            func.count().label("cantidad"),
        )
        .join(EtiquetaEnvio, EtiquetaEnvio.pistoleado_operador_id == Operador.id)
        .filter(
            EtiquetaEnvio.fecha_envio == fecha_filtro,
            EtiquetaEnvio.pistoleado_at.isnot(None),
        )
    )
    if logistica_id is not None:
        op_rows = op_rows.filter(EtiquetaEnvio.logistica_id == logistica_id)
    op_rows = op_rows.group_by(Operador.nombre).all()
    por_operador = {row.nombre: row.cantidad for row in op_rows}

    # Pistoleadas cuyo pedido ERP está "En Preparación"
    # shipping_id → MercadoLibreOrderShipping.mlshippingid → mlo_id
    # → SaleOrderHeader.mlo_id → ssos_id → SaleOrderStatus.ssos_name
    ssos_preparacion = db.query(SaleOrderStatus.ssos_id).filter(SaleOrderStatus.ssos_name == "En Preparación").first()
    en_preparacion = 0
    if ssos_preparacion:
        en_prep_q = (
            db.query(func.count())
            .select_from(EtiquetaEnvio)
            .join(
                MercadoLibreOrderShipping,
                MercadoLibreOrderShipping.mlshippingid == EtiquetaEnvio.shipping_id,
            )
            .join(
                SaleOrderHeader,
                SaleOrderHeader.mlo_id == MercadoLibreOrderShipping.mlo_id,
            )
            .filter(
                EtiquetaEnvio.fecha_envio == fecha_filtro,
                EtiquetaEnvio.pistoleado_at.isnot(None),
                SaleOrderHeader.ssos_id == ssos_preparacion.ssos_id,
            )
        )
        if logistica_id is not None:
            en_prep_q = en_prep_q.filter(EtiquetaEnvio.logistica_id == logistica_id)
        en_preparacion = en_prep_q.scalar() or 0

    return PistoleadoStatsResponse(
        total_etiquetas=total,
        pistoleadas=pistoleadas,
        pendientes=pendientes,
        porcentaje=porcentaje,
        en_preparacion=en_preparacion,
        por_caja=por_caja,
        por_operador=por_operador,
    )


@router.delete(
    "/etiquetas-envio/pistolear/{shipping_id}",
    response_model=dict,
    summary="Deshacer pistoleado (ANULAR)",
)
def deshacer_pistoleado(
    shipping_id: str,
    operador_id: int = Query(..., description="Operador que ejecuta la anulación"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Revierte un pistoleado por error (comando ANULAR).
    Pone pistoleado_at, pistoleado_caja, pistoleado_operador_id en NULL.
    Registra actividad 'despistoleado' con el estado anterior.
    """
    _check_permiso(db, current_user, "envios_flex.pistoleado")

    # Validar operador activo
    operador = (
        db.query(Operador)
        .filter(
            Operador.id == operador_id,
            Operador.activo.is_(True),
        )
        .first()
    )
    if not operador:
        raise HTTPException(404, "Operador no encontrado o inactivo")

    etiqueta = (
        db.query(EtiquetaEnvio)
        .filter(
            EtiquetaEnvio.shipping_id == shipping_id,
        )
        .first()
    )
    if not etiqueta:
        raise HTTPException(404, f"Etiqueta {shipping_id} no encontrada")

    if etiqueta.pistoleado_at is None:
        raise HTTPException(400, f"Etiqueta {shipping_id} no está pistoleada")

    # Guardar estado anterior para auditoría
    op_previo = db.query(Operador).filter(Operador.id == etiqueta.pistoleado_operador_id).first()
    estado_anterior = {
        "shipping_id": shipping_id,
        "pistoleado_at": str(etiqueta.pistoleado_at),
        "pistoleado_caja": etiqueta.pistoleado_caja,
        "pistoleado_operador_id": etiqueta.pistoleado_operador_id,
        "pistoleado_operador_nombre": op_previo.nombre if op_previo else None,
    }

    # Limpiar pistoleado
    etiqueta.pistoleado_at = None
    etiqueta.pistoleado_caja = None
    etiqueta.pistoleado_operador_id = None

    # Registrar actividad
    actividad = OperadorActividad(
        operador_id=operador_id,
        usuario_id=current_user.id,
        tab_key="pistoleado",
        accion="despistoleado",
        detalle=estado_anterior,
    )
    db.add(actividad)

    db.commit()

    return {"ok": True, "shipping_id": shipping_id, "anulado_por": operador.nombre}


# ── Re-enrichment manual ────────────────────────────────────────


class ReEnriquecerRequest(BaseModel):
    """Body para re-enriquecer etiquetas."""

    fecha_desde: Optional[date] = None
    fecha_hasta: Optional[date] = None
    shipping_ids: Optional[List[str]] = None


@router.post(
    "/etiquetas-envio/re-enriquecer",
    summary="Re-enriquece etiquetas desde ml_previews con fallback HTTP",
)
async def re_enriquecer_etiquetas(
    body: ReEnriquecerRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Re-enriquece etiquetas en dos fases:
    1. Batch rápido: lee ml_previews directo (1 query para todos)
    2. Fallback HTTP: los que no están en ml_previews los busca
       uno por uno vía el proxy ml-webhook (~200ms c/u)

    Modos de uso:
    - Por fecha: {fecha_desde, fecha_hasta} → re-enriquece todas en ese rango
    - Por IDs:   {shipping_ids: ["123", "456"]} → re-enriquece esas específicas
    - Sin filtro: {} → re-enriquece todo lo de hoy

    Requiere permiso envios_flex.config.
    """
    _check_permiso(db, current_user, "envios_flex.config")

    # Determinar qué etiquetas re-enriquecer
    if body.shipping_ids:
        ids = body.shipping_ids
    else:
        desde = body.fecha_desde or date.today()
        hasta = body.fecha_hasta or date.today()

        etiquetas = (
            db.query(EtiquetaEnvio.shipping_id)
            .filter(
                EtiquetaEnvio.fecha_envio >= desde,
                EtiquetaEnvio.fecha_envio <= hasta,
            )
            .all()
        )
        ids = [e.shipping_id for e in etiquetas]

    if not ids:
        return {
            "actualizadas": 0,
            "sin_preview": 0,
            "fallback_ok": 0,
            "fallback_errores": 0,
            "total": 0,
            "mensaje": "No hay etiquetas para re-enriquecer",
        }

    # Fase 1: batch desde ml_previews (rápido)
    resultado_db = re_enriquecer_desde_db(ids)
    ids_sin_preview = resultado_db.get("ids_sin_preview", [])

    # Fase 2: fallback HTTP para los que no estaban en ml_previews
    fallback_ok = 0
    fallback_errores = 0
    if ids_sin_preview:
        resultado_http = await re_enriquecer_por_http(ids_sin_preview)
        fallback_ok = resultado_http["actualizadas"]
        fallback_errores = resultado_http["errores"]

    total_actualizadas = resultado_db["actualizadas"] + fallback_ok
    return {
        "actualizadas": total_actualizadas,
        "sin_preview": resultado_db["sin_preview"],
        "fallback_ok": fallback_ok,
        "fallback_errores": fallback_errores,
        "total": len(ids),
        "mensaje": (
            f"Re-enriquecidas {total_actualizadas} de {len(ids)} etiquetas "
            f"({resultado_db['actualizadas']} por DB, {fallback_ok} por HTTP, "
            f"{fallback_errores} errores)"
        ),
    }


# ── Impresión de etiquetas ZPL ──────────────────────────────────

# Errores de ML traducidos al español
_ML_LABEL_ERRORS: dict = {
    "NOT_PRINTABLE_STATUS": "El envío no está listo para imprimir (ya fue despachado, entregado o cancelado)",
    "invalid_shipment_ff_public": "Los envíos Fulfillment no permiten imprimir etiquetas desde acá",
    "invalid_shipment_mode": "Este envío no es de tipo ME2 (MercadoEnvíos 2)",
    "invalid_shipment_caller": "Usuario no autorizado para este envío",
}


@router.get(
    "/etiquetas-envio/{shipping_id}/etiqueta",
    summary="Obtiene la etiqueta ZPL de un envío desde ML",
)
async def obtener_etiqueta_zpl(
    shipping_id: str,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """
    Obtiene la etiqueta ZPL de un envío desde MercadoLibre vía ml-webhook proxy.

    Devuelve {ok: true, zpl: "^XA..."} si la etiqueta está disponible,
    o {ok: false, error: "...", code: "..."} con error descriptivo en español.

    Solo se puede imprimir si el envío está en ready_to_ship / ready_to_print o printed.
    Requiere permiso envios_flex.ver.
    """
    _check_permiso(db, current_user, "envios_flex.ver")

    # Verificar que la etiqueta existe en nuestro sistema
    etiqueta = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id == shipping_id).first()
    if not etiqueta:
        raise HTTPException(status_code=404, detail="Etiqueta no encontrada")

    resultado = await fetch_shipment_label_zpl(shipping_id)

    if not resultado["ok"]:
        # Traducir error de ML al español
        code = resultado.get("code", "")
        error_es = _ML_LABEL_ERRORS.get(code, resultado.get("error", "Error desconocido"))
        return {"ok": False, "error": error_es, "code": code}

    return resultado


# ── Impresión de etiquetas ZPL para envíos manuales ─────────────────


import logging

_logger = logging.getLogger(__name__)


@router.get(
    "/etiquetas-envio/{shipping_id}/etiqueta-manual",
    summary="Genera etiqueta ZPL local para un envío manual",
)
async def generar_etiqueta_manual_zpl(
    shipping_id: str,
    num_bultos: int = Query(1, ge=1),
    tipo_envio_manual: Optional[str] = Query(None),
    tipo_domicilio_manual: Optional[str] = Query(None),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Response:
    """
    Genera etiquetas ZPL a partir del template local (etiqueta.zpl)
    para envíos manuales (es_manual=True).

    Usa los datos del envío manual (destinatario, dirección, CP, ciudad,
    observaciones). Si el envío tiene soh_id y bra_id, obtiene los items
    del pedido ERP para incluir SKUs y cantidad.

    Parámetros:
    - shipping_id: ID del envío manual (ej: MAN_20260123_001)
    - num_bultos: Número de bultos (genera una etiqueta por bulto, 1-10)
    - tipo_envio_manual: Override del tipo de envío (ej: "Domicilio")
    - tipo_domicilio_manual: Override del tipo de domicilio (Particular/Comercial/Sucursal)
    """
    _check_permiso(db, current_user, "envios_flex.ver")

    etiqueta = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id == shipping_id).first()
    if not etiqueta:
        raise HTTPException(status_code=404, detail="Etiqueta no encontrada")

    if not etiqueta.es_manual:
        raise HTTPException(
            status_code=400,
            detail="Este envío no es manual. Usá el endpoint de etiquetas ML.",
        )

    # ── Obtener items del pedido ERP (si hay soh_id + bra_id) ────────
    cantidad_total = 0
    skus_concatenados = "N/A"
    id_pedido = "N/A"
    orden_tn = "N/A"

    if etiqueta.manual_soh_id and etiqueta.manual_bra_id:
        id_pedido = str(etiqueta.manual_soh_id)

        items_query = (
            db.query(
                SaleOrderDetail.item_id,
                SaleOrderDetail.sod_qty,
                TBItem.item_code,
            )
            .outerjoin(
                TBItem,
                and_(
                    SaleOrderDetail.item_id == TBItem.item_id,
                    SaleOrderDetail.comp_id == TBItem.comp_id,
                ),
            )
            .filter(
                and_(
                    SaleOrderDetail.soh_id == etiqueta.manual_soh_id,
                    SaleOrderDetail.bra_id == etiqueta.manual_bra_id,
                    func.coalesce(
                        SaleOrderDetail.item_id,
                        SaleOrderDetail.sod_item_id_origin,
                    ).notin_([2953, 2954]),
                )
            )
            .all()
        )

        cantidad_total = sum(float(i.sod_qty) if i.sod_qty else 0 for i in items_query)
        skus_list = [i.item_code for i in items_query if i.item_code]
        skus_concatenados = " - ".join(skus_list) if skus_list else "N/A"

    # ── Datos de dirección del envío manual ──────────────────────────
    destinatario = etiqueta.manual_receiver_name or "N/A"
    calle = etiqueta.manual_street_name or ""
    numero = etiqueta.manual_street_number or ""
    direccion = f"{calle} {numero}".strip() or "N/A"
    codigo_postal = etiqueta.manual_zip_code or "N/A"
    ciudad = etiqueta.manual_city_name or "N/A"
    observaciones = etiqueta.manual_comment or "N/A"
    # Teléfono: manual_phone tiene prioridad, fallback a TBCustomer
    telefono = etiqueta.manual_phone or None

    if not telefono and etiqueta.manual_cust_id:
        from app.models.tb_customer import TBCustomer

        cliente = (
            db.query(TBCustomer.cust_phone1, TBCustomer.cust_cellphone)
            .filter(TBCustomer.cust_id == etiqueta.manual_cust_id)
            .first()
        )
        if cliente:
            telefono = cliente.cust_cellphone or cliente.cust_phone1 or None

    telefono = telefono or "N/A"

    # ── Tipo de envío y domicilio ────────────────────────────────────
    tipo_envio = tipo_envio_manual or "Domicilio"
    tipo_domicilio = tipo_domicilio_manual or "Particular"

    # ── Leer template ZPL ────────────────────────────────────────────
    template_path = FilePath(__file__).parent.parent.parent.parent / "templates" / "etiqueta.zpl"

    try:
        with open(template_path, "r", encoding="utf-8") as f:
            zpl_template = f.read()
    except FileNotFoundError:
        _logger.error(f"Template ZPL no encontrado en: {template_path}")
        raise HTTPException(status_code=500, detail="Template de etiqueta no encontrado")

    # ── Contexto para template ───────────────────────────────────────
    bra_id = etiqueta.manual_bra_id or 0
    soh_id = etiqueta.manual_soh_id or 0

    context = {
        "CANTIDAD_ITEMS_PEDIDO": str(int(cantidad_total)) if cantidad_total else "0",
        "SKUS_CONCATENADOS": skus_concatenados[:50],
        "ID_PEDIDO": id_pedido,
        "ORDEN_TN": orden_tn,
        "TIPO_ENVIO_ETIQUETA": tipo_envio,
        "NOMBRE_DESTINATARIO": destinatario,
        "TELEFONO_DESTINATARIO": telefono,
        "DIRECCION_CALLE": direccion,
        "OBSERVACIONES": observaciones,
        "CODIGO_POSTAL": codigo_postal,
        "BARRIO": ciudad,
        "TIPO_DOMICILIO": tipo_domicilio,
        "TOTAL_BULTOS": str(num_bultos),
    }

    # ── Generar etiquetas (una por bulto) ────────────────────────────
    zpl_labels = []
    for i in range(1, num_bultos + 1):
        label_context = context.copy()
        label_context["BULTO_ACTUAL"] = str(i)
        label_context["CODIGO_ENVIO"] = f"{bra_id}-{soh_id}-{i}"

        # QR data: JSON para pistoleado
        qr_obj = {
            "id": shipping_id,
            "bulto": i,
            "total_bultos": num_bultos,
        }
        if soh_id:
            qr_obj["soh_id"] = soh_id
        label_context["QR_DATA"] = json.dumps(qr_obj, separators=(",", ":"))

        rendered_zpl = zpl_template
        for key, value in label_context.items():
            rendered_zpl = rendered_zpl.replace(f"{{{{{key}}}}}", str(value))

        zpl_labels.append(rendered_zpl)

    # ── Remito de transporte (etiqueta adicional al final) ─────────
    if etiqueta.transporte_id:
        transporte = db.query(Transporte).filter(Transporte.id == etiqueta.transporte_id).first()
        if transporte:
            # Obtener nombre de logística si tiene
            logistica_nombre = "N/A"
            if etiqueta.logistica_id:
                logistica = db.query(Logistica).filter(Logistica.id == etiqueta.logistica_id).first()
                if logistica:
                    logistica_nombre = logistica.nombre

            remito_template_path = (
                FilePath(__file__).parent.parent.parent.parent / "templates" / "remito_transporte.zpl"
            )
            try:
                with open(remito_template_path, "r", encoding="utf-8") as f:
                    remito_template = f.read()

                remito_context = {
                    "FECHA_ENVIO": etiqueta.fecha_envio or "N/A",
                    "SHIPPING_ID": shipping_id,
                    "TRANSPORTE_NOMBRE": transporte.nombre or "N/A",
                    "TRANSPORTE_DIRECCION": transporte.direccion or "N/A",
                    "TRANSPORTE_CP": transporte.cp or "N/A",
                    "TRANSPORTE_LOCALIDAD": transporte.localidad or "N/A",
                    "TRANSPORTE_TELEFONO": transporte.telefono or "N/A",
                    "TRANSPORTE_HORARIO": transporte.horario or "N/A",
                    "NOMBRE_DESTINATARIO": destinatario,
                    "DIRECCION_CLIENTE": direccion,
                    "CP_CLIENTE": codigo_postal,
                    "CIUDAD_CLIENTE": ciudad,
                    "TELEFONO_DESTINATARIO": telefono,
                    "ID_PEDIDO": id_pedido,
                    "CANTIDAD_ITEMS": str(int(cantidad_total)) if cantidad_total else "0",
                    "SKUS_CONCATENADOS": skus_concatenados[:50],
                    "OBSERVACIONES": observaciones,
                    "TOTAL_BULTOS": str(num_bultos),
                    "BULTOS_PLURAL": "S" if num_bultos != 1 else "",
                    "LOGISTICA_NOMBRE": logistica_nombre,
                }

                rendered_remito = remito_template
                for key, value in remito_context.items():
                    rendered_remito = rendered_remito.replace(f"{{{{{key}}}}}", str(value))

                zpl_labels.append(rendered_remito)
                _logger.info(f"Remito de transporte '{transporte.nombre}' agregado para envío {shipping_id}")
            except FileNotFoundError:
                _logger.warning(f"Template remito_transporte.zpl no encontrado, se omite remito para {shipping_id}")

    full_zpl = "\n".join(zpl_labels)

    _logger.info(f"Generadas {num_bultos} etiquetas ZPL para envío manual {shipping_id}")

    return Response(
        content=full_zpl,
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename=etiqueta_manual_{shipping_id}.txt"},
    )


# ── Geocodificación masiva ───────────────────────────────────────────


class GeocodificarRequest(BaseModel):
    """IDs de etiquetas a geocodificar."""

    shipping_ids: List[str]


class GeocodificarResponse(BaseModel):
    """Resultado de geocodificación masiva."""

    total: int
    geocodificados: int
    ya_tenian: int
    sin_resultado: int
    errores: int


@router.post(
    "/etiquetas-envio/geocodificar",
    response_model=GeocodificarResponse,
    summary="Geocodificar etiquetas (o re-geocodificar con coords de transporte)",
)
async def geocodificar_etiquetas(
    body: GeocodificarRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> GeocodificarResponse:
    """
    Geocodifica etiquetas que no tienen lat/lng, o actualiza las que tienen
    transporte asignado para que apunten a la dirección del transporte.

    Para cada etiqueta:
      1. Si tiene transporte → SIEMPRE usar coords del transporte (aunque ya
         tenga lat/lng propias). Esto corrige envíos que apuntan a la
         dirección del cliente cuando deberían apuntar al transporte.
      2. Si NO tiene transporte y ya tiene coords → skip (ya_tenian).
      3. Geocodificar dirección del cliente (manual, enriquecida, o ML).
    """
    _check_permiso(db, current_user, "envios_flex.asignar_logistica")

    if len(body.shipping_ids) > 200:
        raise HTTPException(status_code=400, detail="Máximo 200 etiquetas por request")

    etiquetas = db.query(EtiquetaEnvio).filter(EtiquetaEnvio.shipping_id.in_(body.shipping_ids)).all()

    geocodificados = 0
    ya_tenian = 0
    sin_resultado = 0
    errores = 0

    for etiqueta in etiquetas:
        try:
            lat, lng = None, None

            # ── Con transporte: SIEMPRE usar coords del transporte ──
            if etiqueta.transporte_id:
                transporte = db.query(Transporte).filter(Transporte.id == etiqueta.transporte_id).first()
                if transporte:
                    if transporte.latitud and transporte.longitud:
                        lat, lng = transporte.latitud, transporte.longitud
                    elif transporte.direccion:
                        ciudad_transp = transporte.localidad or "Buenos Aires"
                        coords = await geocode_address(transporte.direccion, ciudad=ciudad_transp, db=db)
                        if coords:
                            lat, lng = coords
                            transporte.latitud = lat
                            transporte.longitud = lng

                if lat is not None and lng is not None:
                    # Actualizar aunque ya tuviera coords (pueden ser del cliente)
                    if etiqueta.latitud == lat and etiqueta.longitud == lng:
                        ya_tenian += 1
                    else:
                        etiqueta.latitud = lat
                        etiqueta.longitud = lng
                        geocodificados += 1
                        logger.info(
                            "Geocoding (transporte) %s → (%.6f, %.6f)",
                            etiqueta.shipping_id,
                            lat,
                            lng,
                        )
                    continue

                # Transporte sin coords ni dirección → caer al fallback del cliente
                # (no hacemos continue, dejamos que siga abajo)

            # ── Sin transporte: skip si ya tiene coordenadas ──
            if etiqueta.latitud and etiqueta.longitud:
                ya_tenian += 1
                continue

            # ── Fallback: dirección del cliente (manual o enriquecida) ──
            direccion = None
            ciudad = "Buenos Aires"
            zip_code = None

            if etiqueta.es_manual and etiqueta.manual_street_name:
                direccion = f"{etiqueta.manual_street_name} {etiqueta.manual_street_number or ''}".strip()
                ciudad = etiqueta.manual_city_name or "Buenos Aires"
                zip_code = etiqueta.manual_zip_code
            elif etiqueta.direccion_completa:
                direccion = etiqueta.direccion_completa
            else:
                # Buscar en ML shipping como último recurso
                ml_ship = (
                    db.query(MercadoLibreOrderShipping)
                    .filter(MercadoLibreOrderShipping.mlshippingid == etiqueta.shipping_id)
                    .first()
                )
                if ml_ship and ml_ship.mlstreet_name:
                    direccion = f"{ml_ship.mlstreet_name} {ml_ship.mlstreet_number or ''}".strip()
                    ciudad = ml_ship.mlcity_name or "Buenos Aires"
                    zip_code = ml_ship.mlzip_code

            if direccion:
                coords = await geocode_address(direccion, ciudad=ciudad, zip_code=zip_code, db=db)
                if coords:
                    lat, lng = coords

            if lat is not None and lng is not None:
                etiqueta.latitud = lat
                etiqueta.longitud = lng
                geocodificados += 1
                logger.info("Geocoding OK %s → (%.6f, %.6f)", etiqueta.shipping_id, lat, lng)
            else:
                sin_resultado += 1
                logger.warning("Geocoding sin resultado para %s", etiqueta.shipping_id)

        except Exception:
            logger.exception("Error geocodificando %s", etiqueta.shipping_id)
            errores += 1

    db.commit()

    return GeocodificarResponse(
        total=len(etiquetas),
        geocodificados=geocodificados,
        ya_tenian=ya_tenian,
        sin_resultado=sin_resultado,
        errores=errores,
    )


# ── Smart Polling ─────────────────────────────────────────────


class CheckUpdatesResponse(BaseModel):
    """Respuesta ligera para polling: count + timestamp del último cambio."""

    count: int = Field(description="Total de etiquetas que matchean los filtros base")
    last_updated: Optional[str] = Field(None, description="Timestamp ISO del último updated_at")

    model_config = ConfigDict(from_attributes=True)


@router.get(
    "/etiquetas-envio/check-updates",
    response_model=CheckUpdatesResponse,
    summary="Check ligero para polling — count + last_updated",
)
def check_updates(
    fecha_envio: Optional[date] = Query(None, description="Fecha de envío exacta"),
    fecha_desde: Optional[date] = Query(None, description="Desde fecha (inclusive)"),
    fecha_hasta: Optional[date] = Query(None, description="Hasta fecha (inclusive)"),
    logistica_id: Optional[int] = Query(None, description="Filtrar por logística"),
    sin_logistica: bool = Query(False, description="Solo sin logística"),
    solo_outlet: bool = Query(False, description="Solo outlet"),
    solo_turbo: bool = Query(False, description="Solo turbo"),
    pistoleado: Optional[str] = Query(None, pattern="^(si|no)$", description="Filtrar por pistoleado: si/no"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> CheckUpdatesResponse:
    """
    Endpoint ultra-ligero para smart polling (cada ~10s).

    Devuelve solo COUNT(*) + MAX(updated_at) sobre EtiquetaEnvio
    con filtros básicos (fecha, logística, outlet, turbo, pistoleado).

    NO hace JOINs pesados (ML shipping, cordón, estado ERP, search).
    Si cualquier etiqueta dentro del rango cambia, el frontend recarga.

    Filtros que requieren JOINs (cordon, sin_cordon, mlstatus, ssos_id,
    search) se omiten intencionalmente — el COUNT puede diferir del
    total visible, pero last_updated siempre detectará cambios.
    """
    _check_permiso(db, current_user, "envios_flex.ver")

    query = db.query(
        func.count(EtiquetaEnvio.shipping_id).label("count"),
        func.max(EtiquetaEnvio.updated_at).label("last_updated"),
    )

    # ── Filtros directos sobre EtiquetaEnvio (sin JOINs) ──
    if fecha_envio:
        query = query.filter(EtiquetaEnvio.fecha_envio == fecha_envio)
    else:
        if fecha_desde:
            query = query.filter(EtiquetaEnvio.fecha_envio >= fecha_desde)
        if fecha_hasta:
            query = query.filter(EtiquetaEnvio.fecha_envio <= fecha_hasta)

    if logistica_id is not None:
        query = query.filter(EtiquetaEnvio.logistica_id == logistica_id)

    if sin_logistica:
        query = query.filter(EtiquetaEnvio.logistica_id.is_(None))

    if solo_outlet:
        query = query.filter(EtiquetaEnvio.es_outlet.is_(True))

    if solo_turbo:
        query = query.filter(EtiquetaEnvio.es_turbo.is_(True))

    if pistoleado == "si":
        query = query.filter(EtiquetaEnvio.pistoleado_at.isnot(None))
    elif pistoleado == "no":
        query = query.filter(EtiquetaEnvio.pistoleado_at.is_(None))

    row = query.one()

    return CheckUpdatesResponse(
        count=row.count,
        last_updated=row.last_updated.isoformat() if row.last_updated else None,
    )


# ── Export manuales (datos editados por el usuario) ──────────


class ExportManualEnvio(BaseModel):
    """Un envío manual con datos editados para exportar."""

    numero_tracking: str = ""
    fecha_venta: str = ""
    valor_declarado: str = ""
    peso_declarado: str = ""
    destinatario: str = ""
    telefono: str = ""
    direccion: str = ""
    localidad: str = ""
    codigo_postal: str = ""
    observaciones: str = ""
    total_a_cobrar: str = ""
    logistica_inversa: str = ""

    model_config = ConfigDict(from_attributes=True)


class ExportManualesRequest(BaseModel):
    """Request body para exportar envíos manuales."""

    envios: List[ExportManualEnvio]


EXPORT_MANUALES_COLUMNS = [
    ("numero_tracking", "Numero de tracking"),
    ("fecha_venta", "Fecha de venta"),
    ("valor_declarado", "Valor declarado"),
    ("peso_declarado", "Peso declarado"),
    ("destinatario", "Destinatario"),
    ("telefono", "Teléfono de contacto"),
    ("direccion", "Dirección"),
    ("localidad", "Localidad"),
    ("codigo_postal", "Código postal"),
    ("observaciones", "Observaciones"),
    ("total_a_cobrar", "4 Total a cobrar"),
    ("logistica_inversa", "1 Logistica Inversa"),
]


@router.post(
    "/etiquetas-envio/export-manuales",
    summary="Exportar envíos manuales editados a Excel (XLSX)",
    response_class=StreamingResponse,
)
def exportar_manuales(
    body: ExportManualesRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> StreamingResponse:
    """
    Genera un XLSX con los datos de envíos manuales editados por el usuario.
    Los datos vienen pre-rellenados desde el frontend y pueden haber sido
    modificados antes de exportar. No consulta la DB.
    """
    _check_permiso(db, current_user, "envios_flex.exportar")

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Envíos Manuales"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Headers
    for col_idx, (_, label) in enumerate(EXPORT_MANUALES_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Datos
    for row_idx, envio in enumerate(body.envios, start=2):
        envio_dict = envio.model_dump()
        for col_idx, (key, _) in enumerate(EXPORT_MANUALES_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=envio_dict.get(key, ""))

    # Anchos de columna
    from openpyxl.utils import get_column_letter

    col_widths = [18, 14, 14, 14, 25, 18, 35, 20, 12, 30, 16, 16]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"envios_manuales_{date.today().isoformat()}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

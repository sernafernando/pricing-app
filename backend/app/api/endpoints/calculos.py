from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel, ConfigDict, Field
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.core.database import get_db
from app.models.calculo_pricing import CalculoPricing
from app.models.usuario import Usuario
from app.api.deps import get_current_user

router = APIRouter()


class CalculoRequest(BaseModel):
    descripcion: str = Field(..., min_length=1, max_length=500)
    ean: Optional[str] = Field(None, max_length=50)
    costo: float = Field(..., gt=0)
    moneda_costo: str = Field(..., pattern="^(ARS|USD)$")
    iva: float
    comision_ml: float = Field(..., ge=0, le=100)
    costo_envio: float = Field(default=0, ge=0)
    precio_final: float = Field(..., gt=0)
    markup_porcentaje: Optional[float] = None
    limpio: Optional[float] = None
    comision_total: Optional[float] = None
    tipo_cambio_usado: Optional[float] = None
    cantidad: int = Field(default=0, ge=0)
    precios_cuotas: Optional[dict] = None  # JSONB para cuotas


class CalculoResponse(BaseModel):
    id: int
    usuario_id: int
    descripcion: str
    ean: Optional[str]
    costo: float
    moneda_costo: str
    iva: float
    comision_ml: float
    costo_envio: float
    precio_final: float
    markup_porcentaje: Optional[float]
    limpio: Optional[float]
    comision_total: Optional[float]
    tipo_cambio_usado: Optional[float]
    cantidad: int
    precios_cuotas: Optional[dict] = None
    fecha_creacion: datetime
    fecha_modificacion: datetime

    model_config = ConfigDict(from_attributes=True)


class CantidadUpdate(BaseModel):
    cantidad: int = Field(..., ge=0)


@router.post("/calculos", response_model=CalculoResponse)
async def crear_calculo(
    calculo: CalculoRequest, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)
):
    """Crea un nuevo cálculo de pricing guardado"""

    if not calculo.descripcion and not calculo.ean:
        raise HTTPException(status_code=400, detail="Debe proporcionar al menos descripción o EAN")

    nuevo_calculo = CalculoPricing(
        usuario_id=current_user.id,
        descripcion=calculo.descripcion,
        ean=calculo.ean,
        costo=calculo.costo,
        moneda_costo=calculo.moneda_costo,
        iva=calculo.iva,
        comision_ml=calculo.comision_ml,
        costo_envio=calculo.costo_envio,
        precio_final=calculo.precio_final,
        markup_porcentaje=calculo.markup_porcentaje,
        limpio=calculo.limpio,
        comision_total=calculo.comision_total,
        tipo_cambio_usado=calculo.tipo_cambio_usado,
        cantidad=calculo.cantidad,
        precios_cuotas=calculo.precios_cuotas,  # Guardar JSONB de cuotas
    )

    db.add(nuevo_calculo)
    db.commit()
    db.refresh(nuevo_calculo)

    return nuevo_calculo


@router.get("/calculos", response_model=List[CalculoResponse])
async def listar_calculos(db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    """Lista todos los cálculos del usuario"""

    calculos = (
        db.query(CalculoPricing)
        .filter(CalculoPricing.usuario_id == current_user.id)
        .order_by(CalculoPricing.fecha_creacion.desc())
        .all()
    )

    return calculos


@router.get("/calculos/{calculo_id}", response_model=CalculoResponse)
async def obtener_calculo(
    calculo_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)
):
    """Obtiene un cálculo específico"""

    calculo = (
        db.query(CalculoPricing)
        .filter(CalculoPricing.id == calculo_id, CalculoPricing.usuario_id == current_user.id)
        .first()
    )

    if not calculo:
        raise HTTPException(status_code=404, detail="Cálculo no encontrado")

    return calculo


@router.put("/calculos/{calculo_id}", response_model=CalculoResponse)
async def actualizar_calculo(
    calculo_id: int,
    calculo: CalculoRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Actualiza un cálculo existente"""

    if not calculo.descripcion and not calculo.ean:
        raise HTTPException(status_code=400, detail="Debe proporcionar al menos descripción o EAN")

    calculo_db = (
        db.query(CalculoPricing)
        .filter(CalculoPricing.id == calculo_id, CalculoPricing.usuario_id == current_user.id)
        .first()
    )

    if not calculo_db:
        raise HTTPException(status_code=404, detail="Cálculo no encontrado")

    calculo_db.descripcion = calculo.descripcion
    calculo_db.ean = calculo.ean
    calculo_db.costo = calculo.costo
    calculo_db.moneda_costo = calculo.moneda_costo
    calculo_db.iva = calculo.iva
    calculo_db.comision_ml = calculo.comision_ml
    calculo_db.costo_envio = calculo.costo_envio
    calculo_db.precio_final = calculo.precio_final
    calculo_db.markup_porcentaje = calculo.markup_porcentaje
    calculo_db.limpio = calculo.limpio
    calculo_db.comision_total = calculo.comision_total
    calculo_db.tipo_cambio_usado = calculo.tipo_cambio_usado
    calculo_db.cantidad = calculo.cantidad
    calculo_db.precios_cuotas = calculo.precios_cuotas  # ← AGREGAR ESTO
    calculo_db.fecha_modificacion = datetime.now()

    db.commit()
    db.refresh(calculo_db)

    return calculo_db


@router.patch("/calculos/{calculo_id}/cantidad")
async def actualizar_cantidad(
    calculo_id: int,
    cantidad_data: CantidadUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Actualiza solo la cantidad de un cálculo (endpoint rápido)"""

    calculo = (
        db.query(CalculoPricing)
        .filter(CalculoPricing.id == calculo_id, CalculoPricing.usuario_id == current_user.id)
        .first()
    )

    if not calculo:
        raise HTTPException(status_code=404, detail="Cálculo no encontrado")

    calculo.cantidad = cantidad_data.cantidad
    calculo.fecha_modificacion = datetime.now()

    db.commit()
    db.refresh(calculo)

    return {"mensaje": "Cantidad actualizada", "cantidad": calculo.cantidad}


@router.delete("/calculos/{calculo_id}")
async def eliminar_calculo(
    calculo_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)
):
    """Elimina un cálculo"""

    calculo = (
        db.query(CalculoPricing)
        .filter(CalculoPricing.id == calculo_id, CalculoPricing.usuario_id == current_user.id)
        .first()
    )

    if not calculo:
        raise HTTPException(status_code=404, detail="Cálculo no encontrado")

    db.delete(calculo)
    db.commit()

    return {"mensaje": "Cálculo eliminado correctamente"}


class AccionMasivaRequest(BaseModel):
    calculo_ids: List[int]


@router.post("/calculos/acciones/eliminar-masivo")
async def eliminar_calculos_masivo(
    request: AccionMasivaRequest, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)
):
    """Elimina múltiples cálculos"""

    if not request.calculo_ids:
        raise HTTPException(status_code=400, detail="No se proporcionaron IDs")

    calculos = (
        db.query(CalculoPricing)
        .filter(CalculoPricing.id.in_(request.calculo_ids), CalculoPricing.usuario_id == current_user.id)
        .all()
    )

    count = len(calculos)

    for calculo in calculos:
        db.delete(calculo)

    db.commit()

    return {"mensaje": f"{count} cálculos eliminados", "count": count}


# ========== CALCULAR PRECIOS DE CUOTAS ==========


class CalcularCuotasRequest(BaseModel):
    costo: float = Field(..., gt=0)
    moneda_costo: str = Field(..., pattern="^(ARS|USD)$")
    iva: float
    envio: float = Field(default=0, ge=0)
    markup_objetivo: float = Field(..., description="Markup objetivo en porcentaje (ej: 15.5)")
    tipo_cambio: Optional[float] = None
    grupo_id: int = Field(default=1, description="ID del grupo de comisión (1-13)")
    adicional_markup: float = Field(default=4.0, description="Markup adicional para cuotas en puntos porcentuales")


class PrecioCuotaResponse(BaseModel):
    cuotas: int
    pricelist_id: int
    precio: float
    comision_base_pct: float
    comision_total: float
    limpio: float
    markup_real: float


@router.post("/calculos/calcular-cuotas", response_model=List[PrecioCuotaResponse])
async def calcular_precios_cuotas(
    request: CalcularCuotasRequest, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)
):
    """
    Calcula precios de cuotas (3, 6, 9, 12) usando cálculo convergente.

    Dado un markup objetivo (ej: 15%), calcula el precio necesario para cada plan de cuotas
    de manera que mantengan el mismo markup después de aplicar las comisiones específicas de cada plan.
    """
    print(f"🔍 Request recibido: {request}")
    print(f"🔍 User: {current_user.username if current_user else 'None'}")

    from app.services.pricing_calculator import calcular_precio_producto

    # Configuración de cuotas: nombre -> (cantidad_cuotas, pricelist_id)
    cuotas_config = [
        (3, 17),  # ML PREMIUM 3C
        (6, 14),  # ML PREMIUM 6C
        (9, 13),  # ML PREMIUM 9C
        (12, 23),  # ML PREMIUM 12C
    ]

    resultados = []
    errores = []

    for cuotas, pricelist_id in cuotas_config:
        try:
            resultado = calcular_precio_producto(
                db=db,
                costo=request.costo,
                moneda_costo=request.moneda_costo,
                iva=request.iva,
                envio=request.envio,
                grupo_id=request.grupo_id,  # ← Usar grupo_id directo
                pricelist_id=pricelist_id,
                markup_objetivo=request.markup_objetivo,
                tipo_cambio=request.tipo_cambio,
                adicional_markup=request.adicional_markup,
            )

            if "error" not in resultado:
                resultados.append(
                    PrecioCuotaResponse(
                        cuotas=cuotas,
                        pricelist_id=pricelist_id,
                        precio=round(resultado["precio"], 2),
                        comision_base_pct=resultado["comision_base_pct"],
                        comision_total=resultado["comision_total"],
                        limpio=resultado["limpio"],
                        markup_real=resultado["markup_real"],
                    )
                )
            else:
                errores.append(f"{cuotas}C: {resultado['error']}")
        except Exception as e:
            # Si falla el cálculo de una cuota, continuar con las demás
            error_msg = f"{cuotas}C: {str(e)}"
            print(f"❌ Error calculando cuotas {cuotas}: {e}")
            import traceback

            traceback.print_exc()
            errores.append(error_msg)
            continue

    if not resultados:
        error_detail = f"No se pudieron calcular precios de cuotas. Errores: {'; '.join(errores)}"
        print(f"❌ {error_detail}")
        raise HTTPException(status_code=400, detail=error_detail)

    return resultados


@router.patch("/calculos/{calculo_id}/cuotas")
async def actualizar_cuotas_calculo(
    calculo_id: int,
    precios_cuotas: dict,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Actualiza los precios de cuotas de un cálculo guardado"""
    calculo = (
        db.query(CalculoPricing)
        .filter(CalculoPricing.id == calculo_id, CalculoPricing.usuario_id == current_user.id)
        .first()
    )

    if not calculo:
        raise HTTPException(404, "Cálculo no encontrado")

    # Actualizar campo JSONB
    calculo.precios_cuotas = precios_cuotas.get("precios_cuotas")
    db.commit()

    return {"mensaje": "Cuotas actualizadas correctamente"}


@router.get("/calculos/exportar/excel")
async def exportar_calculos_excel(
    filtro: Optional[str] = None,  # 'todos', 'con_cantidad', 'seleccionados'
    ids: Optional[str] = None,  # comma-separated IDs for 'seleccionados'
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """
    Exporta cálculos a Excel con multi-sheet:
    - Sheet 1: Resumen (precio de lista)
    - Sheet 2: Cuotas (detalle de 3, 6, 9, 12 cuotas)
    """

    query = db.query(CalculoPricing).filter(CalculoPricing.usuario_id == current_user.id)

    # Aplicar filtros
    if filtro == "con_cantidad":
        query = query.filter(CalculoPricing.cantidad > 0)
    elif filtro == "seleccionados" and ids:
        id_list = [int(id.strip()) for id in ids.split(",") if id.strip()]
        query = query.filter(CalculoPricing.id.in_(id_list))

    calculos = query.order_by(CalculoPricing.fecha_creacion.desc()).all()

    # Crear workbook de Excel
    wb = Workbook()

    # ========== SHEET 1: RESUMEN ==========
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Headers Sheet Resumen
    headers_resumen = [
        "ID",
        "Fecha",
        "Descripción",
        "EAN",
        "Cantidad",
        "Costo",
        "Moneda",
        "IVA %",
        "Precio Lista",
        "Markup %",
        "Limpio",
        "Tiene Cuotas",
    ]

    for col, header in enumerate(headers_resumen, start=1):
        cell = ws_resumen.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Datos Resumen
    for row_idx, calc in enumerate(calculos, start=2):
        ws_resumen.cell(row=row_idx, column=1, value=calc.id)
        ws_resumen.cell(row=row_idx, column=2, value=calc.fecha_creacion.strftime("%d/%m/%Y"))
        ws_resumen.cell(row=row_idx, column=3, value=calc.descripcion)
        ws_resumen.cell(row=row_idx, column=4, value=calc.ean or "")
        ws_resumen.cell(row=row_idx, column=5, value=calc.cantidad)
        ws_resumen.cell(row=row_idx, column=6, value=float(calc.costo))
        ws_resumen.cell(row=row_idx, column=7, value=calc.moneda_costo)
        ws_resumen.cell(row=row_idx, column=8, value=float(calc.iva))
        ws_resumen.cell(row=row_idx, column=9, value=float(calc.precio_final))
        ws_resumen.cell(row=row_idx, column=10, value=float(calc.markup_porcentaje) if calc.markup_porcentaje else "")
        ws_resumen.cell(row=row_idx, column=11, value=float(calc.limpio) if calc.limpio else "")
        ws_resumen.cell(row=row_idx, column=12, value="Sí" if calc.precios_cuotas else "No")

    # Anchos de columna Resumen
    ws_resumen.column_dimensions["A"].width = 8
    ws_resumen.column_dimensions["B"].width = 12
    ws_resumen.column_dimensions["C"].width = 40
    ws_resumen.column_dimensions["D"].width = 15
    ws_resumen.column_dimensions["E"].width = 10
    ws_resumen.column_dimensions["F"].width = 12
    ws_resumen.column_dimensions["G"].width = 8
    ws_resumen.column_dimensions["H"].width = 8
    ws_resumen.column_dimensions["I"].width = 15
    ws_resumen.column_dimensions["J"].width = 12
    ws_resumen.column_dimensions["K"].width = 15
    ws_resumen.column_dimensions["L"].width = 12

    # ========== SHEET 2: CUOTAS ==========
    ws_cuotas = wb.create_sheet("Cuotas")

    # Headers Sheet Cuotas
    headers_cuotas = [
        "Calc ID",
        "Descripción",
        "Plan",
        "Precio",
        "Comisión %",
        "Comisión Total",
        "Limpio",
        "Markup Real %",
        "Adicional %",
    ]

    for col, header in enumerate(headers_cuotas, start=1):
        cell = ws_cuotas.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Datos Cuotas
    row_idx_cuotas = 2
    for calc in calculos:
        if calc.precios_cuotas and "cuotas" in calc.precios_cuotas:
            adicional = calc.precios_cuotas.get("adicional_markup", 4.0)

            for cuota_data in calc.precios_cuotas["cuotas"]:
                ws_cuotas.cell(row=row_idx_cuotas, column=1, value=calc.id)
                ws_cuotas.cell(row=row_idx_cuotas, column=2, value=calc.descripcion)
                ws_cuotas.cell(row=row_idx_cuotas, column=3, value=f"{cuota_data['cuotas']}C")
                ws_cuotas.cell(row=row_idx_cuotas, column=4, value=float(cuota_data["precio"]))
                ws_cuotas.cell(row=row_idx_cuotas, column=5, value=float(cuota_data["comision_base_pct"]))
                ws_cuotas.cell(row=row_idx_cuotas, column=6, value=float(cuota_data["comision_total"]))
                ws_cuotas.cell(row=row_idx_cuotas, column=7, value=float(cuota_data["limpio"]))
                ws_cuotas.cell(row=row_idx_cuotas, column=8, value=float(cuota_data["markup_real"]))
                ws_cuotas.cell(row=row_idx_cuotas, column=9, value=float(adicional))
                row_idx_cuotas += 1

    # Anchos de columna Cuotas
    ws_cuotas.column_dimensions["A"].width = 10
    ws_cuotas.column_dimensions["B"].width = 40
    ws_cuotas.column_dimensions["C"].width = 8
    ws_cuotas.column_dimensions["D"].width = 15
    ws_cuotas.column_dimensions["E"].width = 12
    ws_cuotas.column_dimensions["F"].width = 15
    ws_cuotas.column_dimensions["G"].width = 15
    ws_cuotas.column_dimensions["H"].width = 15
    ws_cuotas.column_dimensions["I"].width = 12

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Nombre de archivo según filtro
    filename = f"calculos_pricing_{filtro or 'todos'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

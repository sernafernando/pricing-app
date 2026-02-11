"""
Endpoints para informes de cuentas corrientes (proveedores y clientes).
- Proveedores: export 26 del ERP
- Clientes: export 29 del ERP

Flujo:
  POST /sync  → llama al ERP vía gbp-parser, truncate-and-reload en tabla local
  GET  /proveedores | /clientes → solo lee la tabla local (con filtros)
  GET  /exportar → genera XLSX desde la tabla local (con filtros)
"""

from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func as sa_func
from sqlalchemy.orm import Session
from typing import Optional
import httpx

from app.core.database import get_db
from app.core.exceptions import api_error, ErrorCode
from app.core.logging import get_logger
from app.api.deps import get_current_user
from app.models.usuario import Usuario
from app.models.cuenta_corriente_proveedor import CuentaCorrienteProveedor
from app.models.cuenta_corriente_cliente import CuentaCorrienteCliente
from app.models.tb_branch import TBBranch

router = APIRouter()
logger = get_logger(__name__)

# URL interna del gbp-parser (mismo servidor)
GBP_PARSER_URL = "http://localhost:8002/api/gbp-parser"

# Mapeo tipo → (modelo, export_id, campo_id_erp, campo_nombre_erp)
_TIPO_CONFIG = {
    "proveedores": {
        "model": CuentaCorrienteProveedor,
        "export_id": 26,
        "field_map": {
            "bra_id": "BraID",
            "id_proveedor": "ID_Proveedor",
            "proveedor": "Proveedor",
        },
        "name_col": CuentaCorrienteProveedor.proveedor,
        "id_label": "ID Proveedor",
        "name_label": "Proveedor",
    },
    "clientes": {
        "model": CuentaCorrienteCliente,
        "export_id": 29,
        "field_map": {
            "bra_id": "BraID",
            "id_cliente": "ID_Cliente",
            "cliente": "Cliente",
        },
        "name_col": CuentaCorrienteCliente.cliente,
        "id_label": "ID Cliente",
        "name_label": "Cliente",
    },
}


def _parse_decimal(value: str) -> float:
    """Convierte string a float, tolerando vacíos y formatos raros del ERP."""
    if not value or not str(value).strip():
        return 0.0
    try:
        return float(str(value).strip().replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


async def _fetch_from_erp(export_id: int) -> list[dict]:
    """Obtiene datos del ERP llamando al gbp-parser con el export indicado."""
    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.get(GBP_PARSER_URL, params={"intExpgr_id": export_id})

    if response.status_code != 200:
        raise api_error(
            502,
            ErrorCode.INTERNAL_ERROR,
            f"Error del gbp-parser: {response.status_code} - {response.text[:200]}",
        )

    data = response.json()

    if not isinstance(data, list):
        raise api_error(502, ErrorCode.INTERNAL_ERROR, "Respuesta inesperada del ERP")

    return data


def _build_branch_map(db: Session) -> dict[int, str]:
    """Construye un diccionario bra_id -> bra_desc de sucursales activas."""
    branches = (
        db.query(TBBranch.bra_id, TBBranch.bra_desc)
        .filter(TBBranch.bra_disabled == False)  # noqa: E712
        .all()
    )
    return {b.bra_id: b.bra_desc or f"Sucursal {b.bra_id}" for b in branches}


def _serialize_results(
    resultados: list,
    tipo: str,
    branch_map: dict[int, str],
) -> list[dict]:
    """Serializa los resultados de la query a dicts para la respuesta JSON."""
    data = []
    for r in resultados:
        item = {
            "id": r.id,
            "bra_id": r.bra_id,
            "sucursal": branch_map.get(r.bra_id, f"Sucursal {r.bra_id}"),
            "monto_total": float(r.monto_total) if r.monto_total else 0,
            "monto_abonado": float(r.monto_abonado) if r.monto_abonado else 0,
            "pendiente": float(r.pendiente) if r.pendiente else 0,
        }
        if tipo == "proveedores":
            item["id_proveedor"] = r.id_proveedor
            item["proveedor"] = r.proveedor
        else:
            item["id_cliente"] = r.id_cliente
            item["cliente"] = r.cliente
        data.append(item)
    return data


def _get_synced_at(db: Session, tipo: str) -> Optional[str]:
    """Obtiene el timestamp de la última sincronización para un tipo dado."""
    cfg = _TIPO_CONFIG[tipo]
    model = cfg["model"]
    result = db.query(sa_func.max(model.synced_at)).scalar()
    return result.isoformat() if result else None


def _query_filtered(
    db: Session,
    tipo: str,
    buscar: Optional[str],
    sucursal: Optional[int],
):
    """Arma la query filtrada sobre la tabla local."""
    cfg = _TIPO_CONFIG[tipo]
    model = cfg["model"]
    name_col = cfg["name_col"]

    query = db.query(model)
    if buscar:
        query = query.filter(name_col.ilike(f"%{buscar}%"))
    if sucursal is not None:
        query = query.filter(model.bra_id == sucursal)
    return query.order_by(name_col).all()


# ---------------------------------------------------------------------------
# Sucursales disponibles (para el filtro del frontend)
# ---------------------------------------------------------------------------


@router.get("/cuentas-corrientes/sucursales")
async def listar_sucursales_cc(
    db: Session = Depends(get_db),
    _user: Usuario = Depends(get_current_user),
) -> list[dict]:
    """Retorna las sucursales activas para el filtro de cuentas corrientes."""
    sucursales = (
        db.query(TBBranch.bra_id, TBBranch.bra_desc)
        .filter(TBBranch.bra_disabled == False)  # noqa: E712
        .order_by(TBBranch.bra_desc)
        .all()
    )
    return [{"bra_id": s.bra_id, "bra_desc": s.bra_desc} for s in sucursales]


# ---------------------------------------------------------------------------
# Sincronizar desde ERP (POST — solo cuando el usuario pide "Actualizar")
# ---------------------------------------------------------------------------


@router.post("/cuentas-corrientes/sync")
async def sincronizar_cuentas_corrientes(
    tipo: str = Query(..., description="'proveedores' o 'clientes'"),
    db: Session = Depends(get_db),
    _user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Llama al ERP vía gbp-parser y hace truncate-and-reload en la tabla local.
    Si falla y hay datos previos, no tira error (los datos viejos quedan).
    """
    if tipo not in _TIPO_CONFIG:
        raise api_error(422, ErrorCode.VALIDATION_ERROR, "tipo debe ser 'proveedores' o 'clientes'")

    cfg = _TIPO_CONFIG[tipo]
    model = cfg["model"]
    export_id = cfg["export_id"]
    field_map = cfg["field_map"]

    try:
        rows = await _fetch_from_erp(export_id)

        db.query(model).delete()
        db.flush()

        registros = []
        for row in rows:
            kwargs = {
                "monto_total": _parse_decimal(row.get("Monto_Total", "0")),
                "monto_abonado": _parse_decimal(row.get("Monto_Abonado", "0")),
                "pendiente": _parse_decimal(row.get("Pendiente", "0")),
            }

            for model_field, erp_field in field_map.items():
                raw = row.get(erp_field, "")
                if model_field == "bra_id" or model_field.startswith("id_"):
                    kwargs[model_field] = int(raw or 0)
                else:
                    kwargs[model_field] = str(raw).strip()

            registros.append(model(**kwargs))

        db.bulk_save_objects(registros)
        db.commit()
        logger.info("CC %s sincronizadas: %d registros", tipo, len(registros))

        return {"ok": True, "registros": len(registros)}

    except Exception as e:
        db.rollback()
        logger.error("Error sincronizando CC %s: %s", tipo, e, exc_info=True)

        cached = db.query(model).count()
        if cached == 0:
            raise api_error(
                502,
                ErrorCode.INTERNAL_ERROR,
                f"Error al consultar ERP y no hay datos previos: {str(e)}",
            )
        logger.warning("Sync CC %s falló, hay %d registros previos en cache", tipo, cached)
        return {"ok": False, "registros": cached, "cached": True}


# ---------------------------------------------------------------------------
# Listar (GET — solo lee la tabla local con filtros)
# ---------------------------------------------------------------------------


@router.get("/cuentas-corrientes/proveedores")
async def listar_cuentas_corrientes_proveedores(
    db: Session = Depends(get_db),
    _user: Usuario = Depends(get_current_user),
    buscar: Optional[str] = Query(None, description="Buscar por nombre de proveedor"),
    sucursal: Optional[int] = Query(None, description="Filtrar por bra_id de sucursal"),
) -> dict:
    """Lee cuentas corrientes de proveedores desde la tabla local (sin llamar al ERP)."""
    branch_map = _build_branch_map(db)
    resultados = _query_filtered(db, "proveedores", buscar, sucursal)
    return {
        "total": len(resultados),
        "synced_at": _get_synced_at(db, "proveedores"),
        "data": _serialize_results(resultados, "proveedores", branch_map),
    }


@router.get("/cuentas-corrientes/clientes")
async def listar_cuentas_corrientes_clientes(
    db: Session = Depends(get_db),
    _user: Usuario = Depends(get_current_user),
    buscar: Optional[str] = Query(None, description="Buscar por nombre de cliente"),
    sucursal: Optional[int] = Query(None, description="Filtrar por bra_id de sucursal"),
) -> dict:
    """Lee cuentas corrientes de clientes desde la tabla local (sin llamar al ERP)."""
    branch_map = _build_branch_map(db)
    resultados = _query_filtered(db, "clientes", buscar, sucursal)
    return {
        "total": len(resultados),
        "synced_at": _get_synced_at(db, "clientes"),
        "data": _serialize_results(resultados, "clientes", branch_map),
    }


# ---------------------------------------------------------------------------
# Exportar XLSX (con filtros aplicados, desde tabla local)
# ---------------------------------------------------------------------------


@router.get("/cuentas-corrientes/exportar")
async def exportar_cuentas_corrientes(
    tipo: str = Query(..., description="'proveedores' o 'clientes'"),
    db: Session = Depends(get_db),
    _user: Usuario = Depends(get_current_user),
    buscar: Optional[str] = Query(None, description="Buscar por nombre"),
    sucursal: Optional[int] = Query(None, description="Filtrar por bra_id"),
) -> StreamingResponse:
    """Exporta las cuentas corrientes (con filtros) a XLSX desde la tabla local."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    if tipo not in _TIPO_CONFIG:
        raise api_error(422, ErrorCode.VALIDATION_ERROR, "tipo debe ser 'proveedores' o 'clientes'")

    cfg = _TIPO_CONFIG[tipo]
    branch_map = _build_branch_map(db)
    resultados = _query_filtered(db, tipo, buscar, sucursal)

    id_field_label = cfg["id_label"]
    name_label = cfg["name_label"]

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"CC {tipo.capitalize()}"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    money_fmt = "#,##0.00"

    # Headers
    headers = ["Sucursal", id_field_label, name_label, "Monto Total", "Abonado", "Pendiente"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data
    for row_idx, r in enumerate(resultados, 2):
        sucursal_nombre = branch_map.get(r.bra_id, f"Sucursal {r.bra_id}")
        id_valor = r.id_proveedor if tipo == "proveedores" else r.id_cliente
        nombre_valor = r.proveedor if tipo == "proveedores" else r.cliente

        ws.cell(row=row_idx, column=1, value=sucursal_nombre)
        ws.cell(row=row_idx, column=2, value=id_valor)
        ws.cell(row=row_idx, column=3, value=nombre_valor)

        cell_total = ws.cell(row=row_idx, column=4, value=float(r.monto_total or 0))
        cell_total.number_format = money_fmt

        cell_abonado = ws.cell(row=row_idx, column=5, value=float(r.monto_abonado or 0))
        cell_abonado.number_format = money_fmt

        cell_pendiente = ws.cell(row=row_idx, column=6, value=float(r.pendiente or 0))
        cell_pendiente.number_format = money_fmt

    # Fila de totales
    total_row = len(resultados) + 2
    ws.cell(row=total_row, column=3, value="TOTALES").font = Font(bold=True)

    for col_idx in (4, 5, 6):
        col_letter = chr(64 + col_idx)  # D, E, F
        cell = ws.cell(
            row=total_row,
            column=col_idx,
            value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})",
        )
        cell.number_format = money_fmt
        cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    # Escribir a buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"cuentas_corrientes_{tipo}_{timestamp}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

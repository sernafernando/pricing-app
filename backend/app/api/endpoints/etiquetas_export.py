"""
Endpoints de exportación de etiquetas de envío.

Incluye:
- GET /etiquetas-envio/export (XLSX general)
- POST /etiquetas-envio/export-manuales (XLS Lightdata)
"""

import io
from datetime import date
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, case, cast, func, Numeric, or_
from sqlalchemy.orm import Session, aliased

from app.core.database import get_db
from app.api.deps import get_current_user
from app.models.usuario import Usuario
from app.models.etiqueta_envio import EtiquetaEnvio
from app.models.logistica import Logistica
from app.models.codigo_postal_cordon import CodigoPostalCordon
from app.models.sale_order_status import SaleOrderStatus
from app.models.operador import Operador
from app.models.logistica_costo_cordon import LogisticaCostoCordon

from app.api.endpoints.etiquetas_shared import (
    _check_permiso,
    _build_costo_case,
    _get_lluvia_config,
    _soh_status_subquery,
    _manual_soh_status_subquery,
    _facturado_ml_subquery,
    _facturado_manual_subquery,
    _shipping_dedup_subquery,
    EXPORT_COLUMNS,
    EXPORT_MANUALES_COLUMNS,
    ExportManualesRequest,
)

router = APIRouter()


@router.get(
    "/etiquetas-envio/export",
    summary="Exportar etiquetas a Excel (XLSX)",
    response_class=StreamingResponse,
)
def exportar_etiquetas(
    fecha_desde: date = Query(..., description="Desde fecha (inclusive)"),
    fecha_hasta: date = Query(..., description="Hasta fecha (inclusive)"),
    columnas: Optional[str] = Query(
        None,
        description="Columnas a incluir (comma-separated). Si no se especifica, todas.",
    ),
    cordon: Optional[str] = Query(None, description="Filtrar por cordón"),
    logistica_id: Optional[int] = Query(None, description="Filtrar por logística"),
    sin_logistica: bool = Query(False, description="Solo sin logística asignada"),
    mlstatus: Optional[str] = Query(None, description="Filtrar por estado ML"),
    solo_outlet: bool = Query(False, description="Solo etiquetas de productos outlet"),
    solo_turbo: bool = Query(False, description="Solo etiquetas de envíos turbo"),
    search: Optional[str] = Query(None, description="Buscar por shipping_id, destinatario o dirección"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> StreamingResponse:
    """
    Exporta etiquetas filtradas a un archivo Excel (.xlsx).
    Soporta selección de columnas y todos los filtros de la vista.
    """
    _check_permiso(db, current_user, "envios_flex.exportar")

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    # Validar columnas solicitadas
    if columnas:
        cols_solicitadas = [c.strip() for c in columnas.split(",") if c.strip()]
        invalidas = [c for c in cols_solicitadas if c not in EXPORT_COLUMNS]
        if invalidas:
            raise HTTPException(400, f"Columnas inválidas: {', '.join(invalidas)}")
    else:
        cols_solicitadas = list(EXPORT_COLUMNS.keys())

    # Pre-filtrar shipping_ids por rango de fechas (performance)
    ids_fecha_exp = (
        db.query(EtiquetaEnvio.shipping_id)
        .filter(
            EtiquetaEnvio.fecha_envio >= fecha_desde,
            EtiquetaEnvio.fecha_envio <= fecha_hasta,
        )
        .scalar_subquery()
    )

    # Reusar subquery deduplicada de estado ERP
    soh_sub = _soh_status_subquery(db, shipping_ids_sub=ids_fecha_exp)
    manual_soh_sub = _manual_soh_status_subquery(db, shipping_ids_sub=ids_fecha_exp)
    facturado_ml_exp = _facturado_ml_subquery(db, shipping_ids_sub=ids_fecha_exp)
    facturado_manual_exp = _facturado_manual_subquery(db, shipping_ids_sub=ids_fecha_exp)
    ManualSaleOrderStatus = aliased(SaleOrderStatus)

    hoy_export = date.today()
    max_costo_exp = (
        db.query(
            LogisticaCostoCordon.logistica_id.label("costo_logistica_id"),
            LogisticaCostoCordon.cordon.label("costo_cordon"),
            func.max(LogisticaCostoCordon.id).label("max_id"),
        )
        .filter(LogisticaCostoCordon.vigente_desde <= hoy_export)
        .group_by(LogisticaCostoCordon.logistica_id, LogisticaCostoCordon.cordon)
        .subquery()
    )

    costo_exp = (
        db.query(
            LogisticaCostoCordon.logistica_id.label("costo_log_id"),
            LogisticaCostoCordon.cordon.label("costo_cordon_val"),
            LogisticaCostoCordon.costo.label("costo_valor"),
            LogisticaCostoCordon.costo_turbo.label("costo_turbo_valor"),
        )
        .join(
            max_costo_exp,
            LogisticaCostoCordon.id == max_costo_exp.c.max_id,
        )
        .subquery()
    )

    cordon_norm_exp = func.replace(CodigoPostalCordon.cordon, "ó", "o")

    # Lluvia offset config
    lluvia_tipo_e, lluvia_valor_e = _get_lluvia_config(db)

    # Subquery deduplicada: una fila por mlshippingid (evita duplicados por items)
    shipping_exp = _shipping_dedup_subquery(db, shipping_ids_sub=ids_fecha_exp)

    # COALESCE para envíos manuales en export
    exp_receiver = func.coalesce(EtiquetaEnvio.manual_receiver_name, shipping_exp.c.mlreceiver_name)
    exp_street = func.coalesce(EtiquetaEnvio.manual_street_name, shipping_exp.c.mlstreet_name)
    exp_street_num = func.coalesce(EtiquetaEnvio.manual_street_number, shipping_exp.c.mlstreet_number)
    exp_zip = func.coalesce(EtiquetaEnvio.manual_zip_code, shipping_exp.c.mlzip_code)
    exp_city = func.coalesce(EtiquetaEnvio.manual_city_name, shipping_exp.c.mlcity_name)
    exp_status = func.coalesce(EtiquetaEnvio.manual_status, shipping_exp.c.mlstatus)

    query = (
        db.query(
            EtiquetaEnvio.shipping_id,
            EtiquetaEnvio.fecha_envio,
            EtiquetaEnvio.pistoleado_at,
            EtiquetaEnvio.pistoleado_caja,
            Operador.nombre.label("pistoleado_operador_nombre"),
            Logistica.nombre.label("logistica_nombre"),
            exp_receiver.label("mlreceiver_name"),
            exp_street.label("mlstreet_name"),
            exp_street_num.label("mlstreet_number"),
            exp_zip.label("mlzip_code"),
            exp_city.label("mlcity_name"),
            exp_status.label("mlstatus"),
            CodigoPostalCordon.cordon,
            case(
                (
                    func.coalesce(SaleOrderStatus.ssos_name, ManualSaleOrderStatus.ssos_name).isnot(None),
                    func.coalesce(SaleOrderStatus.ssos_name, ManualSaleOrderStatus.ssos_name),
                ),
                (
                    or_(
                        facturado_ml_exp.c.shipping_id_str.isnot(None),
                        facturado_manual_exp.c.shipping_id_str.isnot(None),
                    ),
                    "Facturado",
                ),
                else_=None,
            ).label("ssos_name"),
            func.coalesce(
                cast(EtiquetaEnvio.costo_override, Numeric(12, 2)),
                _build_costo_case(
                    costo_exp.c.costo_turbo_valor,
                    costo_exp.c.costo_valor,
                    lluvia_tipo_e,
                    lluvia_valor_e,
                ),
            ).label("costo_envio"),
            EtiquetaEnvio.es_turbo,
            EtiquetaEnvio.es_lluvia,
            EtiquetaEnvio.flag_envio,
            EtiquetaEnvio.flag_envio_motivo,
        )
        .outerjoin(Logistica, EtiquetaEnvio.logistica_id == Logistica.id)
        .outerjoin(
            shipping_exp,
            EtiquetaEnvio.shipping_id == shipping_exp.c.mlshippingid,
        )
        .outerjoin(
            CodigoPostalCordon,
            exp_zip == CodigoPostalCordon.codigo_postal,
        )
        .outerjoin(soh_sub, soh_sub.c.shipping_id_str == EtiquetaEnvio.shipping_id)
        .outerjoin(SaleOrderStatus, soh_sub.c.soh_ssos_id == SaleOrderStatus.ssos_id)
        .outerjoin(manual_soh_sub, manual_soh_sub.c.shipping_id_str == EtiquetaEnvio.shipping_id)
        .outerjoin(ManualSaleOrderStatus, manual_soh_sub.c.manual_ssos_id == ManualSaleOrderStatus.ssos_id)
        .outerjoin(facturado_ml_exp, facturado_ml_exp.c.shipping_id_str == EtiquetaEnvio.shipping_id)
        .outerjoin(facturado_manual_exp, facturado_manual_exp.c.shipping_id_str == EtiquetaEnvio.shipping_id)
        .outerjoin(Operador, EtiquetaEnvio.pistoleado_operador_id == Operador.id)
        .outerjoin(
            costo_exp,
            and_(
                costo_exp.c.costo_log_id == EtiquetaEnvio.logistica_id,
                costo_exp.c.costo_cordon_val == cordon_norm_exp,
            ),
        )
    )

    # Filtros
    query = query.filter(
        EtiquetaEnvio.fecha_envio >= fecha_desde,
        EtiquetaEnvio.fecha_envio <= fecha_hasta,
    )

    if cordon:
        query = query.filter(CodigoPostalCordon.cordon == cordon)
    if logistica_id is not None:
        query = query.filter(EtiquetaEnvio.logistica_id == logistica_id)
    if sin_logistica:
        query = query.filter(EtiquetaEnvio.logistica_id.is_(None))
    if solo_outlet:
        query = query.filter(EtiquetaEnvio.es_outlet.is_(True))
    if solo_turbo:
        query = query.filter(EtiquetaEnvio.es_turbo.is_(True))
    if mlstatus:
        query = query.filter(exp_status == mlstatus)
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            (EtiquetaEnvio.shipping_id.ilike(search_term))
            | (exp_receiver.ilike(search_term))
            | (exp_street.ilike(search_term))
            | (exp_city.ilike(search_term))
        )

    query = query.order_by(EtiquetaEnvio.fecha_envio, EtiquetaEnvio.shipping_id.desc())
    rows = query.all()

    # ── Generar XLSX ──────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Envíos Flex"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Headers
    headers = [EXPORT_COLUMNS[c] for c in cols_solicitadas]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Mapeo de columna → valor
    def get_cell_value(col_key: str, row: object) -> object:
        """Extrae el valor de una columna para una fila de resultados."""
        if col_key == "shipping_id":
            return row.shipping_id
        elif col_key == "fecha_envio":
            return row.fecha_envio
        elif col_key == "destinatario":
            return row.mlreceiver_name or ""
        elif col_key == "direccion":
            parts = [row.mlstreet_name or "", row.mlstreet_number or ""]
            return " ".join(p for p in parts if p).strip()
        elif col_key == "cp":
            return row.mlzip_code or ""
        elif col_key == "localidad":
            return row.mlcity_name or ""
        elif col_key == "cordon":
            return row.cordon or ""
        elif col_key == "logistica":
            return row.logistica_nombre or ""
        elif col_key == "costo_envio":
            return float(row.costo_envio) if row.costo_envio is not None else ""
        elif col_key == "estado_ml":
            return row.mlstatus or ""
        elif col_key == "estado_erp":
            return row.ssos_name or ""
        elif col_key == "pistoleado":
            if row.pistoleado_at:
                ts = str(row.pistoleado_at)[:16]  # YYYY-MM-DD HH:MM
                operador = row.pistoleado_operador_nombre or ""
                return f"{ts} — {operador}" if operador else ts
            return ""
        elif col_key == "caja":
            return row.pistoleado_caja or ""
        elif col_key == "turbo":
            return "Turbo" if row.es_turbo else ""
        elif col_key == "lluvia":
            return "Lluvia" if row.es_lluvia else ""
        elif col_key == "flag_envio":
            labels = {
                "mal_pasado": "Mal pasado",
                "envio_cancelado": "Cancelado",
                "duplicado": "Duplicado",
                "otro": "Otro",
            }
            return labels.get(row.flag_envio, row.flag_envio or "")
        elif col_key == "flag_envio_motivo":
            return row.flag_envio_motivo or ""
        return ""

    # Datos
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_key in enumerate(cols_solicitadas, start=1):
            ws.cell(row=row_idx, column=col_idx, value=get_cell_value(col_key, row))

    # Anchos automáticos (estimados)
    col_widths = {
        "shipping_id": 16,
        "fecha_envio": 14,
        "destinatario": 25,
        "direccion": 35,
        "cp": 8,
        "localidad": 20,
        "cordon": 12,
        "logistica": 18,
        "costo_envio": 14,
        "estado_ml": 18,
        "estado_erp": 18,
        "pistoleado": 22,
        "caja": 14,
        "turbo": 8,
        "lluvia": 8,
        "flag_envio": 14,
        "flag_envio_motivo": 30,
    }
    for col_idx, col_key in enumerate(cols_solicitadas, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = col_widths.get(col_key, 15)

    # Generar archivo en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"envios_flex_{fecha_desde}_{fecha_hasta}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post(
    "/etiquetas-envio/export-manuales",
    summary="Exportar envíos manuales editados a Excel (.xls) para Lightdata",
    response_class=StreamingResponse,
)
def exportar_manuales(
    body: ExportManualesRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> StreamingResponse:
    """
    Genera un .xls (BIFF8) con los datos de envíos manuales editados por el
    usuario.  El formato replica exactamente el template de Lightdata:
    - Formato .xls (no .xlsx)
    - Sheet llamada "Simple"
    - Columna B (Fecha de venta) con formato Text (@)
    - Headers sin negrita, sin fondo, Calibri 11
    - 500 filas pre-formateadas (columna B con formato @)
    """
    _check_permiso(db, current_user, "envios_flex.exportar")

    import xlwt

    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Simple")

    # -- Estilos que matchean el template de Lightdata --
    # Header: Calibri 11, sin bold, sin fondo, negro
    header_style = xlwt.XFStyle()
    header_font = xlwt.Font()
    header_font.name = "Calibri"
    header_font.height = 220  # 11pt * 20
    header_font.bold = False
    header_style.font = header_font

    # Estilo normal para celdas de texto
    text_style = xlwt.XFStyle()
    text_font = xlwt.Font()
    text_font.name = "Calibri"
    text_font.height = 220
    text_style.font = text_font

    # Estilo para columna B (Fecha de venta): formato Text (@)
    date_text_style = xlwt.XFStyle()
    date_text_font = xlwt.Font()
    date_text_font.name = "Calibri"
    date_text_font.height = 220
    date_text_style.font = date_text_font
    date_text_style.num_format_str = "@"

    # Estilo para header de columna B (también @)
    header_date_style = xlwt.XFStyle()
    header_date_font = xlwt.Font()
    header_date_font.name = "Calibri"
    header_date_font.height = 220
    header_date_font.bold = False
    header_date_style.font = header_date_font
    header_date_style.num_format_str = "@"

    # -- Headers (row 0) --
    for col_idx, (_, label) in enumerate(EXPORT_MANUALES_COLUMNS):
        if col_idx == 1:  # Columna B: Fecha de venta
            ws.write(0, col_idx, label, header_date_style)
        else:
            ws.write(0, col_idx, label, header_style)

    # -- Datos --
    # IMPORTANTE: Lightdata valida que TODAS las celdas tengan formato
    # explícito.  Escribir siempre, incluso vacías (string vacío "").
    for row_idx, envio in enumerate(body.envios, start=1):
        envio_dict = envio.model_dump()
        for col_idx, (key, _) in enumerate(EXPORT_MANUALES_COLUMNS):
            raw = envio_dict.get(key, "") or ""

            # Columna B (Fecha de venta): formato Text (@), dd/mm/yyyy
            if col_idx == 1:
                fecha_str = str(raw)
                # Convertir ISO (yyyy-mm-dd) a dd/mm/yyyy para Lightdata
                if fecha_str and len(fecha_str) == 10 and "-" in fecha_str:
                    try:
                        dt = date.fromisoformat(fecha_str)
                        fecha_str = dt.strftime("%d/%m/%Y")
                    except (ValueError, TypeError):
                        pass
                ws.write(row_idx, col_idx, fecha_str, date_text_style)
                continue

            # Observaciones: Lightdata rechaza celda vacía, poner "-"
            if key == "observaciones" and not raw:
                raw = "-"

            # Todas las demás columnas: texto con formato General
            ws.write(row_idx, col_idx, str(raw), text_style)

    # Pre-formatear filas vacías restantes (columna B con @) hasta 500 filas
    # como hace el template de Lightdata
    data_rows = len(body.envios)
    for row_idx in range(data_rows + 1, 500):
        ws.write(row_idx, 1, "", date_text_style)

    # -- Anchos de columna (en unidades de 1/256 del ancho de un carácter) --
    col_widths_chars = [18, 14, 14, 14, 25, 18, 35, 20, 12, 30, 16, 16]
    for col_idx, w in enumerate(col_widths_chars):
        ws.col(col_idx).width = w * 256

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"envios_manuales_{date.today().isoformat()}.xls"

    return StreamingResponse(
        output,
        media_type="application/vnd.ms-excel",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

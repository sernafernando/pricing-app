from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import or_, func, and_, select, tuple_
from sqlalchemy.sql.elements import ColumnElement
from typing import Optional
from app.core.database import get_db
from app.models.producto import ProductoERP, ProductoPricing
from app.models.usuario import Usuario
from app.api.deps import get_current_user
from fastapi.responses import Response
import logging

from app.api.endpoints.productos_shared import (  # noqa: F401
    ExportRebateRequest,
    batch_colores,
    color_slot,
    filtro_colores,
    join_color_layer,
    resolver_layer_activo,
)

logger = logging.getLogger(__name__)

router = APIRouter()


def _cargar_precios_mlwebhook(mla_ids: list[str]) -> dict[str, float]:
    """
    Consulta ml_previews (DB webhook) para obtener el precio publicado en ML
    de un lote de MLAs. Retorna {mla_id: precio}. Si la DB no está disponible,
    retorna dict vacío sin romper el export.
    """
    from sqlalchemy import text
    from app.core.database import get_mlwebhook_engine

    result: dict[str, float] = {}
    if not mla_ids:
        return result
    try:
        engine = get_mlwebhook_engine()
        resources = [f"/items/{mla}" for mla in mla_ids]
        with engine.connect() as conn:
            rows = conn.execute(
                text("""
                    SELECT DISTINCT ON (resource) resource, price
                    FROM ml_previews
                    WHERE resource = ANY(:resources)
                      AND price IS NOT NULL
                    ORDER BY resource, last_updated DESC
                """),
                {"resources": resources},
            ).fetchall()
        for row in rows:
            mla = row.resource.replace("/items/", "")
            if row.price:
                result[mla] = float(row.price)
    except Exception as e:
        logger.warning("No se pudo consultar ml_previews para precios: %s", e)
    return result


def _parsear_tiendas_oficiales_mla(csv: Optional[str]) -> Optional[tuple[list[int], bool]]:
    """
    Parsea CSV de tiendas oficiales para filtrar MLAs.

    Acepta IDs numéricos y el literal 'sin_tienda'.
    Retorna (lista_ids, incluir_sin_tienda) o None si el filtro no debe aplicarse.

    Raises HTTPException(400) si encuentra un token inválido (no numérico y distinto de 'sin_tienda').
    """
    if not csv:
        return None
    tokens = [t.strip() for t in csv.split(",") if t.strip()]
    if not tokens:
        return None
    ids: list[int] = []
    incluir_sin_tienda = False
    for token in tokens:
        if token == "sin_tienda":
            # 'sin_tienda' es sentinel para mlp_official_store_id IS NULL (productos publicados sin tienda oficial).
            incluir_sin_tienda = True
        else:
            try:
                ids.append(int(token))
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail=f"Token inválido en tiendas_oficiales: '{token}'. Esperado: int o 'sin_tienda'.",
                )
    return (ids, incluir_sin_tienda)


def _build_filtro_tiendas_oficiales_mla(
    parsed: tuple[list[int], bool],
) -> Optional[ColumnElement[bool]]:
    """
    Construye la expresión SQLAlchemy para filtrar MLAs por tienda oficial.
    Retorna una expresión que se puede usar en .filter() sobre MercadoLibreItemPublicado.

    parsed: tupla (lista_ids, incluir_sin_tienda) del parser.

    El retorno puede ser:
      - `BinaryExpression` (1 sola condición: `.in_(...)` o `.is_(None)`)
      - `BooleanClauseList` (2 condiciones combinadas con `or_(...)`)
      - `None` (si parsed no produce condiciones, ej. ([], False))

    Ambas variantes concretas heredan de `ColumnElement[bool]` en SQLAlchemy 2.0+,
    que es el ancestro común correcto para tipar el retorno.
    """
    from app.models.mercadolibre_item_publicado import MercadoLibreItemPublicado

    ids, incluir_sin_tienda = parsed
    condiciones = []
    if ids:
        condiciones.append(MercadoLibreItemPublicado.mlp_official_store_id.in_(ids))
    if incluir_sin_tienda:
        condiciones.append(MercadoLibreItemPublicado.mlp_official_store_id.is_(None))
    if not condiciones:
        return None
    if len(condiciones) == 1:
        return condiciones[0]
    return or_(*condiciones)


def _apply_search_filter(query, search: Optional[str]):
    """Aplica filtro de búsqueda sobre ProductoERP (descripcion, marca, codigo).

    Soporta los mismos operadores que el listado de productos:
      - `*valor`       → termina en `valor`
      - `valor*`       → comienza con `valor`
      - `campo:valor`  → match sobre `ean`/`codigo`/`marca` (exacto) o `desc`/`descripcion` (normalizado)
      - texto plano    → contiene (normalizado: sin guiones ni espacios, case-insensitive)

    Retorna el query con el filtro aplicado, o el query original si `search` es vacío.
    """
    if not search:
        return query

    search_filter = None

    if ":" in search and not search.startswith("*") and not search.endswith("*"):
        parts = search.split(":", 1)
        if len(parts) == 2:
            field, value = parts[0].strip().lower(), parts[1].strip()
            if field in ("ean", "codigo"):
                search_filter = and_(
                    ProductoERP.codigo.isnot(None),
                    ProductoERP.codigo != "",
                    func.upper(ProductoERP.codigo) == value.upper(),
                )
            elif field == "marca":
                search_filter = and_(
                    ProductoERP.marca.isnot(None),
                    ProductoERP.marca != "",
                    func.upper(ProductoERP.marca) == value.upper(),
                )
            elif field in ("desc", "descripcion"):
                value_normalized = value.replace("-", "").replace(" ", "").upper()
                search_filter = and_(
                    ProductoERP.descripcion.isnot(None),
                    func.replace(func.replace(func.upper(ProductoERP.descripcion), "-", ""), " ", "").like(
                        f"%{value_normalized}%"
                    ),
                )
    elif search.startswith("*") and not search.endswith("*"):
        value = search[1:].upper()
        search_filter = or_(
            and_(ProductoERP.descripcion.isnot(None), func.upper(ProductoERP.descripcion).like(f"%{value}")),
            and_(ProductoERP.marca.isnot(None), func.upper(ProductoERP.marca).like(f"%{value}")),
            and_(ProductoERP.codigo.isnot(None), func.upper(ProductoERP.codigo).like(f"%{value}")),
        )
    elif search.endswith("*") and not search.startswith("*"):
        value = search[:-1].upper()
        search_filter = or_(
            and_(ProductoERP.descripcion.isnot(None), func.upper(ProductoERP.descripcion).like(f"{value}%")),
            and_(ProductoERP.marca.isnot(None), func.upper(ProductoERP.marca).like(f"{value}%")),
            and_(ProductoERP.codigo.isnot(None), func.upper(ProductoERP.codigo).like(f"{value}%")),
        )
    else:
        search_normalized = search.replace("-", "").replace(" ", "").upper()
        search_filter = or_(
            func.replace(func.replace(func.upper(ProductoERP.descripcion), "-", ""), " ", "").like(
                f"%{search_normalized}%"
            ),
            func.replace(func.replace(func.upper(ProductoERP.marca), "-", ""), " ", "").like(f"%{search_normalized}%"),
            func.replace(func.upper(ProductoERP.codigo), "-", "").like(f"%{search_normalized}%"),
        )

    if search_filter is not None:
        query = query.filter(search_filter)
    return query


@router.post("/productos/exportar-rebate")
def exportar_rebate(
    request: ExportRebateRequest, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)
):
    """Exporta productos con rebate a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from datetime import datetime, date
    from calendar import monthrange
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from app.models.publicacion_ml import PublicacionML
    from app.models.mla_banlist import MLABanlist
    from app.models.mercadolibre_item_publicado import MercadoLibreItemPublicado

    # Obtener MLAs baneados
    mlas_baneados = db.query(MLABanlist.mla).filter(MLABanlist.activo == True).all()
    mlas_baneados_set = {mla[0] for mla in mlas_baneados}

    # Filtro de tiendas oficiales (MLA): parsear CSV una sola vez al inicio.
    # `tiendas_oficiales` (plural) opera a nivel MLA, distinto de filtros["tienda_oficial"]
    # (singular, scope producto). Si parser devuelve None → no aplica filtro.
    parsed_tiendas_oficiales = _parsear_tiendas_oficiales_mla(request.tiendas_oficiales)
    filtro_tiendas_oficiales_mla = (
        _build_filtro_tiendas_oficiales_mla(parsed_tiendas_oficiales) if parsed_tiendas_oficiales is not None else None
    )
    # Subquery de MLAs (mlp_publicationID) permitidos. Se usa en .filter(PublicacionML.mla.in_(...))
    # porque la query primaria de Rebate es sobre PublicacionML, no sobre MercadoLibreItemPublicado.
    subq_mlas_permitidos = None
    if filtro_tiendas_oficiales_mla is not None:
        subq_mlas_permitidos = (
            db.query(MercadoLibreItemPublicado.mlp_publicationID)
            .filter(
                filtro_tiendas_oficiales_mla,
                MercadoLibreItemPublicado.mlp_publicationID.isnot(None),
            )
            .subquery()
        )

    # Fechas por defecto
    hoy = date.today()
    fecha_desde = request.fecha_desde
    fecha_hasta = request.fecha_hasta
    if not fecha_desde:
        fecha_desde = hoy.strftime("%Y-%m-%d")
    if not fecha_hasta:
        ultimo_dia = monthrange(hoy.year, hoy.month)[1]
        fecha_hasta = f"{hoy.year}-{hoy.month:02d}-{ultimo_dia:02d}"

    # Determinar pricelist_id según tipo_cuotas
    pricelist_map = {"clasica": 4, "3": 17, "6": 14, "9": 13, "12": 23}
    pricelist_id = pricelist_map.get(request.tipo_cuotas, 4)

    # Mapeo lista web → PVP equivalente para buscar MLAs en ambas listas
    pvp_equivalente_map = {"clasica": 12, "3": 18, "6": 19, "9": 20, "12": 21}
    pricelist_pvp_equivalente = pvp_equivalente_map.get(request.tipo_cuotas)

    # Construir query con filtros
    layer_activo = resolver_layer_activo(
        request.filtros.get("equipo_id") if request.filtros else None, current_user, db
    )
    query = (
        db.query(ProductoERP, ProductoPricing)
        .join(ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id)
        .filter(ProductoPricing.participa_rebate == True, ProductoPricing.out_of_cards != True)
    )
    query = join_color_layer(query, layer_activo)

    # Aplicar filtros si existen
    if request.filtros:
        filtros = request.filtros

        query = _apply_search_filter(query, filtros.get("search"))

        if filtros.get("con_stock") is not None:
            query = query.filter(ProductoERP.stock > 0 if filtros["con_stock"] else ProductoERP.stock == 0)

        if filtros.get("con_precio") is not None:
            if filtros["con_precio"]:
                query = query.filter(ProductoPricing.precio_lista_ml.isnot(None))
            else:
                query = query.filter(ProductoPricing.precio_lista_ml.is_(None))

        if filtros.get("marcas"):
            marcas_list = [m.strip().upper() for m in filtros["marcas"].split(",")]
            query = query.filter(func.upper(ProductoERP.marca).in_(marcas_list))

        if filtros.get("subcategorias"):
            subcat_list = [int(s.strip()) for s in filtros["subcategorias"].split(",")]
            query = query.filter(ProductoERP.subcategoria_id.in_(subcat_list))

        if filtros.get("con_oferta") is not None:
            # Filtro de oferta si es necesario
            pass

        if filtros.get("con_web_transf") is not None:
            if filtros["con_web_transf"]:
                query = query.filter(ProductoPricing.participa_web_transferencia == True)
            else:
                query = query.filter(
                    or_(
                        ProductoPricing.participa_web_transferencia == False,
                        ProductoPricing.participa_web_transferencia.is_(None),
                    )
                )

        # Filtro de colores (lee del layer de equipo activo, ver productos_shared)
        query = filtro_colores(query, filtros.get("colores"), color_slot(None))

        if filtros.get("pms"):
            from app.models.marca_pm import MarcaPM

            pms_ids = [int(pm) for pm in filtros["pms"].split(",")]
            pares_pm = db.query(MarcaPM.marca, MarcaPM.categoria).filter(MarcaPM.usuario_id.in_(pms_ids)).all()
            if pares_pm:
                pares_upper = [(m.upper(), c.upper()) for m, c in pares_pm]
                query = query.filter(
                    tuple_(func.upper(ProductoERP.marca), func.upper(ProductoERP.categoria)).in_(pares_upper)
                )
            else:
                query = query.filter(ProductoERP.item_id == -1)

        if filtros.get("con_rebate") is not None:
            if filtros["con_rebate"]:
                query = query.filter(ProductoPricing.participa_rebate == True)
            else:
                query = query.filter(
                    or_(ProductoPricing.participa_rebate == False, ProductoPricing.participa_rebate.is_(None))
                )

        if filtros.get("out_of_cards") is not None:
            if filtros["out_of_cards"]:
                query = query.filter(ProductoPricing.out_of_cards == True)
            else:
                query = query.filter(or_(ProductoPricing.out_of_cards == False, ProductoPricing.out_of_cards.is_(None)))

        if filtros.get("markup_clasica_positivo") is not None:
            if filtros["markup_clasica_positivo"]:
                query = query.filter(ProductoPricing.markup > 0)
            else:
                query = query.filter(ProductoPricing.markup <= 0)

        if filtros.get("markup_rebate_positivo") is not None:
            if filtros["markup_rebate_positivo"]:
                query = query.filter(ProductoPricing.markup_rebate > 0)
            else:
                query = query.filter(ProductoPricing.markup_rebate <= 0)

        if filtros.get("markup_oferta_positivo") is not None:
            if filtros["markup_oferta_positivo"]:
                query = query.filter(ProductoPricing.mejor_oferta_markup > 0)
            else:
                query = query.filter(ProductoPricing.mejor_oferta_markup <= 0)

        if filtros.get("markup_web_transf_positivo") is not None:
            if filtros["markup_web_transf_positivo"]:
                query = query.filter(ProductoPricing.markup_web_real > 0)
            else:
                query = query.filter(ProductoPricing.markup_web_real <= 0)

        # Filtros de auditoría
        if (
            filtros.get("audit_usuarios")
            or filtros.get("audit_tipos_accion")
            or filtros.get("audit_fecha_desde")
            or filtros.get("audit_fecha_hasta")
        ):
            from app.models.auditoria import Auditoria

            # Subquery para obtener item_ids que cumplen con los filtros de auditoría
            audit_query = db.query(Auditoria.item_id).distinct()

            if filtros.get("audit_usuarios"):
                usuarios_ids = [int(u) for u in filtros["audit_usuarios"].split(",")]
                audit_query = audit_query.filter(Auditoria.usuario_id.in_(usuarios_ids))

            if filtros.get("audit_tipos_accion"):
                tipos_list = filtros["audit_tipos_accion"].split(",")
                audit_query = audit_query.filter(Auditoria.tipo_accion.in_(tipos_list))

            if filtros.get("audit_fecha_desde"):
                from datetime import datetime, timedelta

                try:
                    fecha_desde_dt = datetime.strptime(filtros["audit_fecha_desde"], "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    try:
                        fecha_desde_dt = datetime.strptime(filtros["audit_fecha_desde"], "%Y-%m-%d %H:%M")
                    except ValueError:
                        try:
                            fecha_desde_dt = datetime.strptime(filtros["audit_fecha_desde"], "%Y-%m-%d")
                        except ValueError:
                            from datetime import date

                            fecha_desde_dt = datetime.combine(date.today(), datetime.min.time())

                # Convertir de hora local (ART = UTC-3) a UTC sumando 3 horas
                fecha_desde_dt = fecha_desde_dt + timedelta(hours=3)
                audit_query = audit_query.filter(Auditoria.fecha >= fecha_desde_dt)

            if filtros.get("audit_fecha_hasta"):
                from datetime import datetime, timedelta

                try:
                    fecha_hasta_dt = datetime.strptime(filtros["audit_fecha_hasta"], "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    try:
                        fecha_hasta_dt = datetime.strptime(filtros["audit_fecha_hasta"], "%Y-%m-%d %H:%M")
                    except ValueError:
                        try:
                            fecha_hasta_dt = datetime.strptime(filtros["audit_fecha_hasta"], "%Y-%m-%d")
                            fecha_hasta_dt = fecha_hasta_dt.replace(hour=23, minute=59, second=59)
                        except ValueError:
                            from datetime import date

                            fecha_hasta_dt = datetime.combine(date.today(), datetime.max.time())

                # Convertir de hora local (ART = UTC-3) a UTC sumando 3 horas
                fecha_hasta_dt = fecha_hasta_dt + timedelta(hours=3)
                audit_query = audit_query.filter(Auditoria.fecha <= fecha_hasta_dt)

            item_ids_auditados = [item_id for (item_id,) in audit_query.all()]
            if item_ids_auditados:
                query = query.filter(ProductoERP.item_id.in_(item_ids_auditados))
            else:
                # Si no hay items que cumplan con los filtros de auditoría, no devolver nada
                query = query.filter(ProductoERP.item_id == -1)

        # Filtro de estado de publicaciones MLA
        if filtros.get("estado_mla"):
            from app.models.mercadolibre_item_publicado import MercadoLibreItemPublicado

            estado_mla_val = filtros["estado_mla"]

            if estado_mla_val == "activa":
                # Tienen al menos una publicación activa
                items_activos_subquery = (
                    db.query(MercadoLibreItemPublicado.item_id)
                    .filter(
                        MercadoLibreItemPublicado.mlp_id.isnot(None),
                        or_(
                            MercadoLibreItemPublicado.optval_statusId == 2,
                            MercadoLibreItemPublicado.optval_statusId.is_(None),
                        ),
                    )
                    .distinct()
                    .subquery()
                )

                query = query.filter(ProductoERP.item_id.in_(select(items_activos_subquery.c.item_id)))

            elif estado_mla_val == "pausada":
                # Tienen publicaciones pero ninguna activa
                items_con_publis = (
                    db.query(MercadoLibreItemPublicado.item_id)
                    .filter(MercadoLibreItemPublicado.mlp_id.isnot(None))
                    .distinct()
                    .subquery()
                )

                items_activos = (
                    db.query(MercadoLibreItemPublicado.item_id)
                    .filter(
                        MercadoLibreItemPublicado.mlp_id.isnot(None),
                        or_(
                            MercadoLibreItemPublicado.optval_statusId == 2,
                            MercadoLibreItemPublicado.optval_statusId.is_(None),
                        ),
                    )
                    .distinct()
                    .subquery()
                )

                query = query.filter(
                    ProductoERP.item_id.in_(select(items_con_publis.c.item_id)),
                    ~ProductoERP.item_id.in_(select(items_activos.c.item_id)),
                )

    # Filtro de estado_mla directo del request (fuera del dict filtros)
    if request.estado_mla:
        from app.models.mercadolibre_item_publicado import MercadoLibreItemPublicado

        if request.estado_mla == "activa":
            # Tienen al menos una publicación activa
            items_activos_subquery = (
                db.query(MercadoLibreItemPublicado.item_id)
                .filter(
                    MercadoLibreItemPublicado.mlp_id.isnot(None),
                    or_(
                        MercadoLibreItemPublicado.optval_statusId == 2,
                        MercadoLibreItemPublicado.optval_statusId.is_(None),
                    ),
                )
                .distinct()
                .subquery()
            )

            query = query.filter(ProductoERP.item_id.in_(select(items_activos_subquery.c.item_id)))

        elif request.estado_mla == "pausada":
            # Tienen publicaciones pero ninguna activa
            items_con_publis = (
                db.query(MercadoLibreItemPublicado.item_id)
                .filter(MercadoLibreItemPublicado.mlp_id.isnot(None))
                .distinct()
                .subquery()
            )

            items_activos = (
                db.query(MercadoLibreItemPublicado.item_id)
                .filter(
                    MercadoLibreItemPublicado.mlp_id.isnot(None),
                    or_(
                        MercadoLibreItemPublicado.optval_statusId == 2,
                        MercadoLibreItemPublicado.optval_statusId.is_(None),
                    ),
                )
                .distinct()
                .subquery()
            )

            query = query.filter(
                ProductoERP.item_id.in_(select(items_con_publis.c.item_id)),
                ~ProductoERP.item_id.in_(select(items_activos.c.item_id)),
            )

    productos = query.all()

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Rebate Export"

    # Headers
    if request.formato == "nuevo":
        headers = [
            "MLA",
            "",
            "",
            "",
            "TIPO DE OFERTA",
            "DESDE",
            "HASTA",
            "PVP LLENO",
            "PVP SELLER",
        ]
    else:
        # Formato tradicional (original)
        headers = [
            "REBATE",
            "MARCA",
            "DESDE",
            "HASTA",
            "TIPO DE OFERTA",
            "CATEGORÍA",
            "DESCRIPCIÓN DE LA PUBLICACIÓN",
            "TIPO DE PUBLICACIÓN",
            "STOCK",
            "FULL",
            "MLAs",
            "PVP LLENO",
            "PVP SELLER",
        ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Header columna T para formato nuevo
    if request.formato == "nuevo":
        cell = ws.cell(row=1, column=20, value="REBATE %")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Precargar todos los MLA IDs de los productos del export para fallback batch a ml_previews
    item_ids_export = [p[0].item_id for p in productos]
    all_mla_ids = [
        mla_id
        for (mla_id,) in db.query(PublicacionML.mla)
        .filter(PublicacionML.item_id.in_(item_ids_export), PublicacionML.activo == True)
        .all()
    ]
    mlwebhook_prices: dict[str, float] | None = None  # lazy: se carga solo si se necesita

    # Datos
    row = 2
    for producto_erp, producto_pricing in productos:
        # Buscar MLAs de la lista seleccionada
        query_mlas = db.query(PublicacionML).filter(
            PublicacionML.item_id == producto_erp.item_id,
            PublicacionML.pricelist_id == pricelist_id,
            PublicacionML.activo == True,
        )
        if subq_mlas_permitidos is not None:
            query_mlas = query_mlas.filter(PublicacionML.mla.in_(select(subq_mlas_permitidos)))
        mlas = query_mlas.all()

        # Para cuotas: también traer MLAs de la pricelist PVP equivalente y unificar
        if pricelist_pvp_equivalente:
            query_mlas_pvp = db.query(PublicacionML).filter(
                PublicacionML.item_id == producto_erp.item_id,
                PublicacionML.pricelist_id == pricelist_pvp_equivalente,
                PublicacionML.activo == True,
            )
            if subq_mlas_permitidos is not None:
                query_mlas_pvp = query_mlas_pvp.filter(PublicacionML.mla.in_(select(subq_mlas_permitidos)))
            mlas_pvp = query_mlas_pvp.all()
            # Deduplicar por MLA (el código MLA es único, priorizar el de cuotas)
            mlas_existentes = {m.mla for m in mlas}
            for m in mlas_pvp:
                if m.mla not in mlas_existentes:
                    mlas.append(m)
                    mlas_existentes.add(m.mla)

        # Si no tiene MLAs, skip
        if not mlas:
            continue

        # Obtener precio de la lista seleccionada de PrecioML
        from app.models.precio_ml import PrecioML

        precio_lista = (
            db.query(PrecioML)
            .filter(PrecioML.item_id == producto_erp.item_id, PrecioML.pricelist_id == pricelist_id)
            .first()
        )

        precio_pricelist = float(precio_lista.precio) if precio_lista and precio_lista.precio else 0

        # También obtener precio de la pricelist PVP equivalente (para MLAs que vengan de PVP)
        precio_pricelist_pvp = 0
        if pricelist_pvp_equivalente:
            precio_lista_pvp = (
                db.query(PrecioML)
                .filter(PrecioML.item_id == producto_erp.item_id, PrecioML.pricelist_id == pricelist_pvp_equivalente)
                .first()
            )
            precio_pricelist_pvp = float(precio_lista_pvp.precio) if precio_lista_pvp and precio_lista_pvp.precio else 0

        # NO skipear acá: si precios_ml no tiene registro, el fallback por MLA
        # (mlp_lastPriceInformedByML) más abajo resuelve el PVP LLENO real de ML

        porcentaje_rebate = float(producto_pricing.porcentaje_rebate or 3.8)

        if pricelist_id == 4:
            # Clásica: necesita precio_lista_ml como base
            if not producto_pricing.precio_lista_ml:
                continue
            precio_base = float(producto_pricing.precio_lista_ml)
            # PVP SELLER = precio_lista_ml / (1 - rebate%)
            pvp_seller = precio_base / (1 - porcentaje_rebate / 100)
            # PVP LLENO se determina por MLA (web o pvp) más abajo
        else:
            # Cuotas: usar el precio editado de la columna de cuotas (ProductoPricing)
            # como base para PVP SELLER, y el precio de la pricelist (PrecioML) como PVP LLENO
            cuotas_campo_map = {
                "3": "precio_3_cuotas",
                "6": "precio_6_cuotas",
                "9": "precio_9_cuotas",
                "12": "precio_12_cuotas",
            }
            campo_cuota = cuotas_campo_map.get(request.tipo_cuotas)
            precio_cuota_editado = float(getattr(producto_pricing, campo_cuota, None) or 0) if campo_cuota else 0

            if precio_cuota_editado == 0:
                # Si no tiene precio de cuota editado, skip este producto
                continue

            porcentaje_cuotas = (
                request.porcentaje_rebate_override
                if request.porcentaje_rebate_override is not None
                else porcentaje_rebate
            )
            # PVP SELLER = precio editado de cuotas / (1 - rebate%)
            pvp_seller = precio_cuota_editado / (1 - porcentaje_cuotas / 100)

        # Una fila por cada MLA (excluyendo los baneados)
        for mla in mlas:
            # Saltar si el MLA está en la banlist
            if mla.mla in mlas_baneados_set:
                continue

            # Determinar PVP LLENO según la pricelist del MLA (con fallback)
            es_mla_pvp = pricelist_pvp_equivalente and mla.pricelist_id == pricelist_pvp_equivalente
            if es_mla_pvp:
                precio_base_lleno = precio_pricelist_pvp if precio_pricelist_pvp > 0 else precio_pricelist
            else:
                precio_base_lleno = precio_pricelist if precio_pricelist > 0 else precio_pricelist_pvp

            # Fallback: si no hay PrecioML para ninguna lista, usar el precio real del MLA
            if precio_base_lleno == 0:
                # 1) Intentar desde tb_mercadolibre_items_publicados
                mla_publicado = (
                    db.query(MercadoLibreItemPublicado)
                    .filter(MercadoLibreItemPublicado.mlp_publicationID == mla.mla)
                    .first()
                )
                if mla_publicado and mla_publicado.mlp_lastPriceInformedByML:
                    precio_base_lleno = float(mla_publicado.mlp_lastPriceInformedByML)

            if precio_base_lleno == 0:
                # 2) Intentar desde ml_previews (DB webhook) — precio real publicado en ML
                if mlwebhook_prices is None:
                    mlwebhook_prices = _cargar_precios_mlwebhook(all_mla_ids)
                precio_webhook = mlwebhook_prices.get(mla.mla, 0)
                if precio_webhook > 0:
                    precio_base_lleno = precio_webhook

            if pricelist_id == 4:
                pvp_lleno = precio_base_lleno
            else:
                offset_lleno = request.offset_pvp_lleno if request.offset_pvp_lleno is not None else 0
                pvp_lleno = precio_base_lleno * (1 + offset_lleno / 100)

            # Si no hay PVP LLENO válido para este MLA, skip
            if not pvp_lleno or pvp_lleno == 0:
                continue

            if request.formato == "nuevo":
                # Formato DXI (9 columnas + col T = rebate%)
                ws.cell(row=row, column=1, value=mla.mla)
                ws.cell(row=row, column=2, value="")
                ws.cell(row=row, column=3, value="")
                ws.cell(row=row, column=4, value="")
                ws.cell(row=row, column=5, value="DxI")
                ws.cell(row=row, column=6, value=fecha_desde)
                ws.cell(row=row, column=7, value=fecha_hasta)
                ws.cell(row=row, column=8, value=round(pvp_lleno, 2))
                ws.cell(row=row, column=9, value=round(pvp_seller, 2))
                # Columna T (20): porcentaje de rebate formateado (ej: "3,80%")
                rebate_mostrar = porcentaje_cuotas if pricelist_id != 4 else porcentaje_rebate
                ws.cell(row=row, column=20, value=f"{rebate_mostrar:.2f}%".replace(".", ","))
            else:
                # Formato tradicional
                rebate_mostrar = porcentaje_cuotas if pricelist_id != 4 else porcentaje_rebate
                ws.cell(row=row, column=1, value=f"{rebate_mostrar}%")
                ws.cell(row=row, column=2, value=producto_erp.marca or "")
                ws.cell(row=row, column=3, value=fecha_desde)
                ws.cell(row=row, column=4, value=fecha_hasta)
                ws.cell(row=row, column=5, value="DxI")
                ws.cell(row=row, column=6, value="")  # Categoría vacía
                ws.cell(row=row, column=7, value=mla.item_title or producto_erp.descripcion or "")
                ws.cell(row=row, column=8, value="Clásica" if pricelist_id == 4 else f"{request.tipo_cuotas} Cuotas")
                ws.cell(row=row, column=9, value=producto_erp.stock)
                ws.cell(row=row, column=10, value="FALSE")
                ws.cell(row=row, column=11, value=mla.mla)
                ws.cell(row=row, column=12, value=pvp_lleno)
                ws.cell(row=row, column=13, value=round(pvp_seller, 2))

            row += 1

    # Ocultar columnas vacías J-S (10 a 19) en formato nuevo
    if request.formato == "nuevo":
        for col_letter in ["J", "K", "L", "M", "N", "O", "P", "Q", "R", "S"]:
            ws.column_dimensions[col_letter].hidden = True

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=rebate_export_{hoy.strftime('%Y%m%d')}.xlsx"},
    )


@router.get("/exportar-web-transferencia")
def exportar_web_transferencia(
    porcentaje_adicional: float = Query(0, description="Porcentaje adicional a sumar"),
    currency_id: int = Query(1, description="ID de moneda: 1=ARS, 2=USD"),
    offset_dolar: float = Query(0, description="Offset en pesos para ajustar el dólar"),
    search: Optional[str] = None,
    con_stock: Optional[bool] = None,
    con_precio: Optional[bool] = None,
    marcas: Optional[str] = None,
    subcategorias: Optional[str] = None,
    con_rebate: Optional[bool] = None,
    con_oferta: Optional[bool] = None,
    con_web_transf: Optional[bool] = None,
    tiendanube_con_descuento: Optional[bool] = None,
    tiendanube_sin_descuento: Optional[bool] = None,
    tiendanube_no_publicado: Optional[bool] = None,
    markup_clasica_positivo: Optional[bool] = None,
    markup_rebate_positivo: Optional[bool] = None,
    markup_oferta_positivo: Optional[bool] = None,
    markup_web_transf_positivo: Optional[bool] = None,
    out_of_cards: Optional[bool] = None,
    colores: Optional[str] = None,
    pms: Optional[str] = None,
    audit_usuarios: Optional[str] = None,
    audit_tipos_accion: Optional[str] = None,
    audit_fecha_desde: Optional[str] = None,
    audit_fecha_hasta: Optional[str] = None,
    estado_mla: Optional[str] = None,
    equipo_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Exporta precios de Web Transferencia en formato Excel con filtros opcionales"""
    from io import BytesIO
    from openpyxl import Workbook
    from app.models.tipo_cambio import TipoCambio

    layer_activo = resolver_layer_activo(equipo_id, current_user, db)

    # Obtener productos con precio web transferencia
    query = (
        db.query(ProductoERP.item_id, ProductoERP.codigo, ProductoPricing.precio_web_transferencia)
        .join(ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id)
        .filter(
            ProductoPricing.participa_web_transferencia == True, ProductoPricing.precio_web_transferencia.isnot(None)
        )
    )
    query = join_color_layer(query, layer_activo)

    # FILTRADO POR AUDITORÍA (igual que en el listado principal)
    if audit_usuarios or audit_tipos_accion or audit_fecha_desde or audit_fecha_hasta:
        from app.models.auditoria import Auditoria
        from datetime import datetime

        # Construir filtros de auditoría base
        filtros_audit = [Auditoria.item_id.isnot(None)]

        if audit_usuarios:
            usuarios_ids = [int(u) for u in audit_usuarios.split(",")]
            filtros_audit.append(Auditoria.usuario_id.in_(usuarios_ids))

        if audit_tipos_accion:
            tipos_list = audit_tipos_accion.split(",")
            filtros_audit.append(Auditoria.tipo_accion.in_(tipos_list))

        if audit_fecha_desde:
            try:
                fecha_desde_dt = datetime.strptime(audit_fecha_desde, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    fecha_desde_dt = datetime.strptime(audit_fecha_desde, "%Y-%m-%d %H:%M")
                except ValueError:
                    try:
                        fecha_desde_dt = datetime.strptime(audit_fecha_desde, "%Y-%m-%d")
                    except ValueError:
                        from datetime import date

                        fecha_desde_dt = datetime.combine(date.today(), datetime.min.time())

            # Convertir de hora local (ART = UTC-3) a UTC sumando 3 horas
            from datetime import timedelta

            fecha_desde_dt = fecha_desde_dt + timedelta(hours=3)

            filtros_audit.append(Auditoria.fecha >= fecha_desde_dt)

        if audit_fecha_hasta:
            try:
                fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, "%Y-%m-%d %H:%M")
                except ValueError:
                    try:
                        fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, "%Y-%m-%d")
                        fecha_hasta_dt = fecha_hasta_dt.replace(hour=23, minute=59, second=59)
                    except ValueError:
                        from datetime import date

                        fecha_hasta_dt = datetime.combine(date.today(), datetime.max.time())

            # Convertir de hora local (ART = UTC-3) a UTC sumando 3 horas
            from datetime import timedelta

            fecha_hasta_dt = fecha_hasta_dt + timedelta(hours=3)

            filtros_audit.append(Auditoria.fecha <= fecha_hasta_dt)

        # Obtener productos que tienen auditorías cumpliendo los criterios
        audit_query = db.query(Auditoria.item_id).filter(and_(*filtros_audit))
        item_ids_audit = [item_id for (item_id,) in audit_query.distinct().all()]

        if item_ids_audit:
            query = query.filter(ProductoERP.item_id.in_(item_ids_audit))
        else:
            # Si no hay productos con las auditorías filtradas, retornar vacío
            wb = Workbook()
            ws = wb.active
            ws.append(["No se encontraron productos con los filtros aplicados"])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=web_transferencia_vacia.xlsx"},
            )

    # Aplicar filtros básicos
    query = _apply_search_filter(query, search)

    if con_stock:
        query = query.filter(ProductoERP.stock > 0)

    if con_precio:
        query = query.filter(ProductoPricing.precio_lista_ml.isnot(None))

    if marcas:
        marcas_list = [m.strip().upper() for m in marcas.split(",")]
        query = query.filter(func.upper(ProductoERP.marca).in_(marcas_list))

    if subcategorias:
        subcats_list = [int(s) for s in subcategorias.split(",")]
        query = query.filter(ProductoERP.subcategoria_id.in_(subcats_list))

    # Filtro por PMs (Product Managers) - filtra por pares (marca, categoria)
    if pms:
        from app.models.marca_pm import MarcaPM

        pm_ids = [int(pm.strip()) for pm in pms.split(",")]

        pares_pm = db.query(MarcaPM.marca, MarcaPM.categoria).filter(MarcaPM.usuario_id.in_(pm_ids)).all()

        if pares_pm:
            pares_upper = [(m.upper(), c.upper()) for m, c in pares_pm]
            query = query.filter(
                tuple_(func.upper(ProductoERP.marca), func.upper(ProductoERP.categoria)).in_(pares_upper)
            )
        else:
            # Si el PM no tiene marcas asignadas, no hay productos
            wb = Workbook()
            ws = wb.active
            ws.append(["No hay productos para los PMs seleccionados"])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=web_transferencia_vacia.xlsx"},
            )

    # Filtro por colores (lee del layer de equipo activo, ver productos_shared)
    query = filtro_colores(query, colores, color_slot(None))

    # Filtros booleanos avanzados
    if con_rebate is not None:
        if con_rebate:
            query = query.filter(ProductoPricing.participa_rebate == True)
        else:
            query = query.filter(
                or_(ProductoPricing.participa_rebate == False, ProductoPricing.participa_rebate.is_(None))
            )

    if con_oferta is not None:
        # Este filtro requiere join con ofertas, se aplicará después
        pass

    if out_of_cards is not None:
        if out_of_cards:
            query = query.filter(ProductoPricing.out_of_cards == True)
        else:
            query = query.filter(or_(ProductoPricing.out_of_cards == False, ProductoPricing.out_of_cards.is_(None)))

    # Filtros de Tienda Nube
    if tiendanube_con_descuento:
        query = query.filter(ProductoPricing.descuento_tiendanube.isnot(None), ProductoPricing.descuento_tiendanube > 0)

    if tiendanube_sin_descuento:
        query = query.filter(
            (ProductoPricing.descuento_tiendanube.is_(None)) | (ProductoPricing.descuento_tiendanube == 0)
        )

    if tiendanube_no_publicado:
        # Productos con stock pero NO en Tienda Nube
        from app.models.tienda_nube_producto import TiendaNubeProducto
        from sqlalchemy.sql import exists

        subquery = exists().where(
            and_(TiendaNubeProducto.item_id == ProductoERP.item_id, TiendaNubeProducto.activo == True)
        )
        query = query.filter(and_(ProductoERP.stock > 0, ~subquery))

    # Filtro de estado de publicaciones MLA
    if estado_mla:
        from app.models.mercadolibre_item_publicado import MercadoLibreItemPublicado

        if estado_mla == "activa":
            # Tienen al menos una publicación activa
            items_activos_subquery = (
                db.query(MercadoLibreItemPublicado.item_id)
                .filter(
                    MercadoLibreItemPublicado.mlp_id.isnot(None),
                    or_(
                        MercadoLibreItemPublicado.optval_statusId == 2,
                        MercadoLibreItemPublicado.optval_statusId.is_(None),
                    ),
                )
                .distinct()
                .subquery()
            )

            query = query.filter(ProductoERP.item_id.in_(select(items_activos_subquery.c.item_id)))

        elif estado_mla == "pausada":
            # Tienen publicaciones pero ninguna activa
            items_con_publis = (
                db.query(MercadoLibreItemPublicado.item_id)
                .filter(MercadoLibreItemPublicado.mlp_id.isnot(None))
                .distinct()
                .subquery()
            )

            items_activos = (
                db.query(MercadoLibreItemPublicado.item_id)
                .filter(
                    MercadoLibreItemPublicado.mlp_id.isnot(None),
                    or_(
                        MercadoLibreItemPublicado.optval_statusId == 2,
                        MercadoLibreItemPublicado.optval_statusId.is_(None),
                    ),
                )
                .distinct()
                .subquery()
            )

            query = query.filter(
                ProductoERP.item_id.in_(select(items_con_publis.c.item_id)),
                ~ProductoERP.item_id.in_(select(items_activos.c.item_id)),
            )

    productos = query.all()

    # Aplicar filtros de markup y oferta (requieren cálculos, se hacen después de la query)
    if (
        markup_clasica_positivo is not None
        or markup_rebate_positivo is not None
        or markup_oferta_positivo is not None
        or markup_web_transf_positivo is not None
        or con_oferta is not None
    ):
        from app.models.oferta_ml import OfertaML
        from app.models.publicacion_ml import PublicacionML
        from datetime import date

        productos_filtrados = []
        hoy = date.today()

        for producto in productos:
            item_id = producto[0]
            incluir = True

            # Obtener ProductoERP para el item_id
            producto_erp = db.query(ProductoERP).filter(ProductoERP.item_id == item_id).first()
            if not producto_erp:
                continue

            producto_pricing = db.query(ProductoPricing).filter(ProductoPricing.item_id == item_id).first()
            if not producto_pricing:
                continue

            # Filtro de markup clásica
            if markup_clasica_positivo is not None and incluir:
                markup = producto_pricing.markup_calculado if producto_pricing else None
                if markup is not None:
                    if markup_clasica_positivo and markup < 0:
                        incluir = False
                    elif not markup_clasica_positivo and markup >= 0:
                        incluir = False
                else:
                    incluir = False

            # Filtro de markup web transferencia
            if markup_web_transf_positivo is not None and incluir:
                if producto_pricing and producto_pricing.markup_web_real is not None:
                    markup_web = float(producto_pricing.markup_web_real)
                    if markup_web_transf_positivo and markup_web < 0:
                        incluir = False
                    elif not markup_web_transf_positivo and markup_web >= 0:
                        incluir = False
                else:
                    incluir = False

            # Filtro de oferta
            if con_oferta is not None and incluir:
                pubs = db.query(PublicacionML).filter(PublicacionML.item_id == item_id).all()
                tiene_oferta = False
                for pub in pubs:
                    oferta = (
                        db.query(OfertaML)
                        .filter(
                            OfertaML.mla == pub.mla,
                            OfertaML.fecha_desde <= hoy,
                            OfertaML.fecha_hasta >= hoy,
                            OfertaML.pvp_seller.isnot(None),
                        )
                        .first()
                    )
                    if oferta:
                        tiene_oferta = True
                        break
                if con_oferta and not tiene_oferta:
                    incluir = False
                elif not con_oferta and tiene_oferta:
                    incluir = False

            if incluir:
                productos_filtrados.append(producto)

        productos = productos_filtrados

    # Obtener dólar venta si currency_id es 2
    dolar_ajustado = None
    if currency_id == 2:
        tipo_cambio = db.query(TipoCambio).order_by(TipoCambio.id.desc()).first()
        if tipo_cambio:
            dolar_ajustado = float(tipo_cambio.venta) + offset_dolar

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Web Transferencia"

    # Header
    ws.append(["Código/EAN", "Precio", "ID Moneda"])

    # Datos - todo como texto
    for item_id, codigo, precio_base in productos:
        # Aplicar porcentaje adicional
        precio_final = float(precio_base) * (1 + porcentaje_adicional / 100)

        # Si es USD, dividir por dólar ajustado
        if currency_id == 2 and dolar_ajustado:
            precio_final = precio_final / dolar_ajustado
            # Para USD, redondear a 2 decimales
            precio_str = f"{precio_final:.2f}"
        else:
            # Para ARS, redondear a múltiplo de 10
            precio_final = round(precio_final / 10) * 10
            precio_str = str(int(precio_final))

        ws.append([str(codigo), precio_str, str(currency_id)])

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=web_transferencia.xlsx"},
    )


@router.get("/exportar-clasica")
def exportar_clasica(
    porcentaje_adicional: float = Query(0, description="Porcentaje adicional sobre rebate"),
    tipo_cuotas: str = Query(
        "clasica", description="Tipo de cuotas: clasica, 3, 6, 9, 12, pvp, pvp_3, pvp_6, pvp_9, pvp_12"
    ),
    currency_id: int = Query(1, description="ID de moneda: 1=ARS, 2=USD"),
    offset_dolar: float = Query(0, description="Offset en pesos para ajustar el dólar"),
    search: Optional[str] = None,
    con_stock: Optional[bool] = None,
    con_precio: Optional[bool] = None,
    marcas: Optional[str] = None,
    subcategorias: Optional[str] = None,
    con_rebate: Optional[bool] = None,
    con_oferta: Optional[bool] = None,
    con_web_transf: Optional[bool] = None,
    tiendanube_con_descuento: Optional[bool] = None,
    tiendanube_sin_descuento: Optional[bool] = None,
    tiendanube_no_publicado: Optional[bool] = None,
    markup_clasica_positivo: Optional[bool] = None,
    markup_rebate_positivo: Optional[bool] = None,
    markup_oferta_positivo: Optional[bool] = None,
    markup_web_transf_positivo: Optional[bool] = None,
    out_of_cards: Optional[bool] = None,
    colores: Optional[str] = None,
    pms: Optional[str] = None,
    audit_usuarios: Optional[str] = None,
    audit_tipos_accion: Optional[str] = None,
    audit_fecha_desde: Optional[str] = None,
    audit_fecha_hasta: Optional[str] = None,
    con_mla: Optional[bool] = None,
    estado_mla: Optional[str] = None,
    nuevos_ultimos_7_dias: Optional[bool] = None,
    tienda_oficial: Optional[str] = None,
    tiendas_oficiales: Optional[str] = Query(
        None,
        description=(
            "CSV de IDs de tiendas oficiales con literal 'sin_tienda'. "
            "Filtra columnas MLA dinámicas (mlp_official_store_id). "
            "Ej: 'sin_tienda,57997,2645'. "
            "Distinto de 'tienda_oficial' (filtro a nivel producto)."
        ),
    ),
    equipo_id: Optional[int] = None,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Exporta precios de Clásica. Si tiene rebate activo, aplica % sobre precio rebate."""
    from app.services.permisos_service import verificar_permiso

    # Verificar permiso
    if not verificar_permiso(db, current_user, "productos.exportar_clasica"):
        raise HTTPException(status_code=403, detail="No tienes permiso para exportar lista de precios clásica")
    from io import BytesIO
    from openpyxl import Workbook
    from app.models.tipo_cambio import TipoCambio

    layer_activo = resolver_layer_activo(equipo_id, current_user, db)

    # Obtener productos con precio clásica y precios con cuotas
    query = (
        db.query(
            ProductoERP.item_id,
            ProductoERP.codigo,
            ProductoPricing.precio_lista_ml,
            ProductoPricing.participa_rebate,
            ProductoPricing.porcentaje_rebate,
            ProductoPricing.precio_3_cuotas,
            ProductoPricing.precio_6_cuotas,
            ProductoPricing.precio_9_cuotas,
            ProductoPricing.precio_12_cuotas,
            ProductoPricing.precio_pvp,
            ProductoPricing.precio_pvp_3_cuotas,
            ProductoPricing.precio_pvp_6_cuotas,
            ProductoPricing.precio_pvp_9_cuotas,
            ProductoPricing.precio_pvp_12_cuotas,
        )
        .join(ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id)
        .filter(ProductoPricing.precio_lista_ml.isnot(None))
    )
    query = join_color_layer(query, layer_activo)

    # FILTRADO POR AUDITORÍA (igual que en el listado principal)
    if audit_usuarios or audit_tipos_accion or audit_fecha_desde or audit_fecha_hasta:
        from app.models.auditoria import Auditoria
        from datetime import datetime

        # Construir filtros de auditoría base
        filtros_audit = [Auditoria.item_id.isnot(None)]

        if audit_usuarios:
            usuarios_ids = [int(u) for u in audit_usuarios.split(",")]
            filtros_audit.append(Auditoria.usuario_id.in_(usuarios_ids))

        if audit_tipos_accion:
            tipos_list = audit_tipos_accion.split(",")
            filtros_audit.append(Auditoria.tipo_accion.in_(tipos_list))

        if audit_fecha_desde:
            try:
                fecha_desde_dt = datetime.strptime(audit_fecha_desde, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    fecha_desde_dt = datetime.strptime(audit_fecha_desde, "%Y-%m-%d %H:%M")
                except ValueError:
                    try:
                        fecha_desde_dt = datetime.strptime(audit_fecha_desde, "%Y-%m-%d")
                    except ValueError:
                        from datetime import date

                        fecha_desde_dt = datetime.combine(date.today(), datetime.min.time())

            # Convertir de hora local (ART = UTC-3) a UTC sumando 3 horas
            from datetime import timedelta

            fecha_desde_dt = fecha_desde_dt + timedelta(hours=3)

            filtros_audit.append(Auditoria.fecha >= fecha_desde_dt)

        if audit_fecha_hasta:
            try:
                fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, "%Y-%m-%d %H:%M")
                except ValueError:
                    try:
                        fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, "%Y-%m-%d")
                        fecha_hasta_dt = fecha_hasta_dt.replace(hour=23, minute=59, second=59)
                    except ValueError:
                        from datetime import date

                        fecha_hasta_dt = datetime.combine(date.today(), datetime.max.time())

            # Convertir de hora local (ART = UTC-3) a UTC sumando 3 horas
            from datetime import timedelta

            fecha_hasta_dt = fecha_hasta_dt + timedelta(hours=3)

            filtros_audit.append(Auditoria.fecha <= fecha_hasta_dt)

        # Obtener productos que tienen auditorías cumpliendo los criterios
        audit_query = db.query(Auditoria.item_id).filter(and_(*filtros_audit))
        item_ids_audit = [item_id for (item_id,) in audit_query.distinct().all()]

        if item_ids_audit:
            query = query.filter(ProductoERP.item_id.in_(item_ids_audit))
        else:
            # Si no hay productos con las auditorías filtradas, retornar vacío
            from io import BytesIO
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.append(["No se encontraron productos con los filtros aplicados"])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=exportacion_clasica_vacia.xlsx"},
            )

    # Aplicar filtros básicos (con soporte para operadores *, +, :)
    query = _apply_search_filter(query, search)

    if con_stock:
        query = query.filter(ProductoERP.stock > 0)

    if con_precio:
        query = query.filter(ProductoPricing.precio_lista_ml.isnot(None))

    if marcas:
        marcas_list = [m.strip().upper() for m in marcas.split(",")]
        query = query.filter(func.upper(ProductoERP.marca).in_(marcas_list))

    if subcategorias:
        subcats_list = [int(s) for s in subcategorias.split(",")]
        query = query.filter(ProductoERP.subcategoria_id.in_(subcats_list))

    # Filtro por PMs (Product Managers) - filtra por pares (marca, categoria)
    if pms:
        from app.models.marca_pm import MarcaPM

        pm_ids = [int(pm.strip()) for pm in pms.split(",")]

        pares_pm = db.query(MarcaPM.marca, MarcaPM.categoria).filter(MarcaPM.usuario_id.in_(pm_ids)).all()

        if pares_pm:
            pares_upper = [(m.upper(), c.upper()) for m, c in pares_pm]
            query = query.filter(
                tuple_(func.upper(ProductoERP.marca), func.upper(ProductoERP.categoria)).in_(pares_upper)
            )
        else:
            # Si el PM no tiene marcas asignadas, no hay productos
            from io import BytesIO
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.append(["No hay productos para los PMs seleccionados"])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=exportacion_clasica_vacia.xlsx"},
            )

    # Filtro por colores (lee del layer de equipo activo, ver productos_shared)
    query = filtro_colores(query, colores, color_slot(None))

    # Filtros booleanos avanzados
    if con_rebate is not None:
        if con_rebate:
            query = query.filter(ProductoPricing.participa_rebate == True)
        else:
            query = query.filter(
                or_(ProductoPricing.participa_rebate == False, ProductoPricing.participa_rebate.is_(None))
            )

    if con_web_transf is not None:
        if con_web_transf:
            query = query.filter(ProductoPricing.participa_web_transferencia == True)
        else:
            query = query.filter(
                or_(
                    ProductoPricing.participa_web_transferencia == False,
                    ProductoPricing.participa_web_transferencia.is_(None),
                )
            )

    if out_of_cards is not None:
        if out_of_cards:
            query = query.filter(ProductoPricing.out_of_cards == True)
        else:
            query = query.filter(or_(ProductoPricing.out_of_cards == False, ProductoPricing.out_of_cards.is_(None)))

    # Filtros de Tienda Nube
    if tiendanube_con_descuento:
        query = query.filter(ProductoPricing.descuento_tiendanube.isnot(None), ProductoPricing.descuento_tiendanube > 0)

    if tiendanube_sin_descuento:
        query = query.filter(
            (ProductoPricing.descuento_tiendanube.is_(None)) | (ProductoPricing.descuento_tiendanube == 0)
        )

    if tiendanube_no_publicado:
        # Productos con stock pero NO en Tienda Nube
        from app.models.tienda_nube_producto import TiendaNubeProducto
        from sqlalchemy.sql import exists

        subquery = exists().where(
            and_(TiendaNubeProducto.item_id == ProductoERP.item_id, TiendaNubeProducto.activo == True)
        )
        query = query.filter(and_(ProductoERP.stock > 0, ~subquery))

    # Filtro de estado de publicaciones MLA
    if estado_mla:
        from app.models.mercadolibre_item_publicado import MercadoLibreItemPublicado

        if estado_mla == "activa":
            # Tienen al menos una publicación activa
            items_activos_subquery = (
                db.query(MercadoLibreItemPublicado.item_id)
                .filter(
                    MercadoLibreItemPublicado.mlp_id.isnot(None),
                    or_(
                        MercadoLibreItemPublicado.optval_statusId == 2,
                        MercadoLibreItemPublicado.optval_statusId.is_(None),
                    ),
                )
                .distinct()
                .subquery()
            )

            query = query.filter(ProductoERP.item_id.in_(select(items_activos_subquery.c.item_id)))

        elif estado_mla == "pausada":
            # Tienen publicaciones pero ninguna activa
            items_con_publis = (
                db.query(MercadoLibreItemPublicado.item_id)
                .filter(MercadoLibreItemPublicado.mlp_id.isnot(None))
                .distinct()
                .subquery()
            )

            items_activos = (
                db.query(MercadoLibreItemPublicado.item_id)
                .filter(
                    MercadoLibreItemPublicado.mlp_id.isnot(None),
                    or_(
                        MercadoLibreItemPublicado.optval_statusId == 2,
                        MercadoLibreItemPublicado.optval_statusId.is_(None),
                    ),
                )
                .distinct()
                .subquery()
            )

            query = query.filter(
                ProductoERP.item_id.in_(select(items_con_publis.c.item_id)),
                ~ProductoERP.item_id.in_(select(items_activos.c.item_id)),
            )

    productos = query.all()

    # Aplicar filtros de markup y oferta (requieren cálculos, se hacen después de la query)
    if (
        markup_clasica_positivo is not None
        or markup_rebate_positivo is not None
        or markup_oferta_positivo is not None
        or markup_web_transf_positivo is not None
        or con_oferta is not None
    ):
        from app.services.pricing_calculator import (
            obtener_tipo_cambio_actual,
            convertir_a_pesos,
            obtener_grupo_subcategoria,
            obtener_comision_base,
            calcular_comision_ml_total,
            calcular_limpio,
            calcular_markup,
        )
        from app.models.oferta_ml import OfertaML
        from app.models.publicacion_ml import PublicacionML
        from datetime import date

        productos_filtrados = []
        hoy = date.today()

        for producto in productos:
            item_id = producto[0]
            incluir = True

            # Obtener ProductoERP para el item_id
            producto_erp = db.query(ProductoERP).filter(ProductoERP.item_id == item_id).first()
            if not producto_erp:
                continue

            producto_pricing = db.query(ProductoPricing).filter(ProductoPricing.item_id == item_id).first()
            if not producto_pricing:
                continue

            # Filtro de markup clásica
            if markup_clasica_positivo is not None and incluir:
                markup = producto_pricing.markup_calculado if producto_pricing else None
                if markup is not None:
                    if markup_clasica_positivo and markup < 0:
                        incluir = False
                    elif not markup_clasica_positivo and markup >= 0:
                        incluir = False
                else:
                    incluir = False

            # Filtro de markup rebate
            if markup_rebate_positivo is not None and incluir:
                if producto_pricing and producto_pricing.participa_rebate and producto_pricing.precio_lista_ml:
                    try:
                        precio_rebate = float(producto_pricing.precio_lista_ml) * (
                            1 + float(producto_pricing.porcentaje_rebate or 3.8) / 100
                        )
                        tipo_cambio = (
                            obtener_tipo_cambio_actual(db, "USD") if producto_erp.moneda_costo == "USD" else None
                        )
                        costo_ars = convertir_a_pesos(producto_erp.costo, producto_erp.moneda_costo, tipo_cambio)
                        grupo_id = obtener_grupo_subcategoria(db, producto_erp.subcategoria_id)
                        comision_base = obtener_comision_base(db, 4, grupo_id)  # 4 = Clásica
                        if comision_base:
                            comisiones = calcular_comision_ml_total(
                                precio_rebate, comision_base, producto_erp.iva, db=db
                            )
                            limpio = calcular_limpio(
                                precio_rebate,
                                producto_erp.iva,
                                producto_erp.envio or 0,
                                comisiones["comision_total"],
                                db=db,
                                grupo_id=grupo_id,
                            )
                            markup_rebate = calcular_markup(limpio, costo_ars) * 100
                            if markup_rebate_positivo and markup_rebate < 0:
                                incluir = False
                            elif not markup_rebate_positivo and markup_rebate >= 0:
                                incluir = False
                        else:
                            incluir = False
                    except Exception as e:
                        logger.warning("Error calculando markup rebate para item %s: %s", item_id, e)
                        incluir = False
                else:
                    incluir = False

            # Filtro de oferta
            if con_oferta is not None and incluir:
                pubs = db.query(PublicacionML).filter(PublicacionML.item_id == item_id).all()
                tiene_oferta = False
                for pub in pubs:
                    oferta = (
                        db.query(OfertaML)
                        .filter(
                            OfertaML.mla == pub.mla,
                            OfertaML.fecha_desde <= hoy,
                            OfertaML.fecha_hasta >= hoy,
                            OfertaML.pvp_seller.isnot(None),
                        )
                        .first()
                    )
                    if oferta:
                        tiene_oferta = True
                        break
                if con_oferta and not tiene_oferta:
                    incluir = False
                elif not con_oferta and tiene_oferta:
                    incluir = False

            # Filtro de markup oferta
            if markup_oferta_positivo is not None and incluir:
                # Implementación similar a markup rebate pero con precio de oferta
                # Por simplicidad, si llegó hasta acá y tiene oferta, se incluye
                pass

            # Filtro de markup web transferencia
            if markup_web_transf_positivo is not None and incluir:
                if producto_pricing and producto_pricing.markup_web_real is not None:
                    markup_web = float(producto_pricing.markup_web_real)
                    if markup_web_transf_positivo and markup_web < 0:
                        incluir = False
                    elif not markup_web_transf_positivo and markup_web >= 0:
                        incluir = False
                else:
                    incluir = False

            if incluir:
                productos_filtrados.append(producto)

        productos = productos_filtrados

    # Obtener dólar venta si currency_id es 2
    dolar_ajustado = None
    if currency_id == 2:
        tipo_cambio = db.query(TipoCambio).order_by(TipoCambio.id.desc()).first()
        if tipo_cambio:
            dolar_ajustado = float(tipo_cambio.venta) + offset_dolar

    # Determinar qué prli_ids corresponden al tipo de cuotas seleccionado
    # Cada tipo tiene una lista Web y una PVP (ambas representan el mismo precio)
    # Mapeo tipo_cuotas -> [prli_id_web, prli_id_pvp]
    tipo_cuotas_to_prli = {
        "clasica": [4, 12],  # Clásica Web + PVP
        "3": [17, 18],  # 3 Cuotas Web + PVP
        "6": [14, 19],  # 6 Cuotas Web + PVP
        "9": [13, 20],  # 9 Cuotas Web + PVP
        "12": [23, 21],  # 12 Cuotas Web + PVP
        "pvp": [12],  # PVP Base (solo lista 12)
        "pvp_3": [18],  # PVP 3 Cuotas (solo lista 18)
        "pvp_6": [19],  # PVP 6 Cuotas (solo lista 19)
        "pvp_9": [20],  # PVP 9 Cuotas (solo lista 20)
        "pvp_12": [21],  # PVP 12 Cuotas (solo lista 21)
    }

    prli_ids_seleccionados = tipo_cuotas_to_prli.get(tipo_cuotas, [])

    # Obtener MLA IDs para cada producto de la lista seleccionada
    from app.models.mercadolibre_item_publicado import MercadoLibreItemPublicado

    # Filtro de tiendas oficiales (MLA): se inyecta como AND directo en la query
    # primaria de MercadoLibreItemPublicado (no necesita subquery porque la query
    # ya está sobre ese modelo). Distinto de `tienda_oficial` (singular, scope producto).
    parsed_tiendas_oficiales = _parsear_tiendas_oficiales_mla(tiendas_oficiales)
    filtro_tiendas_oficiales_mla = (
        _build_filtro_tiendas_oficiales_mla(parsed_tiendas_oficiales) if parsed_tiendas_oficiales is not None else None
    )

    # Crear diccionario: item_id -> [mla_ids]
    mla_por_item = {}

    if prli_ids_seleccionados:
        # Consultar publicaciones de AMBAS listas (Web y PVP) para el tipo seleccionado
        # optval_statusId: 2 = Publicada, 3 = Pausada, 5 = Finalizada, 6 = Pausada Forzada, 10 = Des-Enlazada
        item_ids = [p[0] for p in productos]

        query_publicaciones = db.query(
            MercadoLibreItemPublicado.item_id, MercadoLibreItemPublicado.mlp_publicationID
        ).filter(
            MercadoLibreItemPublicado.item_id.in_(item_ids),
            MercadoLibreItemPublicado.prli_id.in_(prli_ids_seleccionados),
            MercadoLibreItemPublicado.mlp_id.isnot(None),
            # Incluir publicadas (2), pausadas (3), pausadas forzadas (6)
            # Excluir finalizadas (5) y des-enlazadas (10)
            or_(
                MercadoLibreItemPublicado.optval_statusId == 2,
                MercadoLibreItemPublicado.optval_statusId == 3,
                MercadoLibreItemPublicado.optval_statusId == 6,
                MercadoLibreItemPublicado.optval_statusId.is_(None),
            ),
        )
        if filtro_tiendas_oficiales_mla is not None:
            query_publicaciones = query_publicaciones.filter(filtro_tiendas_oficiales_mla)

        publicaciones = query_publicaciones.all()

        # Agrupar MLAs por item_id
        for item_id, mla_id in publicaciones:
            if item_id not in mla_por_item:
                mla_por_item[item_id] = []
            mla_por_item[item_id].append(mla_id)

    # Determinar el número máximo de MLAs que tiene cualquier producto
    max_mlas = max([len(mlas) for mlas in mla_por_item.values()]) if mla_por_item else 0

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = tipo_cuotas.title()

    # Header - Columnas base + una columna por cada MLA
    header = ["Código/EAN", "Precio", "ID Moneda"]
    for i in range(max_mlas):
        header.append(f"MLA {i + 1}")
    ws.append(header)

    # Datos
    for (
        item_id,
        codigo,
        precio_clasica,
        participa_rebate,
        porcentaje_rebate,
        precio_3,
        precio_6,
        precio_9,
        precio_12,
        precio_pvp,
        precio_pvp_3,
        precio_pvp_6,
        precio_pvp_9,
        precio_pvp_12,
    ) in productos:
        # Determinar qué precio usar según tipo_cuotas
        if tipo_cuotas == "clasica":
            # Si tiene rebate activo, calcular precio rebate y aplicar % adicional
            if participa_rebate and porcentaje_rebate:
                precio_rebate = precio_clasica * (1 + float(porcentaje_rebate) / 100)
                precio_exportar = precio_rebate * (1 + porcentaje_adicional / 100)
            else:
                # Si no tiene rebate, usar precio clásica sin modificar
                precio_exportar = precio_clasica
        elif tipo_cuotas == "3":
            # Si no hay precio de 3 cuotas, saltar este producto
            if not precio_3:
                continue
            precio_exportar = float(precio_3)
        elif tipo_cuotas == "6":
            # Si no hay precio de 6 cuotas, saltar este producto
            if not precio_6:
                continue
            precio_exportar = float(precio_6)
        elif tipo_cuotas == "9":
            # Si no hay precio de 9 cuotas, saltar este producto
            if not precio_9:
                continue
            precio_exportar = float(precio_9)
        elif tipo_cuotas == "12":
            # Si no hay precio de 12 cuotas, saltar este producto
            if not precio_12:
                continue
            precio_exportar = float(precio_12)
        elif tipo_cuotas == "pvp":
            # Usar precio PVP clásico (sin cuotas)
            if not precio_pvp:
                continue
            precio_exportar = float(precio_pvp)
        elif tipo_cuotas == "pvp_3":
            # Si no hay precio PVP 3 cuotas, saltar este producto
            if not precio_pvp_3:
                continue
            precio_exportar = float(precio_pvp_3)
        elif tipo_cuotas == "pvp_6":
            # Si no hay precio PVP 6 cuotas, saltar este producto
            if not precio_pvp_6:
                continue
            precio_exportar = float(precio_pvp_6)
        elif tipo_cuotas == "pvp_9":
            # Si no hay precio PVP 9 cuotas, saltar este producto
            if not precio_pvp_9:
                continue
            precio_exportar = float(precio_pvp_9)
        elif tipo_cuotas == "pvp_12":
            # Si no hay precio PVP 12 cuotas, saltar este producto
            if not precio_pvp_12:
                continue
            precio_exportar = float(precio_pvp_12)
        else:
            precio_exportar = precio_clasica

        # Si es USD, dividir por dólar ajustado
        if currency_id == 2 and dolar_ajustado:
            precio_final = precio_exportar / dolar_ajustado
            # Para USD, redondear a 2 decimales
            precio_str = f"{precio_final:.2f}"
        elif tipo_cuotas.startswith("pvp"):
            # Para PVP en ARS, exportar precio exacto sin redondear
            precio_str = f"{float(precio_exportar):.2f}"
        else:
            # Para ARS (clásica/cuotas), redondear a múltiplo de 10
            precio_final = round(precio_exportar / 10) * 10
            precio_str = str(int(precio_final))

        # Obtener MLAs de la lista seleccionada para este item
        mlas = mla_por_item.get(item_id, [])

        # Crear fila con columnas base
        fila = [str(codigo), precio_str, str(currency_id)]

        # Agregar cada MLA en su propia columna
        for i in range(max_mlas):
            if i < len(mlas):
                fila.append(mlas[i])
            else:
                fila.append("")  # Columna vacía si no hay MLA

        ws.append(fila)

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=clasica.xlsx"},
    )


@router.get("/exportar-vista-actual")
def exportar_vista_actual(
    page: int = 1,
    page_size: int = 10000,
    search: Optional[str] = None,
    con_stock: Optional[bool] = None,
    con_precio: Optional[bool] = None,
    marcas: Optional[str] = None,
    subcategorias: Optional[str] = None,
    con_rebate: Optional[bool] = None,
    con_oferta: Optional[bool] = None,
    con_web_transf: Optional[bool] = None,
    tiendanube_con_descuento: Optional[bool] = None,
    tiendanube_sin_descuento: Optional[bool] = None,
    tiendanube_no_publicado: Optional[bool] = None,
    markup_clasica_positivo: Optional[bool] = None,
    markup_rebate_positivo: Optional[bool] = None,
    markup_oferta_positivo: Optional[bool] = None,
    markup_web_transf_positivo: Optional[bool] = None,
    out_of_cards: Optional[bool] = None,
    colores: Optional[str] = None,
    pms: Optional[str] = None,
    audit_usuarios: Optional[str] = None,
    audit_tipos_accion: Optional[str] = None,
    audit_fecha_desde: Optional[str] = None,
    audit_fecha_hasta: Optional[str] = None,
    estado_mla: Optional[str] = None,
    equipo_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Exporta la vista actual de productos a Excel con todos los datos"""
    try:
        import logging

        logger = logging.getLogger(__name__)
        logger.info(
            f"Exportar vista actual - Filtros TN: con_descuento={tiendanube_con_descuento}, sin_descuento={tiendanube_sin_descuento}, no_publicado={tiendanube_no_publicado}"
        )

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        from fastapi.responses import StreamingResponse

        layer_activo = resolver_layer_activo(equipo_id, current_user, db)

        # Usar la misma lógica de obtener_productos para filtrar
        query = db.query(ProductoERP, ProductoPricing).outerjoin(
            ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id
        )
        query = join_color_layer(query, layer_activo)

        # Aplicar todos los filtros (reutilizar la lógica del endpoint obtener_productos)
        query = _apply_search_filter(query, search)

        if con_stock is not None:
            query = query.filter(ProductoERP.stock > 0 if con_stock else ProductoERP.stock == 0)

        if con_precio is not None:
            if con_precio:
                query = query.filter(ProductoPricing.precio_lista_ml.isnot(None))
            else:
                query = query.filter(ProductoPricing.precio_lista_ml.is_(None))

        if marcas:
            marcas_list = [m.strip().upper() for m in marcas.split(",")]
            query = query.filter(func.upper(ProductoERP.marca).in_(marcas_list))

        if subcategorias:
            subcat_list = [int(s.strip()) for s in subcategorias.split(",")]
            query = query.filter(ProductoERP.subcategoria_id.in_(subcat_list))

        if con_rebate is not None:
            if con_rebate:
                query = query.filter(ProductoPricing.participa_rebate == True)
            else:
                query = query.filter(
                    or_(ProductoPricing.participa_rebate == False, ProductoPricing.participa_rebate.is_(None))
                )

        if con_oferta is not None:
            if con_oferta:
                query = query.filter(ProductoPricing.precio_3_cuotas.isnot(None))
            else:
                query = query.filter(ProductoPricing.precio_3_cuotas.is_(None))

        if con_web_transf is not None:
            if con_web_transf:
                query = query.filter(ProductoPricing.participa_web_transferencia == True)
            else:
                query = query.filter(
                    or_(
                        ProductoPricing.participa_web_transferencia == False,
                        ProductoPricing.participa_web_transferencia.is_(None),
                    )
                )

        if tiendanube_con_descuento:
            logger.info("Aplicando filtro: tiendanube_con_descuento")
            query = query.filter(
                ProductoPricing.descuento_tiendanube.isnot(None), ProductoPricing.descuento_tiendanube > 0
            )

        if tiendanube_sin_descuento:
            logger.info("Aplicando filtro: tiendanube_sin_descuento")
            query = query.filter(
                or_(ProductoPricing.descuento_tiendanube.is_(None), ProductoPricing.descuento_tiendanube == 0)
            )

        if tiendanube_no_publicado:
            logger.info("Aplicando filtro: tiendanube_no_publicado")
            # Productos con stock pero NO en Tienda Nube
            from app.models.tienda_nube_producto import TiendaNubeProducto
            from sqlalchemy.sql import exists

            subquery = exists().where(
                and_(TiendaNubeProducto.item_id == ProductoERP.item_id, TiendaNubeProducto.activo == True)
            )
            query = query.filter(and_(ProductoERP.stock > 0, ~subquery))

        if out_of_cards is not None:
            if out_of_cards:
                query = query.filter(ProductoPricing.out_of_cards == True)
            else:
                query = query.filter(or_(ProductoPricing.out_of_cards == False, ProductoPricing.out_of_cards.is_(None)))

        if markup_clasica_positivo is not None:
            if markup_clasica_positivo:
                query = query.filter(ProductoPricing.markup_calculado > 0)
            else:
                query = query.filter(ProductoPricing.markup_calculado <= 0)

        if markup_rebate_positivo is not None:
            if markup_rebate_positivo:
                query = query.filter(ProductoPricing.markup_rebate > 0)
            else:
                query = query.filter(ProductoPricing.markup_rebate <= 0)

        if markup_oferta_positivo is not None:
            if markup_oferta_positivo:
                query = query.filter(ProductoPricing.markup_oferta > 0)
            else:
                query = query.filter(ProductoPricing.markup_oferta <= 0)

        if markup_web_transf_positivo is not None:
            if markup_web_transf_positivo:
                query = query.filter(ProductoPricing.markup_web_real > 0)
            else:
                query = query.filter(ProductoPricing.markup_web_real <= 0)

        # Filtro de colores (lee del layer de equipo activo, ver productos_shared)
        query = filtro_colores(query, colores, color_slot(None))

        if pms:
            from app.models.marca_pm import MarcaPM

            pms_ids = [int(pm) for pm in pms.split(",")]
            pares_pm = db.query(MarcaPM.marca, MarcaPM.categoria).filter(MarcaPM.usuario_id.in_(pms_ids)).all()
            if pares_pm:
                pares_upper = [(m.upper(), c.upper()) for m, c in pares_pm]
                query = query.filter(
                    tuple_(func.upper(ProductoERP.marca), func.upper(ProductoERP.categoria)).in_(pares_upper)
                )
            else:
                query = query.filter(ProductoERP.item_id == -1)

        # Filtros de auditoría
        if audit_usuarios or audit_tipos_accion or audit_fecha_desde or audit_fecha_hasta:
            from app.models.auditoria import Auditoria
            from datetime import datetime, timedelta

            subquery_filters = []
            if audit_usuarios:
                usuarios_list = [int(u.strip()) for u in audit_usuarios.split(",")]
                subquery_filters.append(Auditoria.usuario_id.in_(usuarios_list))
            if audit_tipos_accion:
                tipos_list = audit_tipos_accion.split(",")
                subquery_filters.append(Auditoria.tipo_accion.in_(tipos_list))
            if audit_fecha_desde:
                try:
                    fecha_inicio = datetime.strptime(audit_fecha_desde, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    try:
                        fecha_inicio = datetime.strptime(audit_fecha_desde, "%Y-%m-%d %H:%M")
                    except ValueError:
                        try:
                            fecha_inicio = datetime.strptime(audit_fecha_desde, "%Y-%m-%d")
                        except ValueError:
                            from datetime import date

                            fecha_inicio = datetime.combine(date.today(), datetime.min.time())
                subquery_filters.append(Auditoria.fecha >= fecha_inicio)
            if audit_fecha_hasta:
                try:
                    fecha_fin = datetime.strptime(audit_fecha_hasta, "%Y-%m-%d %H:%M:%S") + timedelta(days=1)
                except ValueError:
                    try:
                        fecha_fin = datetime.strptime(audit_fecha_hasta, "%Y-%m-%d %H:%M") + timedelta(days=1)
                    except ValueError:
                        try:
                            fecha_fin = datetime.strptime(audit_fecha_hasta, "%Y-%m-%d") + timedelta(days=1)
                        except ValueError:
                            from datetime import date

                            fecha_fin = datetime.combine(date.today(), datetime.max.time())
                subquery_filters.append(Auditoria.fecha < fecha_fin)

            if subquery_filters:
                from sqlalchemy import exists

                query = query.filter(exists().where(and_(Auditoria.item_id == ProductoERP.item_id, *subquery_filters)))

        # Filtro de estado de publicaciones MLA
        if estado_mla:
            from app.models.mercadolibre_item_publicado import MercadoLibreItemPublicado

            if estado_mla == "activa":
                # Tienen al menos una publicación activa
                items_activos_subquery = (
                    db.query(MercadoLibreItemPublicado.item_id)
                    .filter(
                        MercadoLibreItemPublicado.mlp_id.isnot(None),
                        or_(
                            MercadoLibreItemPublicado.optval_statusId == 2,
                            MercadoLibreItemPublicado.optval_statusId.is_(None),
                        ),
                    )
                    .distinct()
                    .subquery()
                )

                query = query.filter(ProductoERP.item_id.in_(select(items_activos_subquery.c.item_id)))

            elif estado_mla == "pausada":
                # Tienen publicaciones pero ninguna activa
                items_con_publis = (
                    db.query(MercadoLibreItemPublicado.item_id)
                    .filter(MercadoLibreItemPublicado.mlp_id.isnot(None))
                    .distinct()
                    .subquery()
                )

                items_activos = (
                    db.query(MercadoLibreItemPublicado.item_id)
                    .filter(
                        MercadoLibreItemPublicado.mlp_id.isnot(None),
                        or_(
                            MercadoLibreItemPublicado.optval_statusId == 2,
                            MercadoLibreItemPublicado.optval_statusId.is_(None),
                        ),
                    )
                    .distinct()
                    .subquery()
                )

                query = query.filter(
                    ProductoERP.item_id.in_(select(items_con_publis.c.item_id)),
                    ~ProductoERP.item_id.in_(select(items_activos.c.item_id)),
                )

        # Ejecutar query
        productos = query.limit(page_size).offset((page - 1) * page_size).all()
        logger.info(f"Se encontraron {len(productos)} productos para exportar")

        # Batch-fetch colors for the active layer (T-color-layer).
        colores_activo_export = batch_colores(db, [p.item_id for p, _ in productos], layer_activo)

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Vista Actual"

        # Encabezados
        headers = [
            "Código",
            "Descripción",
            "Marca",
            "Stock",
            "Costo",
            "Precio Clásica",
            "Markup Clásica (%)",
            "Precio Rebate",
            "Markup Rebate (%)",
            "Precio Oferta",
            "Markup Oferta (%)",
            "Precio Web Transf",
            "Markup Web (%)",
            "Precio TN",
            "Descuento TN (%)",
            "Publicado TN",
            "Out of Cards",
            "Color",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Datos
        row_num = 2
        for producto_erp, producto_pricing in productos:
            ws.cell(row=row_num, column=1, value=producto_erp.codigo or "")
            ws.cell(row=row_num, column=2, value=producto_erp.descripcion or "")
            ws.cell(row=row_num, column=3, value=producto_erp.marca or "")
            ws.cell(row=row_num, column=4, value=producto_erp.stock or 0)
            ws.cell(row=row_num, column=5, value=float(producto_erp.costo) if producto_erp.costo else 0)

            if producto_pricing:
                ws.cell(
                    row=row_num,
                    column=6,
                    value=float(producto_pricing.precio_lista_ml) if producto_pricing.precio_lista_ml else None,
                )
                ws.cell(
                    row=row_num,
                    column=7,
                    value=float(producto_pricing.markup_calculado) if producto_pricing.markup_calculado else None,
                )

                # Calcular precio rebate dinámicamente (misma fórmula que el listado)
                precio_rebate = None
                if producto_pricing.participa_rebate and producto_pricing.precio_lista_ml:
                    porcentaje_rebate = float(producto_pricing.porcentaje_rebate or 3.8)
                    precio_rebate = float(producto_pricing.precio_lista_ml) / (1 - porcentaje_rebate / 100)
                ws.cell(row=row_num, column=8, value=precio_rebate)

                ws.cell(
                    row=row_num,
                    column=9,
                    value=float(producto_pricing.markup_rebate) if producto_pricing.markup_rebate else None,
                )
                ws.cell(
                    row=row_num,
                    column=10,
                    value=float(producto_pricing.precio_3_cuotas) if producto_pricing.precio_3_cuotas else None,
                )
                ws.cell(
                    row=row_num,
                    column=11,
                    value=float(producto_pricing.markup_oferta) if producto_pricing.markup_oferta else None,
                )
                ws.cell(
                    row=row_num,
                    column=12,
                    value=float(producto_pricing.precio_web_transferencia)
                    if producto_pricing.precio_web_transferencia
                    else None,
                )
                ws.cell(
                    row=row_num,
                    column=13,
                    value=float(producto_pricing.markup_web_real) if producto_pricing.markup_web_real else None,
                )

                # Tienda Nube
                ws.cell(
                    row=row_num,
                    column=14,
                    value=float(producto_pricing.precio_tiendanube) if producto_pricing.precio_tiendanube else None,
                )
                ws.cell(
                    row=row_num,
                    column=15,
                    value=float(producto_pricing.descuento_tiendanube)
                    if producto_pricing.descuento_tiendanube
                    else None,
                )
                ws.cell(row=row_num, column=16, value="Sí" if producto_pricing.publicado_tiendanube else "No")

                ws.cell(row=row_num, column=17, value="Sí" if producto_pricing.out_of_cards else "No")
                _color_activo_row = colores_activo_export.get(producto_erp.item_id)
                ws.cell(row=row_num, column=18, value=(_color_activo_row.color_ml if _color_activo_row else None) or "")

            row_num += 1

        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception as e:
                    logger.warning("Error calculando ancho de columna %s: %s", column_letter, e)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=vista_actual.xlsx"},
        )
    except Exception as e:
        logger.error("Error en exportar_vista_actual: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al exportar vista actual: {str(e)}")


@router.get("/exportar-lista-gremio")
def exportar_lista_gremio(
    search: Optional[str] = None,
    con_stock: Optional[bool] = None,
    marcas: Optional[str] = None,
    subcategorias: Optional[str] = None,
    colores: Optional[str] = None,
    con_precio_gremio: Optional[bool] = None,
    currency_id: int = 1,  # 1=ARS, 2=USD
    offset_dolar: float = 0,  # Offset para el tipo de cambio
    equipo_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Exporta Lista Gremio a Excel con precios calculados. Soporta ARS y USD."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from sqlalchemy import text
    from app.models.markup_tienda import MarkupTiendaBrand, MarkupTiendaProducto
    from app.services.pricing_calculator import obtener_constantes_pricing, obtener_tipo_cambio_actual

    try:
        # Obtener constantes y tipo de cambio
        constantes = obtener_constantes_pricing(db)
        varios_porcentaje = constantes.get("varios", 7)
        tipo_cambio = obtener_tipo_cambio_actual(db, "USD")

        # Ajustar tipo de cambio con offset
        tipo_cambio_ajustado = tipo_cambio + offset_dolar if tipo_cambio else None

        # Cargar markups de tienda
        markups_marca = db.query(MarkupTiendaBrand).filter(MarkupTiendaBrand.activo == True).all()
        markups_marca_dict = {m.brand_desc.upper(): m.markup_porcentaje for m in markups_marca if m.brand_desc}

        markups_producto = db.query(MarkupTiendaProducto).filter(MarkupTiendaProducto.activo == True).all()
        markups_producto_dict = {m.item_id: m.markup_porcentaje for m in markups_producto}

        # Cargar overrides manuales
        from app.models.precio_gremio_override import PrecioGremioOverride

        precio_gremio_overrides = {}
        overrides = db.query(PrecioGremioOverride).all()
        precio_gremio_overrides = {o.item_id: o for o in overrides}

        # Cargar nombres de subcategorías
        subcats_result = db.execute(text("SELECT subcat_id, subcat_desc FROM tb_subcategory"))
        subcats_dict = {row.subcat_id: row.subcat_desc for row in subcats_result}

        layer_activo = resolver_layer_activo(equipo_id, current_user, db)

        # Query base
        query = db.query(ProductoERP, ProductoPricing).outerjoin(
            ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id
        )
        query = join_color_layer(query, layer_activo)

        # Aplicar filtros
        query = _apply_search_filter(query, search)

        if con_stock is not None:
            query = query.filter(ProductoERP.stock > 0 if con_stock else ProductoERP.stock == 0)

        if marcas:
            marcas_list = [m.strip().upper() for m in marcas.split(",")]
            query = query.filter(func.upper(ProductoERP.marca).in_(marcas_list))

        if subcategorias:
            subcat_list = [int(s.strip()) for s in subcategorias.split(",")]
            query = query.filter(ProductoERP.subcategoria_id.in_(subcat_list))

        # Filtro de colores (vista tienda, lee del layer de equipo activo)
        query = filtro_colores(query, colores, color_slot("tienda"))

        # Ejecutar query
        results = query.order_by(ProductoERP.marca, ProductoERP.codigo).all()

        # Función para convertir a pesos
        def convertir_a_pesos(costo, moneda):
            if costo is None:
                return None
            costo_float = float(costo)
            if moneda and moneda.upper() == "USD" and tipo_cambio:
                return costo_float * tipo_cambio
            return costo_float

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista Gremio"

        # Headers - cambiar según moneda
        moneda_texto = "USD" if currency_id == 2 else "ARS"
        headers = [
            "Marca",
            "Categoría",
            "Subcategoría",
            "Código",
            "Descripción",
            "Stock",
            f"Precio Gremio {moneda_texto} s/IVA",
            f"Precio Gremio {moneda_texto} c/IVA",
        ]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Datos
        row_num = 2
        for producto_erp, producto_pricing in results:
            # Calcular costo en ARS
            costo_ars = convertir_a_pesos(producto_erp.costo, producto_erp.moneda_costo)

            # Calcular precio gremio - Verificar override manual primero
            precio_gremio_sin_iva = None
            precio_gremio_con_iva = None
            markup_gremio = None

            # Si existe override manual, usar esos precios
            if producto_erp.item_id in precio_gremio_overrides:
                override = precio_gremio_overrides[producto_erp.item_id]
                precio_gremio_sin_iva = float(override.precio_gremio_sin_iva_manual)
                precio_gremio_con_iva = float(override.precio_gremio_con_iva_manual)
            else:
                # Calcular automáticamente según reglas
                # Primero buscar markup por producto
                if producto_erp.item_id in markups_producto_dict:
                    markup_gremio = markups_producto_dict[producto_erp.item_id]
                # Si no, buscar por marca
                elif producto_erp.marca and producto_erp.marca.upper() in markups_marca_dict:
                    markup_gremio = markups_marca_dict[producto_erp.marca.upper()]

                if markup_gremio is not None and costo_ars and costo_ars > 0:
                    precio_gremio_sin_iva = costo_ars * (1 + varios_porcentaje / 100) * (1 + markup_gremio / 100)
                    iva_producto = producto_erp.iva if producto_erp.iva else 21.0
                    precio_gremio_con_iva = precio_gremio_sin_iva * (1 + iva_producto / 100)

            # Convertir a USD si es necesario
            if currency_id == 2 and tipo_cambio_ajustado and tipo_cambio_ajustado > 0:
                if precio_gremio_sin_iva:
                    precio_gremio_sin_iva = precio_gremio_sin_iva / tipo_cambio_ajustado
                if precio_gremio_con_iva:
                    precio_gremio_con_iva = precio_gremio_con_iva / tipo_cambio_ajustado

            # Filtro de solo productos con precio gremio
            if con_precio_gremio and precio_gremio_sin_iva is None:
                continue

            ws.cell(row=row_num, column=1, value=producto_erp.marca or "")
            ws.cell(row=row_num, column=2, value=producto_erp.categoria or "")
            ws.cell(row=row_num, column=3, value=subcats_dict.get(producto_erp.subcategoria_id, "") or "")
            ws.cell(row=row_num, column=4, value=producto_erp.codigo or "")
            ws.cell(row=row_num, column=5, value=producto_erp.descripcion or "")
            ws.cell(row=row_num, column=6, value=producto_erp.stock or 0)
            ws.cell(row=row_num, column=7, value=round(precio_gremio_sin_iva, 2) if precio_gremio_sin_iva else None)
            ws.cell(row=row_num, column=8, value=round(precio_gremio_con_iva, 2) if precio_gremio_con_iva else None)

            row_num += 1

        # Ajustar anchos de columna
        column_widths = [15, 20, 20, 15, 50, 10, 18, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=lista_gremio.xlsx"},
        )

    except Exception as e:
        logger.error("Error en exportar_lista_gremio: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al exportar lista gremio: {str(e)}")


@router.get("/exportar-lista-sugerido")
def exportar_lista_sugerido(
    search: Optional[str] = None,
    con_stock: Optional[bool] = None,
    marcas: Optional[str] = None,
    subcategorias: Optional[str] = None,
    colores: Optional[str] = None,
    con_precio_sugerido: Optional[bool] = None,
    currency_id: int = 1,  # 1=ARS, 2=USD
    offset_dolar: float = 0,
    equipo_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Exporta Lista Precio Sugerido a Excel. Fórmula: costo * (1+varios%) * (1 + (markup_clasica + markup_sugerido)%)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from sqlalchemy import text
    from app.models.markup_tienda import MarkupTiendaBrand, MarkupTiendaProducto
    from app.services.pricing_calculator import obtener_constantes_pricing, obtener_tipo_cambio_actual

    try:
        # Obtener constantes y tipo de cambio
        constantes = obtener_constantes_pricing(db)
        varios_porcentaje = constantes.get("varios", 7)
        tipo_cambio = obtener_tipo_cambio_actual(db, "USD")

        # Ajustar tipo de cambio con offset
        tipo_cambio_ajustado = tipo_cambio + offset_dolar if tipo_cambio else None

        # Cargar markups sugerido de tienda
        markups_marca = db.query(MarkupTiendaBrand).filter(MarkupTiendaBrand.activo == True).all()
        markups_sugerido_marca_dict = {
            m.brand_desc.upper(): m.markup_sugerido
            for m in markups_marca
            if m.brand_desc and m.markup_sugerido is not None
        }

        markups_producto = db.query(MarkupTiendaProducto).filter(MarkupTiendaProducto.activo == True).all()
        markups_sugerido_producto_dict = {
            m.item_id: m.markup_sugerido for m in markups_producto if m.markup_sugerido is not None
        }

        # Cargar nombres de subcategorías
        subcats_result = db.execute(text("SELECT subcat_id, subcat_desc FROM tb_subcategory"))
        subcats_dict = {row.subcat_id: row.subcat_desc for row in subcats_result}

        layer_activo = resolver_layer_activo(equipo_id, current_user, db)

        # Query base
        query = db.query(ProductoERP, ProductoPricing).outerjoin(
            ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id
        )
        query = join_color_layer(query, layer_activo)

        # Aplicar filtros
        query = _apply_search_filter(query, search)

        if con_stock is not None:
            query = query.filter(ProductoERP.stock > 0 if con_stock else ProductoERP.stock == 0)

        if marcas:
            marcas_list = [m.strip().upper() for m in marcas.split(",")]
            query = query.filter(func.upper(ProductoERP.marca).in_(marcas_list))

        if subcategorias:
            subcat_list = [int(s.strip()) for s in subcategorias.split(",")]
            query = query.filter(ProductoERP.subcategoria_id.in_(subcat_list))

        # Filtro de colores (vista tienda, lee del layer de equipo activo)
        query = filtro_colores(query, colores, color_slot("tienda"))

        # Ejecutar query
        results = query.order_by(ProductoERP.marca, ProductoERP.codigo).all()

        # Función para convertir a pesos
        def convertir_a_pesos(costo, moneda):
            if costo is None:
                return None
            costo_float = float(costo)
            if moneda and moneda.upper() == "USD" and tipo_cambio:
                return costo_float * tipo_cambio
            return costo_float

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista Sugerido"

        # Headers
        moneda_texto = "USD" if currency_id == 2 else "ARS"
        headers = [
            "Marca",
            "Categoría",
            "Subcategoría",
            "Código",
            "Descripción",
            "Stock",
            "Markup Clásica %",
            "Markup Sugerido %",
            "Markup Total %",
            f"Precio Sugerido {moneda_texto} s/IVA",
            f"Precio Sugerido {moneda_texto} c/IVA",
        ]
        header_fill = PatternFill(start_color="4A7C59", end_color="4A7C59", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Datos
        row_num = 2
        for producto_erp, producto_pricing in results:
            # Calcular costo en ARS
            costo_ars = convertir_a_pesos(producto_erp.costo, producto_erp.moneda_costo)

            # Resolver markup_sugerido (producto > marca)
            markup_sugerido_valor = None
            if producto_erp.item_id in markups_sugerido_producto_dict:
                markup_sugerido_valor = markups_sugerido_producto_dict[producto_erp.item_id]
            elif producto_erp.marca and producto_erp.marca.upper() in markups_sugerido_marca_dict:
                markup_sugerido_valor = markups_sugerido_marca_dict[producto_erp.marca.upper()]

            # markup_clasica viene de markup_calculado
            markup_clasica = producto_pricing.markup_calculado if producto_pricing else None

            # Calcular precio sugerido
            precio_sugerido_sin_iva = None
            precio_sugerido_con_iva = None
            markup_total = None

            if markup_clasica is not None and costo_ars and costo_ars > 0:
                # Si no hay markup_sugerido configurado, usar 0 (precio = solo markup_clasica)
                effective_sugerido = markup_sugerido_valor if markup_sugerido_valor is not None else 0.0
                markup_total = markup_clasica + effective_sugerido
                precio_sugerido_sin_iva = costo_ars * (1 + varios_porcentaje / 100) * (1 + markup_total / 100)
                iva_producto = producto_erp.iva if producto_erp.iva else 21.0
                precio_sugerido_con_iva = precio_sugerido_sin_iva * (1 + iva_producto / 100)

            # Convertir a USD si es necesario
            if currency_id == 2 and tipo_cambio_ajustado and tipo_cambio_ajustado > 0:
                if precio_sugerido_sin_iva:
                    precio_sugerido_sin_iva = precio_sugerido_sin_iva / tipo_cambio_ajustado
                if precio_sugerido_con_iva:
                    precio_sugerido_con_iva = precio_sugerido_con_iva / tipo_cambio_ajustado

            # Filtro de solo productos con precio sugerido
            if con_precio_sugerido and precio_sugerido_sin_iva is None:
                continue

            ws.cell(row=row_num, column=1, value=producto_erp.marca or "")
            ws.cell(row=row_num, column=2, value=producto_erp.categoria or "")
            ws.cell(row=row_num, column=3, value=subcats_dict.get(producto_erp.subcategoria_id, "") or "")
            ws.cell(row=row_num, column=4, value=producto_erp.codigo or "")
            ws.cell(row=row_num, column=5, value=producto_erp.descripcion or "")
            ws.cell(row=row_num, column=6, value=producto_erp.stock or 0)
            ws.cell(row=row_num, column=7, value=round(markup_clasica, 2) if markup_clasica is not None else None)
            ws.cell(
                row=row_num,
                column=8,
                value=round(markup_sugerido_valor, 2) if markup_sugerido_valor is not None else None,
            )
            ws.cell(row=row_num, column=9, value=round(markup_total, 2) if markup_total is not None else None)
            ws.cell(
                row=row_num, column=10, value=round(precio_sugerido_sin_iva, 2) if precio_sugerido_sin_iva else None
            )
            ws.cell(
                row=row_num, column=11, value=round(precio_sugerido_con_iva, 2) if precio_sugerido_con_iva else None
            )

            row_num += 1

        # Ajustar anchos de columna
        column_widths = [15, 20, 20, 15, 50, 10, 16, 16, 14, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=lista_sugerido.xlsx"},
        )

    except Exception as e:
        logger.error("Error en exportar_lista_sugerido: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al exportar lista sugerido: {str(e)}")


@router.get("/exportar-lista-web-transferencia")
def exportar_lista_web_transferencia(
    search: Optional[str] = None,
    con_stock: Optional[bool] = None,
    marcas: Optional[str] = None,
    subcategorias: Optional[str] = None,
    colores: Optional[str] = None,
    currency_id: int = 1,  # 1=ARS, 2=USD
    offset_dolar: float = 0,  # Offset para el tipo de cambio
    equipo_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Exporta Lista Web Transferencia a Excel con formato de lista (como Lista Gremio).
    Usa el precio_web_transferencia almacenado. Soporta ARS y USD."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from sqlalchemy import text
    from app.services.pricing_calculator import obtener_tipo_cambio_actual

    try:
        # Obtener tipo de cambio
        tipo_cambio = obtener_tipo_cambio_actual(db, "USD")

        # Ajustar tipo de cambio con offset
        tipo_cambio_ajustado = tipo_cambio + offset_dolar if tipo_cambio else None

        # Cargar nombres de subcategorías
        subcats_result = db.execute(text("SELECT subcat_id, subcat_desc FROM tb_subcategory"))
        subcats_dict = {row.subcat_id: row.subcat_desc for row in subcats_result}

        layer_activo = resolver_layer_activo(equipo_id, current_user, db)

        # Query base: solo productos con web transferencia activa y precio definido
        query = (
            db.query(ProductoERP, ProductoPricing)
            .join(ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id)
            .filter(
                ProductoPricing.participa_web_transferencia == True,
                ProductoPricing.precio_web_transferencia.isnot(None),
            )
        )
        query = join_color_layer(query, layer_activo)

        # Aplicar filtros
        query = _apply_search_filter(query, search)

        if con_stock is not None:
            query = query.filter(ProductoERP.stock > 0 if con_stock else ProductoERP.stock == 0)

        if marcas:
            marcas_list = [m.strip().upper() for m in marcas.split(",")]
            query = query.filter(func.upper(ProductoERP.marca).in_(marcas_list))

        if subcategorias:
            subcat_list = [int(s.strip()) for s in subcategorias.split(",")]
            query = query.filter(ProductoERP.subcategoria_id.in_(subcat_list))

        # Filtro de colores (vista tienda, lee del layer de equipo activo)
        query = filtro_colores(query, colores, color_slot("tienda"))

        # Ejecutar query
        results = query.order_by(ProductoERP.marca, ProductoERP.codigo).all()

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista Web Transferencia"

        # Headers
        moneda_texto = "USD" if currency_id == 2 else "ARS"
        headers = [
            "Marca",
            "Categoría",
            "Subcategoría",
            "Código",
            "Descripción",
            "Stock",
            f"Precio Web Transf. {moneda_texto} s/IVA",
            f"Precio Web Transf. {moneda_texto} c/IVA",
        ]
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Datos
        row_num = 2
        for producto_erp, producto_pricing in results:
            precio_con_iva = float(producto_pricing.precio_web_transferencia)

            # Derivar precio sin IVA
            iva_producto = producto_erp.iva if producto_erp.iva else 21.0
            precio_sin_iva = precio_con_iva / (1 + iva_producto / 100)

            # Convertir a USD si es necesario
            if currency_id == 2 and tipo_cambio_ajustado and tipo_cambio_ajustado > 0:
                precio_sin_iva = precio_sin_iva / tipo_cambio_ajustado
                precio_con_iva = precio_con_iva / tipo_cambio_ajustado

            ws.cell(row=row_num, column=1, value=producto_erp.marca or "")
            ws.cell(row=row_num, column=2, value=producto_erp.categoria or "")
            ws.cell(row=row_num, column=3, value=subcats_dict.get(producto_erp.subcategoria_id, "") or "")
            ws.cell(row=row_num, column=4, value=producto_erp.codigo or "")
            ws.cell(row=row_num, column=5, value=producto_erp.descripcion or "")
            ws.cell(row=row_num, column=6, value=producto_erp.stock or 0)
            ws.cell(row=row_num, column=7, value=round(precio_sin_iva, 2))
            ws.cell(row=row_num, column=8, value=round(precio_con_iva, 2))

            row_num += 1

        # Ajustar anchos de columna
        column_widths = [15, 20, 20, 15, 50, 10, 22, 22]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=lista_web_transferencia.xlsx"},
        )

    except Exception as e:
        logger.error("Error en exportar_lista_web_transferencia: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al exportar lista web transferencia: {str(e)}")

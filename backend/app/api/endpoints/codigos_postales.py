"""
Endpoints para gestión de códigos postales y cordones de envío.

Lista los CPs únicos que aparecen en envíos self_service de ML,
permite asignar cordón (CABA, Cordón 1, Cordón 2, Cordón 3) y
subir un XLSX masivo para importar el mapeo CP → Cordón.
La localidad se popula automáticamente desde mlcity_name de shipping.
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from sqlalchemy import func, distinct, case, literal
from typing import List, Optional
from pydantic import BaseModel, ConfigDict, Field
from io import BytesIO

from app.core.database import get_db
from app.api.deps import get_current_user
from app.models.usuario import Usuario
from app.models.mercadolibre_order_shipping import MercadoLibreOrderShipping
from app.models.codigo_postal_cordon import CodigoPostalCordon

router = APIRouter()

# Cordones válidos
CORDONES_VALIDOS = ["CABA", "Cordón 1", "Cordón 2", "Cordón 3"]


# ── Schemas ──────────────────────────────────────────────────────────


class CodigoPostalResponse(BaseModel):
    """Un CP con su localidad y cordón asignado."""

    codigo_postal: str
    localidad: Optional[str] = None
    cordon: Optional[str] = None  # None = "Sin Asignar"
    cantidad_envios: int = 0

    model_config = ConfigDict(from_attributes=True)


class ActualizarCordonRequest(BaseModel):
    """Payload para cambiar el cordón de un CP."""

    cordon: Optional[str] = Field(
        None,
        description="Cordón a asignar. None para desasignar.",
    )


class ImportResultResponse(BaseModel):
    """Resultado de la importación XLSX."""

    total_filas: int
    actualizados: int
    creados: int
    errores: int
    detalle_errores: List[str] = []


class EstadisticasCordonResponse(BaseModel):
    """Estadísticas de distribución de cordones."""

    total_cps: int
    sin_asignar: int
    por_cordon: dict[str, int]


# ── Endpoints ────────────────────────────────────────────────────────


@router.get(
    "/codigos-postales",
    response_model=List[CodigoPostalResponse],
    summary="Listar CPs de envíos self_service con cordón",
)
def listar_codigos_postales(
    cordon: Optional[str] = Query(None, description="Filtrar por cordón"),
    sin_asignar: bool = Query(False, description="Solo CPs sin cordón asignado"),
    search: Optional[str] = Query(None, description="Buscar por CP o localidad"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> List[CodigoPostalResponse]:
    """
    Devuelve todos los CPs únicos que aparecen en envíos con
    ml_logistic_type = 'self_service', junto con la localidad más
    frecuente de ML y el cordón asignado (si existe).
    """

    # Subquery: CPs distintos de self_service con localidad más frecuente y cantidad
    # Usamos una ventana para obtener la localidad más frecuente por CP
    shipping_cps = (
        db.query(
            MercadoLibreOrderShipping.mlzip_code.label("cp"),
            func.mode().within_group(MercadoLibreOrderShipping.mlcity_name).label("localidad_frecuente"),
            func.count().label("cantidad_envios"),
        )
        .filter(
            MercadoLibreOrderShipping.ml_logistic_type == "self_service",
            MercadoLibreOrderShipping.mlzip_code.isnot(None),
            MercadoLibreOrderShipping.mlzip_code != "",
        )
        .group_by(MercadoLibreOrderShipping.mlzip_code)
        .subquery()
    )

    # Join con la tabla de cordones
    query = db.query(
        shipping_cps.c.cp.label("codigo_postal"),
        func.coalesce(
            CodigoPostalCordon.localidad,
            shipping_cps.c.localidad_frecuente,
        ).label("localidad"),
        CodigoPostalCordon.cordon,
        shipping_cps.c.cantidad_envios,
    ).outerjoin(
        CodigoPostalCordon,
        CodigoPostalCordon.codigo_postal == shipping_cps.c.cp,
    )

    # Filtros
    if cordon:
        if cordon not in CORDONES_VALIDOS:
            raise HTTPException(400, f"Cordón inválido. Válidos: {CORDONES_VALIDOS}")
        query = query.filter(CodigoPostalCordon.cordon == cordon)

    if sin_asignar:
        query = query.filter(CodigoPostalCordon.cordon.is_(None))

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            (shipping_cps.c.cp.ilike(search_term)) | (shipping_cps.c.localidad_frecuente.ilike(search_term))
        )

    # Ordenar: sin asignar primero, luego por CP
    query = query.order_by(
        case(
            (CodigoPostalCordon.cordon.is_(None), literal(0)),
            else_=literal(1),
        ),
        shipping_cps.c.cp,
    )

    rows = query.all()

    return [
        CodigoPostalResponse(
            codigo_postal=row.codigo_postal,
            localidad=row.localidad,
            cordon=row.cordon,
            cantidad_envios=row.cantidad_envios,
        )
        for row in rows
    ]


@router.put(
    "/codigos-postales/{codigo_postal}/cordon",
    response_model=CodigoPostalResponse,
    summary="Asignar o cambiar el cordón de un CP",
)
def actualizar_cordon(
    codigo_postal: str,
    payload: ActualizarCordonRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> CodigoPostalResponse:
    """
    Asigna o cambia el cordón de un código postal.
    Si el CP no existe en cp_cordones, lo crea.
    Si cordon es None, desasigna el cordón.
    """
    if payload.cordon is not None and payload.cordon not in CORDONES_VALIDOS:
        raise HTTPException(400, f"Cordón inválido. Válidos: {CORDONES_VALIDOS}")

    # Buscar o crear registro
    registro = db.query(CodigoPostalCordon).filter(CodigoPostalCordon.codigo_postal == codigo_postal).first()

    if registro:
        registro.cordon = payload.cordon
    else:
        # Obtener la localidad más frecuente de ML shipping
        localidad = (
            db.query(func.mode().within_group(MercadoLibreOrderShipping.mlcity_name))
            .filter(
                MercadoLibreOrderShipping.mlzip_code == codigo_postal,
                MercadoLibreOrderShipping.mlcity_name.isnot(None),
            )
            .scalar()
        )

        registro = CodigoPostalCordon(
            codigo_postal=codigo_postal,
            localidad=localidad,
            cordon=payload.cordon,
        )
        db.add(registro)

    db.commit()
    db.refresh(registro)

    # Obtener cantidad_envios para la respuesta
    cantidad = (
        db.query(func.count())
        .filter(
            MercadoLibreOrderShipping.ml_logistic_type == "self_service",
            MercadoLibreOrderShipping.mlzip_code == codigo_postal,
        )
        .scalar()
    ) or 0

    return CodigoPostalResponse(
        codigo_postal=registro.codigo_postal,
        localidad=registro.localidad,
        cordon=registro.cordon,
        cantidad_envios=cantidad,
    )


@router.post(
    "/codigos-postales/import-xlsx",
    response_model=ImportResultResponse,
    summary="Importar mapeo CP → Cordón desde XLSX",
)
def importar_xlsx(
    file: UploadFile = File(..., description="Archivo XLSX con columnas CP y Cordon"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> ImportResultResponse:
    """
    Importa un archivo XLSX con columnas CP y Cordon.
    Para cada fila:
    - Si el CP ya existe, actualiza el cordón.
    - Si no existe, lo crea con el cordón indicado.
    - Si el cordón está vacío, lo ignora.
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "El archivo debe ser .xlsx o .xls")

    try:
        from openpyxl import load_workbook

        contents = file.file.read()
        wb = load_workbook(BytesIO(contents), read_only=True)
        ws = wb.active

        if ws is None:
            raise HTTPException(400, "El archivo no tiene hojas activas")

    except ImportError:
        raise HTTPException(500, "openpyxl no está instalado en el servidor")
    except Exception as e:
        raise HTTPException(400, f"Error leyendo el archivo: {str(e)}")

    # Leer header para encontrar las columnas
    header = [
        str(cell.value).strip().upper() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    # Buscar columnas CP y Cordon (flexible)
    cp_col = None
    cordon_col = None

    for idx, col_name in enumerate(header):
        if col_name in ("CP", "CODIGO_POSTAL", "CODIGO POSTAL", "CÓDIGO POSTAL", "COD_POSTAL"):
            cp_col = idx
        elif col_name in ("CORDON", "CORDÓN", "CORD", "ZONA"):
            cordon_col = idx

    if cp_col is None:
        raise HTTPException(
            400,
            f"No se encontró la columna de código postal. "
            f"Columnas encontradas: {header}. "
            f"Se esperaba: CP, CODIGO_POSTAL, etc.",
        )
    if cordon_col is None:
        raise HTTPException(
            400,
            f"No se encontró la columna de cordón. "
            f"Columnas encontradas: {header}. "
            f"Se esperaba: CORDON, CORDÓN, ZONA, etc.",
        )

    # Procesar filas
    total_filas = 0
    actualizados = 0
    creados = 0
    errores = 0
    detalle_errores: List[str] = []

    # Cargar CPs existentes en memoria para bulk operation
    existentes = {r.codigo_postal: r for r in db.query(CodigoPostalCordon).all()}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        total_filas += 1

        try:
            cp_value = row[cp_col].value
            cordon_value = row[cordon_col].value

            if cp_value is None:
                continue

            # Normalizar CP (puede venir como número)
            cp_str = str(int(cp_value) if isinstance(cp_value, float) else cp_value).strip()

            if not cp_str:
                continue

            # Normalizar cordón
            cordon_str = str(cordon_value).strip() if cordon_value else None

            if cordon_str and cordon_str not in CORDONES_VALIDOS:
                errores += 1
                detalle_errores.append(f"Fila {row_idx}: Cordón '{cordon_str}' no válido para CP {cp_str}")
                continue

            if cp_str in existentes:
                existentes[cp_str].cordon = cordon_str
                actualizados += 1
            else:
                nuevo = CodigoPostalCordon(
                    codigo_postal=cp_str,
                    cordon=cordon_str,
                )
                db.add(nuevo)
                existentes[cp_str] = nuevo
                creados += 1

        except Exception as e:
            errores += 1
            detalle_errores.append(f"Fila {row_idx}: {str(e)}")

    wb.close()

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Error guardando en base de datos: {str(e)}")

    # Después de importar, popular localidades desde ML shipping
    _popular_localidades_desde_shipping(db)

    return ImportResultResponse(
        total_filas=total_filas,
        actualizados=actualizados,
        creados=creados,
        errores=errores,
        detalle_errores=detalle_errores[:50],  # Limitar a 50 errores
    )


@router.get(
    "/codigos-postales/estadisticas",
    response_model=EstadisticasCordonResponse,
    summary="Estadísticas de distribución de cordones",
)
def estadisticas_cordones(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> EstadisticasCordonResponse:
    """Devuelve cuántos CPs hay por cada cordón y sin asignar."""

    # Total de CPs únicos en self_service
    total_cps = (
        db.query(func.count(distinct(MercadoLibreOrderShipping.mlzip_code)))
        .filter(
            MercadoLibreOrderShipping.ml_logistic_type == "self_service",
            MercadoLibreOrderShipping.mlzip_code.isnot(None),
            MercadoLibreOrderShipping.mlzip_code != "",
        )
        .scalar()
    ) or 0

    # Distribución por cordón
    cordon_counts = (
        db.query(
            CodigoPostalCordon.cordon,
            func.count().label("cantidad"),
        )
        .filter(CodigoPostalCordon.cordon.isnot(None))
        .group_by(CodigoPostalCordon.cordon)
        .all()
    )

    por_cordon = {row.cordon: row.cantidad for row in cordon_counts}
    asignados = sum(por_cordon.values())
    sin_asignar = total_cps - asignados

    return EstadisticasCordonResponse(
        total_cps=total_cps,
        sin_asignar=max(0, sin_asignar),
        por_cordon=por_cordon,
    )


# ── Helpers ──────────────────────────────────────────────────────────


def _popular_localidades_desde_shipping(db: Session) -> None:
    """
    Para los CPs en cp_cordones que no tienen localidad,
    busca la localidad más frecuente en ML shipping y la asigna.
    """
    sin_localidad = db.query(CodigoPostalCordon).filter(CodigoPostalCordon.localidad.is_(None)).all()

    for registro in sin_localidad:
        localidad = (
            db.query(func.mode().within_group(MercadoLibreOrderShipping.mlcity_name))
            .filter(
                MercadoLibreOrderShipping.mlzip_code == registro.codigo_postal,
                MercadoLibreOrderShipping.mlcity_name.isnot(None),
                MercadoLibreOrderShipping.mlcity_name != "",
            )
            .scalar()
        )

        if localidad:
            registro.localidad = localidad

    db.commit()

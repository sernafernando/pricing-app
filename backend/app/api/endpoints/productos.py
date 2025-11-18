from fastapi import APIRouter, Depends, HTTPException, Query, Body, Request
from sqlalchemy.orm import Session
from sqlalchemy import or_, func, and_, select
from typing import Optional, List
from app.core.database import get_db
from app.models.producto import ProductoERP, ProductoPricing
from app.models.usuario import Usuario
from pydantic import BaseModel, Field
from datetime import datetime, date
from app.models.auditoria_precio import AuditoriaPrecio
from app.api.deps import get_current_user
from fastapi.responses import Response
from decimal import Decimal
import logging

logger = logging.getLogger(__name__)

router = APIRouter()

class ProductoResponse(BaseModel):
    item_id: int
    codigo: str
    descripcion: str
    marca: Optional[str]
    categoria: Optional[str]
    subcategoria_id: Optional[int]
    moneda_costo: Optional[str]
    costo: float
    costo_ars: Optional[float]
    iva: float
    stock: int
    precio_lista_ml: Optional[float]
    markup: Optional[float]
    usuario_modifico: Optional[str]
    fecha_modificacion: Optional[datetime]
    tiene_precio: bool
    necesita_revision: bool
    participa_rebate: Optional[bool] = False
    porcentaje_rebate: Optional[float] = 3.8
    precio_rebate: Optional[float] = None
    markup_rebate: Optional[float] = None
    participa_web_transferencia: Optional[bool] = False
    porcentaje_markup_web: Optional[float] = 6.0
    precio_web_transferencia: Optional[float] = None
    markup_web_real: Optional[float] = None
    preservar_porcentaje_web: Optional[bool] = False
    mejor_oferta_precio: Optional[float] = None
    mejor_oferta_monto_rebate: Optional[float] = None
    mejor_oferta_pvp_seller: Optional[float] = None
    mejor_oferta_markup: Optional[float] = None
    mejor_oferta_porcentaje_rebate: Optional[float] = None
    mejor_oferta_fecha_hasta: Optional[date] = None
    out_of_cards: Optional[bool] = False
    color_marcado: Optional[str] = None
    precio_3_cuotas: Optional[float] = None
    precio_6_cuotas: Optional[float] = None
    precio_9_cuotas: Optional[float] = None
    precio_12_cuotas: Optional[float] = None
    markup_3_cuotas: Optional[float] = None
    markup_6_cuotas: Optional[float] = None
    markup_9_cuotas: Optional[float] = None
    markup_12_cuotas: Optional[float] = None

    # Configuración individual
    recalcular_cuotas_auto: Optional[bool] = None
    markup_adicional_cuotas_custom: Optional[float] = None

    # Estado de catálogo ML
    catalog_status: Optional[str] = None
    has_catalog: Optional[bool] = None
    catalog_price_to_win: Optional[float] = None
    catalog_winner_price: Optional[float] = None

    # Precios Tienda Nube
    tn_price: Optional[float] = None  # Precio normal
    tn_promotional_price: Optional[float] = None  # Precio promocional
    tn_has_promotion: Optional[bool] = None  # Si tiene promoción activa

    class Config:
        from_attributes = True

class ProductoListResponse(BaseModel):
    total: int
    page: int
    page_size: int
    productos: List[ProductoResponse]

class PrecioUpdate(BaseModel):
    precio_lista_final: Optional[float] = None
    precio_contado_final: Optional[float] = None
    comentario: Optional[str] = None

class RebateUpdate(BaseModel):
    participa_rebate: bool
    porcentaje_rebate: float = Field(ge=0, le=100)

@router.get("/productos", response_model=ProductoListResponse)
async def listar_productos(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=10000),
    search: Optional[str] = None,
    categoria: Optional[str] = None,
    marcas: Optional[str] = None,
    subcategorias: Optional[str] = None,
    con_stock: Optional[bool] = None,
    con_precio: Optional[bool] = None,
    orden_campos: Optional[str] = None,
    orden_direcciones: Optional[str] = None,
    audit_usuarios: Optional[str] = None,
    audit_tipos_accion: Optional[str] = None,
    audit_fecha_desde: Optional[str] = None,
    audit_fecha_hasta: Optional[str] = None,
    con_rebate: Optional[bool] = None,
    con_oferta: Optional[bool] = None,
    con_web_transf: Optional[bool] = None,
    tn_con_descuento: Optional[bool] = None,
    tn_sin_descuento: Optional[bool] = None,
    tn_no_publicado: Optional[bool] = None,
    markup_clasica_positivo: Optional[bool] = None,
    markup_rebate_positivo: Optional[bool] = None,
    markup_oferta_positivo: Optional[bool] = None,
    markup_web_transf_positivo: Optional[bool] = None,
    out_of_cards: Optional[bool] = None,
    colores: Optional[str] = None,
    pms: Optional[str] = None,
    con_mla: Optional[bool] = None,
    nuevos_ultimos_7_dias: Optional[bool] = None,
    db: Session = Depends(get_db)
):
    query = db.query(ProductoERP, ProductoPricing).outerjoin(
        ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id
    )

    # FILTRADO POR AUDITORÍA
    if audit_usuarios or audit_tipos_accion or audit_fecha_desde or audit_fecha_hasta:
        from app.models.auditoria import Auditoria
        from datetime import datetime
        from sqlalchemy.sql import exists

        # Construir filtros de auditoría base
        filtros_audit = [Auditoria.item_id.isnot(None)]

        if audit_usuarios:
            usuarios_ids = [int(u) for u in audit_usuarios.split(',')]
            filtros_audit.append(Auditoria.usuario_id.in_(usuarios_ids))

        if audit_tipos_accion:
            tipos_list = audit_tipos_accion.split(',')
            filtros_audit.append(Auditoria.tipo_accion.in_(tipos_list))

        fecha_desde_dt = None
        if audit_fecha_desde:
            try:
                # Intentar con segundos
                fecha_desde_dt = datetime.strptime(audit_fecha_desde, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    # Intentar sin segundos (formato datetime-local de HTML5)
                    fecha_desde_dt = datetime.strptime(audit_fecha_desde, '%Y-%m-%d %H:%M')
                except ValueError:
                    try:
                        # Intentar solo fecha (poner hora en 00:00:00)
                        fecha_desde_dt = datetime.strptime(audit_fecha_desde, '%Y-%m-%d')
                    except ValueError:
                        # Si falla todo, usar fecha de hoy
                        from datetime import date
                        fecha_desde_dt = datetime.combine(date.today(), datetime.min.time())

            # Convertir de hora local (ART = UTC-3) a UTC sumando 3 horas
            # porque la base de datos guarda fechas en UTC
            from datetime import timedelta
            fecha_desde_dt = fecha_desde_dt + timedelta(hours=3)

            filtros_audit.append(Auditoria.fecha >= fecha_desde_dt)

        fecha_hasta_dt = None
        if audit_fecha_hasta:
            try:
                # Intentar con segundos
                fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    # Intentar sin segundos (formato datetime-local de HTML5)
                    fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, '%Y-%m-%d %H:%M')
                except ValueError:
                    try:
                        # Solo fecha: poner hora al final del día
                        fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, '%Y-%m-%d')
                        fecha_hasta_dt = fecha_hasta_dt.replace(hour=23, minute=59, second=59)
                    except ValueError:
                        from datetime import date
                        fecha_hasta_dt = datetime.combine(date.today(), datetime.max.time())

            # Convertir de hora local (ART = UTC-3) a UTC sumando 3 horas
            from datetime import timedelta
            fecha_hasta_dt = fecha_hasta_dt + timedelta(hours=3)

            filtros_audit.append(Auditoria.fecha <= fecha_hasta_dt)

        # Obtener productos que tienen auditorías cumpliendo los criterios
        audit_query = db.query(Auditoria.item_id).filter(and_(*filtros_audit))
        item_ids = [item_id for (item_id,) in audit_query.distinct().all()]

        if item_ids:
            query = query.filter(ProductoERP.item_id.in_(item_ids))
        else:
            return ProductoListResponse(total=0, page=page, page_size=page_size, productos=[])

    if search:
        search_normalized = search.replace('-', '').replace(' ', '').upper()
        query = query.filter(
            or_(
                func.replace(func.replace(func.upper(ProductoERP.descripcion), '-', ''), ' ', '').like(f"%{search_normalized}%"),
                func.replace(func.replace(func.upper(ProductoERP.marca), '-', ''), ' ', '').like(f"%{search_normalized}%"),
                func.replace(func.upper(ProductoERP.codigo), '-', '').like(f"%{search_normalized}%")
            )
        )

    if categoria:
        query = query.filter(ProductoERP.categoria == categoria)

    if subcategorias:
        subcat_list = [int(s.strip()) for s in subcategorias.split(',')]
        query = query.filter(ProductoERP.subcategoria_id.in_(subcat_list))
    
    if marcas:
        marcas_list = [m.strip().upper() for m in marcas.split(',')]
        query = query.filter(func.upper(ProductoERP.marca).in_(marcas_list))

    # Filtro por PMs (Product Managers)
    if pms:
        from app.models.marca_pm import MarcaPM
        pm_ids = [int(pm.strip()) for pm in pms.split(',')]

        # Obtener marcas asignadas a esos PMs
        marcas_pm = db.query(MarcaPM.marca).filter(MarcaPM.usuario_id.in_(pm_ids)).all()
        marcas_asignadas = [m[0] for m in marcas_pm]

        if marcas_asignadas:
            query = query.filter(func.upper(ProductoERP.marca).in_([m.upper() for m in marcas_asignadas]))
        else:
            # Si no hay marcas asignadas, retornar vacío
            return ProductoListResponse(total=0, page=page, page_size=page_size, productos=[])

    if con_stock is not None:
        query = query.filter(ProductoERP.stock > 0 if con_stock else ProductoERP.stock == 0)
    
    if con_precio is not None:
        if con_precio:
            query = query.filter(ProductoPricing.precio_lista_ml.isnot(None))
        else:
            query = query.filter(ProductoPricing.precio_lista_ml.is_(None))

    # Filtros de valores específicos
    if con_rebate is not None:
        if con_rebate:
            query = query.filter(ProductoPricing.participa_rebate == True)
        else:
            query = query.filter(
                or_(
                    ProductoPricing.participa_rebate == False,
                    ProductoPricing.participa_rebate.is_(None)
                )
            )

    if con_web_transf is not None:
        if con_web_transf:
            query = query.filter(ProductoPricing.participa_web_transferencia == True)
        else:
            query = query.filter(
                or_(
                    ProductoPricing.participa_web_transferencia == False,
                    ProductoPricing.participa_web_transferencia.is_(None)
                )
            )

    # Filtros de Tienda Nube
    if tn_con_descuento or tn_sin_descuento or tn_no_publicado:
        from app.models.tienda_nube_producto import TiendaNubeProducto

        if tn_con_descuento:
            # Productos que tienen promotional_price (con descuento)
            query = query.join(
                TiendaNubeProducto,
                and_(
                    ProductoERP.item_id == TiendaNubeProducto.item_id,
                    TiendaNubeProducto.activo == True,
                    TiendaNubeProducto.promotional_price.isnot(None),
                    TiendaNubeProducto.promotional_price > 0
                )
            )
        elif tn_sin_descuento:
            # Productos publicados pero sin promotional_price
            query = query.join(
                TiendaNubeProducto,
                and_(
                    ProductoERP.item_id == TiendaNubeProducto.item_id,
                    TiendaNubeProducto.activo == True,
                    or_(
                        TiendaNubeProducto.promotional_price.is_(None),
                        TiendaNubeProducto.promotional_price == 0
                    )
                )
            )
        elif tn_no_publicado:
            # Productos con stock pero NO en Tienda Nube
            from sqlalchemy.sql import exists
            subquery = exists().where(
                and_(
                    TiendaNubeProducto.item_id == ProductoERP.item_id,
                    TiendaNubeProducto.activo == True
                )
            )
            query = query.filter(
                and_(
                    ProductoERP.stock > 0,
                    ~subquery
                )
            )

    # Filtros de markup
    if markup_clasica_positivo is not None:
        if markup_clasica_positivo:
            query = query.filter(ProductoPricing.markup_calculado > 0)
        else:
            query = query.filter(ProductoPricing.markup_calculado < 0)

    if markup_rebate_positivo is not None:
        if markup_rebate_positivo:
            query = query.filter(ProductoPricing.markup_rebate > 0)
        else:
            query = query.filter(ProductoPricing.markup_rebate < 0)

    if markup_oferta_positivo is not None:
        if markup_oferta_positivo:
            query = query.filter(ProductoPricing.markup_oferta > 0)
        else:
            query = query.filter(ProductoPricing.markup_oferta < 0)

    if markup_web_transf_positivo is not None:
        if markup_web_transf_positivo:
            query = query.filter(ProductoPricing.markup_web_real > 0)
        else:
            query = query.filter(ProductoPricing.markup_web_real < 0)

    if out_of_cards is not None:
        if out_of_cards:
            query = query.filter(ProductoPricing.out_of_cards == True)
        else:
            query = query.filter(
                or_(
                    ProductoPricing.out_of_cards == False,
                    ProductoPricing.out_of_cards.is_(None)
                )
            )

    # Filtro de oferta (ofertas vigentes en MercadoLibre)
    if con_oferta is not None:
        from app.models.oferta_ml import OfertaML
        from app.models.publicacion_ml import PublicacionML
        from datetime import date

        hoy_date = date.today()

        items_con_oferta_vigente_subquery = db.query(PublicacionML.item_id).join(
            OfertaML, PublicacionML.mla == OfertaML.mla
        ).filter(
            OfertaML.fecha_desde <= hoy_date,
            OfertaML.fecha_hasta >= hoy_date,
            OfertaML.pvp_seller.isnot(None)
        ).distinct().subquery()

        if con_oferta:
            query = query.filter(ProductoERP.item_id.in_(items_con_oferta_vigente_subquery))
        else:
            query = query.filter(~ProductoERP.item_id.in_(items_con_oferta_vigente_subquery))

    # Filtro de colores
    if colores:
        colores_list = colores.split(',')

        # Verificar si se está filtrando por "sin color"
        if 'sin_color' in colores_list:
            # Remover 'sin_color' de la lista
            colores_con_valor = [c for c in colores_list if c != 'sin_color']

            if colores_con_valor:
                # Si hay otros colores además de sin_color, buscar ambos
                query = query.filter(
                    or_(
                        ProductoPricing.color_marcado.in_(colores_con_valor),
                        ProductoPricing.color_marcado.is_(None)
                    )
                )
            else:
                # Solo sin_color: productos sin color asignado
                query = query.filter(ProductoPricing.color_marcado.is_(None))
        else:
            # Filtro normal por colores específicos
            query = query.filter(ProductoPricing.color_marcado.in_(colores_list))

    # Filtro de MLA (con/sin publicación) - usar subconsultas para evitar conflictos de join
    if con_mla is not None:
        from app.models.mercadolibre_item_publicado import MercadoLibreItemPublicado
        from app.models.item_sin_mla_banlist import ItemSinMLABanlist

        if con_mla:
            # Con MLA: tienen publicación activa
            items_con_mla_subquery = db.query(MercadoLibreItemPublicado.item_id).filter(
                MercadoLibreItemPublicado.mlp_id.isnot(None),
                or_(
                    MercadoLibreItemPublicado.optval_statusId == 2,
                    MercadoLibreItemPublicado.optval_statusId.is_(None)
                )
            ).distinct().subquery()

            query = query.filter(ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)))
        else:
            # Sin MLA: no tienen publicación (excluye banlist)
            items_con_mla_subquery = db.query(MercadoLibreItemPublicado.item_id).filter(
                MercadoLibreItemPublicado.mlp_id.isnot(None),
                or_(
                    MercadoLibreItemPublicado.optval_statusId == 2,
                    MercadoLibreItemPublicado.optval_statusId.is_(None)
                )
            ).distinct().subquery()

            items_en_banlist_subquery = db.query(ItemSinMLABanlist.item_id).subquery()

            query = query.filter(
                ~ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)),
                ~ProductoERP.item_id.in_(select(items_en_banlist_subquery.c.item_id))
            )

    # Filtro de productos nuevos (últimos 7 días)
    if nuevos_ultimos_7_dias:
        from datetime import datetime, timedelta
        fecha_limite = datetime.now() - timedelta(days=7)
        query = query.filter(ProductoERP.fecha_sync >= fecha_limite)

    # Ordenamiento
    orden_requiere_calculo = False
    if orden_campos and orden_direcciones:
        campos = orden_campos.split(',')
        direcciones = orden_direcciones.split(',')

        for campo, direccion in zip(campos, direcciones):
            # Mapeo de campos del frontend a columnas de la DB
            if campo == 'item_id':
                col = ProductoERP.item_id
            elif campo == 'codigo':
                col = ProductoERP.codigo
            elif campo == 'descripcion':
                col = ProductoERP.descripcion
            elif campo == 'marca':
                col = ProductoERP.marca
            elif campo == 'moneda_costo':
                col = ProductoERP.moneda_costo
            elif campo == 'costo':
                col = ProductoERP.costo
            elif campo == 'stock':
                col = ProductoERP.stock
            elif campo == 'precio_lista_ml':
                col = ProductoPricing.precio_lista_ml
            elif campo == 'markup' or campo == 'precio_clasica':
                col = ProductoPricing.markup_calculado
            elif campo == 'precio_rebate':
                # Markup rebate requiere cálculo dinámico
                orden_requiere_calculo = True
                continue
            elif campo == 'mejor_oferta':
                # Mejor oferta requiere cálculo dinámico
                orden_requiere_calculo = True
                continue
            elif campo == 'web_transf':
                col = ProductoPricing.markup_web_real
            else:
                continue

            if direccion == 'asc':
                query = query.order_by(col.asc().nullslast())
            else:
                query = query.order_by(col.desc().nullslast())

    # Contar total y paginar
    # Los filtros de markup ahora se aplican en SQL, no necesitamos traer todo
    total_productos = None
    if not orden_requiere_calculo:
        total_productos = query.count()
        offset = (page - 1) * page_size
        results = query.offset(offset).limit(page_size).all()
    else:
        # Solo si hay ordenamiento que requiere cálculo dinámico
        results = query.all()
        total_antes_filtro = len(results)

    from app.models.oferta_ml import OfertaML
    from app.models.publicacion_ml import PublicacionML
    from app.services.pricing_calculator import (
        obtener_tipo_cambio_actual,
        convertir_a_pesos,
        obtener_grupo_subcategoria,
        obtener_comision_base,
        calcular_comision_ml_total,
        calcular_limpio,
        calcular_markup
    )
    from datetime import date
    
    hoy = date.today()
    
    productos = []
    for producto_erp, producto_pricing in results:
        costo_ars = producto_erp.costo if producto_erp.moneda_costo == "ARS" else None
        # Buscar mejor oferta vigente
        mejor_oferta_precio = None
        mejor_oferta_monto = None
        mejor_oferta_pvp = None
        mejor_oferta_markup = None
        mejor_oferta_porcentaje = None
        mejor_oferta_fecha_hasta = None
        
        # Buscar publicación del producto
        pubs = db.query(PublicacionML).filter(PublicacionML.item_id == producto_erp.item_id).all()
                                
        mejor_oferta = None
        mejor_pub = None
        
        for pub in pubs:
            # Buscar oferta vigente para esta publicación
            oferta = db.query(OfertaML).filter(
                OfertaML.mla == pub.mla,
                OfertaML.fecha_desde <= hoy,
                OfertaML.fecha_hasta >= hoy,
                OfertaML.pvp_seller.isnot(None)
            ).order_by(OfertaML.fecha_desde.desc()).first()
            
            if oferta:
                # Tomar la primera que encuentre (o implementar lógica para elegir la mejor)
                if not mejor_oferta:
                    mejor_oferta = oferta
                    mejor_pub = pub
        
        
        if mejor_oferta and mejor_pub:
            mejor_oferta_precio = float(mejor_oferta.precio_final) if mejor_oferta.precio_final else None
            mejor_oferta_pvp = float(mejor_oferta.pvp_seller) if mejor_oferta.pvp_seller else None
            mejor_oferta_porcentaje = float(mejor_oferta.aporte_meli_porcentaje) if mejor_oferta.aporte_meli_porcentaje else None  # ← AGREGAR
            mejor_oferta_fecha_hasta = mejor_oferta.fecha_hasta
            
            # Calcular monto rebate
            if mejor_oferta_precio and mejor_oferta_pvp:
                mejor_oferta_monto = mejor_oferta_pvp - mejor_oferta_precio
            
            # Calcular markup de la oferta
            if mejor_oferta_pvp and mejor_oferta_pvp > 0:
                tipo_cambio = None
                if producto_erp.moneda_costo == "USD":
                    tipo_cambio = obtener_tipo_cambio_actual(db, "USD")
                
                costo_calc = convertir_a_pesos(producto_erp.costo, producto_erp.moneda_costo, tipo_cambio)
                grupo_id = obtener_grupo_subcategoria(db, producto_erp.subcategoria_id)
                comision_base = obtener_comision_base(db, mejor_pub.pricelist_id, grupo_id)
                
                if comision_base:
                    comisiones = calcular_comision_ml_total(
                        mejor_oferta_pvp,
                        comision_base,
                        producto_erp.iva,
                        db=db
                    )
                    limpio = calcular_limpio(
                        mejor_oferta_pvp,
                        producto_erp.iva,
                        producto_erp.envio or 0,
                        comisiones["comision_total"],
                        db=db,
                        grupo_id=grupo_id
                    )
                    mejor_oferta_markup = calcular_markup(limpio, costo_calc)

        # Calcular precio_rebate y markup_rebate
        precio_rebate = None
        markup_rebate = None
        if producto_pricing and producto_pricing.precio_lista_ml and producto_pricing.participa_rebate:
            porcentaje_rebate_val = float(producto_pricing.porcentaje_rebate if producto_pricing.porcentaje_rebate is not None else 3.8)
            precio_rebate = float(producto_pricing.precio_lista_ml) / (1 - porcentaje_rebate_val / 100)

            # Calcular markup del rebate
            tipo_cambio_rebate = None
            if producto_erp.moneda_costo == "USD":
                tipo_cambio_rebate = obtener_tipo_cambio_actual(db, "USD")

            costo_rebate = convertir_a_pesos(producto_erp.costo, producto_erp.moneda_costo, tipo_cambio_rebate)
            grupo_id_rebate = obtener_grupo_subcategoria(db, producto_erp.subcategoria_id)
            comision_base_rebate = obtener_comision_base(db, 4, grupo_id_rebate)  # Lista clásica

            if comision_base_rebate and precio_rebate > 0:
                comisiones_rebate = calcular_comision_ml_total(
                    precio_rebate,
                    comision_base_rebate,
                    producto_erp.iva,
                    db=db
                )
                limpio_rebate = calcular_limpio(
                    precio_rebate,
                    producto_erp.iva,
                    producto_erp.envio or 0,
                    comisiones_rebate["comision_total"],
                    db=db,
                    grupo_id=grupo_id_rebate
                )
                markup_rebate = calcular_markup(limpio_rebate, costo_rebate) * 100

        # Si el producto tiene rebate y está out_of_cards, replicar el rebate a mejor_oferta
        if producto_pricing and producto_pricing.out_of_cards and precio_rebate is not None and markup_rebate is not None:
            # Replicar datos del rebate a mejor_oferta
            mejor_oferta_precio = precio_rebate
            mejor_oferta_pvp = precio_rebate  # El PVP es el mismo que el precio rebate
            mejor_oferta_markup = markup_rebate / 100  # Convertir de porcentaje a decimal
            mejor_oferta_porcentaje = None  # No hay aporte de Meli en rebate
            mejor_oferta_monto = None  # No hay monto de rebate en este caso
            mejor_oferta_fecha_hasta = None  # No aplica fecha para rebate

        # Calcular markups para precios de cuotas
        markup_3_cuotas = None
        markup_6_cuotas = None
        markup_9_cuotas = None
        markup_12_cuotas = None

        if producto_pricing:
            cuotas_config = [
                (producto_pricing.precio_3_cuotas, 17, '3_cuotas'),
                (producto_pricing.precio_6_cuotas, 14, '6_cuotas'),
                (producto_pricing.precio_9_cuotas, 13, '9_cuotas'),
                (producto_pricing.precio_12_cuotas, 23, '12_cuotas')
            ]

            for precio_cuota, pricelist_id, nombre_cuota in cuotas_config:
                if precio_cuota and float(precio_cuota) > 0:
                    try:
                        tipo_cambio_cuota = None
                        if producto_erp.moneda_costo == "USD":
                            tipo_cambio_cuota = obtener_tipo_cambio_actual(db, "USD")

                        costo_cuota = convertir_a_pesos(producto_erp.costo, producto_erp.moneda_costo, tipo_cambio_cuota)
                        grupo_id_cuota = obtener_grupo_subcategoria(db, producto_erp.subcategoria_id)
                        comision_base_cuota = obtener_comision_base(db, pricelist_id, grupo_id_cuota)

                        if comision_base_cuota:
                            comisiones_cuota = calcular_comision_ml_total(
                                float(precio_cuota),
                                comision_base_cuota,
                                producto_erp.iva,
                                db=db
                            )
                            limpio_cuota = calcular_limpio(
                                float(precio_cuota),
                                producto_erp.iva,
                                producto_erp.envio or 0,
                                comisiones_cuota["comision_total"],
                                db=db,
                                grupo_id=grupo_id_cuota
                            )
                            markup_calculado = calcular_markup(limpio_cuota, costo_cuota) * 100

                            if nombre_cuota == '3_cuotas':
                                markup_3_cuotas = markup_calculado
                            elif nombre_cuota == '6_cuotas':
                                markup_6_cuotas = markup_calculado
                            elif nombre_cuota == '9_cuotas':
                                markup_9_cuotas = markup_calculado
                            elif nombre_cuota == '12_cuotas':
                                markup_12_cuotas = markup_calculado
                    except Exception:
                        # Si hay error calculando el markup, simplemente no lo mostramos
                        pass

        producto_obj = ProductoResponse(
            item_id=producto_erp.item_id,
            codigo=producto_erp.codigo,
            descripcion=producto_erp.descripcion,
            marca=producto_erp.marca,
            categoria=producto_erp.categoria,
            subcategoria_id=producto_erp.subcategoria_id,
            moneda_costo=producto_erp.moneda_costo,
            costo=producto_erp.costo,
            costo_ars=costo_ars,
            iva=producto_erp.iva,
            stock=producto_erp.stock,
            precio_lista_ml=producto_pricing.precio_lista_ml if producto_pricing else None,
            markup=producto_pricing.markup_calculado if producto_pricing else None,
            usuario_modifico=None,
            fecha_modificacion=producto_pricing.fecha_modificacion if producto_pricing else None,
            tiene_precio=producto_pricing.precio_lista_ml is not None if producto_pricing else False,
            necesita_revision=False,
            participa_rebate=producto_pricing.participa_rebate if producto_pricing else False,
            porcentaje_rebate=float(producto_pricing.porcentaje_rebate) if producto_pricing and producto_pricing.porcentaje_rebate is not None else 3.8,
            precio_rebate=precio_rebate,
            markup_rebate=markup_rebate,
            participa_web_transferencia=producto_pricing.participa_web_transferencia if producto_pricing else False,
            porcentaje_markup_web=float(producto_pricing.porcentaje_markup_web) if producto_pricing and producto_pricing.porcentaje_markup_web else 6.0,
            precio_web_transferencia=float(producto_pricing.precio_web_transferencia) if producto_pricing and producto_pricing.precio_web_transferencia else None,
            markup_web_real=float(producto_pricing.markup_web_real) if producto_pricing and producto_pricing.markup_web_real else None,
            preservar_porcentaje_web=producto_pricing.preservar_porcentaje_web if producto_pricing else False,
            mejor_oferta_precio=mejor_oferta_precio,
            mejor_oferta_monto_rebate=mejor_oferta_monto,
            mejor_oferta_pvp_seller=mejor_oferta_pvp,
            mejor_oferta_markup=mejor_oferta_markup,
            mejor_oferta_porcentaje_rebate=mejor_oferta_porcentaje,
            mejor_oferta_fecha_hasta=mejor_oferta_fecha_hasta,
            out_of_cards=producto_pricing.out_of_cards if producto_pricing else False,
            color_marcado=producto_pricing.color_marcado if producto_pricing else None,
            precio_3_cuotas=float(producto_pricing.precio_3_cuotas) if producto_pricing and producto_pricing.precio_3_cuotas else None,
            precio_6_cuotas=float(producto_pricing.precio_6_cuotas) if producto_pricing and producto_pricing.precio_6_cuotas else None,
            precio_9_cuotas=float(producto_pricing.precio_9_cuotas) if producto_pricing and producto_pricing.precio_9_cuotas else None,
            precio_12_cuotas=float(producto_pricing.precio_12_cuotas) if producto_pricing and producto_pricing.precio_12_cuotas else None,
            markup_3_cuotas=markup_3_cuotas,
            markup_6_cuotas=markup_6_cuotas,
            markup_9_cuotas=markup_9_cuotas,
            markup_12_cuotas=markup_12_cuotas,
            recalcular_cuotas_auto=producto_pricing.recalcular_cuotas_auto if producto_pricing else None,
            markup_adicional_cuotas_custom=float(producto_pricing.markup_adicional_cuotas_custom) if producto_pricing and producto_pricing.markup_adicional_cuotas_custom else None,
            catalog_status=None,  # Se llenará después
            has_catalog=None,  # Se llenará después
        )

        # Los filtros de markup ahora se aplican en SQL
        # Solo agregamos el producto a la lista
        productos.append(producto_obj)

    # Obtener catalog status de los productos con publicaciones ML
    if productos:
        from sqlalchemy import text
        item_ids = [p.item_id for p in productos]

        # Obtener MLAs de estos items
        mla_query = db.query(PublicacionML.item_id, PublicacionML.mla).filter(
            PublicacionML.item_id.in_(item_ids)
        ).all()

        # Crear diccionario item_id -> [mla_ids]
        item_to_mlas = {}
        all_mlas = []
        for item_id, mla in mla_query:
            if item_id not in item_to_mlas:
                item_to_mlas[item_id] = []
            item_to_mlas[item_id].append(mla)
            all_mlas.append(mla)

        # Consultar catalog status de estos MLAs
        if all_mlas:
            catalog_statuses = db.execute(text("""
                SELECT mla, catalog_product_id, status, price_to_win, winner_price
                FROM v_ml_catalog_status_latest
                WHERE mla = ANY(:mla_ids)
            """), {"mla_ids": all_mlas}).fetchall()

            # Crear diccionario mla -> datos de catálogo
            mla_to_catalog = {}
            for mla, catalog_id, status, price_to_win, winner_price in catalog_statuses:
                mla_to_catalog[mla] = {
                    'status': status,
                    'price_to_win': float(price_to_win) if price_to_win else None,
                    'winner_price': float(winner_price) if winner_price else None
                }

            # Asignar status a productos
            for producto in productos:
                if producto.item_id in item_to_mlas:
                    mlas = item_to_mlas[producto.item_id]
                    # Si tiene catálogo, tomar el primer status encontrado
                    for mla in mlas:
                        if mla in mla_to_catalog:
                            catalog_data = mla_to_catalog[mla]
                            producto.catalog_status = catalog_data['status']
                            producto.catalog_price_to_win = catalog_data['price_to_win']
                            producto.catalog_winner_price = catalog_data['winner_price']
                            producto.has_catalog = True
                            break

    # Obtener precios de Tienda Nube
    if productos:
        item_ids = [p.item_id for p in productos]

        tn_precios = db.execute(text("""
            SELECT
                item_id,
                price,
                promotional_price,
                CASE WHEN promotional_price IS NOT NULL AND promotional_price > 0 THEN true ELSE false END as has_promotion
            FROM tienda_nube_productos
            WHERE item_id = ANY(:item_ids)
            AND activo = true
        """), {"item_ids": item_ids}).fetchall()

        # Crear diccionario item_id -> precios TN
        tn_dict = {}
        for item_id, price, promo_price, has_promo in tn_precios:
            tn_dict[item_id] = {
                'price': float(price) if price else None,
                'promotional_price': float(promo_price) if promo_price else None,
                'has_promotion': has_promo
            }

        # Asignar precios TN a productos
        for producto in productos:
            if producto.item_id in tn_dict:
                tn_data = tn_dict[producto.item_id]
                producto.tn_price = tn_data['price']
                producto.tn_promotional_price = tn_data['promotional_price']
                producto.tn_has_promotion = tn_data['has_promotion']

    # Si aplicamos ordenamiento dinámico, necesitamos paginar manualmente
    if orden_requiere_calculo:
        # Ordenamiento dinámico si es necesario
        if orden_requiere_calculo and orden_campos and orden_direcciones:
            campos = orden_campos.split(',')
            direcciones = orden_direcciones.split(',')

            # Ordenar por cada columna con su dirección (en orden inverso para aplicar prioridad correcta)
            for i in range(len(campos) - 1, -1, -1):
                campo = campos[i]
                direccion = direcciones[i]
                reverse = (direccion == 'desc')

                if campo in ['precio_rebate', 'mejor_oferta', 'precio_clasica', 'web_transf']:
                    def get_sort_value(prod, campo=campo):
                        if campo == 'precio_rebate':
                            val = prod.markup_rebate
                        elif campo == 'mejor_oferta':
                            val = prod.mejor_oferta_markup
                            if val is not None:
                                val = val * 100
                        elif campo == 'precio_clasica':
                            val = prod.markup
                        elif campo == 'web_transf':
                            val = prod.markup_web_real
                        else:
                            val = None
                        return (val is None, val if val is not None else float('-inf'))

                    productos.sort(key=get_sort_value, reverse=reverse)

        total = len(productos)
        offset = (page - 1) * page_size
        productos = productos[offset:offset + page_size]
    else:
        # Si no hay filtros dinámicos, usar el total pre-calculado
        total = total_productos if total_productos is not None else len(productos)

    return ProductoListResponse(total=total, page=page, page_size=page_size, productos=productos)

@router.get("/productos/precios-listas")
async def listar_productos_con_precios_listas(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    search: Optional[str] = None,
    categoria: Optional[str] = None,
    marca: Optional[str] = None,
    con_stock: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Lista productos con sus precios en todas las listas de ML"""
    from app.models.precio_ml import PrecioML
    from app.models.publicacion_ml import PublicacionML

    # Query base
    query = db.query(ProductoERP).outerjoin(
        ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id
    )

    # Filtros
    if search:
        search_normalized = search.replace('-', '').replace(' ', '').upper()
        query = query.filter(
            or_(
                func.replace(func.replace(func.upper(ProductoERP.descripcion), '-', ''), ' ', '').like(f"%{search_normalized}%"),
                func.replace(func.replace(func.upper(ProductoERP.marca), '-', ''), ' ', '').like(f"%{search_normalized}%"),
                func.replace(func.upper(ProductoERP.codigo), '-', '').like(f"%{search_normalized}%")
            )
        )

    if categoria:
        query = query.filter(ProductoERP.categoria == categoria)
    if marca:
        query = query.filter(ProductoERP.marca == marca)
    if con_stock is not None:
        query = query.filter(ProductoERP.stock > 0 if con_stock else ProductoERP.stock == 0)

    total = query.count()
    offset = (page - 1) * page_size
    results = query.offset(offset).limit(page_size).all()

    productos = []
    for producto_erp in results:
        # Obtener precios de todas las listas directamente por item_id
        precios_listas = {}
    
        for pricelist_id in [4, 17, 14, 13, 23]:
            precio_ml = db.query(PrecioML).filter(
                PrecioML.item_id == producto_erp.item_id,
                PrecioML.pricelist_id == pricelist_id
            ).first()
        
            if precio_ml:
                precios_listas[pricelist_id] = {
                    "precio": float(precio_ml.precio) if precio_ml.precio else None,
                    "mla": precio_ml.mla,
                    "cotizacion_dolar": float(precio_ml.cotizacion_dolar) if precio_ml.cotizacion_dolar else None
                }
        
        productos.append({
            "item_id": producto_erp.item_id,
            "codigo": producto_erp.codigo,
            "descripcion": producto_erp.descripcion,
            "marca": producto_erp.marca,
            "categoria": producto_erp.categoria,
            "stock": producto_erp.stock,
            "costo": float(producto_erp.costo),
            "moneda_costo": producto_erp.moneda_costo,
            "precios_listas": precios_listas
        })
    
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "productos": productos
    }


@router.get("/productos/{item_id}", response_model=ProductoResponse)
async def obtener_producto(item_id: int, db: Session = Depends(get_db)):
    result = db.query(ProductoERP, ProductoPricing).outerjoin(
        ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id
    ).filter(ProductoERP.item_id == item_id).first()

    if not result:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    producto_erp, producto_pricing = result
    costo_ars = producto_erp.costo if producto_erp.moneda_costo == "ARS" else None

    return ProductoResponse(
        item_id=producto_erp.item_id,
        codigo=producto_erp.codigo,
        descripcion=producto_erp.descripcion,
        marca=producto_erp.marca,
        categoria=producto_erp.categoria,
        subcategoria_id=producto_erp.subcategoria_id,
        moneda_costo=producto_erp.moneda_costo,
        costo=producto_erp.costo,
        costo_ars=costo_ars,
        iva=producto_erp.iva,
        stock=producto_erp.stock,
        precio_lista_ml=producto_pricing.precio_lista_ml if producto_pricing else None,
        markup=producto_pricing.markup_calculado if producto_pricing else None,
        usuario_modifico=None,
        fecha_modificacion=producto_pricing.fecha_modificacion if producto_pricing else None,
        tiene_precio=producto_pricing.precio_lista_ml is not None if producto_pricing else False,
        necesita_revision=False
    )

@router.get("/stats")
async def obtener_estadisticas(
    # Filtros de búsqueda
    search: Optional[str] = None,
    con_stock: Optional[bool] = None,
    con_precio: Optional[bool] = None,
    marcas: Optional[str] = None,
    subcategorias: Optional[str] = None,
    con_rebate: Optional[bool] = None,
    con_oferta: Optional[bool] = None,
    con_web_transf: Optional[bool] = None,
    markup_clasica_positivo: Optional[bool] = None,
    markup_rebate_positivo: Optional[bool] = None,
    markup_oferta_positivo: Optional[bool] = None,
    markup_web_transf_positivo: Optional[bool] = None,
    out_of_cards: Optional[bool] = None,
    colores: Optional[str] = None,
    product_managers: Optional[str] = None,
    audit_usuarios: Optional[str] = None,
    audit_tipos_accion: Optional[str] = None,
    audit_fecha_desde: Optional[str] = None,
    audit_fecha_hasta: Optional[str] = None,
    con_mla: Optional[bool] = None,
    nuevos_ultimos_7_dias: Optional[bool] = None,
    db: Session = Depends(get_db)
):
    """
    Obtiene estadísticas de productos según filtros aplicados.
    Si no se aplican filtros, devuelve estadísticas globales.
    """
    from datetime import datetime, timedelta
    from app.models.auditoria_precio import AuditoriaPrecio
    from app.models.item_sin_mla_banlist import ItemSinMLABanlist
    from app.models.mercadolibre_item_publicado import MercadoLibreItemPublicado

    # Query base - seleccionar ambos ProductoERP y ProductoPricing
    query = db.query(ProductoERP, ProductoPricing).outerjoin(
        ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id
    )

    # Aplicar filtros de búsqueda
    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            or_(
                ProductoERP.codigo.ilike(search_pattern),
                ProductoERP.descripcion.ilike(search_pattern),
                ProductoERP.marca.ilike(search_pattern)
            )
        )

    # Filtro de stock
    if con_stock is not None:
        if con_stock:
            query = query.filter(ProductoERP.stock > 0)
        else:
            query = query.filter(ProductoERP.stock == 0)

    # Filtro de precio
    if con_precio is not None:
        if con_precio:
            query = query.filter(ProductoPricing.precio_lista_ml.isnot(None))
        else:
            query = query.filter(
                or_(
                    ProductoPricing.precio_lista_ml.is_(None),
                    ProductoPricing.id.is_(None)
                )
            )

    # Filtro de marcas
    if marcas:
        marcas_list = marcas.split(',')
        query = query.filter(ProductoERP.marca.in_(marcas_list))

    # Filtro de subcategorías
    if subcategorias:
        subcategorias_list = [int(s) for s in subcategorias.split(',')]
        query = query.filter(ProductoERP.subcategoria_id.in_(subcategorias_list))

    # Filtro de rebate
    if con_rebate is not None:
        if con_rebate:
            query = query.filter(
                ProductoPricing.participa_rebate == True,
                ProductoPricing.precio_lista_ml.isnot(None)
            )
        else:
            query = query.filter(
                or_(
                    ProductoPricing.participa_rebate == False,
                    ProductoPricing.participa_rebate.is_(None)
                )
            )

    # Filtro de mejor oferta
    if con_oferta is not None:
        if con_oferta:
            query = query.filter(
                ProductoPricing.precio_3_cuotas.isnot(None)
            )
        else:
            query = query.filter(
                or_(
                    ProductoPricing.precio_3_cuotas.is_(None),
                    ProductoPricing.id.is_(None)
                )
            )

    # Filtro de web transferencia
    if con_web_transf is not None:
        if con_web_transf:
            query = query.filter(
                ProductoPricing.participa_web_transferencia == True,
                ProductoPricing.precio_web_transferencia.isnot(None)
            )
        else:
            query = query.filter(
                or_(
                    ProductoPricing.participa_web_transferencia == False,
                    ProductoPricing.participa_web_transferencia.is_(None)
                )
            )

    # Filtros de markup
    if markup_clasica_positivo is not None:
        if markup_clasica_positivo:
            query = query.filter(ProductoPricing.markup_calculado > 0)
        else:
            query = query.filter(
                or_(
                    ProductoPricing.markup_calculado <= 0,
                    ProductoPricing.markup_calculado.is_(None)
                )
            )

    # Filtro out of cards
    if out_of_cards is not None:
        if out_of_cards:
            query = query.filter(ProductoPricing.out_of_cards == True)
        else:
            query = query.filter(
                or_(
                    ProductoPricing.out_of_cards == False,
                    ProductoPricing.out_of_cards.is_(None)
                )
            )

    # Filtro de colores
    if colores:
        colores_list = colores.split(',')
        if 'sin_color' in colores_list:
            colores_con_valor = [c for c in colores_list if c != 'sin_color']
            if colores_con_valor:
                query = query.filter(
                    or_(
                        ProductoPricing.color_marcado.in_(colores_con_valor),
                        ProductoPricing.color_marcado.is_(None)
                    )
                )
            else:
                query = query.filter(ProductoPricing.color_marcado.is_(None))
        else:
            query = query.filter(ProductoPricing.color_marcado.in_(colores_list))

    # Filtro de Product Managers
    if product_managers:
        pm_list = product_managers.split(',')
        pm_ints = [int(pm) for pm in pm_list]
        query = query.filter(ProductoERP.subcategoria_id.in_(
            db.query(Subcategoria.id).filter(Subcategoria.pm_id.in_(pm_ints))
        ))

    # Filtros de auditoría
    if audit_usuarios or audit_tipos_accion or audit_fecha_desde or audit_fecha_hasta:
        subquery_auditoria = db.query(AuditoriaPrecio.item_id).distinct()

        if audit_usuarios:
            usuarios_list = [int(u) for u in audit_usuarios.split(',')]
            subquery_auditoria = subquery_auditoria.filter(AuditoriaPrecio.usuario_id.in_(usuarios_list))

        if audit_tipos_accion:
            tipos_list = audit_tipos_accion.split(',')
            subquery_auditoria = subquery_auditoria.filter(AuditoriaPrecio.tipo_accion.in_(tipos_list))

        if audit_fecha_desde:
            fecha_desde_dt = datetime.fromisoformat(audit_fecha_desde)
            fecha_desde_dt = fecha_desde_dt + timedelta(hours=3)
            subquery_auditoria = subquery_auditoria.filter(AuditoriaPrecio.fecha_accion >= fecha_desde_dt)

        if audit_fecha_hasta:
            fecha_hasta_dt = datetime.fromisoformat(audit_fecha_hasta)
            fecha_hasta_dt = fecha_hasta_dt + timedelta(hours=3)
            subquery_auditoria = subquery_auditoria.filter(AuditoriaPrecio.fecha_accion <= fecha_hasta_dt)

        query = query.filter(ProductoERP.item_id.in_(subquery_auditoria))

    # Filtro de MLA (con/sin publicación) - usar subconsultas para evitar conflictos de join
    if con_mla is not None:
        if con_mla:
            # Con MLA: tienen publicación activa
            items_con_mla_subquery = db.query(MercadoLibreItemPublicado.item_id).filter(
                MercadoLibreItemPublicado.mlp_id.isnot(None),
                or_(
                    MercadoLibreItemPublicado.optval_statusId == 2,
                    MercadoLibreItemPublicado.optval_statusId.is_(None)
                )
            ).distinct().subquery()

            query = query.filter(ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)))
        else:
            # Sin MLA: no tienen publicación (excluye banlist)
            items_con_mla_subquery = db.query(MercadoLibreItemPublicado.item_id).filter(
                MercadoLibreItemPublicado.mlp_id.isnot(None),
                or_(
                    MercadoLibreItemPublicado.optval_statusId == 2,
                    MercadoLibreItemPublicado.optval_statusId.is_(None)
                )
            ).distinct().subquery()

            items_en_banlist_subquery = db.query(ItemSinMLABanlist.item_id).subquery()

            query = query.filter(
                ~ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)),
                ~ProductoERP.item_id.in_(select(items_en_banlist_subquery.c.item_id))
            )

    # Filtro de productos nuevos (últimos 7 días)
    if nuevos_ultimos_7_dias:
        from datetime import timezone
        fecha_limite = datetime.now(timezone.utc) - timedelta(days=7)
        query = query.filter(ProductoERP.fecha_sync >= fecha_limite)

    # ESTADÍSTICAS CALCULADAS
    # Las estadísticas son un desglose de los productos YA filtrados
    # Usar COUNT SQL para mejor rendimiento en lugar de iterar en Python

    from datetime import date, timezone
    from app.models.oferta_ml import OfertaML
    from app.models.publicacion_ml import PublicacionML

    # Total según filtros
    total_filtrado = query.count()

    hoy = date.today()
    fecha_limite_nuevos = datetime.now(timezone.utc) - timedelta(days=7)

    # Subquery para items con MLA
    items_con_mla_subquery = db.query(MercadoLibreItemPublicado.item_id).filter(
        MercadoLibreItemPublicado.mlp_id.isnot(None),
        or_(
            MercadoLibreItemPublicado.optval_statusId == 2,
            MercadoLibreItemPublicado.optval_statusId.is_(None)
        )
    ).distinct().subquery()

    # Subquery para items en banlist
    items_en_banlist_subquery = db.query(ItemSinMLABanlist.item_id).subquery()

    # Subquery para items con oferta vigente
    items_con_oferta_subquery = db.query(PublicacionML.item_id).join(
        OfertaML, PublicacionML.mla == OfertaML.mla
    ).filter(
        OfertaML.fecha_desde <= hoy,
        OfertaML.fecha_hasta >= hoy,
        OfertaML.pvp_seller.isnot(None)
    ).distinct().subquery()

    # Con stock
    total_con_stock = query.filter(ProductoERP.stock > 0).count()

    # Con precio
    total_con_precio = query.filter(ProductoPricing.precio_lista_ml.isnot(None)).count()

    # Nuevos (últimos 7 días)
    nuevos = query.filter(ProductoERP.fecha_sync >= fecha_limite_nuevos).count()

    # Nuevos sin precio
    nuevos_sin_precio = query.filter(
        ProductoERP.fecha_sync >= fecha_limite_nuevos,
        or_(
            ProductoPricing.precio_lista_ml.is_(None),
            ProductoPricing.item_id.is_(None)
        )
    ).count()

    # Con stock sin precio
    stock_sin_precio = query.filter(
        ProductoERP.stock > 0,
        or_(
            ProductoPricing.precio_lista_ml.is_(None),
            ProductoPricing.item_id.is_(None)
        )
    ).count()

    # Sin MLA (no en banlist)
    sin_mla_count = query.filter(
        ~ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)),
        ~ProductoERP.item_id.in_(select(items_en_banlist_subquery.c.item_id))
    ).count()

    # Sin MLA con stock
    sin_mla_con_stock = query.filter(
        ~ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)),
        ~ProductoERP.item_id.in_(select(items_en_banlist_subquery.c.item_id)),
        ProductoERP.stock > 0
    ).count()

    # Sin MLA sin stock
    sin_mla_sin_stock = query.filter(
        ~ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)),
        ~ProductoERP.item_id.in_(select(items_en_banlist_subquery.c.item_id)),
        ProductoERP.stock == 0
    ).count()

    # Sin MLA nuevos
    sin_mla_nuevos = query.filter(
        ~ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)),
        ~ProductoERP.item_id.in_(select(items_en_banlist_subquery.c.item_id)),
        ProductoERP.fecha_sync >= fecha_limite_nuevos
    ).count()

    # Mejor oferta sin rebate
    mejor_oferta_sin_rebate = query.filter(
        ProductoERP.item_id.in_(select(items_con_oferta_subquery.c.item_id)),
        or_(
            ProductoPricing.participa_rebate.is_(False),
            ProductoPricing.participa_rebate.is_(None),
            ProductoPricing.item_id.is_(None)
        )
    ).count()

    # Markup negativo clásica
    markup_negativo_clasica = query.filter(
        ProductoPricing.markup_calculado < 0
    ).count()

    # Markup negativo rebate
    markup_negativo_rebate = query.filter(
        ProductoPricing.markup_rebate < 0
    ).count()

    # Markup negativo oferta
    markup_negativo_oferta = query.filter(
        ProductoPricing.markup_oferta < 0
    ).count()

    # Markup negativo web
    markup_negativo_web = query.filter(
        ProductoPricing.markup_web_real < 0
    ).count()

    return {
        "total_productos": total_filtrado,
        "nuevos_ultimos_7_dias": nuevos,
        "nuevos_sin_precio": nuevos_sin_precio,
        "con_stock_sin_precio": stock_sin_precio,
        "sin_mla_no_banlist": sin_mla_count,
        "sin_mla_con_stock": sin_mla_con_stock,
        "sin_mla_sin_stock": sin_mla_sin_stock,
        "sin_mla_nuevos": sin_mla_nuevos,
        "mejor_oferta_sin_rebate": mejor_oferta_sin_rebate,
        "markup_negativo_clasica": markup_negativo_clasica,
        "markup_negativo_rebate": markup_negativo_rebate,
        "markup_negativo_oferta": markup_negativo_oferta,
        "markup_negativo_web": markup_negativo_web,
        "con_stock": total_con_stock,
        "con_precio": total_con_precio
    }

@router.get("/stats-dinamicos")
async def obtener_stats_dinamicos(
    search: Optional[str] = None,
    categoria: Optional[str] = None,
    marcas: Optional[str] = None,
    subcategorias: Optional[str] = None,
    con_stock: Optional[bool] = None,
    con_precio: Optional[bool] = None,
    con_rebate: Optional[bool] = None,
    con_oferta: Optional[bool] = None,
    con_web_transf: Optional[bool] = None,
    tn_con_descuento: Optional[bool] = None,
    tn_sin_descuento: Optional[bool] = None,
    tn_no_publicado: Optional[bool] = None,
    markup_clasica_positivo: Optional[bool] = None,
    markup_rebate_positivo: Optional[bool] = None,
    markup_oferta_positivo: Optional[bool] = None,
    markup_web_transf_positivo: Optional[bool] = None,
    out_of_cards: Optional[bool] = None,
    colores: Optional[str] = None,
    pms: Optional[str] = None,
    con_mla: Optional[bool] = None,
    nuevos_ultimos_7_dias: Optional[bool] = None,
    db: Session = Depends(get_db)
):
    """
    Obtiene estadísticas dinámicas de productos según filtros aplicados.
    Las estadísticas se calculan SOLO sobre los productos que cumplen con los filtros.
    """
    from datetime import datetime, timedelta, date, timezone
    from app.models.oferta_ml import OfertaML
    from app.models.publicacion_ml import PublicacionML
    from app.models.item_sin_mla_banlist import ItemSinMLABanlist
    from app.models.mercadolibre_item_publicado import MercadoLibreItemPublicado
    from app.models.subcategoria import Subcategoria

    # Query base - igual que en /productos
    query = db.query(ProductoERP, ProductoPricing).outerjoin(
        ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id
    )

    # APLICAR TODOS LOS FILTROS (copiado del endpoint /productos)

    # Filtro de búsqueda
    if search:
        search_normalized = search.replace('-', '').replace(' ', '').upper()
        query = query.filter(
            or_(
                func.replace(func.replace(func.upper(ProductoERP.descripcion), '-', ''), ' ', '').like(f"%{search_normalized}%"),
                func.replace(func.replace(func.upper(ProductoERP.marca), '-', ''), ' ', '').like(f"%{search_normalized}%"),
                func.replace(func.upper(ProductoERP.codigo), '-', '').like(f"%{search_normalized}%")
            )
        )

    # Filtros básicos
    if categoria:
        query = query.filter(ProductoERP.categoria == categoria)

    if marcas:
        marcas_list = marcas.split(',')
        query = query.filter(ProductoERP.marca.in_(marcas_list))

    if subcategorias:
        subcat_list = [int(s) for s in subcategorias.split(',')]
        query = query.filter(ProductoERP.subcategoria_id.in_(subcat_list))

    if con_stock is not None:
        if con_stock:
            query = query.filter(ProductoERP.stock > 0)
        else:
            query = query.filter(ProductoERP.stock == 0)

    if con_precio is not None:
        if con_precio:
            query = query.filter(ProductoPricing.precio_lista_ml.isnot(None))
        else:
            query = query.filter(
                or_(
                    ProductoPricing.precio_lista_ml.is_(None),
                    ProductoPricing.item_id.is_(None)
                )
            )

    # Filtros de participación
    if con_rebate is not None:
        if con_rebate:
            query = query.filter(ProductoPricing.participa_rebate == True)
        else:
            query = query.filter(
                or_(
                    ProductoPricing.participa_rebate == False,
                    ProductoPricing.participa_rebate.is_(None),
                    ProductoPricing.item_id.is_(None)
                )
            )

    if con_web_transf is not None:
        if con_web_transf:
            query = query.filter(ProductoPricing.participa_web_transferencia == True)
        else:
            query = query.filter(
                or_(
                    ProductoPricing.participa_web_transferencia == False,
                    ProductoPricing.participa_web_transferencia.is_(None),
                    ProductoPricing.item_id.is_(None)
                )
            )

    # Filtros de Tienda Nube
    if tn_con_descuento or tn_sin_descuento or tn_no_publicado:
        from app.models.tienda_nube_producto import TiendaNubeProducto

        if tn_con_descuento:
            query = query.join(
                TiendaNubeProducto,
                and_(
                    ProductoERP.item_id == TiendaNubeProducto.item_id,
                    TiendaNubeProducto.activo == True,
                    TiendaNubeProducto.promotional_price.isnot(None),
                    TiendaNubeProducto.promotional_price > 0
                )
            )
        elif tn_sin_descuento:
            query = query.join(
                TiendaNubeProducto,
                and_(
                    ProductoERP.item_id == TiendaNubeProducto.item_id,
                    TiendaNubeProducto.activo == True,
                    or_(
                        TiendaNubeProducto.promotional_price.is_(None),
                        TiendaNubeProducto.promotional_price == 0
                    )
                )
            )
        elif tn_no_publicado:
            from sqlalchemy.sql import exists
            subquery = exists().where(
                and_(
                    TiendaNubeProducto.item_id == ProductoERP.item_id,
                    TiendaNubeProducto.activo == True
                )
            )
            query = query.filter(
                and_(
                    ProductoERP.stock > 0,
                    ~subquery
                )
            )

    # Filtros de markup
    if markup_clasica_positivo is not None:
        if markup_clasica_positivo:
            query = query.filter(ProductoPricing.markup_calculado > 0)
        else:
            query = query.filter(ProductoPricing.markup_calculado < 0)

    if markup_rebate_positivo is not None:
        if markup_rebate_positivo:
            query = query.filter(ProductoPricing.markup_rebate > 0)
        else:
            query = query.filter(ProductoPricing.markup_rebate < 0)

    if markup_oferta_positivo is not None:
        if markup_oferta_positivo:
            query = query.filter(ProductoPricing.markup_oferta > 0)
        else:
            query = query.filter(ProductoPricing.markup_oferta < 0)

    if markup_web_transf_positivo is not None:
        if markup_web_transf_positivo:
            query = query.filter(ProductoPricing.markup_web_real > 0)
        else:
            query = query.filter(ProductoPricing.markup_web_real < 0)

    if out_of_cards is not None:
        if out_of_cards:
            query = query.filter(ProductoPricing.out_of_cards == True)
        else:
            query = query.filter(
                or_(
                    ProductoPricing.out_of_cards == False,
                    ProductoPricing.out_of_cards.is_(None)
                )
            )

    # Filtro de oferta
    if con_oferta is not None:
        from datetime import date
        hoy = date.today()

        items_con_oferta_subquery = db.query(PublicacionML.item_id).join(
            OfertaML, PublicacionML.mla == OfertaML.mla
        ).filter(
            OfertaML.fecha_desde <= hoy,
            OfertaML.fecha_hasta >= hoy,
            OfertaML.pvp_seller.isnot(None)
        ).distinct().subquery()

        if con_oferta:
            query = query.filter(ProductoERP.item_id.in_(items_con_oferta_subquery))
        else:
            query = query.filter(~ProductoERP.item_id.in_(items_con_oferta_subquery))

    # Filtro de colores
    if colores:
        colores_list = colores.split(',')
        query = query.filter(ProductoPricing.color_marcado.in_(colores_list))

    # Filtro de PMs
    if pms:
        pm_list = pms.split(',')
        pm_ints = [int(pm) for pm in pm_list]
        query = query.filter(ProductoERP.subcategoria_id.in_(
            db.query(Subcategoria.id).filter(Subcategoria.pm_id.in_(pm_ints))
        ))

    # Filtro de MLA
    if con_mla is not None:
        if con_mla:
            items_con_mla_subquery = db.query(MercadoLibreItemPublicado.item_id).filter(
                MercadoLibreItemPublicado.mlp_id.isnot(None),
                or_(
                    MercadoLibreItemPublicado.optval_statusId == 2,
                    MercadoLibreItemPublicado.optval_statusId.is_(None)
                )
            ).distinct().subquery()

            query = query.filter(ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)))
        else:
            items_con_mla_subquery = db.query(MercadoLibreItemPublicado.item_id).filter(
                MercadoLibreItemPublicado.mlp_id.isnot(None),
                or_(
                    MercadoLibreItemPublicado.optval_statusId == 2,
                    MercadoLibreItemPublicado.optval_statusId.is_(None)
                )
            ).distinct().subquery()

            items_en_banlist_subquery = db.query(ItemSinMLABanlist.item_id).subquery()

            query = query.filter(
                ~ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)),
                ~ProductoERP.item_id.in_(select(items_en_banlist_subquery.c.item_id))
            )

    # Filtro de productos nuevos
    if nuevos_ultimos_7_dias:
        fecha_limite = datetime.now(timezone.utc) - timedelta(days=7)
        query = query.filter(ProductoERP.fecha_sync >= fecha_limite)

    # CALCULAR ESTADÍSTICAS SOBRE PRODUCTOS FILTRADOS

    hoy = date.today()
    fecha_limite_nuevos = datetime.now(timezone.utc) - timedelta(days=7)

    # Subqueries para cálculos
    items_con_mla_subquery = db.query(MercadoLibreItemPublicado.item_id).filter(
        MercadoLibreItemPublicado.mlp_id.isnot(None),
        or_(
            MercadoLibreItemPublicado.optval_statusId == 2,
            MercadoLibreItemPublicado.optval_statusId.is_(None)
        )
    ).distinct().subquery()

    items_en_banlist_subquery = db.query(ItemSinMLABanlist.item_id).subquery()

    items_con_oferta_subquery = db.query(PublicacionML.item_id).join(
        OfertaML, PublicacionML.mla == OfertaML.mla
    ).filter(
        OfertaML.fecha_desde <= hoy,
        OfertaML.fecha_hasta >= hoy,
        OfertaML.pvp_seller.isnot(None)
    ).distinct().subquery()

    # Total según filtros
    total_filtrado = query.count()

    # Estadísticas con COUNT SQL
    total_con_stock = query.filter(ProductoERP.stock > 0).count()
    total_con_precio = query.filter(ProductoPricing.precio_lista_ml.isnot(None)).count()
    nuevos = query.filter(ProductoERP.fecha_sync >= fecha_limite_nuevos).count()

    nuevos_sin_precio = query.filter(
        ProductoERP.fecha_sync >= fecha_limite_nuevos,
        or_(
            ProductoPricing.precio_lista_ml.is_(None),
            ProductoPricing.item_id.is_(None)
        )
    ).count()

    stock_sin_precio = query.filter(
        ProductoERP.stock > 0,
        or_(
            ProductoPricing.precio_lista_ml.is_(None),
            ProductoPricing.item_id.is_(None)
        )
    ).count()

    sin_mla_count = query.filter(
        ~ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)),
        ~ProductoERP.item_id.in_(select(items_en_banlist_subquery.c.item_id))
    ).count()

    sin_mla_con_stock = query.filter(
        ~ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)),
        ~ProductoERP.item_id.in_(select(items_en_banlist_subquery.c.item_id)),
        ProductoERP.stock > 0
    ).count()

    sin_mla_sin_stock = query.filter(
        ~ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)),
        ~ProductoERP.item_id.in_(select(items_en_banlist_subquery.c.item_id)),
        ProductoERP.stock == 0
    ).count()

    sin_mla_nuevos = query.filter(
        ~ProductoERP.item_id.in_(select(items_con_mla_subquery.c.item_id)),
        ~ProductoERP.item_id.in_(select(items_en_banlist_subquery.c.item_id)),
        ProductoERP.fecha_sync >= fecha_limite_nuevos
    ).count()

    mejor_oferta_sin_rebate = query.filter(
        ProductoERP.item_id.in_(select(items_con_oferta_subquery.c.item_id)),
        or_(
            ProductoPricing.participa_rebate.is_(False),
            ProductoPricing.participa_rebate.is_(None),
            ProductoPricing.item_id.is_(None)
        )
    ).count()

    markup_negativo_clasica = query.filter(ProductoPricing.markup_calculado < 0).count()
    markup_negativo_rebate = query.filter(ProductoPricing.markup_rebate < 0).count()
    markup_negativo_oferta = query.filter(ProductoPricing.markup_oferta < 0).count()
    markup_negativo_web = query.filter(ProductoPricing.markup_web_real < 0).count()

    return {
        "total_productos": total_filtrado,
        "nuevos_ultimos_7_dias": nuevos,
        "nuevos_sin_precio": nuevos_sin_precio,
        "con_stock_sin_precio": stock_sin_precio,
        "sin_mla_no_banlist": sin_mla_count,
        "sin_mla_con_stock": sin_mla_con_stock,
        "sin_mla_sin_stock": sin_mla_sin_stock,
        "sin_mla_nuevos": sin_mla_nuevos,
        "mejor_oferta_sin_rebate": mejor_oferta_sin_rebate,
        "markup_negativo_clasica": markup_negativo_clasica,
        "markup_negativo_rebate": markup_negativo_rebate,
        "markup_negativo_oferta": markup_negativo_oferta,
        "markup_negativo_web": markup_negativo_web,
        "con_stock": total_con_stock,
        "con_precio": total_con_precio
    }


@router.get("/categorias")
async def listar_categorias(db: Session = Depends(get_db)):
    categorias = db.query(ProductoERP.categoria).distinct().order_by(ProductoERP.categoria).all()
    return {"categorias": [c[0] for c in categorias if c[0]]}

@router.get("/marcas")
async def listar_marcas(
    search: Optional[str] = None,
    con_stock: Optional[bool] = None,
    con_precio: Optional[bool] = None,
    subcategorias: Optional[str] = None,
    con_rebate: Optional[bool] = None,
    con_oferta: Optional[bool] = None,
    con_web_transf: Optional[bool] = None,
    markup_clasica_positivo: Optional[bool] = None,
    markup_rebate_positivo: Optional[bool] = None,
    markup_oferta_positivo: Optional[bool] = None,
    markup_web_transf_positivo: Optional[bool] = None,
    out_of_cards: Optional[bool] = None,
    colores: Optional[str] = None,
    audit_usuarios: Optional[str] = None,
    audit_tipos_accion: Optional[str] = None,
    audit_fecha_desde: Optional[str] = None,
    audit_fecha_hasta: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Lista marcas disponibles según filtros activos"""

    # Query base igual que en el endpoint de listar productos
    query = db.query(ProductoERP.marca).distinct().join(
        ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id, isouter=True
    )

    # Aplicar filtros (reutilizar la lógica del endpoint de listar productos)
    if search:
        query = query.filter(
            or_(
                ProductoERP.codigo.ilike(f'%{search}%'),
                ProductoERP.descripcion.ilike(f'%{search}%'),
                ProductoERP.marca.ilike(f'%{search}%')
            )
        )

    if con_stock is not None:
        if con_stock:
            query = query.filter(ProductoERP.stock > 0)
        else:
            query = query.filter(ProductoERP.stock == 0)

    if con_precio is not None:
        if con_precio:
            query = query.filter(ProductoPricing.precio_lista_ml.isnot(None))
        else:
            query = query.filter(ProductoPricing.precio_lista_ml.is_(None))

    if subcategorias:
        subcat_list = [int(s.strip()) for s in subcategorias.split(',') if s.strip()]
        query = query.filter(ProductoERP.subcategoria_id.in_(subcat_list))

    if con_rebate is not None:
        query = query.filter(ProductoPricing.participa_rebate == con_rebate)

    if con_web_transf is not None:
        query = query.filter(ProductoPricing.participa_web_transferencia == con_web_transf)

    if out_of_cards is not None:
        if out_of_cards:
            query = query.filter(ProductoPricing.out_of_cards == True)
        else:
            query = query.filter(
                or_(
                    ProductoPricing.out_of_cards == False,
                    ProductoPricing.out_of_cards.is_(None)
                )
            )

    if colores:
        colores_list = colores.split(',')
        query = query.filter(ProductoPricing.color_marcado.in_(colores_list))

    if markup_clasica_positivo is not None:
        if markup_clasica_positivo:
            query = query.filter(ProductoPricing.markup_calculado > 0)
        else:
            query = query.filter(ProductoPricing.markup_calculado < 0)

    marcas = query.order_by(ProductoERP.marca).all()
    return {"marcas": [m[0] for m in marcas if m[0]]}

@router.get("/productos/{item_id}/ofertas-vigentes")
async def obtener_ofertas_vigentes(item_id: int, db: Session = Depends(get_db)):
    from app.models.publicacion_ml import PublicacionML
    from app.models.oferta_ml import OfertaML
    from app.services.pricing_calculator import (
        obtener_tipo_cambio_actual,
        convertir_a_pesos,
        obtener_grupo_subcategoria,
        obtener_comision_base,
        calcular_comision_ml_total,
        calcular_limpio,
        calcular_markup
    )
    
    producto = db.query(ProductoERP).filter(ProductoERP.item_id == item_id).first()
    if not producto:
        return {"item_id": item_id, "publicaciones": []}
    
    tipo_cambio = None
    if producto.moneda_costo == "USD":
        tipo_cambio = obtener_tipo_cambio_actual(db, "USD")
    costo_ars = convertir_a_pesos(producto.costo, producto.moneda_costo, tipo_cambio)
    
    publicaciones = db.query(PublicacionML).filter(PublicacionML.item_id == item_id).all()
    if not publicaciones:
        return {"item_id": item_id, "publicaciones": []}
    
    hoy = date.today()
    resultado = []
    
    for pub in publicaciones:
        oferta = db.query(OfertaML).filter(
            OfertaML.mla == pub.mla,
            OfertaML.fecha_desde <= hoy,
            OfertaML.fecha_hasta >= hoy
        ).first()
        
        markup_oferta = None
        if oferta and oferta.pvp_seller and oferta.pvp_seller > 0:
            grupo_id = obtener_grupo_subcategoria(db, producto.subcategoria_id)
            comision_base = obtener_comision_base(db, pub.pricelist_id, grupo_id)
            
            if comision_base:
                comisiones = calcular_comision_ml_total(
                    oferta.pvp_seller,
                    comision_base,
                    producto.iva,
                    db=db
                )
                limpio = calcular_limpio(
                    oferta.pvp_seller,
                    producto.iva,
                    producto.envio or 0,
                    comisiones["comision_total"],
                    db=db,
                    grupo_id=grupo_id
                )
                markup_oferta = round(calcular_markup(limpio, costo_ars) * 100, 2)
        
        resultado.append({
            "mla": pub.mla,
            "item_title": pub.item_title,
            "pricelist_id": pub.pricelist_id,
            "lista_nombre": pub.lista_nombre,
            "tiene_oferta": oferta is not None,
            "oferta": {
                "precio_final": oferta.precio_final,
                "pvp_seller": oferta.pvp_seller,
                "markup_oferta": markup_oferta,
                "aporte_meli_pesos": oferta.aporte_meli_pesos,
                "aporte_meli_porcentaje": oferta.aporte_meli_porcentaje,
                "fecha_desde": oferta.fecha_desde.isoformat(),
                "fecha_hasta": oferta.fecha_hasta.isoformat(),
            } if oferta else None
        })
    
    return {
        "item_id": item_id,
        "total_publicaciones": len(resultado),
        "con_oferta": sum(1 for r in resultado if r["tiene_oferta"]),
        "publicaciones": resultado
    }
    
@router.patch("/productos/{producto_id}/precio")
async def actualizar_precio(
    producto_id: int,
    datos: PrecioUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Actualiza precio de un producto y registra en auditoría"""
    
    producto = db.query(ProductoPricing).filter(ProductoPricing.id == producto_id).first()
    if not producto:
        raise HTTPException(404, "Producto no encontrado")
    
    # Guardar valores anteriores para auditoría
    precio_ant = producto.precio_lista_final
    contado_ant = producto.precio_contado_final
    
    # Actualizar precios
    if datos.precio_lista_final is not None:
        producto.precio_lista_final = datos.precio_lista_final
    if datos.precio_contado_final is not None:
        producto.precio_contado_final = datos.precio_contado_final
    
    # Registrar en auditoría SOLO si cambió algún precio
    if (datos.precio_lista_final is not None and precio_ant != datos.precio_lista_final) or \
       (datos.precio_contado_final is not None and contado_ant != datos.precio_contado_final):
        
        auditoria = AuditoriaPrecio(
            producto_id=producto_id,
            usuario_id=current_user.id,
            precio_anterior=precio_ant,
            precio_contado_anterior=contado_ant,
            precio_nuevo=producto.precio_lista_final,
            precio_contado_nuevo=producto.precio_contado_final,
            comentario=datos.comentario if hasattr(datos, 'comentario') else None
        )
        db.add(auditoria)
    
    db.commit()
    db.refresh(producto)
    
    return producto


@router.patch("/productos/{item_id}/rebate")
async def actualizar_rebate(
    item_id: int,
    datos: RebateUpdate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Actualiza configuración de rebate de un producto"""
    from app.services.auditoria_service import registrar_auditoria
    from app.models.auditoria import TipoAccion

    pricing = db.query(ProductoPricing).filter(ProductoPricing.item_id == item_id).first()

    # Guardar valores anteriores
    valores_anteriores = {
        "participa_rebate": pricing.participa_rebate if pricing else False,
        "porcentaje_rebate": float(pricing.porcentaje_rebate) if pricing and pricing.porcentaje_rebate is not None else None
    }

    if not pricing:
        pricing = ProductoPricing(
            item_id=item_id,
            participa_rebate=datos.participa_rebate,
            porcentaje_rebate=datos.porcentaje_rebate,
            usuario_id=current_user.id
        )
        db.add(pricing)
    else:
        pricing.participa_rebate = datos.participa_rebate
        pricing.porcentaje_rebate = datos.porcentaje_rebate
        pricing.fecha_modificacion = datetime.now()

    db.commit()
    db.refresh(pricing)

    # Determinar tipo de acción
    if datos.participa_rebate and not valores_anteriores["participa_rebate"]:
        tipo_accion = TipoAccion.ACTIVAR_REBATE
    elif not datos.participa_rebate and valores_anteriores["participa_rebate"]:
        tipo_accion = TipoAccion.DESACTIVAR_REBATE
    else:
        tipo_accion = TipoAccion.MODIFICAR_PORCENTAJE_REBATE

    # Registrar auditoría
    registrar_auditoria(
        db=db,
        usuario_id=current_user.id,
        tipo_accion=tipo_accion,
        item_id=item_id,
        valores_anteriores=valores_anteriores,
        valores_nuevos={
            "participa_rebate": datos.participa_rebate,
            "porcentaje_rebate": datos.porcentaje_rebate
        }
    )

    return {
        "item_id": item_id,
        "participa_rebate": datos.participa_rebate,
        "porcentaje_rebate": datos.porcentaje_rebate
    }
        
class ExportRebateRequest(BaseModel):
    fecha_desde: Optional[str] = None
    fecha_hasta: Optional[str] = None
    filtros: Optional[dict] = None

@router.post("/productos/exportar-rebate")
async def exportar_rebate(
    request: ExportRebateRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Exporta productos con rebate a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from datetime import datetime, date
    from calendar import monthrange
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from app.models.publicacion_ml import PublicacionML
    from app.models.mla_banlist import MLABanlist

    # Obtener MLAs baneados
    mlas_baneados = db.query(MLABanlist.mla).filter(MLABanlist.activo == True).all()
    mlas_baneados_set = {mla[0] for mla in mlas_baneados}

    # Fechas por defecto
    hoy = date.today()
    fecha_desde = request.fecha_desde
    fecha_hasta = request.fecha_hasta
    if not fecha_desde:
        fecha_desde = hoy.strftime('%Y-%m-%d')
    if not fecha_hasta:
        ultimo_dia = monthrange(hoy.year, hoy.month)[1]
        fecha_hasta = f"{hoy.year}-{hoy.month:02d}-{ultimo_dia:02d}"

    # Construir query con filtros
    query = db.query(ProductoERP, ProductoPricing).join(
        ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id
    ).filter(
        ProductoPricing.participa_rebate == True,
        ProductoPricing.out_of_cards != True
    )

    # Aplicar filtros si existen
    if request.filtros:
        filtros = request.filtros

        if filtros.get('search'):
            search_normalized = filtros['search'].replace('-', '').replace(' ', '').upper()
            query = query.filter(
                or_(
                    func.replace(func.replace(func.upper(ProductoERP.descripcion), '-', ''), ' ', '').like(f"%{search_normalized}%"),
                    func.replace(func.replace(func.upper(ProductoERP.marca), '-', ''), ' ', '').like(f"%{search_normalized}%"),
                    func.replace(func.upper(ProductoERP.codigo), '-', '').like(f"%{search_normalized}%")
                )
            )

        if filtros.get('con_stock') is not None:
            query = query.filter(ProductoERP.stock > 0 if filtros['con_stock'] else ProductoERP.stock == 0)

        if filtros.get('con_precio') is not None:
            if filtros['con_precio']:
                query = query.filter(ProductoPricing.precio_lista_ml.isnot(None))
            else:
                query = query.filter(ProductoPricing.precio_lista_ml.is_(None))

        if filtros.get('marcas'):
            marcas_list = [m.strip().upper() for m in filtros['marcas'].split(',')]
            query = query.filter(func.upper(ProductoERP.marca).in_(marcas_list))

        if filtros.get('subcategorias'):
            subcat_list = [int(s.strip()) for s in filtros['subcategorias'].split(',')]
            query = query.filter(ProductoERP.subcategoria_id.in_(subcat_list))

        if filtros.get('con_oferta') is not None:
            # Filtro de oferta si es necesario
            pass

        if filtros.get('con_web_transf') is not None:
            if filtros['con_web_transf']:
                query = query.filter(ProductoPricing.participa_web_transferencia == True)
            else:
                query = query.filter(
                    or_(
                        ProductoPricing.participa_web_transferencia == False,
                        ProductoPricing.participa_web_transferencia.is_(None)
                    )
                )

        if filtros.get('colores'):
            colores_list = filtros['colores'].split(',')
            query = query.filter(ProductoPricing.color_marcado.in_(colores_list))

        if filtros.get('pms'):
            from app.models.marca_pm import MarcaPM
            pms_ids = [int(pm) for pm in filtros['pms'].split(',')]
            marcas_asignadas = db.query(MarcaPM.marca).filter(MarcaPM.usuario_id.in_(pms_ids)).all()
            marcas_list = [m[0].upper() for m in marcas_asignadas]
            if marcas_list:
                query = query.filter(func.upper(ProductoERP.marca).in_(marcas_list))

        if filtros.get('con_rebate') is not None:
            if filtros['con_rebate']:
                query = query.filter(ProductoPricing.participa_rebate == True)
            else:
                query = query.filter(
                    or_(
                        ProductoPricing.participa_rebate == False,
                        ProductoPricing.participa_rebate.is_(None)
                    )
                )

        if filtros.get('out_of_cards') is not None:
            if filtros['out_of_cards']:
                query = query.filter(ProductoPricing.out_of_cards == True)
            else:
                query = query.filter(
                    or_(
                        ProductoPricing.out_of_cards == False,
                        ProductoPricing.out_of_cards.is_(None)
                    )
                )

        if filtros.get('markup_clasica_positivo') is not None:
            if filtros['markup_clasica_positivo']:
                query = query.filter(ProductoPricing.markup > 0)
            else:
                query = query.filter(ProductoPricing.markup <= 0)

        if filtros.get('markup_rebate_positivo') is not None:
            if filtros['markup_rebate_positivo']:
                query = query.filter(ProductoPricing.markup_rebate > 0)
            else:
                query = query.filter(ProductoPricing.markup_rebate <= 0)

        if filtros.get('markup_oferta_positivo') is not None:
            if filtros['markup_oferta_positivo']:
                query = query.filter(ProductoPricing.mejor_oferta_markup > 0)
            else:
                query = query.filter(ProductoPricing.mejor_oferta_markup <= 0)

        if filtros.get('markup_web_transf_positivo') is not None:
            if filtros['markup_web_transf_positivo']:
                query = query.filter(ProductoPricing.markup_web_real > 0)
            else:
                query = query.filter(ProductoPricing.markup_web_real <= 0)

        # Filtros de auditoría
        if filtros.get('audit_usuarios') or filtros.get('audit_tipos_accion') or filtros.get('audit_fecha_desde') or filtros.get('audit_fecha_hasta'):
            from app.models.auditoria import Auditoria

            # Subquery para obtener item_ids que cumplen con los filtros de auditoría
            audit_query = db.query(Auditoria.item_id).distinct()

            if filtros.get('audit_usuarios'):
                usuarios_ids = [int(u) for u in filtros['audit_usuarios'].split(',')]
                audit_query = audit_query.filter(Auditoria.usuario_id.in_(usuarios_ids))

            if filtros.get('audit_tipos_accion'):
                tipos_list = filtros['audit_tipos_accion'].split(',')
                audit_query = audit_query.filter(Auditoria.tipo_accion.in_(tipos_list))

            if filtros.get('audit_fecha_desde'):
                audit_query = audit_query.filter(Auditoria.fecha >= filtros['audit_fecha_desde'])

            if filtros.get('audit_fecha_hasta'):
                audit_query = audit_query.filter(Auditoria.fecha <= filtros['audit_fecha_hasta'])

            item_ids_auditados = [item_id for (item_id,) in audit_query.all()]
            if item_ids_auditados:
                query = query.filter(ProductoERP.item_id.in_(item_ids_auditados))
            else:
                # Si no hay items que cumplan con los filtros de auditoría, no devolver nada
                query = query.filter(ProductoERP.item_id == -1)

    productos = query.all()
    
    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Rebate Export"
    
    # Headers
    headers = [
        "REBATE", "MARCA", "DESDE", "HASTA", "TIPO DE OFERTA", "CATEGORÍA",
        "DESCRIPCIÓN DE LA PUBLICACIÓN", "TIPO DE PUBLICACIÓN", "STOCK",
        "FULL", "MLAs", "PVP LLENO", "PVP SELLER"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    row = 2
    for producto_erp, producto_pricing in productos:
        # Buscar MLAs de lista clásica (pricelist_id = 4)
        mlas = db.query(PublicacionML).filter(
            PublicacionML.item_id == producto_erp.item_id,
            PublicacionML.pricelist_id == 4,
            PublicacionML.activo == True
        ).all()
        
        # Si no tiene MLAs, skip
        if not mlas:
            continue

        # Obtener precio de lista clásica (pricelist_id = 4) de PrecioML
        from app.models.precio_ml import PrecioML
        precio_clasica = db.query(PrecioML).filter(
            PrecioML.item_id == producto_erp.item_id,
            PrecioML.pricelist_id == 4
        ).first()

        # PVP LLENO = Precio de la lista de precios 4 en MercadoLibre
        pvp_lleno = float(precio_clasica.precio) if precio_clasica and precio_clasica.precio else 0

        if pvp_lleno == 0:
            continue

        # PVP SELLER = Precio con rebate aplicado (mismo cálculo que en línea 369)
        # Basado en precio_lista_ml de ProductoPricing
        if not producto_pricing.precio_lista_ml:
            continue

        porcentaje_rebate = float(producto_pricing.porcentaje_rebate or 3.8)
        pvp_seller = float(producto_pricing.precio_lista_ml) / (1 - porcentaje_rebate / 100)
        
        # Una fila por cada MLA (excluyendo los baneados)
        for mla in mlas:
            # Saltar si el MLA está en la banlist
            if mla.mla in mlas_baneados_set:
                continue

            ws.cell(row=row, column=1, value=f"{porcentaje_rebate}%")
            ws.cell(row=row, column=2, value=producto_erp.marca or "")
            ws.cell(row=row, column=3, value=fecha_desde)
            ws.cell(row=row, column=4, value=fecha_hasta)
            ws.cell(row=row, column=5, value="DXI")
            ws.cell(row=row, column=6, value="")  # Categoría vacía
            ws.cell(row=row, column=7, value=mla.item_title or producto_erp.descripcion or "")
            ws.cell(row=row, column=8, value="Clásica")
            ws.cell(row=row, column=9, value=producto_erp.stock)
            ws.cell(row=row, column=10, value="FALSE")
            ws.cell(row=row, column=11, value=mla.mla)
            ws.cell(row=row, column=12, value=pvp_lleno)
            ws.cell(row=row, column=13, value=round(pvp_seller, 2))
            row += 1
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=rebate_export_{hoy.strftime('%Y%m%d')}.xlsx"}
    )

@router.patch("/productos/{item_id}/web-transferencia")
async def actualizar_web_transferencia(
    item_id: int,
    participa: bool,
    porcentaje_markup: float = 6.0,
    preservar_porcentaje: bool = False,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Activa/desactiva web transferencia y calcula precio"""
    from app.services.pricing_calculator import (
        calcular_precio_web_transferencia,
        obtener_tipo_cambio_actual,
        convertir_a_pesos
    )
    from app.services.auditoria_service import registrar_auditoria
    from app.models.auditoria import TipoAccion

    producto_erp = db.query(ProductoERP).filter(ProductoERP.item_id == item_id).first()
    if not producto_erp:
        raise HTTPException(404, "Producto no encontrado")

    pricing = db.query(ProductoPricing).filter(ProductoPricing.item_id == item_id).first()

    # Guardar valores anteriores
    valores_anteriores = {
        "participa_web_transferencia": pricing.participa_web_transferencia if pricing else False,
        "porcentaje_markup_web": float(pricing.porcentaje_markup_web) if pricing and pricing.porcentaje_markup_web else None,
        "precio_web_transferencia": float(pricing.precio_web_transferencia) if pricing and pricing.precio_web_transferencia else None
    }

    if not pricing:
        pricing = ProductoPricing(
            item_id=item_id,
            participa_web_transferencia=participa,
            porcentaje_markup_web=porcentaje_markup,
            preservar_porcentaje_web=preservar_porcentaje,
            usuario_id=current_user.id
        )
        db.add(pricing)
    else:
        pricing.participa_web_transferencia = participa
        pricing.porcentaje_markup_web = porcentaje_markup
        pricing.preservar_porcentaje_web = preservar_porcentaje
        pricing.fecha_modificacion = datetime.now()

    precio_web = None
    markup_web_real = None

    if participa:
        tipo_cambio = None
        if producto_erp.moneda_costo == "USD":
            tipo_cambio = obtener_tipo_cambio_actual(db, "USD")

        costo_ars = convertir_a_pesos(producto_erp.costo, producto_erp.moneda_costo, tipo_cambio)

        if pricing.markup_calculado is not None and pricing.precio_lista_ml is not None:
            markup_clasica = pricing.markup_calculado / 100
        else:
            markup_clasica = 0

        markup_objetivo = markup_clasica + (porcentaje_markup / 100)

        resultado = calcular_precio_web_transferencia(
            costo_ars=costo_ars,
            iva=producto_erp.iva,
            markup_objetivo=markup_objetivo
        )

        precio_web = resultado["precio"]
        markup_web_real = resultado["markup_real"]
        pricing.precio_web_transferencia = precio_web
        pricing.markup_web_real = markup_web_real
    else:
        pricing.precio_web_transferencia = None
        pricing.markup_web_real = None

    db.commit()
    db.refresh(pricing)

    # Determinar tipo de acción
    if participa and not valores_anteriores["participa_web_transferencia"]:
        tipo_accion = TipoAccion.ACTIVAR_WEB_TRANSFERENCIA
    elif not participa and valores_anteriores["participa_web_transferencia"]:
        tipo_accion = TipoAccion.DESACTIVAR_WEB_TRANSFERENCIA
    else:
        tipo_accion = TipoAccion.MODIFICAR_PRECIO_WEB

    # Registrar auditoría
    registrar_auditoria(
        db=db,
        usuario_id=current_user.id,
        tipo_accion=tipo_accion,
        item_id=item_id,
        valores_anteriores=valores_anteriores,
        valores_nuevos={
            "participa_web_transferencia": participa,
            "porcentaje_markup_web": porcentaje_markup,
            "precio_web_transferencia": precio_web,
            "markup_web_real": markup_web_real
        }
    )

    return {
        "item_id": item_id,
        "participa_web_transferencia": participa,
        "porcentaje_markup_web": porcentaje_markup,
        "precio_web_transferencia": precio_web,
        "markup_web_real": markup_web_real if precio_web else None
    }
    
class CalculoWebMasivoRequest(BaseModel):
        porcentaje_con_precio: float
        porcentaje_sin_precio: float
        filtros: dict = None  # ← AGREGAR
    
@router.post("/productos/calcular-web-masivo")
async def calcular_web_masivo(
    request: CalculoWebMasivoRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Calcula precio web transferencia masivamente"""
    from app.services.pricing_calculator import (
        calcular_precio_web_transferencia,
        obtener_tipo_cambio_actual,
        convertir_a_pesos
    )

    # Obtener productos base
    query = db.query(ProductoERP, ProductoPricing).outerjoin(
        ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id
    )
    
    # Aplicar filtros si existen
    if request.filtros:
        if request.filtros.get('search'):
            search_term = f"%{request.filtros['search']}%"
            query = query.filter(
                (ProductoERP.descripcion.ilike(search_term)) |
                (ProductoERP.codigo.ilike(search_term))
            )
        
        if request.filtros.get('con_stock'):
            query = query.filter(ProductoERP.stock > 0)
        
        if request.filtros.get('con_precio'):
            query = query.filter(ProductoPricing.precio_lista_ml.isnot(None))
        
        if request.filtros.get('marcas'):
            marcas_list = request.filtros['marcas'].split(',')
            query = query.filter(ProductoERP.marca.in_(marcas_list))
        
        if request.filtros.get('subcategorias'):
            subcats_list = [int(s) for s in request.filtros['subcategorias'].split(',')]
            query = query.filter(ProductoERP.subcategoria_id.in_(subcats_list))
    
    productos = query.all()

    procesados = 0

    for producto_erp, producto_pricing in productos:
        # Si el producto tiene preservar_porcentaje_web=True, saltar
        if producto_pricing and producto_pricing.preservar_porcentaje_web:
            continue

        # Determinar markup a usar
        if producto_pricing and producto_pricing.precio_lista_ml:
            # Tiene precio: sumar porcentaje
            markup_base = (producto_pricing.markup_calculado or 0) / 100
            porcentaje_adicional = request.porcentaje_con_precio
        else:
            # No tiene precio: usar porcentaje base
            markup_base = 0
            porcentaje_adicional = request.porcentaje_sin_precio

        markup_objetivo = markup_base + (porcentaje_adicional / 100)

        # Calcular precio
        tipo_cambio = None
        if producto_erp.moneda_costo == "USD":
            tipo_cambio = obtener_tipo_cambio_actual(db, "USD")

        costo_ars = convertir_a_pesos(producto_erp.costo, producto_erp.moneda_costo, tipo_cambio)

        resultado = calcular_precio_web_transferencia(
            costo_ars=costo_ars,
            iva=producto_erp.iva,
            markup_objetivo=markup_objetivo
        )

        # Crear o actualizar pricing
        if not producto_pricing:
            producto_pricing = ProductoPricing(
                item_id=producto_erp.item_id,
                usuario_id=current_user.id
            )
            db.add(producto_pricing)

        producto_pricing.participa_web_transferencia = True
        producto_pricing.porcentaje_markup_web = porcentaje_adicional
        producto_pricing.precio_web_transferencia = resultado["precio"]
        producto_pricing.markup_web_real = resultado["markup_real"]
        producto_pricing.fecha_modificacion = datetime.now()

        procesados += 1

    db.commit()
    
    # Registrar auditoría masiva
    from app.services.auditoria_service import registrar_auditoria
    from app.models.auditoria import TipoAccion
    
    registrar_auditoria(
        db=db,
        usuario_id=current_user.id,
        tipo_accion=TipoAccion.MODIFICACION_MASIVA,
        es_masivo=True,
        productos_afectados=procesados,
        valores_nuevos={
            "accion": "calcular_web_masivo",
            "porcentaje_con_precio": request.porcentaje_con_precio,
            "porcentaje_sin_precio": request.porcentaje_sin_precio,
            "filtros": request.filtros
        },
        comentario="Cálculo masivo de precios web transferencia"
    )

    return {
        "procesados": procesados,
        "porcentaje_con_precio": request.porcentaje_con_precio,
        "porcentaje_sin_precio": request.porcentaje_sin_precio
    }


@router.post("/productos/limpiar-rebate")
async def limpiar_rebate(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Desactiva rebate en todos los productos"""
    from app.services.auditoria_service import registrar_auditoria
    from app.models.auditoria import TipoAccion
    
    count = db.query(ProductoPricing).filter(
        ProductoPricing.participa_rebate == True
    ).count()
    
    db.query(ProductoPricing).update({
        ProductoPricing.participa_rebate: False
    })
    db.commit()

    # Registrar auditoría
    registrar_auditoria(
        db=db,
        usuario_id=current_user.id,
        tipo_accion=TipoAccion.MODIFICACION_MASIVA,
        es_masivo=True,
        productos_afectados=count,
        valores_nuevos={"accion": "limpiar_rebate"},
        comentario="Limpieza masiva de rebate"
    )

    return {
        "mensaje": "Rebate desactivado en todos los productos",
        "productos_actualizados": count
    }

@router.post("/productos/limpiar-web-transferencia")
async def limpiar_web_transferencia(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Desactiva web transferencia en todos los productos"""
    from app.services.auditoria_service import registrar_auditoria
    from app.models.auditoria import TipoAccion
    
    count = db.query(ProductoPricing).filter(
        ProductoPricing.participa_web_transferencia == True
    ).count()
    
    db.query(ProductoPricing).update({
        ProductoPricing.participa_web_transferencia: False,
        ProductoPricing.precio_web_transferencia: None,
        ProductoPricing.markup_web_real: None
    })
    db.commit()

    # Registrar auditoría
    registrar_auditoria(
        db=db,
        usuario_id=current_user.id,
        tipo_accion=TipoAccion.MODIFICACION_MASIVA,
        es_masivo=True,
        productos_afectados=count,
        valores_nuevos={"accion": "limpiar_web_transferencia"},
        comentario="Limpieza masiva de web transferencia"
    )

    return {
        "mensaje": "Web transferencia desactivada en todos los productos",
        "productos_actualizados": count
    }

@router.patch("/productos/{item_id}/color")
async def actualizar_color_producto(
    item_id: int,
    request: dict,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Actualiza el color de marcado de un producto"""

    color = request.get('color')

    # Validar color
    colores_validos = ['rojo', 'naranja', 'amarillo', 'verde', 'azul', 'purpura', 'gris', None]
    if color not in colores_validos:
        raise HTTPException(status_code=400, detail=f"Color inválido: {color}. Válidos: {colores_validos}")

    # Buscar producto pricing
    producto_pricing = db.query(ProductoPricing).filter(
        ProductoPricing.item_id == item_id
    ).first()

    if not producto_pricing:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    color_anterior = producto_pricing.color_marcado
    producto_pricing.color_marcado = color
    db.commit()

    return {
        "mensaje": "Color actualizado",
        "color_anterior": color_anterior,
        "color_nuevo": color
    }

class ConfigCuotasRequest(BaseModel):
    recalcular_cuotas_auto: Optional[bool] = None
    markup_adicional_cuotas_custom: Optional[float] = None

@router.patch("/productos/{item_id}/config-cuotas")
async def actualizar_config_cuotas_producto(
    item_id: int,
    body: ConfigCuotasRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Actualiza la configuración individual de recálculo de cuotas y markup adicional de un producto"""

    recalcular_cuotas_auto = body.recalcular_cuotas_auto
    markup_adicional_cuotas_custom = body.markup_adicional_cuotas_custom

    # Validar markup si se proporciona
    if markup_adicional_cuotas_custom is not None:
        try:
            markup_valor = float(markup_adicional_cuotas_custom)
            if markup_valor < 0 or markup_valor > 100:
                raise ValueError("Markup debe estar entre 0 y 100")
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    # Buscar producto pricing
    producto_pricing = db.query(ProductoPricing).filter(
        ProductoPricing.item_id == item_id
    ).first()

    if not producto_pricing:
        # Crear registro si no existe
        producto_pricing = ProductoPricing(
            item_id=item_id,
            usuario_id=current_user.id
        )
        db.add(producto_pricing)

    # Actualizar configuración
    producto_pricing.recalcular_cuotas_auto = recalcular_cuotas_auto
    producto_pricing.markup_adicional_cuotas_custom = markup_adicional_cuotas_custom
    producto_pricing.usuario_id = current_user.id
    producto_pricing.fecha_modificacion = datetime.now()

    db.commit()
    db.refresh(producto_pricing)

    return {
        "mensaje": "Configuración actualizada",
        "recalcular_cuotas_auto": producto_pricing.recalcular_cuotas_auto,
        "markup_adicional_cuotas_custom": float(producto_pricing.markup_adicional_cuotas_custom) if producto_pricing.markup_adicional_cuotas_custom else None
    }

class ColorLoteRequest(BaseModel):
    item_ids: List[int]
    color: Optional[str] = None

@router.post("/productos/actualizar-color-lote")
async def actualizar_color_productos_lote(
    request: ColorLoteRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Actualiza el color de marcado de múltiples productos"""

    if not request.item_ids:
        raise HTTPException(status_code=400, detail="Debe proporcionar al menos un item_id")

    colores_validos = ['rojo', 'naranja', 'amarillo', 'verde', 'azul', 'purpura', 'gris', None]
    if request.color not in colores_validos:
        raise HTTPException(status_code=400, detail=f"Color inválido")

    count = db.query(ProductoPricing).filter(
        ProductoPricing.item_id.in_(request.item_ids)
    ).update({'color_marcado': request.color}, synchronize_session=False)

    db.commit()

    return {"mensaje": f"{count} productos actualizados", "count": count}

@router.get("/productos/{item_id}/detalle")
async def obtener_detalle_producto(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Obtiene información detallada de un producto (sin datos de ML - se cargan lazy)"""
    from app.services.pricing_calculator import obtener_tipo_cambio_actual, obtener_comision_base, obtener_grupo_subcategoria

    # Producto base
    producto = db.query(ProductoERP).filter(ProductoERP.item_id == item_id).first()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    # Pricing
    pricing = db.query(ProductoPricing).filter(ProductoPricing.item_id == item_id).first()

    # Obtener tipo de cambio
    tipo_cambio = obtener_tipo_cambio_actual(db, "USD")

    # Costo en ARS
    costo_ars = float(producto.costo) * tipo_cambio if producto.moneda_costo == "USD" and tipo_cambio else float(producto.costo)

    # Obtener comisión ML para lista clásica
    grupo_id = obtener_grupo_subcategoria(db, producto.subcategoria_id)
    comision_clasica = obtener_comision_base(db, 1, grupo_id) if grupo_id else None

    # Costo de envío
    costo_envio = float(producto.envio) if producto.envio else 0.0

    # ML data ahora se carga de forma lazy en el endpoint separado /mercadolibre
    # Esto mejora significativamente el tiempo de carga del modal

    # Obtener último proveedor (última compra con puco_id = 10)
    from app.models.tb_supplier import TBSupplier
    from app.models.commercial_transaction import CommercialTransaction
    from app.models.item_transaction import ItemTransaction

    ultimo_proveedor_query = db.query(
        TBSupplier.supp_name,
        ItemTransaction.it_cd
    ).join(
        CommercialTransaction,
        and_(
            CommercialTransaction.comp_id == ItemTransaction.comp_id,
            CommercialTransaction.ct_transaction == ItemTransaction.ct_transaction
        )
    ).join(
        TBSupplier,
        and_(
            TBSupplier.comp_id == CommercialTransaction.comp_id,
            TBSupplier.supp_id == CommercialTransaction.supp_id
        )
    ).filter(
        and_(
            ItemTransaction.puco_id == 10,  # Compras
            ItemTransaction.item_id == item_id,
            CommercialTransaction.supp_id.isnot(None)
        )
    ).order_by(ItemTransaction.it_cd.desc()).first()

    proveedor_info = {
        "nombre": ultimo_proveedor_query.supp_name if ultimo_proveedor_query else None,
        "ultima_compra": ultimo_proveedor_query.it_cd.isoformat() if ultimo_proveedor_query and ultimo_proveedor_query.it_cd else None
    }

    return {
        "producto": {
            "item_id": producto.item_id,
            "codigo": producto.codigo,
            "descripcion": producto.descripcion,
            "marca": producto.marca,
            "categoria": producto.categoria,
            "subcategoria_id": producto.subcategoria_id,
            "stock": producto.stock,
            "moneda_costo": producto.moneda_costo,
            "costo": float(producto.costo),
            "costo_ars": costo_ars,
            "iva": float(producto.iva),
            "costo_envio": costo_envio,
            "tipo_cambio_usado": tipo_cambio
        },
        "pricing": {
            "precio_lista_ml": float(pricing.precio_lista_ml) if pricing and pricing.precio_lista_ml else None,
            "markup": float(pricing.markup_calculado) if pricing and pricing.markup_calculado else None,
            "comision_ml_porcentaje": comision_clasica,
            "participa_rebate": pricing.participa_rebate if pricing else False,
            "porcentaje_rebate": float(pricing.porcentaje_rebate) if pricing and pricing.porcentaje_rebate else None,
            "out_of_cards": pricing.out_of_cards if pricing else False,
            "participa_web_transferencia": pricing.participa_web_transferencia if pricing else False,
            "porcentaje_markup_web": float(pricing.porcentaje_markup_web) if pricing and pricing.porcentaje_markup_web else None,
            "precio_web_transferencia": float(pricing.precio_web_transferencia) if pricing and pricing.precio_web_transferencia else None,
            "markup_web_real": float(pricing.markup_web_real) if pricing and pricing.markup_web_real else None,
            "precio_3_cuotas": float(pricing.precio_3_cuotas) if pricing and pricing.precio_3_cuotas else None,
            "precio_6_cuotas": float(pricing.precio_6_cuotas) if pricing and pricing.precio_6_cuotas else None,
            "precio_9_cuotas": float(pricing.precio_9_cuotas) if pricing and pricing.precio_9_cuotas else None,
            "precio_12_cuotas": float(pricing.precio_12_cuotas) if pricing and pricing.precio_12_cuotas else None,
            "usuario_modifico": pricing.usuario.nombre if pricing and pricing.usuario else None,
            "fecha_modificacion": pricing.fecha_modificacion if pricing else None
        },
        "proveedor": proveedor_info
        # ventas, precios_ml, y publicaciones_ml ahora se cargan de forma lazy en /mercadolibre endpoint
    }

@router.get("/productos/{item_id}/mercadolibre")
async def obtener_datos_ml_producto(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Obtiene solo los datos de MercadoLibre de un producto (lazy loading)"""
    from app.models.publicacion_ml import PublicacionML
    from app.models.venta_ml import VentaML
    from app.services.ml_webhook_client import ml_webhook_client
    from sqlalchemy import text
    from datetime import timedelta, datetime

    # Obtener todas las publicaciones ML del item
    publicaciones_ml_query = db.query(PublicacionML).filter(
        PublicacionML.item_id == item_id,
        PublicacionML.activo == True
    ).all()

    # Crear diccionario base de publicaciones
    publicaciones_dict = {}
    mla_ids = []

    for pub in publicaciones_ml_query:
        publicaciones_dict[pub.mla] = {
            "mla": pub.mla,
            "titulo": pub.item_title,
            "lista_nombre": pub.lista_nombre,
            "pricelist_id": pub.pricelist_id,
            "precio_ml": None,
            "precios": []
        }
        mla_ids.append(pub.mla)

    # Obtener precios de ML para estas publicaciones
    if mla_ids:
        precios_ml_data = db.execute(
            text("""
                SELECT pricelist_id, precio, mla
                FROM precios_ml
                WHERE item_id = :item_id AND mla = ANY(:mla_ids)
            """),
            {"item_id": item_id, "mla_ids": mla_ids}
        ).fetchall()

        for pricelist_id, precio, mla in precios_ml_data:
            if mla and mla in publicaciones_dict:
                publicaciones_dict[mla]["precios"].append({
                    "pricelist_id": pricelist_id,
                    "precio": float(precio) if precio else None
                })

    # Obtener datos de ML via webhook service
    if mla_ids:
        try:
            ml_items = await ml_webhook_client.get_items_batch(mla_ids)
            for mla_id, ml_data in ml_items.items():
                if mla_id in publicaciones_dict:
                    publicaciones_dict[mla_id]["precio_ml"] = float(ml_data.get("price", 0)) if ml_data.get("price") else None
                    publicaciones_dict[mla_id]["catalog_product_id"] = ml_data.get("catalog_product_id")
        except Exception as e:
            logger.error(f"Error consultando ml-webhook: {e}")
            pass

    # Obtener status de catálogo desde la BD
    if mla_ids:
        catalog_statuses = db.execute(text("""
            SELECT mla, catalog_product_id, status, price_to_win, winner_mla, winner_price
            FROM v_ml_catalog_status_latest
            WHERE mla = ANY(:mla_ids)
        """), {"mla_ids": mla_ids}).fetchall()

        for row in catalog_statuses:
            mla, catalog_id, status, ptw, winner, winner_price = row
            if mla in publicaciones_dict:
                publicaciones_dict[mla]["catalog_status"] = status
                publicaciones_dict[mla]["catalog_price_to_win"] = float(ptw) if ptw else None
                publicaciones_dict[mla]["catalog_winner_mla"] = winner
                publicaciones_dict[mla]["catalog_winner_price"] = float(winner_price) if winner_price else None

    # Obtener estado de las publicaciones
    if mla_ids:
        pub_statuses = db.execute(text("""
            SELECT mlp_publicationid, mlp_laststatusid, mlp_active
            FROM tb_mercadolibre_items_publicados
            WHERE mlp_publicationid = ANY(:mla_ids)
        """), {"mla_ids": mla_ids}).fetchall()

        for mla, status_id, is_active in pub_statuses:
            if mla in publicaciones_dict:
                if status_id:
                    status_map = {
                        153: 'active',
                        154: 'paused',
                        155: 'closed',
                        156: 'under_review'
                    }
                    publicaciones_dict[mla]["publication_status"] = status_map.get(status_id, f'status_{status_id}')
                elif is_active is not None:
                    publicaciones_dict[mla]["publication_status"] = 'active' if is_active else 'paused'
                else:
                    publicaciones_dict[mla]["publication_status"] = None

    # Calcular ventas de los últimos 7, 15 y 30 días
    fecha_actual = datetime.now()
    ventas_stats = {}

    for dias in [7, 15, 30]:
        fecha_desde = fecha_actual - timedelta(days=dias)

        ventas_query = db.query(
            func.sum(VentaML.cantidad).label('cantidad_total'),
            func.sum(VentaML.monto_total).label('monto_total'),
            func.count(VentaML.id_venta).label('numero_ventas')
        ).filter(
            and_(
                VentaML.item_id == item_id,
                VentaML.fecha >= fecha_desde,
                VentaML.fecha < fecha_actual + timedelta(days=1)
            )
        ).first()

        ventas_stats[f"ultimos_{dias}_dias"] = {
            "cantidad_vendida": int(ventas_query.cantidad_total or 0),
            "monto_total": float(ventas_query.monto_total or 0),
            "numero_ventas": int(ventas_query.numero_ventas or 0)
        }

    return {
        "publicaciones_ml": sorted(
            publicaciones_dict.values(),
            key=lambda x: (
                {4: 0, 17: 1, 14: 2, 13: 3, 23: 4}.get(x.get('pricelist_id'), 999),
                x.get('mla', '')
            )
        ),
        "ventas": ventas_stats
    }

@router.get("/subcategorias")
async def listar_subcategorias(
    search: Optional[str] = None,
    con_stock: Optional[bool] = None,
    con_precio: Optional[bool] = None,
    marcas: Optional[str] = None,
    con_rebate: Optional[bool] = None,
    con_oferta: Optional[bool] = None,
    con_web_transf: Optional[bool] = None,
    markup_clasica_positivo: Optional[bool] = None,
    markup_rebate_positivo: Optional[bool] = None,
    markup_oferta_positivo: Optional[bool] = None,
    markup_web_transf_positivo: Optional[bool] = None,
    out_of_cards: Optional[bool] = None,
    colores: Optional[str] = None,
    audit_usuarios: Optional[str] = None,
    audit_tipos_accion: Optional[str] = None,
    audit_fecha_desde: Optional[str] = None,
    audit_fecha_hasta: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Lista subcategorías disponibles según filtros activos"""
    from app.models.comision_config import SubcategoriaGrupo
    from collections import defaultdict

    # Query para obtener subcategorias_id disponibles según filtros
    query = db.query(ProductoERP.subcategoria_id).distinct().join(
        ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id, isouter=True
    )

    # Aplicar filtros
    if search:
        query = query.filter(
            or_(
                ProductoERP.codigo.ilike(f'%{search}%'),
                ProductoERP.descripcion.ilike(f'%{search}%'),
                ProductoERP.marca.ilike(f'%{search}%')
            )
        )

    if con_stock is not None:
        if con_stock:
            query = query.filter(ProductoERP.stock > 0)
        else:
            query = query.filter(ProductoERP.stock == 0)

    if con_precio is not None:
        if con_precio:
            query = query.filter(ProductoPricing.precio_lista_ml.isnot(None))
        else:
            query = query.filter(ProductoPricing.precio_lista_ml.is_(None))

    if marcas:
        marcas_list = [m.strip() for m in marcas.split(',') if m.strip()]
        query = query.filter(ProductoERP.marca.in_(marcas_list))

    if con_rebate is not None:
        query = query.filter(ProductoPricing.participa_rebate == con_rebate)

    if con_web_transf is not None:
        query = query.filter(ProductoPricing.participa_web_transferencia == con_web_transf)

    if out_of_cards is not None:
        if out_of_cards:
            query = query.filter(ProductoPricing.out_of_cards == True)
        else:
            query = query.filter(
                or_(
                    ProductoPricing.out_of_cards == False,
                    ProductoPricing.out_of_cards.is_(None)
                )
            )

    if colores:
        colores_list = colores.split(',')
        query = query.filter(ProductoPricing.color_marcado.in_(colores_list))

    if markup_clasica_positivo is not None:
        if markup_clasica_positivo:
            query = query.filter(ProductoPricing.markup_calculado > 0)
        else:
            query = query.filter(ProductoPricing.markup_calculado < 0)

    # Obtener IDs de subcategorías disponibles
    subcat_ids_disponibles = [s[0] for s in query.all() if s[0]]

    # Obtener todas las subcategorías del mapping
    subcats = db.query(SubcategoriaGrupo).filter(
        SubcategoriaGrupo.subcat_id.in_(subcat_ids_disponibles)
    ).order_by(
        SubcategoriaGrupo.nombre_categoria,
        SubcategoriaGrupo.nombre_subcategoria
    ).all()

    # Agrupar por categoría
    agrupadas = defaultdict(list)
    for s in subcats:
        if s.nombre_subcategoria and s.nombre_categoria:
            agrupadas[s.nombre_categoria].append({
                "id": s.subcat_id,
                "nombre": s.nombre_subcategoria,
                "grupo_id": s.grupo_id
            })

    return {
        "categorias": [
            {
                "nombre": cat,
                "subcategorias": subs
            }
            for cat, subs in sorted(agrupadas.items())
        ]
    }
    
@router.post("/sincronizar-subcategorias")
async def sincronizar_subcategorias_endpoint():
    """Sincroniza subcategorías desde el worker"""
    from app.scripts.sync_subcategorias import sincronizar_subcategorias
    sincronizar_subcategorias()
    return {"mensaje": "Subcategorías sincronizadas"}


@router.get("/exportar-web-transferencia")
async def exportar_web_transferencia(
    porcentaje_adicional: float = Query(0, description="Porcentaje adicional a sumar"),
    currency_id: int = Query(1, description="ID de moneda: 1=ARS, 2=USD"),
    offset_dolar: float = Query(0, description="Offset en pesos para ajustar el dólar"),
    search: Optional[str] = None,
    con_stock: Optional[bool] = None,
    con_precio: Optional[bool] = None,
    marcas: Optional[str] = None,
    subcategorias: Optional[str] = None,
    con_rebate: Optional[bool] = None,
    con_oferta: Optional[bool] = None,
    con_web_transf: Optional[bool] = None,
    markup_clasica_positivo: Optional[bool] = None,
    markup_rebate_positivo: Optional[bool] = None,
    markup_oferta_positivo: Optional[bool] = None,
    markup_web_transf_positivo: Optional[bool] = None,
    out_of_cards: Optional[bool] = None,
    colores: Optional[str] = None,
    pms: Optional[str] = None,
    audit_usuarios: Optional[str] = None,
    audit_tipos_accion: Optional[str] = None,
    audit_fecha_desde: Optional[str] = None,
    audit_fecha_hasta: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Exporta precios de Web Transferencia en formato Excel con filtros opcionales"""
    from io import BytesIO
    from openpyxl import Workbook
    from app.models.tipo_cambio import TipoCambio
    
    # Obtener productos con precio web transferencia
    query = db.query(
        ProductoERP.item_id,
        ProductoERP.codigo,
        ProductoPricing.precio_web_transferencia
    ).join(
        ProductoPricing,
        ProductoERP.item_id == ProductoPricing.item_id
    ).filter(
        ProductoPricing.participa_web_transferencia == True,
        ProductoPricing.precio_web_transferencia.isnot(None)
    )

    # FILTRADO POR AUDITORÍA (igual que en el listado principal)
    if audit_usuarios or audit_tipos_accion or audit_fecha_desde or audit_fecha_hasta:
        from app.models.auditoria import Auditoria
        from datetime import datetime

        # Construir filtros de auditoría base
        filtros_audit = [Auditoria.item_id.isnot(None)]

        if audit_usuarios:
            usuarios_ids = [int(u) for u in audit_usuarios.split(',')]
            filtros_audit.append(Auditoria.usuario_id.in_(usuarios_ids))

        if audit_tipos_accion:
            tipos_list = audit_tipos_accion.split(',')
            filtros_audit.append(Auditoria.tipo_accion.in_(tipos_list))

        if audit_fecha_desde:
            try:
                fecha_desde_dt = datetime.strptime(audit_fecha_desde, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    fecha_desde_dt = datetime.strptime(audit_fecha_desde, '%Y-%m-%d %H:%M')
                except ValueError:
                    try:
                        fecha_desde_dt = datetime.strptime(audit_fecha_desde, '%Y-%m-%d')
                    except ValueError:
                        from datetime import date
                        fecha_desde_dt = datetime.combine(date.today(), datetime.min.time())

            # Convertir de hora local (ART = UTC-3) a UTC sumando 3 horas
            from datetime import timedelta
            fecha_desde_dt = fecha_desde_dt + timedelta(hours=3)

            filtros_audit.append(Auditoria.fecha >= fecha_desde_dt)

        if audit_fecha_hasta:
            try:
                fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, '%Y-%m-%d %H:%M')
                except ValueError:
                    try:
                        fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, '%Y-%m-%d')
                        fecha_hasta_dt = fecha_hasta_dt.replace(hour=23, minute=59, second=59)
                    except ValueError:
                        from datetime import date
                        fecha_hasta_dt = datetime.combine(date.today(), datetime.max.time())

            # Convertir de hora local (ART = UTC-3) a UTC sumando 3 horas
            from datetime import timedelta
            fecha_hasta_dt = fecha_hasta_dt + timedelta(hours=3)

            filtros_audit.append(Auditoria.fecha <= fecha_hasta_dt)

        # Obtener productos que tienen auditorías cumpliendo los criterios
        audit_query = db.query(Auditoria.item_id).filter(and_(*filtros_audit))
        item_ids_audit = [item_id for (item_id,) in audit_query.distinct().all()]

        if item_ids_audit:
            query = query.filter(ProductoERP.item_id.in_(item_ids_audit))
        else:
            # Si no hay productos con las auditorías filtradas, retornar vacío
            wb = Workbook()
            ws = wb.active
            ws.append(['No se encontraron productos con los filtros aplicados'])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=web_transferencia_vacia.xlsx"}
            )

    # Aplicar filtros básicos
    if search:
        search_normalized = search.replace('-', '').replace(' ', '').upper()
        query = query.filter(
            or_(
                func.replace(func.replace(func.upper(ProductoERP.descripcion), '-', ''), ' ', '').like(f"%{search_normalized}%"),
                func.replace(func.replace(func.upper(ProductoERP.marca), '-', ''), ' ', '').like(f"%{search_normalized}%"),
                func.replace(func.upper(ProductoERP.codigo), '-', '').like(f"%{search_normalized}%")
            )
        )

    if con_stock:
        query = query.filter(ProductoERP.stock > 0)

    if con_precio:
        query = query.filter(ProductoPricing.precio_lista_ml.isnot(None))

    if marcas:
        marcas_list = [m.strip().upper() for m in marcas.split(',')]
        query = query.filter(func.upper(ProductoERP.marca).in_(marcas_list))

    if subcategorias:
        subcats_list = [int(s) for s in subcategorias.split(',')]
        query = query.filter(ProductoERP.subcategoria_id.in_(subcats_list))

    # Filtro por PMs (Product Managers)
    if pms:
        from app.models.marca_pm import MarcaPM
        pm_ids = [int(pm.strip()) for pm in pms.split(',')]

        # Obtener marcas asignadas a esos PMs
        marcas_pm = db.query(MarcaPM.marca).filter(MarcaPM.usuario_id.in_(pm_ids)).all()
        marcas_asignadas = [m[0] for m in marcas_pm]

        if marcas_asignadas:
            query = query.filter(func.upper(ProductoERP.marca).in_([m.upper() for m in marcas_asignadas]))
        else:
            # Si el PM no tiene marcas asignadas, no hay productos
            wb = Workbook()
            ws = wb.active
            ws.append(['No hay productos para los PMs seleccionados'])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=web_transferencia_vacia.xlsx"}
            )

    # Filtro por colores
    if colores:
        colores_list = [c.strip() for c in colores.split(',')]

        # Verificar si se está filtrando por "sin color"
        if 'sin_color' in colores_list:
            # Remover 'sin_color' de la lista
            colores_con_valor = [c for c in colores_list if c != 'sin_color']

            if colores_con_valor:
                # Si hay otros colores además de sin_color, buscar ambos
                query = query.filter(
                    or_(
                        ProductoPricing.color_marcado.in_(colores_con_valor),
                        ProductoPricing.color_marcado.is_(None)
                    )
                )
            else:
                # Solo sin_color: productos sin color asignado
                query = query.filter(ProductoPricing.color_marcado.is_(None))
        else:
            # Filtro normal por colores específicos
            query = query.filter(ProductoPricing.color_marcado.in_(colores_list))

    # Filtros booleanos avanzados
    if con_rebate is not None:
        if con_rebate:
            query = query.filter(ProductoPricing.participa_rebate == True)
        else:
            query = query.filter(or_(ProductoPricing.participa_rebate == False, ProductoPricing.participa_rebate.is_(None)))

    if con_oferta is not None:
        # Este filtro requiere join con ofertas, se aplicará después
        pass

    if out_of_cards is not None:
        if out_of_cards:
            query = query.filter(ProductoPricing.out_of_cards == True)
        else:
            query = query.filter(or_(ProductoPricing.out_of_cards == False, ProductoPricing.out_of_cards.is_(None)))

    productos = query.all()

    # Aplicar filtros de markup y oferta (requieren cálculos, se hacen después de la query)
    if markup_clasica_positivo is not None or markup_rebate_positivo is not None or markup_oferta_positivo is not None or markup_web_transf_positivo is not None or con_oferta is not None:
        from app.services.pricing_calculator import (
            obtener_tipo_cambio_actual,
            convertir_a_pesos,
            obtener_grupo_subcategoria,
            obtener_comision_base,
            calcular_comision_ml_total,
            calcular_limpio,
            calcular_markup
        )
        from app.models.oferta_ml import OfertaML
        from app.models.publicacion_ml import PublicacionML
        from datetime import date

        productos_filtrados = []
        hoy = date.today()

        for producto in productos:
            item_id = producto[0]
            incluir = True

            # Obtener ProductoERP para el item_id
            producto_erp = db.query(ProductoERP).filter(ProductoERP.item_id == item_id).first()
            if not producto_erp:
                continue

            producto_pricing = db.query(ProductoPricing).filter(ProductoPricing.item_id == item_id).first()
            if not producto_pricing:
                continue

            # Filtro de markup clásica
            if markup_clasica_positivo is not None and incluir:
                markup = producto_pricing.markup_calculado if producto_pricing else None
                if markup is not None:
                    if markup_clasica_positivo and markup < 0:
                        incluir = False
                    elif not markup_clasica_positivo and markup >= 0:
                        incluir = False
                else:
                    incluir = False

            # Filtro de markup web transferencia
            if markup_web_transf_positivo is not None and incluir:
                if producto_pricing and producto_pricing.markup_web_real is not None:
                    markup_web = float(producto_pricing.markup_web_real)
                    if markup_web_transf_positivo and markup_web < 0:
                        incluir = False
                    elif not markup_web_transf_positivo and markup_web >= 0:
                        incluir = False
                else:
                    incluir = False

            # Filtro de oferta
            if con_oferta is not None and incluir:
                pubs = db.query(PublicacionML).filter(PublicacionML.item_id == item_id).all()
                tiene_oferta = False
                for pub in pubs:
                    oferta = db.query(OfertaML).filter(
                        OfertaML.mla == pub.mla,
                        OfertaML.fecha_desde <= hoy,
                        OfertaML.fecha_hasta >= hoy,
                        OfertaML.pvp_seller.isnot(None)
                    ).first()
                    if oferta:
                        tiene_oferta = True
                        break
                if con_oferta and not tiene_oferta:
                    incluir = False
                elif not con_oferta and tiene_oferta:
                    incluir = False

            if incluir:
                productos_filtrados.append(producto)

        productos = productos_filtrados

    # Obtener dólar venta si currency_id es 2
    dolar_ajustado = None
    if currency_id == 2:
        tipo_cambio = db.query(TipoCambio).order_by(TipoCambio.id.desc()).first()
        if tipo_cambio:
            dolar_ajustado = float(tipo_cambio.venta) + offset_dolar

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Web Transferencia"

    # Header
    ws.append(['Código/EAN', 'Precio', 'ID Moneda'])

    # Datos - todo como texto
    for item_id, codigo, precio_base in productos:
        # Aplicar porcentaje adicional
        precio_final = float(precio_base) * (1 + porcentaje_adicional / 100)

        # Si es USD, dividir por dólar ajustado
        if currency_id == 2 and dolar_ajustado:
            precio_final = precio_final / dolar_ajustado
            # Para USD, redondear a 2 decimales
            precio_str = f"{precio_final:.2f}"
        else:
            # Para ARS, redondear a múltiplo de 10
            precio_final = round(precio_final / 10) * 10
            precio_str = str(int(precio_final))

        ws.append([
            str(codigo),
            precio_str,
            str(currency_id)
        ])
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=web_transferencia.xlsx"
        }
    )

@router.get("/exportar-clasica")
async def exportar_clasica(
    porcentaje_adicional: float = Query(0, description="Porcentaje adicional sobre rebate"),
    tipo_cuotas: str = Query("clasica", description="Tipo de cuotas: clasica, 3, 6, 9, 12"),
    currency_id: int = Query(1, description="ID de moneda: 1=ARS, 2=USD"),
    offset_dolar: float = Query(0, description="Offset en pesos para ajustar el dólar"),
    search: Optional[str] = None,
    con_stock: Optional[bool] = None,
    con_precio: Optional[bool] = None,
    marcas: Optional[str] = None,
    subcategorias: Optional[str] = None,
    con_rebate: Optional[bool] = None,
    con_oferta: Optional[bool] = None,
    con_web_transf: Optional[bool] = None,
    markup_clasica_positivo: Optional[bool] = None,
    markup_rebate_positivo: Optional[bool] = None,
    markup_oferta_positivo: Optional[bool] = None,
    markup_web_transf_positivo: Optional[bool] = None,
    out_of_cards: Optional[bool] = None,
    colores: Optional[str] = None,
    pms: Optional[str] = None,
    audit_usuarios: Optional[str] = None,
    audit_tipos_accion: Optional[str] = None,
    audit_fecha_desde: Optional[str] = None,
    audit_fecha_hasta: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Exporta precios de Clásica. Si tiene rebate activo, aplica % sobre precio rebate."""
    from io import BytesIO
    from openpyxl import Workbook
    from app.models.tipo_cambio import TipoCambio
    
    # Obtener productos con precio clásica y precios con cuotas
    query = db.query(
        ProductoERP.item_id,
        ProductoERP.codigo,
        ProductoPricing.precio_lista_ml,
        ProductoPricing.participa_rebate,
        ProductoPricing.porcentaje_rebate,
        ProductoPricing.precio_3_cuotas,
        ProductoPricing.precio_6_cuotas,
        ProductoPricing.precio_9_cuotas,
        ProductoPricing.precio_12_cuotas
    ).join(
        ProductoPricing,
        ProductoERP.item_id == ProductoPricing.item_id
    ).filter(
        ProductoPricing.precio_lista_ml.isnot(None)
    )

    # FILTRADO POR AUDITORÍA (igual que en el listado principal)
    if audit_usuarios or audit_tipos_accion or audit_fecha_desde or audit_fecha_hasta:
        from app.models.auditoria import Auditoria
        from datetime import datetime

        # Construir filtros de auditoría base
        filtros_audit = [Auditoria.item_id.isnot(None)]

        if audit_usuarios:
            usuarios_ids = [int(u) for u in audit_usuarios.split(',')]
            filtros_audit.append(Auditoria.usuario_id.in_(usuarios_ids))

        if audit_tipos_accion:
            tipos_list = audit_tipos_accion.split(',')
            filtros_audit.append(Auditoria.tipo_accion.in_(tipos_list))

        if audit_fecha_desde:
            try:
                fecha_desde_dt = datetime.strptime(audit_fecha_desde, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    fecha_desde_dt = datetime.strptime(audit_fecha_desde, '%Y-%m-%d %H:%M')
                except ValueError:
                    try:
                        fecha_desde_dt = datetime.strptime(audit_fecha_desde, '%Y-%m-%d')
                    except ValueError:
                        from datetime import date
                        fecha_desde_dt = datetime.combine(date.today(), datetime.min.time())

            # Convertir de hora local (ART = UTC-3) a UTC sumando 3 horas
            from datetime import timedelta
            fecha_desde_dt = fecha_desde_dt + timedelta(hours=3)

            filtros_audit.append(Auditoria.fecha >= fecha_desde_dt)

        if audit_fecha_hasta:
            try:
                fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, '%Y-%m-%d %H:%M')
                except ValueError:
                    try:
                        fecha_hasta_dt = datetime.strptime(audit_fecha_hasta, '%Y-%m-%d')
                        fecha_hasta_dt = fecha_hasta_dt.replace(hour=23, minute=59, second=59)
                    except ValueError:
                        from datetime import date
                        fecha_hasta_dt = datetime.combine(date.today(), datetime.max.time())

            # Convertir de hora local (ART = UTC-3) a UTC sumando 3 horas
            from datetime import timedelta
            fecha_hasta_dt = fecha_hasta_dt + timedelta(hours=3)

            filtros_audit.append(Auditoria.fecha <= fecha_hasta_dt)

        # Obtener productos que tienen auditorías cumpliendo los criterios
        audit_query = db.query(Auditoria.item_id).filter(and_(*filtros_audit))
        item_ids_audit = [item_id for (item_id,) in audit_query.distinct().all()]

        if item_ids_audit:
            query = query.filter(ProductoERP.item_id.in_(item_ids_audit))
        else:
            # Si no hay productos con las auditorías filtradas, retornar vacío
            from io import BytesIO
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(['No se encontraron productos con los filtros aplicados'])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=exportacion_clasica_vacia.xlsx"}
            )

    # Aplicar filtros básicos
    if search:
        search_normalized = search.replace('-', '').replace(' ', '').upper()
        query = query.filter(
            or_(
                func.replace(func.replace(func.upper(ProductoERP.descripcion), '-', ''), ' ', '').like(f"%{search_normalized}%"),
                func.replace(func.replace(func.upper(ProductoERP.marca), '-', ''), ' ', '').like(f"%{search_normalized}%"),
                func.replace(func.upper(ProductoERP.codigo), '-', '').like(f"%{search_normalized}%")
            )
        )

    if con_stock:
        query = query.filter(ProductoERP.stock > 0)

    if con_precio:
        query = query.filter(ProductoPricing.precio_lista_ml.isnot(None))

    if marcas:
        marcas_list = [m.strip().upper() for m in marcas.split(',')]
        query = query.filter(func.upper(ProductoERP.marca).in_(marcas_list))

    if subcategorias:
        subcats_list = [int(s) for s in subcategorias.split(',')]
        query = query.filter(ProductoERP.subcategoria_id.in_(subcats_list))

    # Filtro por PMs (Product Managers)
    if pms:
        from app.models.marca_pm import MarcaPM
        pm_ids = [int(pm.strip()) for pm in pms.split(',')]

        # Obtener marcas asignadas a esos PMs
        marcas_pm = db.query(MarcaPM.marca).filter(MarcaPM.usuario_id.in_(pm_ids)).all()
        marcas_asignadas = [m[0] for m in marcas_pm]

        if marcas_asignadas:
            query = query.filter(func.upper(ProductoERP.marca).in_([m.upper() for m in marcas_asignadas]))
        else:
            # Si el PM no tiene marcas asignadas, no hay productos
            from io import BytesIO
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(['No hay productos para los PMs seleccionados'])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=exportacion_clasica_vacia.xlsx"}
            )

    # Filtro por colores
    if colores:
        colores_list = [c.strip() for c in colores.split(',')]

        # Verificar si se está filtrando por "sin color"
        if 'sin_color' in colores_list:
            # Remover 'sin_color' de la lista
            colores_con_valor = [c for c in colores_list if c != 'sin_color']

            if colores_con_valor:
                # Si hay otros colores además de sin_color, buscar ambos
                query = query.filter(
                    or_(
                        ProductoPricing.color_marcado.in_(colores_con_valor),
                        ProductoPricing.color_marcado.is_(None)
                    )
                )
            else:
                # Solo sin_color: productos sin color asignado
                query = query.filter(ProductoPricing.color_marcado.is_(None))
        else:
            # Filtro normal por colores específicos
            query = query.filter(ProductoPricing.color_marcado.in_(colores_list))

    # Filtros booleanos avanzados
    if con_rebate is not None:
        if con_rebate:
            query = query.filter(ProductoPricing.participa_rebate == True)
        else:
            query = query.filter(or_(ProductoPricing.participa_rebate == False, ProductoPricing.participa_rebate.is_(None)))

    if con_web_transf is not None:
        if con_web_transf:
            query = query.filter(ProductoPricing.participa_web_transferencia == True)
        else:
            query = query.filter(or_(ProductoPricing.participa_web_transferencia == False, ProductoPricing.participa_web_transferencia.is_(None)))

    if out_of_cards is not None:
        if out_of_cards:
            query = query.filter(ProductoPricing.out_of_cards == True)
        else:
            query = query.filter(or_(ProductoPricing.out_of_cards == False, ProductoPricing.out_of_cards.is_(None)))

    productos = query.all()

    # Aplicar filtros de markup y oferta (requieren cálculos, se hacen después de la query)
    if markup_clasica_positivo is not None or markup_rebate_positivo is not None or markup_oferta_positivo is not None or markup_web_transf_positivo is not None or con_oferta is not None:
        from app.services.pricing_calculator import (
            obtener_tipo_cambio_actual,
            convertir_a_pesos,
            obtener_grupo_subcategoria,
            obtener_comision_base,
            calcular_comision_ml_total,
            calcular_limpio,
            calcular_markup
        )
        from app.models.oferta_ml import OfertaML
        from app.models.publicacion_ml import PublicacionML
        from datetime import date

        productos_filtrados = []
        hoy = date.today()

        for producto in productos:
            item_id = producto[0]
            incluir = True

            # Obtener ProductoERP para el item_id
            producto_erp = db.query(ProductoERP).filter(ProductoERP.item_id == item_id).first()
            if not producto_erp:
                continue

            producto_pricing = db.query(ProductoPricing).filter(ProductoPricing.item_id == item_id).first()
            if not producto_pricing:
                continue

            # Filtro de markup clásica
            if markup_clasica_positivo is not None and incluir:
                markup = producto_pricing.markup_calculado if producto_pricing else None
                if markup is not None:
                    if markup_clasica_positivo and markup < 0:
                        incluir = False
                    elif not markup_clasica_positivo and markup >= 0:
                        incluir = False
                else:
                    incluir = False

            # Filtro de markup rebate
            if markup_rebate_positivo is not None and incluir:
                if producto_pricing and producto_pricing.participa_rebate and producto_pricing.precio_lista_ml:
                    try:
                        precio_rebate = float(producto_pricing.precio_lista_ml) * (1 + float(producto_pricing.porcentaje_rebate or 3.8) / 100)
                        tipo_cambio = obtener_tipo_cambio_actual(db, "USD") if producto_erp.moneda_costo == "USD" else None
                        costo_ars = convertir_a_pesos(producto_erp.costo, producto_erp.moneda_costo, tipo_cambio)
                        grupo_id = obtener_grupo_subcategoria(db, producto_erp.subcategoria_id)
                        comision_base = obtener_comision_base(db, 4, grupo_id)  # 4 = Clásica
                        if comision_base:
                            comisiones = calcular_comision_ml_total(precio_rebate, comision_base, producto_erp.iva, db=db)
                            limpio = calcular_limpio(precio_rebate, producto_erp.iva, producto_erp.envio or 0, comisiones["comision_total"], db=db, grupo_id=grupo_id)
                            markup_rebate = calcular_markup(limpio, costo_ars) * 100
                            if markup_rebate_positivo and markup_rebate < 0:
                                incluir = False
                            elif not markup_rebate_positivo and markup_rebate >= 0:
                                incluir = False
                        else:
                            incluir = False
                    except:
                        incluir = False
                else:
                    incluir = False

            # Filtro de oferta
            if con_oferta is not None and incluir:
                pubs = db.query(PublicacionML).filter(PublicacionML.item_id == item_id).all()
                tiene_oferta = False
                for pub in pubs:
                    oferta = db.query(OfertaML).filter(
                        OfertaML.mla == pub.mla,
                        OfertaML.fecha_desde <= hoy,
                        OfertaML.fecha_hasta >= hoy,
                        OfertaML.pvp_seller.isnot(None)
                    ).first()
                    if oferta:
                        tiene_oferta = True
                        break
                if con_oferta and not tiene_oferta:
                    incluir = False
                elif not con_oferta and tiene_oferta:
                    incluir = False

            # Filtro de markup oferta
            if markup_oferta_positivo is not None and incluir:
                # Implementación similar a markup rebate pero con precio de oferta
                # Por simplicidad, si llegó hasta acá y tiene oferta, se incluye
                pass

            # Filtro de markup web transferencia
            if markup_web_transf_positivo is not None and incluir:
                if producto_pricing and producto_pricing.markup_web_real is not None:
                    markup_web = float(producto_pricing.markup_web_real)
                    if markup_web_transf_positivo and markup_web < 0:
                        incluir = False
                    elif not markup_web_transf_positivo and markup_web >= 0:
                        incluir = False
                else:
                    incluir = False

            if incluir:
                productos_filtrados.append(producto)

        productos = productos_filtrados

    # Obtener dólar venta si currency_id es 2
    dolar_ajustado = None
    if currency_id == 2:
        tipo_cambio = db.query(TipoCambio).order_by(TipoCambio.id.desc()).first()
        if tipo_cambio:
            dolar_ajustado = float(tipo_cambio.venta) + offset_dolar

    # Determinar qué prli_ids corresponden al tipo de cuotas seleccionado
    # Cada tipo tiene una lista Web y una PVP (ambas representan el mismo precio)
    # Mapeo tipo_cuotas -> [prli_id_web, prli_id_pvp]
    tipo_cuotas_to_prli = {
        "clasica": [4, 12],   # Clásica Web + PVP
        "3": [17, 18],        # 3 Cuotas Web + PVP
        "6": [14, 19],        # 6 Cuotas Web + PVP
        "9": [13, 20],        # 9 Cuotas Web + PVP
        "12": [23, 21]        # 12 Cuotas Web + PVP
    }

    prli_ids_seleccionados = tipo_cuotas_to_prli.get(tipo_cuotas, [])

    # Obtener MLA IDs para cada producto de la lista seleccionada
    from app.models.mercadolibre_item_publicado import MercadoLibreItemPublicado

    # Crear diccionario: item_id -> [mla_ids]
    mla_por_item = {}

    if prli_ids_seleccionados:
        # Consultar publicaciones de AMBAS listas (Web y PVP) para el tipo seleccionado
        # optval_statusId: 2 = Publicada, 3 = Pausada, 5 = Finalizada, 6 = Pausada Forzada, 10 = Des-Enlazada
        item_ids = [p[0] for p in productos]

        publicaciones = db.query(
            MercadoLibreItemPublicado.item_id,
            MercadoLibreItemPublicado.mlp_publicationID
        ).filter(
            MercadoLibreItemPublicado.item_id.in_(item_ids),
            MercadoLibreItemPublicado.prli_id.in_(prli_ids_seleccionados),
            # Publicadas (2) o None (por falta de sync del optval_statusId)
            or_(
                MercadoLibreItemPublicado.optval_statusId == 2,
                MercadoLibreItemPublicado.optval_statusId.is_(None)
            )
        ).all()

        # Agrupar MLAs por item_id
        for item_id, mla_id in publicaciones:
            if item_id not in mla_por_item:
                mla_por_item[item_id] = []
            mla_por_item[item_id].append(mla_id)

    # Determinar el número máximo de MLAs que tiene cualquier producto
    max_mlas = max([len(mlas) for mlas in mla_por_item.values()]) if mla_por_item else 0

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = tipo_cuotas.title()

    # Header - Columnas base + una columna por cada MLA
    header = ['Código/EAN', 'Precio', 'ID Moneda']
    for i in range(max_mlas):
        header.append(f'MLA {i+1}')
    ws.append(header)

    # Datos
    for item_id, codigo, precio_clasica, participa_rebate, porcentaje_rebate, precio_3, precio_6, precio_9, precio_12 in productos:
        # Determinar qué precio usar según tipo_cuotas
        if tipo_cuotas == "clasica":
            # Si tiene rebate activo, calcular precio rebate y aplicar % adicional
            if participa_rebate and porcentaje_rebate:
                precio_rebate = precio_clasica * (1 + float(porcentaje_rebate) / 100)
                precio_exportar = precio_rebate * (1 + porcentaje_adicional / 100)
            else:
                # Si no tiene rebate, usar precio clásica sin modificar
                precio_exportar = precio_clasica
        elif tipo_cuotas == "3":
            # Si no hay precio de 3 cuotas, saltar este producto
            if not precio_3:
                continue
            precio_exportar = float(precio_3)
        elif tipo_cuotas == "6":
            # Si no hay precio de 6 cuotas, saltar este producto
            if not precio_6:
                continue
            precio_exportar = float(precio_6)
        elif tipo_cuotas == "9":
            # Si no hay precio de 9 cuotas, saltar este producto
            if not precio_9:
                continue
            precio_exportar = float(precio_9)
        elif tipo_cuotas == "12":
            # Si no hay precio de 12 cuotas, saltar este producto
            if not precio_12:
                continue
            precio_exportar = float(precio_12)
        else:
            precio_exportar = precio_clasica

        # Si es USD, dividir por dólar ajustado
        if currency_id == 2 and dolar_ajustado:
            precio_final = precio_exportar / dolar_ajustado
            # Para USD, redondear a 2 decimales
            precio_str = f"{precio_final:.2f}"
        else:
            # Para ARS, redondear a múltiplo de 10
            precio_final = round(precio_exportar / 10) * 10
            precio_str = str(int(precio_final))

        # Obtener MLAs de la lista seleccionada para este item
        mlas = mla_por_item.get(item_id, [])

        # Crear fila con columnas base
        fila = [
            str(codigo),
            precio_str,
            str(currency_id)
        ]

        # Agregar cada MLA en su propia columna
        for i in range(max_mlas):
            if i < len(mlas):
                fila.append(mlas[i])
            else:
                fila.append('')  # Columna vacía si no hay MLA

        ws.append(fila)
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=clasica.xlsx"
        }
    )

@router.patch("/productos/{item_id}/out-of-cards")
async def actualizar_out_of_cards(
    item_id: int,
    data: dict,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Actualiza el estado de out_of_cards de un producto"""
    from app.services.auditoria_service import registrar_auditoria
    from app.models.auditoria import TipoAccion
    
    pricing = db.query(ProductoPricing).filter(ProductoPricing.item_id == item_id).first()
    
    if not pricing:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    # Guardar valor anterior
    valor_anterior = pricing.out_of_cards
    valor_nuevo = data.get("out_of_cards", False)
    
    pricing.out_of_cards = valor_nuevo
    db.commit()
    
    # Registrar auditoría
    registrar_auditoria(
        db=db,
        usuario_id=current_user.id,
        tipo_accion=TipoAccion.MARCAR_OUT_OF_CARDS if valor_nuevo else TipoAccion.DESMARCAR_OUT_OF_CARDS,
        item_id=item_id,
        valores_anteriores={"out_of_cards": valor_anterior},
        valores_nuevos={"out_of_cards": valor_nuevo}
    )
    
    return {"status": "success", "out_of_cards": pricing.out_of_cards}

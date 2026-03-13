"""
Router del módulo RRHH — Reportes (Phase 8).

Endpoints de solo lectura para reportes agregados:
- Presentismo mensual
- Sanciones por período
- Vacaciones resumen anual
- Cuenta corriente resumen
- Horas trabajadas (desde fichadas)
- Exportar a Excel (openpyxl)
"""

from __future__ import annotations

import io
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.usuario import Usuario
from app.services.permisos_service import PermisosService
from app.services.rrhh_reportes_service import ReportesService

router = APIRouter(prefix="/rrhh/reportes", tags=["rrhh-reportes"])


# ──────────────────────────────────────────────
# SCHEMAS
# ──────────────────────────────────────────────


class PresentismoEmpleadoRow(BaseModel):
    empleado_id: int
    nombre: str
    legajo: str
    area: str = ""
    puesto: str = ""
    presente: int = 0
    ausente: int = 0
    home_office: int = 0
    vacaciones: int = 0
    art: int = 0
    licencia: int = 0
    franco: int = 0
    feriado: int = 0
    total_registrado: int = 0


class PresentismoMensualResponse(BaseModel):
    mes: int
    anio: int
    area: str | None = None
    fecha_desde: str
    fecha_hasta: str
    total_empleados: int
    items: list[PresentismoEmpleadoRow]


class SancionReporteRow(BaseModel):
    id: int
    empleado_id: int
    empleado_nombre: str
    tipo: str
    fecha: str
    motivo: str
    anulada: bool
    fecha_desde: str | None = None
    fecha_hasta: str | None = None


class ResumenPorTipo(BaseModel):
    tipo: str
    cantidad: int


class ResumenPorEmpleado(BaseModel):
    empleado: str
    cantidad: int


class SancionesReporteResponse(BaseModel):
    fecha_desde: str
    fecha_hasta: str
    total: int
    total_vigentes: int
    total_anuladas: int
    items: list[SancionReporteRow]
    por_tipo: list[ResumenPorTipo]
    por_empleado: list[ResumenPorEmpleado]


class VacacionesEmpleadoRow(BaseModel):
    empleado_id: int
    nombre: str
    legajo: str
    area: str = ""
    antiguedad_anios: int
    dias_correspondientes: int
    dias_gozados: int
    dias_pendientes: int


class VacacionesResumenResponse(BaseModel):
    anio: int
    total_empleados: int
    total_dias_correspondientes: int
    total_dias_gozados: int
    total_dias_pendientes: int
    items: list[VacacionesEmpleadoRow]


class CuentaCorrienteRow(BaseModel):
    empleado_id: int
    nombre: str
    legajo: str
    area: str = ""
    saldo: float


class CuentaCorrienteResumenResponse(BaseModel):
    total_cuentas: int
    total_saldo: float
    con_deuda: int
    con_credito: int
    sin_saldo: int
    items: list[CuentaCorrienteRow]


class DiaTrabajadoDetalle(BaseModel):
    fecha: str
    fichadas: int
    horas: float
    completo: bool


class HorasEmpleadoRow(BaseModel):
    empleado_id: int
    nombre: str
    legajo: str
    total_horas: float
    dias_trabajados: int
    dias_completos: int
    dias_incompletos: int
    detalle: list[DiaTrabajadoDetalle] = []


class HorasTrabajadasResponse(BaseModel):
    mes: int
    anio: int
    empleado_id: int | None = None
    fecha_desde: str
    fecha_hasta: str
    total_empleados: int
    items: list[HorasEmpleadoRow]


# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────

EXPORTABLES = {
    "presentismo-mensual",
    "sanciones-periodo",
    "vacaciones-resumen",
    "cuenta-corriente-resumen",
    "horas-trabajadas",
}


def _check_permiso(db: Session, user: Usuario, codigo: str = "rrhh.ver") -> None:
    """Verifica permiso RRHH o lanza 403."""
    svc = PermisosService(db)
    if not svc.tiene_permiso(user, codigo):
        raise HTTPException(status_code=403, detail=f"Sin permiso: {codigo}")


# ──────────────────────────────────────────────
# ENDPOINTS
# ──────────────────────────────────────────────


@router.get("/presentismo-mensual", response_model=PresentismoMensualResponse)
def reporte_presentismo_mensual(
    mes: int = Query(ge=1, le=12),
    anio: int = Query(ge=2000, le=2100),
    area: str | None = Query(default=None),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> PresentismoMensualResponse:
    """Reporte de presentismo mensual agrupado por empleado."""
    _check_permiso(db, current_user)
    svc = ReportesService(db)
    data = svc.presentismo_mensual(mes, anio, area)
    return PresentismoMensualResponse(**data)


@router.get("/sanciones-periodo", response_model=SancionesReporteResponse)
def reporte_sanciones_periodo(
    fecha_desde: date = Query(),
    fecha_hasta: date = Query(),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> SancionesReporteResponse:
    """Reporte de sanciones en un rango de fechas."""
    _check_permiso(db, current_user)
    if fecha_hasta < fecha_desde:
        raise HTTPException(status_code=400, detail="fecha_hasta debe ser >= fecha_desde")
    svc = ReportesService(db)
    data = svc.sanciones_periodo(fecha_desde, fecha_hasta)
    return SancionesReporteResponse(**data)


@router.get("/vacaciones-resumen", response_model=VacacionesResumenResponse)
def reporte_vacaciones_resumen(
    anio: int = Query(ge=2000, le=2100),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> VacacionesResumenResponse:
    """Resumen de vacaciones para un año (períodos con días pendientes)."""
    _check_permiso(db, current_user)
    svc = ReportesService(db)
    data = svc.vacaciones_resumen(anio)
    return VacacionesResumenResponse(**data)


@router.get("/cuenta-corriente-resumen", response_model=CuentaCorrienteResumenResponse)
def reporte_cuenta_corriente_resumen(
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> CuentaCorrienteResumenResponse:
    """Resumen de cuentas corrientes de empleados (saldos)."""
    _check_permiso(db, current_user)
    svc = ReportesService(db)
    data = svc.cuenta_corriente_resumen()
    return CuentaCorrienteResumenResponse(**data)


@router.get("/horas-trabajadas", response_model=HorasTrabajadasResponse)
def reporte_horas_trabajadas(
    mes: int = Query(ge=1, le=12),
    anio: int = Query(ge=2000, le=2100),
    empleado_id: int | None = Query(default=None),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> HorasTrabajadasResponse:
    """Horas trabajadas calculadas desde fichadas (entrada/salida)."""
    _check_permiso(db, current_user)
    svc = ReportesService(db)
    data = svc.horas_trabajadas(mes, anio, empleado_id)
    return HorasTrabajadasResponse(**data)


@router.get("/exportar/{tipo}")
def exportar_reporte(
    tipo: str,
    mes: int | None = Query(default=None, ge=1, le=12),
    anio: int | None = Query(default=None, ge=2000, le=2100),
    fecha_desde: date | None = Query(default=None),
    fecha_hasta: date | None = Query(default=None),
    area: str | None = Query(default=None),
    empleado_id: int | None = Query(default=None),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """
    Exportar reporte a Excel (XLSX).

    Tipos válidos: presentismo-mensual, sanciones-periodo,
    vacaciones-resumen, cuenta-corriente-resumen, horas-trabajadas.
    """
    _check_permiso(db, current_user)

    if tipo not in EXPORTABLES:
        raise HTTPException(
            status_code=400,
            detail=f"Tipo de reporte inválido. Opciones: {', '.join(sorted(EXPORTABLES))}",
        )

    svc = ReportesService(db)

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl no instalado. Ejecute: pip install openpyxl",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    filename = f"rrhh_{tipo}"

    if tipo == "presentismo-mensual":
        if mes is None or anio is None:
            raise HTTPException(status_code=400, detail="Parámetros requeridos: mes, anio")
        data = svc.presentismo_mensual(mes, anio, area)
        ws.title = f"Presentismo {mes:02d}/{anio}"
        filename = f"presentismo_{anio}_{mes:02d}"
        headers = [
            "Empleado",
            "Legajo",
            "Área",
            "Puesto",
            "Presente",
            "Ausente",
            "Home Office",
            "Vacaciones",
            "ART",
            "Licencia",
            "Franco",
            "Feriado",
            "Total",
        ]
        ws.append(headers)
        for item in data["items"]:
            ws.append(
                [
                    item["nombre"],
                    item["legajo"],
                    item["area"],
                    item["puesto"],
                    item["presente"],
                    item["ausente"],
                    item["home_office"],
                    item["vacaciones"],
                    item["art"],
                    item["licencia"],
                    item["franco"],
                    item["feriado"],
                    item["total_registrado"],
                ]
            )

    elif tipo == "sanciones-periodo":
        if fecha_desde is None or fecha_hasta is None:
            raise HTTPException(
                status_code=400,
                detail="Parámetros requeridos: fecha_desde, fecha_hasta",
            )
        data = svc.sanciones_periodo(fecha_desde, fecha_hasta)
        ws.title = "Sanciones"
        filename = f"sanciones_{fecha_desde}_{fecha_hasta}"
        headers = [
            "Empleado",
            "Tipo",
            "Fecha",
            "Motivo",
            "Anulada",
            "Desde",
            "Hasta",
        ]
        ws.append(headers)
        for item in data["items"]:
            ws.append(
                [
                    item["empleado_nombre"],
                    item["tipo"],
                    item["fecha"],
                    item["motivo"],
                    "Sí" if item["anulada"] else "No",
                    item.get("fecha_desde", ""),
                    item.get("fecha_hasta", ""),
                ]
            )

    elif tipo == "vacaciones-resumen":
        if anio is None:
            raise HTTPException(status_code=400, detail="Parámetro requerido: anio")
        data = svc.vacaciones_resumen(anio)
        ws.title = f"Vacaciones {anio}"
        filename = f"vacaciones_{anio}"
        headers = [
            "Empleado",
            "Legajo",
            "Área",
            "Antigüedad (años)",
            "Días Correspondientes",
            "Días Gozados",
            "Días Pendientes",
        ]
        ws.append(headers)
        for item in data["items"]:
            ws.append(
                [
                    item["nombre"],
                    item["legajo"],
                    item["area"],
                    item["antiguedad_anios"],
                    item["dias_correspondientes"],
                    item["dias_gozados"],
                    item["dias_pendientes"],
                ]
            )

    elif tipo == "cuenta-corriente-resumen":
        data = svc.cuenta_corriente_resumen()
        ws.title = "Cuenta Corriente"
        filename = "cuenta_corriente_resumen"
        headers = ["Empleado", "Legajo", "Área", "Saldo"]
        ws.append(headers)
        for item in data["items"]:
            ws.append(
                [
                    item["nombre"],
                    item["legajo"],
                    item["area"],
                    item["saldo"],
                ]
            )

    elif tipo == "horas-trabajadas":
        if mes is None or anio is None:
            raise HTTPException(status_code=400, detail="Parámetros requeridos: mes, anio")
        data = svc.horas_trabajadas(mes, anio, empleado_id)
        ws.title = f"Horas {mes:02d}/{anio}"
        filename = f"horas_trabajadas_{anio}_{mes:02d}"
        headers = [
            "Empleado",
            "Legajo",
            "Total Horas",
            "Días Trabajados",
            "Días Completos",
            "Días Incompletos",
        ]
        ws.append(headers)
        for item in data["items"]:
            ws.append(
                [
                    item["nombre"],
                    item["legajo"],
                    item["total_horas"],
                    item["dias_trabajados"],
                    item["dias_completos"],
                    item["dias_incompletos"],
                ]
            )

    # Style header row
    from openpyxl.styles import Font, PatternFill

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}.xlsx"',
        },
    )

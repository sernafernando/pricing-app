"""
Router del módulo RRHH - Horas Extras (HE) — Batch 4.

Expone los endpoints definidos en `design.md §6` (21 endpoints):

- Listado, detalle, alta y edición de bloques.
- Transiciones de workflow (aprobar/rechazar/reabrir) individual y bulk.
- Anomalías (`completar-fichada`, `descartar-dia`).
- Recálculo manual con cap configurable (revisión 2 — Q2).
- Liquidación + export Excel (es-AR, revisión 2 — Q1).
- Alertas (listar, marcar leídas).
- Historial append-only.
- Config singleton (GET/PUT).

Convenciones:
- Pydantic v2 (`model_config = ConfigDict(...)`, `@field_validator`).
- Permisos vía `PermisosService(db).tiene_permiso(...)` reutilizado del patrón
  de `rrhh_horarios.py` y `rrhh_vacaciones.py`.
- One-shot Excel exports usan `BytesIO` + `StreamingResponse` con
  `Depends(get_current_user)` (NO long-lived).
- HTTPException con status codes explícitos (403/404/409/422).

ORDEN DE RUTAS:
- FastAPI matchea rutas en orden de declaración. Las rutas estáticas
  (/alertas, /config, /exportar, /bulk/..., /recalcular, /liquidar,
  /historial/{he_id}) van ANTES que las parameterizadas (/{he_id}, etc.)
  para evitar 422 al intentar parsear "alertas" como int.
"""

from __future__ import annotations

import io
import os
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict, Field, field_validator
from sqlalchemy.orm import Session, joinedload

from app.api.deps import get_current_user
from app.core.database import get_db
from app.models.rrhh_empleado import RRHHEmpleado
from app.models.rrhh_horas_extras import (
    EstadoHE,
    GeneradaPorHE,
    RRHHHorasExtras,
    RRHHHorasExtrasAlerta,
    RRHHHorasExtrasConfig,
    RRHHHorasExtrasHistorial,
    TipoDiaHE,
)
from app.models.usuario import Usuario
from app.services.permisos_service import PermisosService
from app.services.rrhh_hikvision_client import ART_TZ
from app.services.rrhh_horas_extras_service import HorasExtrasService

router = APIRouter(prefix="/rrhh/horas-extras", tags=["rrhh-horas-extras"])


# ════════════════════════════════════════════════════════════════════════════
# SCHEMAS Pydantic v2 — T-4.1
# ════════════════════════════════════════════════════════════════════════════


# ─── Response básicos ─────────────────────────────────────────


class FichadaRefSchema(BaseModel):
    """Referencia mínima a una fichada vinculada al bloque."""

    id: int
    timestamp: datetime
    tipo: str
    origen: str
    model_config = ConfigDict(from_attributes=True)


class HorasExtrasResponse(BaseModel):
    """Bloque de HE expuesto al frontend."""

    id: int
    empleado_id: int
    empleado_nombre: str = ""
    empleado_legajo: str = ""
    fecha: date
    turno_esperado_minutos: int
    trabajado_minutos: Optional[int] = None
    extras_minutos: Optional[int] = None
    tipo_dia: str
    porcentaje_recargo: Decimal
    estado: str
    error_tipo: Optional[str] = None
    aprobado_por_nombre: Optional[str] = None
    aprobado_at: Optional[datetime] = None
    motivo_rechazo: Optional[str] = None
    reabierto_por_nombre: Optional[str] = None
    reabierto_at: Optional[datetime] = None
    motivo_reapertura: Optional[str] = None
    liquidacion_periodo: Optional[str] = None
    liquidado_at: Optional[datetime] = None
    generada_por: str
    observaciones: Optional[str] = None
    fichada_entrada: Optional[FichadaRefSchema] = None
    fichada_salida: Optional[FichadaRefSchema] = None
    alertas_no_leidas: int = 0
    created_at: datetime
    updated_at: datetime
    model_config = ConfigDict(from_attributes=True)


class HorasExtrasListResponse(BaseModel):
    """Lista paginada de bloques HE."""

    items: list[HorasExtrasResponse] = []
    total: int = 0
    page: int = 1
    page_size: int = 50


# ─── Create / Update ──────────────────────────────────────────


class HorasExtrasCreate(BaseModel):
    """Bloque manual (sin disparar cron)."""

    empleado_id: int
    fecha: date
    extras_minutos: int = Field(ge=1, le=1440)
    tipo_dia: str = Field(pattern="^(habil_50|sabado_100|domingo_100|feriado_100|manual)$")
    porcentaje_recargo: Decimal = Field(ge=0, le=500)
    observaciones: Optional[str] = Field(default=None, max_length=2000)


class HorasExtrasUpdate(BaseModel):
    """Editar bloque (solo en estados editables)."""

    porcentaje_recargo: Optional[Decimal] = Field(default=None, ge=0, le=500)
    observaciones: Optional[str] = Field(default=None, max_length=2000)


# ─── Workflow actions ─────────────────────────────────────────


class AprobacionRequest(BaseModel):
    """Body de PATCH /{id}/aprobar."""

    porcentaje_override: Optional[Decimal] = Field(default=None, ge=0, le=500)
    observaciones: Optional[str] = Field(default=None, max_length=2000)


class RechazoRequest(BaseModel):
    """Body de PATCH /{id}/rechazar."""

    motivo: str = Field(min_length=3, max_length=2000)


class ReaperturaRequest(BaseModel):
    """Body de PATCH /{id}/reabrir."""

    motivo: str = Field(min_length=3, max_length=2000)


class CompletarFichadaRequest(BaseModel):
    """Body de POST /{id}/completar-fichada."""

    timestamp: datetime
    tipo: str = Field(pattern="^(entrada|salida)$")
    motivo: str = Field(min_length=3, max_length=500)


class DescartarDiaRequest(BaseModel):
    """Body de POST /{id}/descartar-dia."""

    motivo: str = Field(min_length=3, max_length=2000)


class RecalcularRequest(BaseModel):
    """Body de POST /recalcular. Cap configurable validado en endpoint (T-4.2)."""

    fecha_desde: date
    fecha_hasta: date
    empleado_id: Optional[int] = None  # None = todos los activos

    @field_validator("fecha_hasta")
    @classmethod
    def _hasta_ge_desde(cls, v: date, info) -> date:  # type: ignore[no-untyped-def]
        desde = info.data.get("fecha_desde")
        if desde and v < desde:
            raise ValueError("fecha_hasta debe ser >= fecha_desde")
        return v


class BulkAprobarRequest(BaseModel):
    """Body de POST /bulk/aprobar."""

    ids: list[int] = Field(min_length=1, max_length=500)
    porcentaje_override: Optional[Decimal] = Field(default=None, ge=0, le=500)


class BulkRechazarRequest(BaseModel):
    """Body de POST /bulk/rechazar."""

    ids: list[int] = Field(min_length=1, max_length=500)
    motivo: str = Field(min_length=3, max_length=2000)


class LiquidacionRequest(BaseModel):
    """Body de POST /liquidar."""

    periodo: str = Field(pattern="^[0-9]{6}$")  # YYYYMM
    ids: list[int] = Field(min_length=1, max_length=10000)


class LiquidacionResponse(BaseModel):
    """Respuesta de POST /liquidar."""

    periodo: str
    liquidados: int
    rechazados: int
    detalle_rechazos: list[dict[str, Any]] = []


# ─── Config ───────────────────────────────────────────────────


class HorasExtrasConfigSchema(BaseModel):
    """Singleton de configuración del módulo HE."""

    porcentaje_dia_habil: Decimal = Field(ge=0, le=500)
    porcentaje_sabado_pm: Decimal = Field(ge=0, le=500)
    porcentaje_domingo: Decimal = Field(ge=0, le=500)
    porcentaje_feriado: Decimal = Field(ge=0, le=500)
    hora_corte_sabado: time
    tolerancia_extras_minutos: int = Field(ge=0, le=240)
    requiere_aprobacion: bool
    cron_activo: bool
    dias_retencion_alertas: int = Field(ge=1, le=3650)
    cap_dias_recalculo_manual: int = Field(ge=1, le=366)
    updated_at: Optional[datetime] = None
    actualizado_por_nombre: Optional[str] = None
    model_config = ConfigDict(from_attributes=True)


# ─── Alertas / Historial ──────────────────────────────────────


class AlertaResponse(BaseModel):
    """Alerta de divergencia generada por hooks de fichada/turno."""

    id: int
    he_id: int
    tipo: str
    severidad: str
    mensaje: str
    contexto: Optional[dict[str, Any]] = None
    leida_at: Optional[datetime] = None
    leida_por_nombre: Optional[str] = None
    created_at: datetime
    model_config = ConfigDict(from_attributes=True)


class AlertasListResponse(BaseModel):
    """Lista paginada de alertas."""

    items: list[AlertaResponse] = []
    total: int = 0


class HistorialEntryResponse(BaseModel):
    """Entrada del audit trail append-only de un bloque HE."""

    id: int
    he_id: int
    accion: str
    estado_anterior: Optional[str] = None
    estado_nuevo: str
    usuario_id: Optional[int] = None
    usuario_nombre: Optional[str] = None
    motivo: Optional[str] = None
    snapshot: dict[str, Any]
    created_at: datetime
    model_config = ConfigDict(from_attributes=True)


# ════════════════════════════════════════════════════════════════════════════
# HELPERS internos del router
# ════════════════════════════════════════════════════════════════════════════


# Lockfile del cron — design §12 + §14 decisión locked.
# Configurable vía env vars para que dev/staging/prod puedan apuntar a paths distintos.
_CRON_LOCK_PATHS = (
    Path(os.environ.get("RRHH_HE_CRON_LOCK_PRIMARY", "/var/run/pricing-app/rrhh_he_cron.lock")),
    Path(os.environ.get("RRHH_HE_CRON_LOCK_FALLBACK", "/tmp/rrhh_he_cron.lock")),
)


def _check_permiso(db: Session, user: Usuario, codigo: str) -> None:
    """Verifica permiso o lanza 403 (mismo patrón que `rrhh_horarios.py`)."""
    if not PermisosService(db).tiene_permiso(user, codigo):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"Sin permiso: {codigo}",
        )


def _get_bloque_or_404(db: Session, he_id: int) -> RRHHHorasExtras:
    """Carga un bloque HE con relaciones útiles o lanza 404."""
    bloque = (
        db.query(RRHHHorasExtras)
        .options(
            joinedload(RRHHHorasExtras.empleado),
            joinedload(RRHHHorasExtras.fichada_entrada),
            joinedload(RRHHHorasExtras.fichada_salida),
            joinedload(RRHHHorasExtras.aprobado_por),
            joinedload(RRHHHorasExtras.reabierto_por),
            joinedload(RRHHHorasExtras.liquidado_por),
        )
        .filter(RRHHHorasExtras.id == he_id)
        .first()
    )
    if bloque is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Bloque HE {he_id} no encontrado",
        )
    return bloque


def _count_alertas_no_leidas(db: Session, he_id: int) -> int:
    """Devuelve cantidad de alertas no leídas asociadas a un bloque."""
    return (
        db.query(RRHHHorasExtrasAlerta)
        .filter(
            RRHHHorasExtrasAlerta.he_id == he_id,
            RRHHHorasExtrasAlerta.leida_at.is_(None),
        )
        .count()
    )


def _serialize_bloque(db: Session, b: RRHHHorasExtras) -> HorasExtrasResponse:
    """Convierte el ORM a `HorasExtrasResponse` enriquecido con joins."""
    emp = b.empleado
    nombre = f"{emp.apellido}, {emp.nombre}" if emp is not None else ""
    legajo = emp.legajo if emp is not None else ""

    aprobado_por_nombre: Optional[str] = None
    if getattr(b, "aprobado_por", None) is not None:
        aprobado_por_nombre = b.aprobado_por.nombre

    reabierto_por_nombre: Optional[str] = None
    if getattr(b, "reabierto_por", None) is not None:
        reabierto_por_nombre = b.reabierto_por.nombre

    fichada_entrada: Optional[FichadaRefSchema] = None
    if b.fichada_entrada is not None:
        fichada_entrada = FichadaRefSchema(
            id=b.fichada_entrada.id,
            timestamp=b.fichada_entrada.timestamp,
            tipo=b.fichada_entrada.tipo,
            origen=b.fichada_entrada.origen,
        )

    fichada_salida: Optional[FichadaRefSchema] = None
    if b.fichada_salida is not None:
        fichada_salida = FichadaRefSchema(
            id=b.fichada_salida.id,
            timestamp=b.fichada_salida.timestamp,
            tipo=b.fichada_salida.tipo,
            origen=b.fichada_salida.origen,
        )

    return HorasExtrasResponse(
        id=b.id,
        empleado_id=b.empleado_id,
        empleado_nombre=nombre,
        empleado_legajo=legajo or "",
        fecha=b.fecha,
        turno_esperado_minutos=b.turno_esperado_minutos,
        trabajado_minutos=b.trabajado_minutos,
        extras_minutos=b.extras_minutos,
        tipo_dia=b.tipo_dia,
        porcentaje_recargo=b.porcentaje_recargo,
        estado=b.estado,
        error_tipo=b.error_tipo,
        aprobado_por_nombre=aprobado_por_nombre,
        aprobado_at=b.aprobado_at,
        motivo_rechazo=b.motivo_rechazo,
        reabierto_por_nombre=reabierto_por_nombre,
        reabierto_at=b.reabierto_at,
        motivo_reapertura=b.motivo_reapertura,
        liquidacion_periodo=b.liquidacion_periodo,
        liquidado_at=b.liquidado_at,
        generada_por=b.generada_por,
        observaciones=b.observaciones,
        fichada_entrada=fichada_entrada,
        fichada_salida=fichada_salida,
        alertas_no_leidas=_count_alertas_no_leidas(db, b.id),
        created_at=b.created_at,
        updated_at=b.updated_at,
    )


def _get_config_or_500(db: Session) -> RRHHHorasExtrasConfig:
    """Carga el singleton id=1 o lanza 500."""
    cfg = db.query(RRHHHorasExtrasConfig).filter(RRHHHorasExtrasConfig.id == 1).first()
    if cfg is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Configuración de horas extras no inicializada (singleton id=1 ausente)",
        )
    return cfg


def _cron_lock_activo() -> bool:
    """
    Chequea si el cron del módulo está corriendo (lockfile presente).

    No abre el archivo: solo `Path.exists()` para evitar bloquear el lock.
    El cron real (Batch 3) usa `flock` no-bloqueante; si la ruta primaria no
    existe, cae en `/tmp/` (design §12).
    """
    for p in _CRON_LOCK_PATHS:
        try:
            if p.exists():
                return True
        except OSError:
            # Permisos / FS efímero — tratamos como no-lock (best-effort).
            continue
    return False


# ════════════════════════════════════════════════════════════════════════════
# ENDPOINTS — ROOT (sin conflicto con paramétricas)
# ════════════════════════════════════════════════════════════════════════════


@router.get(
    "/",
    response_model=HorasExtrasListResponse,
    summary="Listar bloques de HE con filtros y paginación",
)
def listar_horas_extras(
    empleado_id: Optional[int] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    estado: Optional[str] = Query(
        None,
        description="Estados separados por coma (ej: 'detectada,aprobada')",
    ),
    tipo_dia: Optional[str] = None,
    con_alertas: Optional[bool] = Query(None, description="Si true, solo bloques con alertas no leídas"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> HorasExtrasListResponse:
    """Lista paginada de bloques HE con joins a empleado y conteo de alertas."""
    _check_permiso(db, current_user, "rrhh.ver_horas_extras")

    query = db.query(RRHHHorasExtras)

    if empleado_id is not None:
        query = query.filter(RRHHHorasExtras.empleado_id == empleado_id)
    if fecha_desde is not None:
        query = query.filter(RRHHHorasExtras.fecha >= fecha_desde)
    if fecha_hasta is not None:
        query = query.filter(RRHHHorasExtras.fecha <= fecha_hasta)
    if estado:
        estados = [e.strip() for e in estado.split(",") if e.strip()]
        if estados:
            query = query.filter(RRHHHorasExtras.estado.in_(estados))
    if tipo_dia:
        query = query.filter(RRHHHorasExtras.tipo_dia == tipo_dia)
    if con_alertas:
        # Subquery: he_ids con alertas no leídas.
        subq = db.query(RRHHHorasExtrasAlerta.he_id).filter(RRHHHorasExtrasAlerta.leida_at.is_(None)).subquery()
        query = query.filter(RRHHHorasExtras.id.in_(subq))

    total = query.count()
    offset = (page - 1) * page_size
    bloques = (
        query.options(
            joinedload(RRHHHorasExtras.empleado),
            joinedload(RRHHHorasExtras.fichada_entrada),
            joinedload(RRHHHorasExtras.fichada_salida),
            joinedload(RRHHHorasExtras.aprobado_por),
            joinedload(RRHHHorasExtras.reabierto_por),
        )
        .order_by(RRHHHorasExtras.fecha.desc(), RRHHHorasExtras.id.desc())
        .offset(offset)
        .limit(page_size)
        .all()
    )

    items = [_serialize_bloque(db, b) for b in bloques]
    return HorasExtrasListResponse(items=items, total=total, page=page, page_size=page_size)


@router.post(
    "/",
    response_model=HorasExtrasResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Crear bloque HE manual",
)
def crear_horas_extras(
    data: HorasExtrasCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> HorasExtrasResponse:
    """
    Crea un bloque manual (`generada_por='manual'`, `estado='detectada'`).

    Útil para HE excepcionales donde no hay fichadas o el cron no las captó.
    Audita en historial con `accion='detectada'`.
    """
    _check_permiso(db, current_user, "rrhh.gestionar_horas_extras")

    empleado = db.query(RRHHEmpleado).filter(RRHHEmpleado.id == data.empleado_id).first()
    if empleado is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Empleado {data.empleado_id} no encontrado",
        )

    bloque = RRHHHorasExtras(
        empleado_id=data.empleado_id,
        fecha=data.fecha,
        turno_esperado_minutos=0,
        trabajado_minutos=None,
        extras_minutos=data.extras_minutos,
        tipo_dia=data.tipo_dia,
        porcentaje_recargo=data.porcentaje_recargo,
        estado=EstadoHE.DETECTADA.value,
        generada_por=GeneradaPorHE.MANUAL.value,
        generada_por_id=current_user.id,
        observaciones=data.observaciones,
    )
    db.add(bloque)
    db.flush()

    historial = RRHHHorasExtrasHistorial(
        he_id=bloque.id,
        accion="detectada",
        estado_anterior=None,
        estado_nuevo=bloque.estado,
        usuario_id=current_user.id,
        motivo="Alta manual desde UI",
        snapshot={
            "extras_minutos": bloque.extras_minutos,
            "trabajado_minutos": bloque.trabajado_minutos,
            "turno_esperado_minutos": bloque.turno_esperado_minutos,
            "tipo_dia": bloque.tipo_dia,
            "porcentaje_recargo": str(bloque.porcentaje_recargo),
            "estado": bloque.estado,
            "observaciones": bloque.observaciones,
            "fichada_entrada_id": None,
            "fichada_salida_id": None,
        },
    )
    db.add(historial)
    db.commit()
    db.refresh(bloque)

    bloque = _get_bloque_or_404(db, bloque.id)
    return _serialize_bloque(db, bloque)


# ════════════════════════════════════════════════════════════════════════════
# ENDPOINTS — RUTAS ESTÁTICAS (deben ir ANTES que /{he_id})
# T-4.8 alertas · T-4.4 bulk · T-4.10 config · T-4.7 export/liquidar ·
# T-4.9 historial · T-4.6 recalcular
# ════════════════════════════════════════════════════════════════════════════


@router.get(
    "/alertas",
    response_model=AlertasListResponse,
    summary="Listar alertas (default solo no leídas)",
)
def listar_alertas(
    solo_no_leidas: bool = Query(True, description="Si true, filtra leida_at IS NULL"),
    severidad: Optional[str] = Query(None, pattern="^(info|warning|critical)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> AlertasListResponse:
    """Lista paginada de alertas de divergencia HE."""
    _check_permiso(db, current_user, "rrhh.ver_horas_extras")

    query = db.query(RRHHHorasExtrasAlerta)
    if solo_no_leidas:
        query = query.filter(RRHHHorasExtrasAlerta.leida_at.is_(None))
    if severidad:
        query = query.filter(RRHHHorasExtrasAlerta.severidad == severidad)

    total = query.count()
    offset = (page - 1) * page_size
    alertas = (
        query.options(joinedload(RRHHHorasExtrasAlerta.leida_por))
        .order_by(RRHHHorasExtrasAlerta.created_at.desc())
        .offset(offset)
        .limit(page_size)
        .all()
    )

    items: list[AlertaResponse] = []
    for a in alertas:
        leida_por_nombre = a.leida_por.nombre if a.leida_por is not None else None
        items.append(
            AlertaResponse(
                id=a.id,
                he_id=a.he_id,
                tipo=a.tipo,
                severidad=a.severidad,
                mensaje=a.mensaje,
                contexto=a.contexto,
                leida_at=a.leida_at,
                leida_por_nombre=leida_por_nombre,
                created_at=a.created_at,
            )
        )

    return AlertasListResponse(items=items, total=total)


@router.patch(
    "/alertas/{alerta_id}/leida",
    response_model=AlertaResponse,
    summary="Marcar alerta como leída",
)
def marcar_alerta_leida(
    alerta_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> AlertaResponse:
    """Setea `leida_at = now()`, `leida_por_id = current_user.id`. Idempotente."""
    _check_permiso(db, current_user, "rrhh.ver_horas_extras")

    alerta = (
        db.query(RRHHHorasExtrasAlerta)
        .options(joinedload(RRHHHorasExtrasAlerta.leida_por))
        .filter(RRHHHorasExtrasAlerta.id == alerta_id)
        .first()
    )
    if alerta is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Alerta {alerta_id} no encontrada",
        )

    if alerta.leida_at is None:
        alerta.leida_at = datetime.now(ART_TZ)
        alerta.leida_por_id = current_user.id
        db.commit()
        db.refresh(alerta)

    leida_por_nombre = alerta.leida_por.nombre if alerta.leida_por is not None else None
    return AlertaResponse(
        id=alerta.id,
        he_id=alerta.he_id,
        tipo=alerta.tipo,
        severidad=alerta.severidad,
        mensaje=alerta.mensaje,
        contexto=alerta.contexto,
        leida_at=alerta.leida_at,
        leida_por_nombre=leida_por_nombre,
        created_at=alerta.created_at,
    )


@router.post(
    "/bulk/aprobar",
    summary="Aprobar bulk — los errores no bloquean al resto",
)
def bulk_aprobar(
    data: BulkAprobarRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict[str, Any]:
    """
    Aprueba múltiples bloques. Acumula errores individuales sin abortar.

    Returns:
        `{aprobados: int, fallidos: list[{id, status, detail}]}`.
    """
    _check_permiso(db, current_user, "rrhh.aprobar_horas_extras")

    service = HorasExtrasService(db)
    aprobados = 0
    fallidos: list[dict[str, Any]] = []
    for he_id in data.ids:
        try:
            service.aprobar_bloque(
                he_id=he_id,
                usuario_id=current_user.id,
                porcentaje_override=data.porcentaje_override,
            )
            aprobados += 1
        except HTTPException as exc:
            fallidos.append({"id": he_id, "status": exc.status_code, "detail": exc.detail})
        except Exception as exc:  # noqa: BLE001 — boundary del bulk
            fallidos.append({"id": he_id, "status": 500, "detail": str(exc)})

    return {"aprobados": aprobados, "fallidos": fallidos}


@router.post(
    "/bulk/rechazar",
    summary="Rechazar bulk — los errores no bloquean al resto",
)
def bulk_rechazar(
    data: BulkRechazarRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict[str, Any]:
    """Rechaza múltiples bloques con un único motivo común."""
    _check_permiso(db, current_user, "rrhh.aprobar_horas_extras")

    service = HorasExtrasService(db)
    rechazados = 0
    fallidos: list[dict[str, Any]] = []
    for he_id in data.ids:
        try:
            service.rechazar_bloque(he_id=he_id, usuario_id=current_user.id, motivo=data.motivo)
            rechazados += 1
        except HTTPException as exc:
            fallidos.append({"id": he_id, "status": exc.status_code, "detail": exc.detail})
        except Exception as exc:  # noqa: BLE001
            fallidos.append({"id": he_id, "status": 500, "detail": str(exc)})

    return {"rechazados": rechazados, "fallidos": fallidos}


@router.get(
    "/config",
    response_model=HorasExtrasConfigSchema,
    summary="Leer config singleton del módulo HE",
)
def obtener_config(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> HorasExtrasConfigSchema:
    """Devuelve el singleton id=1 del módulo."""
    _check_permiso(db, current_user, "rrhh.ver_horas_extras")
    cfg = (
        db.query(RRHHHorasExtrasConfig)
        .options(joinedload(RRHHHorasExtrasConfig.actualizado_por))
        .filter(RRHHHorasExtrasConfig.id == 1)
        .first()
    )
    if cfg is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Configuración de horas extras no inicializada (singleton id=1 ausente)",
        )

    actualizado_por_nombre: Optional[str] = None
    if cfg.actualizado_por is not None:
        actualizado_por_nombre = cfg.actualizado_por.nombre

    return HorasExtrasConfigSchema(
        porcentaje_dia_habil=cfg.porcentaje_dia_habil,
        porcentaje_sabado_pm=cfg.porcentaje_sabado_pm,
        porcentaje_domingo=cfg.porcentaje_domingo,
        porcentaje_feriado=cfg.porcentaje_feriado,
        hora_corte_sabado=cfg.hora_corte_sabado,
        tolerancia_extras_minutos=cfg.tolerancia_extras_minutos,
        requiere_aprobacion=cfg.requiere_aprobacion,
        cron_activo=cfg.cron_activo,
        dias_retencion_alertas=cfg.dias_retencion_alertas,
        cap_dias_recalculo_manual=cfg.cap_dias_recalculo_manual,
        updated_at=cfg.updated_at,
        actualizado_por_nombre=actualizado_por_nombre,
    )


@router.put(
    "/config",
    response_model=HorasExtrasConfigSchema,
    summary="Actualizar config singleton del módulo HE",
)
def actualizar_config(
    data: HorasExtrasConfigSchema,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> HorasExtrasConfigSchema:
    """
    Actualiza el singleton. Permiso: `rrhh.config` (mismo que el módulo de
    configuración general — design §6 fila PUT /config).

    Cambios en porcentajes NO recalculan bloques en estados congelados
    (aprobada/rechazada/liquidada/error_fichadas) — locked en design §12.
    """
    _check_permiso(db, current_user, "rrhh.config")

    cfg = db.query(RRHHHorasExtrasConfig).filter(RRHHHorasExtrasConfig.id == 1).first()
    if cfg is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Configuración de horas extras no inicializada (singleton id=1 ausente)",
        )

    cfg.porcentaje_dia_habil = data.porcentaje_dia_habil
    cfg.porcentaje_sabado_pm = data.porcentaje_sabado_pm
    cfg.porcentaje_domingo = data.porcentaje_domingo
    cfg.porcentaje_feriado = data.porcentaje_feriado
    cfg.hora_corte_sabado = data.hora_corte_sabado
    cfg.tolerancia_extras_minutos = data.tolerancia_extras_minutos
    cfg.requiere_aprobacion = data.requiere_aprobacion
    cfg.cron_activo = data.cron_activo
    cfg.dias_retencion_alertas = data.dias_retencion_alertas
    cfg.cap_dias_recalculo_manual = data.cap_dias_recalculo_manual
    cfg.actualizado_por_id = current_user.id

    db.commit()
    db.refresh(cfg)

    actualizado_por_nombre: Optional[str] = None
    if cfg.actualizado_por_id is not None:
        usuario = db.query(Usuario).filter(Usuario.id == cfg.actualizado_por_id).first()
        if usuario is not None:
            actualizado_por_nombre = usuario.nombre

    return HorasExtrasConfigSchema(
        porcentaje_dia_habil=cfg.porcentaje_dia_habil,
        porcentaje_sabado_pm=cfg.porcentaje_sabado_pm,
        porcentaje_domingo=cfg.porcentaje_domingo,
        porcentaje_feriado=cfg.porcentaje_feriado,
        hora_corte_sabado=cfg.hora_corte_sabado,
        tolerancia_extras_minutos=cfg.tolerancia_extras_minutos,
        requiere_aprobacion=cfg.requiere_aprobacion,
        cron_activo=cfg.cron_activo,
        dias_retencion_alertas=cfg.dias_retencion_alertas,
        cap_dias_recalculo_manual=cfg.cap_dias_recalculo_manual,
        updated_at=cfg.updated_at,
        actualizado_por_nombre=actualizado_por_nombre,
    )


@router.get(
    "/exportar",
    summary="Exportar bloques HE a Excel (es-AR, DD/MM/YYYY)",
)
def exportar_excel(
    periodo: Optional[str] = Query(None, pattern="^[0-9]{6}$", description="YYYYMM"),
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    estado: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> StreamingResponse:
    """
    Exporta bloques HE filtrados a XLSX (en memoria — NO long-lived).

    Headers en castellano rioplatense (revisión 2 — Q1):
    Legajo · Apellido y Nombre · CUIL · Fecha · Tipo de día · Minutos extra ·
    % Recargo · Estado · Observaciones · Motivo de rechazo.

    Formato fecha DD/MM/YYYY. Encabezado bold + freeze panes en fila 2.
    Si no hay rows → archivo solo con encabezado, NO error.
    """
    _check_permiso(db, current_user, "rrhh.ver_horas_extras")

    if periodo is None and (fecha_desde is None or fecha_hasta is None):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Se requiere 'periodo' (YYYYMM) o 'fecha_desde' + 'fecha_hasta'",
        )

    query = db.query(RRHHHorasExtras).options(joinedload(RRHHHorasExtras.empleado))

    if periodo is not None:
        query = query.filter(RRHHHorasExtras.liquidacion_periodo == periodo)
        nombre_archivo_periodo = periodo
    else:
        # fecha_desde y fecha_hasta presentes (validado arriba).
        assert fecha_desde is not None and fecha_hasta is not None
        query = query.filter(
            RRHHHorasExtras.fecha >= fecha_desde,
            RRHHHorasExtras.fecha <= fecha_hasta,
        )
        nombre_archivo_periodo = f"{fecha_desde.isoformat()}_{fecha_hasta.isoformat()}"

    if estado:
        estados = [e.strip() for e in estado.split(",") if e.strip()]
        if estados:
            query = query.filter(RRHHHorasExtras.estado.in_(estados))

    bloques = query.order_by(RRHHHorasExtras.fecha.asc(), RRHHHorasExtras.empleado_id.asc()).all()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl no instalado. Ejecute: pip install openpyxl",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horas Extras"

    headers = [
        "Legajo",
        "Apellido y Nombre",
        "CUIL",
        "Fecha",
        "Tipo de día",
        "Minutos extra",
        "% Recargo",
        "Estado",
        "Observaciones",
        "Motivo de rechazo",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Freeze panes en fila 2 (revisión 2 — Q1).
    ws.freeze_panes = "A2"

    # Mapa de display amigable de tipo_dia (es-AR).
    tipo_dia_display = {
        TipoDiaHE.HABIL_50.value: "Día hábil (50%)",
        TipoDiaHE.SABADO_100.value: "Sábado PM (100%)",
        TipoDiaHE.DOMINGO_100.value: "Domingo (100%)",
        TipoDiaHE.FERIADO_100.value: "Feriado (100%)",
        TipoDiaHE.MANUAL.value: "Manual",
    }

    for b in bloques:
        emp = b.empleado
        row = [
            emp.legajo if emp is not None else "",
            f"{emp.apellido}, {emp.nombre}" if emp is not None else "",
            emp.cuil if emp is not None and emp.cuil else "",
            b.fecha.strftime("%d/%m/%Y") if b.fecha else "",
            tipo_dia_display.get(b.tipo_dia, b.tipo_dia or ""),
            b.extras_minutos if b.extras_minutos is not None else "",
            f"{float(b.porcentaje_recargo):.2f}".replace(".", ",") if b.porcentaje_recargo is not None else "",
            b.estado or "",
            b.observaciones or "",
            b.motivo_rechazo or "",
        ]
        ws.append(row)

    # Auto-width.
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"horas_extras_{nombre_archivo_periodo}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/historial/{he_id}",
    response_model=list[HistorialEntryResponse],
    summary="Historial append-only de un bloque HE",
)
def listar_historial(
    he_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> list[HistorialEntryResponse]:
    """Lista el historial completo del bloque ordenado cronológicamente (ASC)."""
    _check_permiso(db, current_user, "rrhh.ver_horas_extras")
    _get_bloque_or_404(db, he_id)

    historial = (
        db.query(RRHHHorasExtrasHistorial)
        .options(joinedload(RRHHHorasExtrasHistorial.usuario))
        .filter(RRHHHorasExtrasHistorial.he_id == he_id)
        .order_by(RRHHHorasExtrasHistorial.created_at.asc())
        .all()
    )

    items: list[HistorialEntryResponse] = []
    for h in historial:
        usuario_nombre = h.usuario.nombre if h.usuario is not None else None
        items.append(
            HistorialEntryResponse(
                id=h.id,
                he_id=h.he_id,
                accion=h.accion,
                estado_anterior=h.estado_anterior,
                estado_nuevo=h.estado_nuevo,
                usuario_id=h.usuario_id,
                usuario_nombre=usuario_nombre,
                motivo=h.motivo,
                snapshot=h.snapshot or {},
                created_at=h.created_at,
            )
        )
    return items


@router.post(
    "/liquidar",
    response_model=LiquidacionResponse,
    summary="Liquidar período (aprobada → liquidada)",
)
def liquidar_periodo(
    data: LiquidacionRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> LiquidacionResponse:
    """Marca como liquidados los IDs aprobados; los demás se acumulan en detalle."""
    _check_permiso(db, current_user, "rrhh.liquidar_horas_extras")

    service = HorasExtrasService(db)
    resultado = service.liquidar_periodo(periodo=data.periodo, ids=data.ids, usuario_id=current_user.id)
    return LiquidacionResponse(**resultado)


@router.post(
    "/recalcular",
    summary="Recálculo manual de un rango (cap configurable, lockfile cron)",
)
def recalcular_periodo(
    data: RecalcularRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> dict[str, int]:
    """
    Trigger manual del cron sobre un rango de fechas.

    Validaciones:
    - Rango <= `cap_dias_recalculo_manual` (default 90, configurable). 422 si excede.
    - Si el lockfile del cron está activo → 409 (race condition, design §12).
    """
    _check_permiso(db, current_user, "rrhh.gestionar_horas_extras")

    cfg = _get_config_or_500(db)
    dias_solicitados = (data.fecha_hasta - data.fecha_desde).days
    if dias_solicitados > cfg.cap_dias_recalculo_manual:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"El rango de recálculo no puede superar {cfg.cap_dias_recalculo_manual} días "
                f"(solicitado: {dias_solicitados})"
            ),
        )

    if _cron_lock_activo():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cron en curso, intentá luego",
        )

    service = HorasExtrasService(db)
    empleado_ids = [data.empleado_id] if data.empleado_id is not None else None
    return service.detectar_he_periodo(
        fecha_desde=data.fecha_desde,
        fecha_hasta=data.fecha_hasta,
        empleado_ids=empleado_ids,
    )


# ════════════════════════════════════════════════════════════════════════════
# ENDPOINTS — RUTAS PARAMETRIZADAS /{he_id} (van AL FINAL)
# T-4.3 detalle/edición · T-4.4 transiciones · T-4.5 anomalías
# ════════════════════════════════════════════════════════════════════════════


@router.get(
    "/{he_id}",
    response_model=HorasExtrasResponse,
    summary="Detalle de un bloque HE",
)
def obtener_horas_extras(
    he_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> HorasExtrasResponse:
    """Obtiene un bloque HE con fichadas y conteo de alertas."""
    _check_permiso(db, current_user, "rrhh.ver_horas_extras")
    bloque = _get_bloque_or_404(db, he_id)
    return _serialize_bloque(db, bloque)


@router.put(
    "/{he_id}",
    response_model=HorasExtrasResponse,
    summary="Editar bloque HE en estados editables",
)
def actualizar_horas_extras(
    he_id: int,
    data: HorasExtrasUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> HorasExtrasResponse:
    """Edita porcentaje y/u observaciones. Solo en estados editables."""
    _check_permiso(db, current_user, "rrhh.gestionar_horas_extras")

    bloque = _get_bloque_or_404(db, he_id)
    estados_editables = (
        EstadoHE.DETECTADA.value,
        EstadoHE.ERROR_FICHADAS.value,
        EstadoHE.PENDIENTE_ASIGNACION_TURNO.value,
    )
    if bloque.estado not in estados_editables:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(f"Bloque {he_id} en estado '{bloque.estado}' no es editable (requiere uno de {estados_editables})"),
        )

    snapshot_anterior = {
        "extras_minutos": bloque.extras_minutos,
        "trabajado_minutos": bloque.trabajado_minutos,
        "turno_esperado_minutos": bloque.turno_esperado_minutos,
        "tipo_dia": bloque.tipo_dia,
        "porcentaje_recargo": str(bloque.porcentaje_recargo) if bloque.porcentaje_recargo is not None else None,
        "estado": bloque.estado,
        "observaciones": bloque.observaciones,
        "fichada_entrada_id": bloque.fichada_entrada_id,
        "fichada_salida_id": bloque.fichada_salida_id,
    }

    cambios: list[str] = []
    if data.porcentaje_recargo is not None and data.porcentaje_recargo != bloque.porcentaje_recargo:
        bloque.porcentaje_recargo = data.porcentaje_recargo
        cambios.append("porcentaje_recargo")
    if data.observaciones is not None:
        bloque.observaciones = data.observaciones
        cambios.append("observaciones")

    if cambios:
        accion = "edicion_porcentaje" if "porcentaje_recargo" in cambios else "edicion_observaciones"
        historial = RRHHHorasExtrasHistorial(
            he_id=bloque.id,
            accion=accion,
            estado_anterior=bloque.estado,
            estado_nuevo=bloque.estado,
            usuario_id=current_user.id,
            motivo=f"Edición de {', '.join(cambios)}",
            snapshot=snapshot_anterior,
        )
        db.add(historial)

    db.commit()
    db.refresh(bloque)
    bloque = _get_bloque_or_404(db, he_id)
    return _serialize_bloque(db, bloque)


@router.patch(
    "/{he_id}/aprobar",
    response_model=HorasExtrasResponse,
    summary="Aprobar bloque HE",
)
def aprobar_bloque(
    he_id: int,
    data: AprobacionRequest = AprobacionRequest(),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> HorasExtrasResponse:
    """Transiciona un bloque desde `detectada` a `aprobada`."""
    _check_permiso(db, current_user, "rrhh.aprobar_horas_extras")

    service = HorasExtrasService(db)
    bloque = service.aprobar_bloque(
        he_id=he_id,
        usuario_id=current_user.id,
        porcentaje_override=data.porcentaje_override,
        observaciones=data.observaciones,
    )
    bloque = _get_bloque_or_404(db, bloque.id)
    return _serialize_bloque(db, bloque)


@router.patch(
    "/{he_id}/rechazar",
    response_model=HorasExtrasResponse,
    summary="Rechazar bloque HE",
)
def rechazar_bloque(
    he_id: int,
    data: RechazoRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> HorasExtrasResponse:
    """Transiciona un bloque a `rechazada` con motivo obligatorio."""
    _check_permiso(db, current_user, "rrhh.aprobar_horas_extras")

    service = HorasExtrasService(db)
    bloque = service.rechazar_bloque(he_id=he_id, usuario_id=current_user.id, motivo=data.motivo)
    bloque = _get_bloque_or_404(db, bloque.id)
    return _serialize_bloque(db, bloque)


@router.patch(
    "/{he_id}/reabrir",
    response_model=HorasExtrasResponse,
    summary="Reabrir bloque HE (aprobada/rechazada/liquidada)",
)
def reabrir_bloque(
    he_id: int,
    data: ReaperturaRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> HorasExtrasResponse:
    """
    Reabre un bloque para corregirlo:
    - `aprobada`/`rechazada` → `detectada` (perm `aprobar`).
    - `liquidada` → `aprobada` (perm `liquidar`).
    """
    bloque = _get_bloque_or_404(db, he_id)
    if bloque.estado == EstadoHE.LIQUIDADA.value:
        _check_permiso(db, current_user, "rrhh.liquidar_horas_extras")
    else:
        _check_permiso(db, current_user, "rrhh.aprobar_horas_extras")

    service = HorasExtrasService(db)
    bloque = service.reabrir_bloque(he_id=he_id, usuario_id=current_user.id, motivo=data.motivo)
    bloque = _get_bloque_or_404(db, bloque.id)
    return _serialize_bloque(db, bloque)


@router.post(
    "/{he_id}/completar-fichada",
    response_model=HorasExtrasResponse,
    summary="Completar fichada faltante (solo desde error_fichadas)",
)
def completar_fichada(
    he_id: int,
    data: CompletarFichadaRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> HorasExtrasResponse:
    """Crea una fichada manual y recalcula el día del bloque (error_fichadas)."""
    _check_permiso(db, current_user, "rrhh.gestionar_horas_extras")

    service = HorasExtrasService(db)
    bloque = service.completar_fichada_faltante(
        he_id=he_id,
        usuario_id=current_user.id,
        timestamp=data.timestamp,
        tipo=data.tipo,
        motivo=data.motivo,
    )
    # Si el resultado es virtual (sin id persistido) devolvemos sin volver a query.
    if bloque.id is None or bloque.id == he_id:
        # Caso virtual: no hubo recálculo persistido. Devolvemos lo que vino.
        return HorasExtrasResponse(
            id=bloque.id or he_id,
            empleado_id=bloque.empleado_id,
            empleado_nombre="",
            empleado_legajo="",
            fecha=bloque.fecha,
            turno_esperado_minutos=bloque.turno_esperado_minutos,
            trabajado_minutos=bloque.trabajado_minutos,
            extras_minutos=bloque.extras_minutos,
            tipo_dia=bloque.tipo_dia,
            porcentaje_recargo=bloque.porcentaje_recargo,
            estado=bloque.estado,
            error_tipo=bloque.error_tipo,
            generada_por=bloque.generada_por,
            observaciones=bloque.observaciones,
            created_at=bloque.created_at or datetime.now(ART_TZ),
            updated_at=bloque.updated_at or datetime.now(ART_TZ),
        )
    bloque = _get_bloque_or_404(db, bloque.id)
    return _serialize_bloque(db, bloque)


@router.post(
    "/{he_id}/descartar-dia",
    response_model=HorasExtrasResponse,
    summary="Descartar día completo (error_fichadas → rechazada)",
)
def descartar_dia(
    he_id: int,
    data: DescartarDiaRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
) -> HorasExtrasResponse:
    """Marca el día como descartado (rechaza con motivo prefijado `[descartado]`)."""
    _check_permiso(db, current_user, "rrhh.aprobar_horas_extras")

    service = HorasExtrasService(db)
    bloque = service.descartar_dia(he_id=he_id, usuario_id=current_user.id, motivo=data.motivo)
    bloque = _get_bloque_or_404(db, bloque.id)
    return _serialize_bloque(db, bloque)

"""
Importar empleados desde Excel (Base_Empleados_Template.xlsx).

Uso:
  cd backend
  source venv/bin/activate
  python -m app.scripts.importar_empleados_excel /ruta/al/archivo.xlsx

Opciones:
  --dry-run    Muestra lo que haría sin escribir en la DB
  --force      Actualiza empleados existentes (match por DNI o legajo)

Mapeo de columnas Excel → Modelo:
  "Apellido y nombre"  → apellido + nombre (split heurístico)
  "DNI"                → dni
  "CUIL"               → cuil
  "Fecha Nacimiento"   → fecha_nacimiento
  "Domicilio Real"     → calle (+ domicilio como texto completo)
  "Localidad"          → localidad
  "Provincia"          → provincia
  "Código Postal"      → codigo_postal
  "Celular"            → telefono
  "Mail"               → email_personal
  "Puesto"             → puesto
  "Área"               → area
  "Fecha Ingreso"      → fecha_ingreso
  "Fecha Baja"         → fecha_egreso + estado="baja"
  "Motivo"             → detalle_baja
  "# de Legajo"        → legajo

  "Nacionalidad"       → datos_custom.nacionalidad
  "Estado Civil"       → datos_custom.estado_civil
"""

import sys
from pathlib import Path

# Agregar backend al path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

import argparse
from datetime import date, datetime

import openpyxl

from app.core.database import SessionLocal
from app.models.rrhh_empleado import RRHHEmpleado


# ──────────────────────────────────────────────
# APELLIDOS COMPUESTOS CONOCIDOS
# ──────────────────────────────────────────────
# Si el nombre completo empieza con alguno de estos,
# se toman las 2 primeras palabras como apellido.
APELLIDOS_COMPUESTOS = {
    "GUTIERREZ GONZALEZ",
    "MENDOZA VALENZUELA",
    "PAREDES CARMONA",
    "BARRIO CANDIDO",
}


def split_nombre_apellido(full_name: str) -> tuple[str, str]:
    """
    Separa 'APELLIDO NOMBRE1 NOMBRE2' en (apellido, nombre).

    Heurística:
    1. Si match con APELLIDOS_COMPUESTOS → 2 palabras = apellido
    2. Si no → primera palabra = apellido, resto = nombre
    """
    parts = full_name.strip().split()
    if len(parts) <= 1:
        return (full_name.strip(), "")

    # Probar apellidos compuestos
    if len(parts) >= 3:
        candidate = f"{parts[0]} {parts[1]}"
        if candidate.upper() in APELLIDOS_COMPUESTOS:
            return (candidate, " ".join(parts[2:]))

    return (parts[0], " ".join(parts[1:]))


def parse_date(val) -> date | None:
    """Convierte datetime de openpyxl a date."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def parse_str(val) -> str | None:
    """Convierte a string limpio o None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def main() -> None:
    parser = argparse.ArgumentParser(description="Importar empleados desde Excel")
    parser.add_argument("archivo", help="Ruta al archivo .xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Solo mostrar, no escribir en DB")
    parser.add_argument("--force", action="store_true", help="Actualizar existentes (match por DNI o legajo)")
    args = parser.parse_args()

    archivo = Path(args.archivo)
    if not archivo.exists():
        print(f"ERROR: Archivo no encontrado: {archivo}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(archivo), read_only=True, data_only=True)
    ws = wb.active

    # Leer headers
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"Columnas encontradas: {headers}")
    print()

    # Mapeo columna → índice (1-based)
    col_map = {}
    for i, h in enumerate(headers, 1):
        if h:
            col_map[h.strip()] = i

    required = ["Apellido y nombre", "DNI", "Fecha Ingreso", "# de Legajo"]
    missing = [r for r in required if r not in col_map]
    if missing:
        print(f"ERROR: Faltan columnas requeridas: {missing}")
        sys.exit(1)

    # Leer filas
    rows_data = []
    for r in range(2, ws.max_row + 1):
        nombre_completo = ws.cell(r, col_map["Apellido y nombre"]).value
        if not nombre_completo:
            continue

        apellido, nombre = split_nombre_apellido(str(nombre_completo).strip())

        dni_raw = ws.cell(r, col_map["DNI"]).value
        dni = str(int(dni_raw)) if isinstance(dni_raw, (int, float)) else parse_str(dni_raw)

        legajo_raw = ws.cell(r, col_map["# de Legajo"]).value
        legajo = str(int(legajo_raw)) if isinstance(legajo_raw, (int, float)) else parse_str(legajo_raw)

        fecha_ingreso = parse_date(ws.cell(r, col_map["Fecha Ingreso"]).value)
        fecha_baja = parse_date(ws.cell(r, col_map.get("Fecha Baja", 999)).value) if "Fecha Baja" in col_map else None

        # Estado: baja si tiene fecha de baja
        estado = "baja" if fecha_baja else "activo"

        # Domicilio completo
        calle = parse_str(ws.cell(r, col_map.get("Domicilio Real", 999)).value) if "Domicilio Real" in col_map else None
        localidad = parse_str(ws.cell(r, col_map.get("Localidad", 999)).value) if "Localidad" in col_map else None
        provincia = parse_str(ws.cell(r, col_map.get("Provincia", 999)).value) if "Provincia" in col_map else None
        domicilio_parts = [p for p in [calle, localidad, provincia] if p]
        domicilio = ", ".join(domicilio_parts) if domicilio_parts else None

        # Datos custom (campos que el modelo no tiene nativamente)
        datos_custom = {}
        if "Nacionalidad" in col_map:
            val = parse_str(ws.cell(r, col_map["Nacionalidad"]).value)
            if val:
                datos_custom["nacionalidad"] = val
        if "Estado Civil" in col_map:
            val = parse_str(ws.cell(r, col_map["Estado Civil"]).value)
            if val:
                datos_custom["estado_civil"] = val

        row = {
            "apellido": apellido,
            "nombre": nombre,
            "dni": dni,
            "cuil": parse_str(ws.cell(r, col_map.get("CUIL", 999)).value) if "CUIL" in col_map else None,
            "fecha_nacimiento": parse_date(ws.cell(r, col_map.get("Fecha Nacimiento", 999)).value)
            if "Fecha Nacimiento" in col_map
            else None,
            "calle": calle,
            "localidad": localidad,
            "provincia": provincia,
            "codigo_postal": parse_str(ws.cell(r, col_map.get("Código Postal", 999)).value)
            if "Código Postal" in col_map
            else None,
            "domicilio": domicilio,
            "telefono": parse_str(ws.cell(r, col_map.get("Celular", 999)).value) if "Celular" in col_map else None,
            "email_personal": parse_str(ws.cell(r, col_map.get("Mail", 999)).value) if "Mail" in col_map else None,
            "puesto": parse_str(ws.cell(r, col_map.get("Puesto", 999)).value) if "Puesto" in col_map else None,
            "area": parse_str(ws.cell(r, col_map.get("Área", 999)).value) if "Área" in col_map else None,
            "legajo": legajo,
            "fecha_ingreso": fecha_ingreso,
            "fecha_egreso": fecha_baja,
            "estado": estado,
            "detalle_baja": parse_str(ws.cell(r, col_map.get("Motivo", 999)).value) if "Motivo" in col_map else None,
            "datos_custom": datos_custom if datos_custom else None,
        }
        rows_data.append(row)

    wb.close()

    if not rows_data:
        print("No se encontraron filas con datos.")
        sys.exit(0)

    # ── Mostrar resumen ──
    print(f"{'=' * 70}")
    print(f"RESUMEN: {len(rows_data)} empleados a importar")
    print(f"{'=' * 70}")
    print()

    activos = sum(1 for r in rows_data if r["estado"] == "activo")
    bajas = sum(1 for r in rows_data if r["estado"] == "baja")
    print(f"  Activos: {activos}")
    print(f"  Bajas: {bajas}")
    print()

    print(f"{'LEGAJO':<8} {'APELLIDO':<25} {'NOMBRE':<25} {'DNI':<12} {'ESTADO':<8} {'ÁREA':<20}")
    print(f"{'-' * 8} {'-' * 25} {'-' * 25} {'-' * 12} {'-' * 8} {'-' * 20}")
    for r in rows_data:
        print(
            f"{r['legajo'] or '???':<8} "
            f"{(r['apellido'] or '')[:25]:<25} "
            f"{(r['nombre'] or '')[:25]:<25} "
            f"{(r['dni'] or ''):<12} "
            f"{r['estado']:<8} "
            f"{(r['area'] or '')[:20]:<20}"
        )

    print()
    print("VERIFICÁ los nombres arriba. Si algún apellido/nombre está mal partido,")
    print("agregá el apellido compuesto a APELLIDOS_COMPUESTOS en este script.")
    print()

    if args.dry_run:
        print(">>> DRY RUN: no se escribió nada en la base de datos.")
        return

    # ── Insertar en DB ──
    db = SessionLocal()
    try:
        created = 0
        updated = 0
        skipped = 0

        for row in rows_data:
            # Buscar existente por DNI o legajo
            existing = None
            if row["dni"]:
                existing = db.query(RRHHEmpleado).filter(RRHHEmpleado.dni == row["dni"]).first()
            if not existing and row["legajo"]:
                existing = db.query(RRHHEmpleado).filter(RRHHEmpleado.legajo == row["legajo"]).first()

            if existing:
                if args.force:
                    # Actualizar
                    for field, value in row.items():
                        if value is not None and hasattr(existing, field):
                            setattr(existing, field, value)
                    updated += 1
                    print(f"  ACTUALIZADO: {row['legajo']} {row['apellido']}, {row['nombre']}")
                else:
                    skipped += 1
                    print(
                        f"  EXISTENTE (skip): {row['legajo']} {row['apellido']}, {row['nombre']} (usar --force para actualizar)"
                    )
                continue

            # Crear nuevo
            emp = RRHHEmpleado(
                nombre=row["nombre"],
                apellido=row["apellido"],
                dni=row["dni"],
                cuil=row["cuil"],
                fecha_nacimiento=row["fecha_nacimiento"],
                calle=row["calle"],
                localidad=row["localidad"],
                provincia=row["provincia"],
                codigo_postal=row["codigo_postal"],
                domicilio=row["domicilio"],
                telefono=row["telefono"],
                email_personal=row["email_personal"],
                puesto=row["puesto"],
                area=row["area"],
                legajo=row["legajo"],
                fecha_ingreso=row["fecha_ingreso"],
                fecha_egreso=row["fecha_egreso"],
                estado=row["estado"],
                detalle_baja=row["detalle_baja"],
                datos_custom=row["datos_custom"],
            )
            db.add(emp)
            created += 1

        db.commit()

        print()
        print(f"{'=' * 70}")
        print("RESULTADO:")
        print(f"  Creados: {created}")
        print(f"  Actualizados: {updated}")
        print(f"  Saltados (existentes): {skipped}")
        print(f"{'=' * 70}")

    except Exception as e:
        db.rollback()
        print(f"ERROR: {e}")
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()

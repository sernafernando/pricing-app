"""
T-8.9 — Test manual: export Excel del período liquidado, formato es-AR.

Verifica que el endpoint `GET /rrhh/horas-extras/exportar?periodo=YYYYMM`:
- Genera un archivo XLSX válido en memoria.
- Headers en castellano rioplatense (Legajo, Apellido y Nombre, CUIL, Fecha,
  Tipo de día, Minutos extra, % Recargo, Estado, Observaciones, Motivo de
  rechazo).
- Fechas en formato DD/MM/YYYY.
- Decimales con coma (es-AR) en columna `% Recargo`.
- Header bold + freeze panes en `A2`.
- Período sin liquidaciones devuelve XLSX con sólo encabezado (NO error).

Estrategia: invocar la función handler `exportar_excel` directamente con un
`Usuario` real de la DB. Luego abrir el `BytesIO` con openpyxl y validar.

Setup (mínimo):
- Crea 2 bloques liquidados con periodo único de prueba (`999912`).
- Llama el handler con periodo=999912.
- Valida headers, formato y filas.
- Llama el handler con periodo=999911 (sin filas) → solo headers.

Uso:
    cd /var/www/html/pricing-app/backend
    python -m app.scripts.test_manual_rrhh_he_export_excel

Spec ref: "Export Excel del período liquidado", revisión 2 Q1.
"""

from __future__ import annotations

import io
import os
import sys
import time as time_mod
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

if __name__ == "__main__":
    backend_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    if backend_path not in sys.path:
        sys.path.insert(0, backend_path)

    from dotenv import load_dotenv

    env_path = Path(backend_path) / ".env"
    load_dotenv(dotenv_path=env_path)

from app.core.database import SessionLocal
from app.models.rrhh_empleado import RRHHEmpleado
from app.models.rrhh_horas_extras import (
    EstadoHE,
    GeneradaPorHE,
    RRHHHorasExtras,
    RRHHHorasExtrasAlerta,
    RRHHHorasExtrasHistorial,
    TipoDiaHE,
)
from app.models.usuario import Usuario
from app.routers.rrhh_horas_extras import exportar_excel


SUFFIX = f"T89_{int(time_mod.time())}"
PERIODO_TEST_CON = "999912"  # Período fictício; bajo el supuesto que no existe en la DB.
PERIODO_TEST_SIN = "999911"

HEADERS_ESPERADOS = [
    "Legajo",
    "Apellido y Nombre",
    "CUIL",
    "Fecha",
    "Tipo de día",
    "Minutos extra",
    "% Recargo",
    "Estado",
    "Observaciones",
    "Motivo de rechazo",
]


def _ok(msg: str) -> None:
    print(f"  PASS - {msg}")


def _fail(msg: str) -> None:
    print(f"  FAIL - {msg}")


def _read_workbook(response):
    import openpyxl  # noqa: F401

    # response es StreamingResponse — body_iterator es un async generator.
    # En la implementación actual, el endpoint guarda BytesIO y lo pasa como iter.
    # Para test inline, tomamos el primer chunk si es bytes.
    chunks: list[bytes] = []
    body_iterator = getattr(response, "body_iterator", None)
    if body_iterator is None:
        raise AssertionError("StreamingResponse no tiene body_iterator")
    # Drenar (sync iteration sobre BytesIO funciona si el endpoint pasa BytesIO
    # directamente — en este router el cierre es seek(0) + StreamingResponse(buf).
    for chunk in body_iterator:
        if isinstance(chunk, bytes):
            chunks.append(chunk)
        elif hasattr(chunk, "read"):
            chunks.append(chunk.read())
        else:
            chunks.append(bytes(chunk))
    data = b"".join(chunks)
    if not data:
        # Fallback: tal vez `response.body` tiene los bytes (StreamingResponse no
        # los expone normalmente, pero en algunos casos si el handler usa BytesIO
        # accesible directamente).
        raise AssertionError("StreamingResponse vacío")
    return openpyxl.load_workbook(io.BytesIO(data))


def main() -> int:
    print(f"Iniciando test_manual_rrhh_he_export_excel (T-8.9) — suffix {SUFFIX}")

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("  SKIP - openpyxl no instalado. pip install openpyxl")
        return 0

    db = SessionLocal()
    fallos = 0
    empleado: RRHHEmpleado | None = None
    bloques_creados: list[int] = []
    fecha_a = date(2026, 4, 5)
    fecha_b = date(2026, 4, 12)

    try:
        usuario = db.query(Usuario).filter(Usuario.id == 1).first()
        if usuario is None:
            print("  SKIP - No existe Usuario(id=1) en la DB; no se puede invocar el handler")
            return 0

        # Setup: empleado + 2 bloques liquidados (insertados directamente).
        empleado = RRHHEmpleado(
            nombre="Test",
            apellido=f"TEST_HE_{SUFFIX}",
            legajo=f"TST{SUFFIX[:14]}",
            estado="activo",
            activo=True,
        )
        db.add(empleado)
        db.commit()
        db.refresh(empleado)

        for fecha in (fecha_a, fecha_b):
            b = RRHHHorasExtras(
                empleado_id=empleado.id,
                fecha=fecha,
                turno_esperado_minutos=480,
                trabajado_minutos=600,
                extras_minutos=120,
                tipo_dia=TipoDiaHE.HABIL_50.value,
                porcentaje_recargo=Decimal("50.00"),
                estado=EstadoHE.LIQUIDADA.value,
                liquidacion_periodo=PERIODO_TEST_CON,
                liquidado_por_id=usuario.id,
                liquidado_at=datetime.now(),
                generada_por=GeneradaPorHE.SISTEMA.value,
            )
            db.add(b)
            db.flush()
            bloques_creados.append(b.id)
        db.commit()
        _ok(f"Setup: 2 bloques liquidados periodo={PERIODO_TEST_CON}")

        # ─── Caso A: período CON liquidaciones ─────────────────────────
        resp = exportar_excel(
            periodo=PERIODO_TEST_CON,
            fecha_desde=None,
            fecha_hasta=None,
            estado=None,
            db=db,
            current_user=usuario,
        )
        wb = _read_workbook(resp)
        ws = wb.active

        # 1. Headers exactos.
        headers_actual = [c.value for c in ws[1]]
        if headers_actual == HEADERS_ESPERADOS:
            _ok(f"Headers en castellano correctos: {headers_actual}")
        else:
            fallos += 1
            _fail(f"Headers no coinciden.\n    esperado: {HEADERS_ESPERADOS}\n    actual:   {headers_actual}")

        # 2. Header bold.
        if all(c.font.bold for c in ws[1]):
            _ok("Header en bold")
        else:
            fallos += 1
            _fail("Header NO está completamente en bold")

        # 3. Freeze panes en A2.
        if ws.freeze_panes == "A2":
            _ok("Freeze panes en A2")
        else:
            fallos += 1
            _fail(f"freeze_panes={ws.freeze_panes!r} (esperaba 'A2')")

        # 4. 2 filas de datos (rows 2 y 3).
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        if len(data_rows) >= 2:
            _ok(f"Datos: {len(data_rows)} filas en el sheet")
        else:
            fallos += 1
            _fail(f"Esperaba >=2 filas; got {len(data_rows)}")

        # 5. Fecha formato DD/MM/YYYY (col 4 = índice 3).
        fechas_ok = []
        for row in data_rows[:2]:
            fecha_cell = row[3]
            if isinstance(fecha_cell, str) and len(fecha_cell) == 10 and fecha_cell[2] == "/" and fecha_cell[5] == "/":
                fechas_ok.append(fecha_cell)
        if len(fechas_ok) == 2:
            _ok(f"Fechas en formato DD/MM/YYYY: {fechas_ok}")
        else:
            fallos += 1
            _fail(f"Fechas no en formato DD/MM/YYYY: {[r[3] for r in data_rows[:2]]}")

        # 6. % Recargo con coma (col 7 = índice 6).
        pcts = [row[6] for row in data_rows[:2]]
        if all(isinstance(p, str) and "," in p for p in pcts):
            _ok(f"% Recargo con coma decimal es-AR: {pcts}")
        else:
            fallos += 1
            _fail(f"% Recargo no usa coma: {pcts}")

        # ─── Caso B: período SIN liquidaciones ────────────────────────
        resp_vacio = exportar_excel(
            periodo=PERIODO_TEST_SIN,
            fecha_desde=None,
            fecha_hasta=None,
            estado=None,
            db=db,
            current_user=usuario,
        )
        wb_v = _read_workbook(resp_vacio)
        ws_v = wb_v.active
        headers_v = [c.value for c in ws_v[1]]
        data_rows_v = list(ws_v.iter_rows(min_row=2, values_only=True))
        if headers_v == HEADERS_ESPERADOS and len(data_rows_v) == 0:
            _ok(f"Período sin liquidaciones ({PERIODO_TEST_SIN}) -> XLSX con headers, 0 filas (no error)")
        else:
            fallos += 1
            _fail(f"Período sin liquidaciones: headers_ok={headers_v == HEADERS_ESPERADOS}, rows={len(data_rows_v)}")

    except Exception as exc:
        fallos += 1
        _fail(f"Excepción inesperada: {exc!r}")
        import traceback

        traceback.print_exc()
    finally:
        try:
            if bloques_creados:
                db.query(RRHHHorasExtrasAlerta).filter(RRHHHorasExtrasAlerta.he_id.in_(bloques_creados)).delete(
                    synchronize_session=False
                )
                db.query(RRHHHorasExtrasHistorial).filter(RRHHHorasExtrasHistorial.he_id.in_(bloques_creados)).delete(
                    synchronize_session=False
                )
                db.query(RRHHHorasExtras).filter(RRHHHorasExtras.id.in_(bloques_creados)).delete(
                    synchronize_session=False
                )
            if empleado is not None:
                row_emp = db.query(RRHHEmpleado).filter(RRHHEmpleado.id == empleado.id).first()
                if row_emp is not None:
                    db.delete(row_emp)
            db.commit()
            print("  Cleanup OK")
        except Exception as cleanup_exc:
            print(f"  WARN - Cleanup falló: {cleanup_exc!r}")
        finally:
            db.close()

    if fallos == 0:
        print("\nResultado: PASS (T-8.9)")
        return 0
    print(f"\nResultado: FAIL ({fallos} aserciones fallidas)")
    return 1


if __name__ == "__main__":
    sys.exit(main())

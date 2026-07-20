"""
Unit tests — read-side color layer refactor (PR3 of productos-color-teams).

Verifies:
- `resolver_layer_activo`: None -> U, explicit U allowed, team member allowed,
  non-member forbidden (403).
- `filtro_colores` parity on both slots (ml/tienda), including the
  `sin_color` sentinel — behavior-identical to the legacy
  `productos_pricing.color_marcado[_tienda]` filter.
- Listing endpoint (`GET /api/productos`) parity: default (no equipo_id)
  reads/filters equal today's legacy-column behavior, since producto_color@U
  is kept in sync with legacy by the PR2 dual-write.
- Read-scoping: a team member requesting `equipo_id=T` sees the team color +
  `color_hint_global`; the default request sees the U (global) color.
- Hint fields present in the listing payload with correct values, including
  `color_hint_equipo_inicial`.
"""

from __future__ import annotations

import pytest
from fastapi import HTTPException

from app.api.endpoints.productos_shared import (
    color_slot,
    filtro_colores,
    resolver_layer_activo,
)
from app.models.equipo import Equipo, EquipoMiembro, ProductoColor
from app.models.producto import ProductoERP, ProductoPricing
from app.models.usuario import AuthProvider, RolUsuario, Usuario
from app.core.security import get_password_hash

# `exportar_lista_gremio` locally imports PrecioGremioOverride inside the
# endpoint body; importing it here ensures its table is registered on
# `Base.metadata` before the session-scoped `engine` fixture's create_all runs.
from app.models.precio_gremio_override import PrecioGremioOverride  # noqa: F401


def _make_usuario(db, username: str, rol_id: int) -> Usuario:
    user = Usuario(
        username=username,
        email=f"{username}@example.com",
        nombre=username,
        password_hash=get_password_hash("TestPass123!"),
        rol=RolUsuario.VENTAS,
        rol_id=rol_id,
        auth_provider=AuthProvider.LOCAL,
        activo=True,
    )
    db.add(user)
    db.flush()
    return user


def _make_producto(db, item_id: int, **kwargs) -> ProductoERP:
    kwargs.setdefault("costo", 100.0)
    kwargs.setdefault("iva", 21.0)
    producto = ProductoERP(
        item_id=item_id,
        codigo=f"COD{item_id}",
        descripcion=kwargs.pop("descripcion", "Producto de prueba"),
        **kwargs,
    )
    db.add(producto)
    db.flush()
    return producto


def _make_equipo(db, nombre: str, es_global: bool = False) -> Equipo:
    equipo = Equipo(nombre=nombre, es_global=es_global)
    db.add(equipo)
    db.flush()
    return equipo


def _add_member(db, equipo_id: int, usuario_id: int) -> EquipoMiembro:
    miembro = EquipoMiembro(equipo_id=equipo_id, usuario_id=usuario_id, rol="miembro")
    db.add(miembro)
    db.flush()
    return miembro


def _global_equipo(db) -> Equipo:
    equipo = db.query(Equipo).filter(Equipo.es_global.is_(True)).first()
    if equipo is None:
        equipo = _make_equipo(db, "Global", es_global=True)
    return equipo


def auth_headers_for(user: Usuario) -> dict:
    from tests.conftest import make_access_token

    token = make_access_token(user)
    return {"Authorization": f"Bearer {token}"}


def _set_legacy_color(db, item_id: int, color: str | None = None, color_tienda: str | None = None) -> None:
    pricing = db.query(ProductoPricing).filter(ProductoPricing.item_id == item_id).first()
    if pricing is None:
        pricing = ProductoPricing(item_id=item_id)
        db.add(pricing)
    if color is not None:
        pricing.color_marcado = color
    if color_tienda is not None:
        pricing.color_marcado_tienda = color_tienda


def _set_producto_color(
    db, equipo_id: int, item_id: int, color_ml: str | None = None, color_tienda: str | None = None
) -> ProductoColor:
    row = db.query(ProductoColor).filter(ProductoColor.equipo_id == equipo_id, ProductoColor.item_id == item_id).first()
    if row is None:
        row = ProductoColor(equipo_id=equipo_id, item_id=item_id)
        db.add(row)
    if color_ml is not None:
        row.color_ml = color_ml
    if color_tienda is not None:
        row.color_tienda = color_tienda
    db.flush()
    return row


# ---------------------------------------------------------------------------
# resolver_layer_activo
# ---------------------------------------------------------------------------


class TestResolverLayerActivo:
    def test_none_resolves_to_global(self, db, rol_ventas) -> None:
        equipo_global = _global_equipo(db)
        user = _make_usuario(db, "r1", rol_ventas.id)

        assert resolver_layer_activo(None, user, db) == equipo_global.id

    def test_explicit_global_allowed(self, db, rol_ventas) -> None:
        equipo_global = _global_equipo(db)
        user = _make_usuario(db, "r2", rol_ventas.id)

        assert resolver_layer_activo(equipo_global.id, user, db) == equipo_global.id

    def test_member_of_team_allowed(self, db, rol_ventas) -> None:
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo R")
        user = _make_usuario(db, "r3", rol_ventas.id)
        _add_member(db, equipo.id, user.id)

        assert resolver_layer_activo(equipo.id, user, db) == equipo.id

    def test_non_member_forbidden(self, db, rol_ventas) -> None:
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo S")
        user = _make_usuario(db, "r4", rol_ventas.id)

        with pytest.raises(HTTPException) as exc:
            resolver_layer_activo(equipo.id, user, db)
        assert exc.value.status_code == 403


# ---------------------------------------------------------------------------
# filtro_colores parity (unit-level, direct query building)
# ---------------------------------------------------------------------------


class TestFiltroColoresParity:
    def _query(self, db):
        return db.query(ProductoERP, ProductoPricing).outerjoin(
            ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id
        )

    def test_ml_slot_specific_colors(self, db, rol_ventas) -> None:
        equipo_global = _global_equipo(db)
        _make_producto(db, 301)
        _make_producto(db, 302)
        _set_producto_color(db, equipo_global.id, 301, color_ml="rojo")
        _set_producto_color(db, equipo_global.id, 302, color_ml="verde")
        db.commit()

        from app.api.endpoints.productos_shared import join_color_layer

        query = join_color_layer(self._query(db), equipo_global.id)
        query = filtro_colores(query, "rojo", color_slot(None))
        results = query.all()

        assert {p.item_id for p, _ in results} == {301}

    def test_ml_slot_sin_color_sentinel(self, db, rol_ventas) -> None:
        equipo_global = _global_equipo(db)
        _make_producto(db, 303)
        _make_producto(db, 304)
        _set_producto_color(db, equipo_global.id, 303, color_ml="rojo")
        # 304 has no ProductoColor row -> outerjoin gives NULL -> matches sin_color
        db.commit()

        from app.api.endpoints.productos_shared import join_color_layer

        query = join_color_layer(self._query(db), equipo_global.id)
        query = filtro_colores(query, "sin_color", color_slot(None))
        results = query.all()

        assert {p.item_id for p, _ in results} == {304}

    def test_ml_slot_sin_color_plus_specific(self, db, rol_ventas) -> None:
        equipo_global = _global_equipo(db)
        _make_producto(db, 305)
        _make_producto(db, 306)
        _make_producto(db, 307)
        _set_producto_color(db, equipo_global.id, 305, color_ml="rojo")
        _set_producto_color(db, equipo_global.id, 306, color_ml="verde")
        # 307: no color
        db.commit()

        from app.api.endpoints.productos_shared import join_color_layer

        query = join_color_layer(self._query(db), equipo_global.id)
        query = filtro_colores(query, "sin_color,rojo", color_slot(None))
        results = query.all()

        assert {p.item_id for p, _ in results} == {305, 307}

    def test_tienda_slot(self, db, rol_ventas) -> None:
        equipo_global = _global_equipo(db)
        _make_producto(db, 308)
        _make_producto(db, 309)
        _set_producto_color(db, equipo_global.id, 308, color_tienda="azul")
        _set_producto_color(db, equipo_global.id, 309, color_tienda="gris")
        db.commit()

        from app.api.endpoints.productos_shared import join_color_layer

        query = join_color_layer(self._query(db), equipo_global.id)
        query = filtro_colores(query, "azul", color_slot("tienda"))
        results = query.all()

        assert {p.item_id for p, _ in results} == {308}

    def test_no_filter_returns_query_unchanged(self, db, rol_ventas) -> None:
        equipo_global = _global_equipo(db)
        _make_producto(db, 310)
        db.commit()

        from app.api.endpoints.productos_shared import join_color_layer

        query = join_color_layer(self._query(db), equipo_global.id)
        query = filtro_colores(query, None, color_slot(None))
        results = query.all()

        assert {p.item_id for p, _ in results} == {310}


# ---------------------------------------------------------------------------
# Listing endpoint parity (default = U layer, must equal today's legacy
# column behavior since producto_color@U is kept in sync by PR2 dual-write)
# ---------------------------------------------------------------------------


class TestListingParityDefaultLayer:
    def test_detail_color_field_matches_legacy_for_default_layer(self, client, db, rol_ventas) -> None:
        """Uses GET /api/productos/{item_id} (single-item detail), which shares
        the same color-layer helpers as the paginated listing endpoints but
        avoids the listing's Tienda Nube batch-fetch raw SQL (`= ANY(:ids)`,
        Postgres-only syntax unsupported by the SQLite test DB)."""
        equipo_global = _global_equipo(db)
        user = _make_usuario(db, "parity1", rol_ventas.id)
        _make_producto(db, 401, stock=5)
        _set_legacy_color(db, 401, color="rojo")
        _set_producto_color(db, equipo_global.id, 401, color_ml="rojo")
        db.commit()

        resp = client.get("/api/productos/401", headers=auth_headers_for(user))
        assert resp.status_code == 200
        item = resp.json()
        assert item["color_marcado"] == "rojo"
        # No team scope requested -> global hint equals the active color.
        assert item["color_hint_global"] == "rojo"
        assert item["color_hint_equipo_inicial"] is None

    def test_listing_filter_by_color_matches_legacy_filter_default_layer(self, db, rol_ventas) -> None:
        """Parity is verified at the query-construction level (same helper the
        `/api/productos` listing endpoint calls): filtro_colores against the
        default (global) layer must select exactly the items whose legacy
        `color_marcado` matched, matching today's behavior. (Full end-to-end
        HTTP coverage of the listing endpoint isn't possible under the SQLite
        test DB — see TestListingParityDefaultLayer docstring above.)"""
        equipo_global = _global_equipo(db)
        _make_producto(db, 402, stock=1)
        _make_producto(db, 403, stock=1)
        _set_legacy_color(db, 402, color="verde")
        _set_producto_color(db, equipo_global.id, 402, color_ml="verde")
        _set_legacy_color(db, 403, color="azul")
        _set_producto_color(db, equipo_global.id, 403, color_ml="azul")
        db.commit()

        from app.api.endpoints.productos_shared import join_color_layer

        query = db.query(ProductoERP, ProductoPricing).outerjoin(
            ProductoPricing, ProductoERP.item_id == ProductoPricing.item_id
        )
        query = join_color_layer(query, equipo_global.id)
        query = filtro_colores(query, "verde", color_slot(None))
        item_ids = {p.item_id for p, _ in query.all()}
        assert 402 in item_ids
        assert 403 not in item_ids


# ---------------------------------------------------------------------------
# Read-scoping across teams
# ---------------------------------------------------------------------------


class TestReadScopingAcrossTeams:
    def test_member_sees_team_color_and_global_hint(self, client, db, rol_ventas) -> None:
        equipo_global = _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Ventas")
        user = _make_usuario(db, "scoped1", rol_ventas.id)
        _add_member(db, equipo.id, user.id)
        _make_producto(db, 501, stock=1)
        _set_producto_color(db, equipo.id, 501, color_ml="violeta")
        _set_producto_color(db, equipo_global.id, 501, color_ml="verde")
        db.commit()

        # Requesting the team scope -> sees team color + global hint.
        resp = client.get(f"/api/productos/501?equipo_id={equipo.id}", headers=auth_headers_for(user))
        assert resp.status_code == 200
        item = resp.json()
        assert item["color_marcado"] == "violeta"
        assert item["color_hint_global"] == "verde"
        assert item["color_hint_equipo_inicial"] == "Equipo Ventas"[0]

        # Default request (no equipo_id) -> sees the global color.
        resp_default = client.get("/api/productos/501", headers=auth_headers_for(user))
        assert resp_default.status_code == 200
        assert resp_default.json()["color_marcado"] == "verde"

    def test_non_member_forbidden_via_listing_endpoint(self, client, db, rol_ventas) -> None:
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Aislado")
        user = _make_usuario(db, "scoped2", rol_ventas.id)
        db.commit()

        resp = client.get(f"/api/productos?equipo_id={equipo.id}", headers=auth_headers_for(user))
        assert resp.status_code == 403


# ---------------------------------------------------------------------------
# Membership 403 wiring on other equipo_id-accepting endpoints (PR3 review
# finding: only /api/productos was covered above).
# ---------------------------------------------------------------------------


class TestMembership403OtherEndpoints:
    def test_export_endpoint_rejects_non_member(self, client, db, rol_ventas) -> None:
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Export Aislado")
        user = _make_usuario(db, "exp1", rol_ventas.id)
        db.commit()

        resp = client.get(
            f"/api/exportar-web-transferencia?equipo_id={equipo.id}",
            headers=auth_headers_for(user),
        )
        assert resp.status_code == 403

        resp_default = client.get("/api/exportar-web-transferencia", headers=auth_headers_for(user))
        assert resp_default.status_code == 200

    def test_marcas_rejects_non_member(self, client, db, rol_ventas) -> None:
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Marcas Aislado")
        user = _make_usuario(db, "marc1", rol_ventas.id)
        db.commit()

        resp = client.get(f"/api/marcas?equipo_id={equipo.id}", headers=auth_headers_for(user))
        assert resp.status_code == 403

        resp_default = client.get("/api/marcas", headers=auth_headers_for(user))
        assert resp_default.status_code == 200

    def test_subcategorias_rejects_non_member(self, client, db, rol_ventas) -> None:
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Subcat Aislado")
        user = _make_usuario(db, "sub1", rol_ventas.id)
        db.commit()

        resp = client.get(f"/api/subcategorias?equipo_id={equipo.id}", headers=auth_headers_for(user))
        assert resp.status_code == 403

        resp_default = client.get("/api/subcategorias", headers=auth_headers_for(user))
        assert resp_default.status_code == 200

    def test_stats_endpoint_rejects_non_member(self, client, db, rol_ventas) -> None:
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Stats Aislado")
        user = _make_usuario(db, "sta1", rol_ventas.id)
        db.commit()

        resp = client.get(f"/api/stats?equipo_id={equipo.id}", headers=auth_headers_for(user))
        assert resp.status_code == 403

        resp_default = client.get("/api/stats", headers=auth_headers_for(user))
        assert resp_default.status_code == 200

    def test_exportar_vista_actual_rejects_non_member_with_403_not_500(self, client, db, rol_ventas) -> None:
        """Guardian Angel round-2 finding: `resolver_layer_activo`'s 403 was
        raised INSIDE a broad `try/except Exception`, re-wrapped as a 500.
        Fixed via `except HTTPException: raise` before the generic handler."""
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Vista Aislado")
        user = _make_usuario(db, "vista1", rol_ventas.id)
        db.commit()

        resp = client.get(f"/api/exportar-vista-actual?equipo_id={equipo.id}", headers=auth_headers_for(user))
        assert resp.status_code == 403

        resp_default = client.get("/api/exportar-vista-actual", headers=auth_headers_for(user))
        assert resp_default.status_code == 200

    def test_exportar_lista_gremio_rejects_non_member_with_403_not_500(self, client, db, rol_ventas) -> None:
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Gremio Aislado")
        user = _make_usuario(db, "gremio1", rol_ventas.id)
        db.commit()

        resp = client.get(f"/api/exportar-lista-gremio?equipo_id={equipo.id}", headers=auth_headers_for(user))
        assert resp.status_code == 403

    def test_exportar_lista_sugerido_rejects_non_member_with_403_not_500(self, client, db, rol_ventas) -> None:
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Sugerido Aislado")
        user = _make_usuario(db, "sugerido1", rol_ventas.id)
        db.commit()

        resp = client.get(f"/api/exportar-lista-sugerido?equipo_id={equipo.id}", headers=auth_headers_for(user))
        assert resp.status_code == 403

    def test_exportar_lista_web_transferencia_rejects_non_member_with_403_not_500(self, client, db, rol_ventas) -> None:
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Web Transf Aislado")
        user = _make_usuario(db, "webtransf1", rol_ventas.id)
        db.commit()

        resp = client.get(
            f"/api/exportar-lista-web-transferencia?equipo_id={equipo.id}", headers=auth_headers_for(user)
        )
        assert resp.status_code == 403

    def test_pricing_post_rejects_non_member(self, client, db, rol_ventas) -> None:
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Pricing Aislado")
        user = _make_usuario(db, "pri1", rol_ventas.id)
        db.commit()

        resp = client.post(
            "/api/productos/calcular-web-masivo",
            json={
                "porcentaje_con_precio": 10.0,
                "porcentaje_sin_precio": 20.0,
                "filtros": {"equipo_id": equipo.id},
            },
            headers=auth_headers_for(user),
        )
        assert resp.status_code == 403

        resp_default = client.post(
            "/api/productos/calcular-web-masivo",
            json={
                "porcentaje_con_precio": 10.0,
                "porcentaje_sin_precio": 20.0,
            },
            headers=auth_headers_for(user),
        )
        assert resp_default.status_code == 200


# ---------------------------------------------------------------------------
# FIX 1 regression: color export must not depend on a productos_pricing row.
# ---------------------------------------------------------------------------


class TestExportColorIndependentOfPricing:
    def test_exportar_vista_actual_includes_color_without_pricing_row(self, client, db, rol_ventas) -> None:
        """Regression for the exportar_vista_actual bug where the color cell
        write lived inside `if producto_pricing:`, silently dropping the
        color for products with a `producto_color` row but no pricing row.

        Unlike some listing/export endpoints, `exportar_vista_actual` doesn't
        use the Postgres-only `= ANY()` raw SQL path, so it's driven fully
        end-to-end here: a product with a color and NO ProductoPricing row
        must still show its color in the exported workbook.
        """
        from io import BytesIO
        from openpyxl import load_workbook

        equipo_global = _global_equipo(db)
        user = _make_usuario(db, "exportcolor1", rol_ventas.id)
        _make_producto(db, 601, stock=1)  # no ProductoPricing row created
        _set_producto_color(db, equipo_global.id, 601, color_ml="amarillo")
        db.commit()

        resp = client.get("/api/exportar-vista-actual", headers=auth_headers_for(user))
        assert resp.status_code == 200

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        color_col = header.index("Color") + 1
        codigo_col = header.index("Código") + 1

        row_values = None
        for row in ws.iter_rows(min_row=2):
            if row[codigo_col - 1].value == "COD601":
                row_values = row
                break

        assert row_values is not None, "Expected product COD601 in the export"
        assert row_values[color_col - 1].value == "amarillo"


# ---------------------------------------------------------------------------
# Guardian Angel review round 2, BLOCKING 1: coerce_equipo_id defensive
# coercion at every filtros-dict extraction site (masivo/export endpoints).
# ---------------------------------------------------------------------------


class TestCoerceEquipoId:
    def test_string_int_resolves_identically_to_int(self, db, rol_ventas) -> None:
        from app.api.endpoints.productos_shared import coerce_equipo_id

        equipo_global = _global_equipo(db)
        assert coerce_equipo_id({"equipo_id": str(equipo_global.id)}) == equipo_global.id
        assert coerce_equipo_id({"equipo_id": equipo_global.id}) == equipo_global.id

    def test_none_and_missing_and_empty_filtros(self) -> None:
        from app.api.endpoints.productos_shared import coerce_equipo_id

        assert coerce_equipo_id(None) is None
        assert coerce_equipo_id({}) is None
        assert coerce_equipo_id({"equipo_id": None}) is None

    def test_garbage_value_raises_422_not_500(self) -> None:
        from app.api.endpoints.productos_shared import coerce_equipo_id

        with pytest.raises(HTTPException) as exc:
            coerce_equipo_id({"equipo_id": "abc"})
        assert exc.value.status_code == 422

    def test_masivo_endpoint_accepts_string_equipo_id(self, client, db, rol_ventas) -> None:
        """String equipo_id in the filtros body resolves the same as int and,
        for a non-member team, still yields a clean 403 (not a 500)."""
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Coerce String")
        user = _make_usuario(db, "coerce1", rol_ventas.id)
        db.commit()

        resp = client.post(
            "/api/productos/calcular-web-masivo",
            json={
                "porcentaje_con_precio": 10.0,
                "porcentaje_sin_precio": 20.0,
                "filtros": {"equipo_id": str(equipo.id)},
            },
            headers=auth_headers_for(user),
        )
        assert resp.status_code == 403

    def test_masivo_endpoint_rejects_garbage_equipo_id_with_422(self, client, db, rol_ventas) -> None:
        _global_equipo(db)
        user = _make_usuario(db, "coerce2", rol_ventas.id)
        db.commit()

        resp = client.post(
            "/api/productos/calcular-web-masivo",
            json={
                "porcentaje_con_precio": 10.0,
                "porcentaje_sin_precio": 20.0,
                "filtros": {"equipo_id": "abc"},
            },
            headers=auth_headers_for(user),
        )
        assert resp.status_code == 422

    def test_export_rebate_endpoint_accepts_string_equipo_id(self, client, db, rol_ventas) -> None:
        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Coerce Export")
        user = _make_usuario(db, "coerce3", rol_ventas.id)
        db.commit()

        resp = client.post(
            "/api/productos/exportar-rebate",
            json={"filtros": {"equipo_id": str(equipo.id)}},
            headers=auth_headers_for(user),
        )
        assert resp.status_code == 403


# ---------------------------------------------------------------------------
# Guardian Angel review round 2, BLOCKING 2: /marcas and /subcategorias now
# support the "sin_color" sentinel like every other listing endpoint. Legacy
# behavior had no such branch at all (see git history of productos_metadata.py
# on main), so adding it is an intentional consistency fix, not a regression.
# ---------------------------------------------------------------------------


class TestMarcasSubcategoriasSinColorSentinel:
    def test_marcas_sin_color_sentinel(self, client, db, rol_ventas) -> None:
        equipo_global = _global_equipo(db)
        user = _make_usuario(db, "marcasin1", rol_ventas.id)
        _make_producto(db, 701, stock=1, marca="MarcaConColor")
        _make_producto(db, 702, stock=1, marca="MarcaSinColor")
        _set_producto_color(db, equipo_global.id, 701, color_ml="rojo")
        # 702 has no ProductoColor row -> matches sin_color sentinel
        db.commit()

        resp = client.get("/api/marcas?colores=sin_color", headers=auth_headers_for(user))
        assert resp.status_code == 200
        nombres = set(resp.json()["marcas"])
        assert "MarcaSinColor" in nombres
        assert "MarcaConColor" not in nombres

    def test_subcategorias_sin_color_sentinel(self, client, db, rol_ventas) -> None:
        from app.models.comision_config import SubcategoriaGrupo

        equipo_global = _global_equipo(db)
        user = _make_usuario(db, "subcatsin1", rol_ventas.id)
        _make_producto(db, 703, stock=1, subcategoria_id=901)
        _make_producto(db, 704, stock=1, subcategoria_id=902)
        _set_producto_color(db, equipo_global.id, 703, color_ml="rojo")
        # 704 has no ProductoColor row -> matches sin_color sentinel
        db.add(
            SubcategoriaGrupo(
                subcat_id=901, grupo_id=1, nombre_subcategoria="Subcat con color", nombre_categoria="Cat"
            )
        )
        db.add(
            SubcategoriaGrupo(
                subcat_id=902, grupo_id=1, nombre_subcategoria="Subcat sin color", nombre_categoria="Cat"
            )
        )
        db.commit()

        resp = client.get("/api/subcategorias?colores=sin_color", headers=auth_headers_for(user))
        assert resp.status_code == 200
        subcats = {
            s["nombre"] for cat in resp.json()["categorias"] for s in cat["subcategorias"]
        }
        assert "Subcat sin color" in subcats
        assert "Subcat con color" not in subcats


# ---------------------------------------------------------------------------
# Guardian Angel review round 2, BLOCKING 3: layer resolution dedup —
# resolver_layer_activo accepts a pre-fetched global_equipo_id to avoid a
# redundant get_global_equipo_id query when the caller already resolved it.
# ---------------------------------------------------------------------------


class TestResolverLayerActivoGlobalIdReuse:
    def test_passing_precomputed_global_id_matches_default_resolution(self, db, rol_ventas) -> None:
        from app.api.endpoints.productos_shared import get_global_equipo_id

        equipo_global = _global_equipo(db)
        user = _make_usuario(db, "dedup1", rol_ventas.id)

        precomputed = get_global_equipo_id(db)
        assert precomputed == equipo_global.id
        assert resolver_layer_activo(None, user, db, precomputed) == equipo_global.id
        assert resolver_layer_activo(equipo_global.id, user, db, precomputed) == equipo_global.id

    def test_member_of_team_allowed_with_precomputed_global_id(self, db, rol_ventas) -> None:
        from app.api.endpoints.productos_shared import get_global_equipo_id

        _global_equipo(db)
        equipo = _make_equipo(db, "Equipo Dedup")
        user = _make_usuario(db, "dedup2", rol_ventas.id)
        _add_member(db, equipo.id, user.id)

        precomputed = get_global_equipo_id(db)
        assert resolver_layer_activo(equipo.id, user, db, precomputed) == equipo.id

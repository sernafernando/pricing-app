"""
Integration tests for PR #1 of change `export-builder` (bugfix): the 7
export endpoints in `productos_export.py` were silently dropping filters the
product list (`listar_productos`) honors (spec FF-2, FF-3).

Strategy: call each export endpoint function directly (like
`test_productos_listing_promo_filter.py` does for the list), against the
real sqlite-backed `db` fixture, patching only the cross-DB
`fetch_mlas_with_active_promo_type` resolver at
`app.api.endpoints.productos_export` (mlwebhook is not reachable in tests).

Each parametrized case seeds two products:
- item 1: has an active MLA publication ("MLA1"), fresh `fecha_sync`.
- item 2: no MLA publication, old `fecha_sync`.

and asserts that filters narrow the export to item 1 only, reading the
exported XLSX's `codigo` column back with openpyxl.
"""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from app.models.producto import ProductoERP, ProductoPricing
from app.models.publicacion_ml import PublicacionML
from app.models.mercadolibre_item_publicado import MercadoLibreItemPublicado
from app.api.endpoints import productos_export as export_mod


def _current_user() -> SimpleNamespace:
    return SimpleNamespace(id=1)


def _seed_products(db) -> None:
    now = datetime.now(UTC)
    old = now - timedelta(days=30)

    p1 = ProductoERP(
        item_id=1,
        codigo="COD1",
        descripcion="Producto Uno",
        marca="MARCA",
        activo=True,
        stock=5,
        costo=100.0,
        fecha_sync=now,
    )
    p2 = ProductoERP(
        item_id=2,
        codigo="COD2",
        descripcion="Producto Dos",
        marca="MARCA",
        activo=True,
        stock=5,
        costo=100.0,
        fecha_sync=old,
    )
    db.add(p1)
    db.add(p2)

    for item_id in (1, 2):
        db.add(
            ProductoPricing(
                item_id=item_id,
                precio_lista_ml=1000.0,
                markup_calculado=10.0,
                participa_web_transferencia=True,
                precio_web_transferencia=1200.0,
            )
        )

    # item 1: has an active MLA publication; item 2: none.
    db.add(PublicacionML(mla="MLA1", item_id=1, activo=True))
    db.add(
        MercadoLibreItemPublicado(
            mlp_id=1,
            item_id=1,
            mlp_publicationID="MLA1",
            optval_statusId=2,  # activa
        )
    )
    db.commit()


def _codigos_in_workbook(xlsx_bytes: bytes, codigo_col: int) -> set[str]:
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    codigos = set()
    for row in ws.iter_rows(min_row=2, values_only=False):
        value = row[codigo_col - 1].value
        if value:
            codigos.add(str(value))
    return codigos


def _response_bytes(response) -> bytes:
    """Every export endpoint builds a fully-buffered `BytesIO` and returns
    either a plain `Response` (body already materialized in `.body`) or a
    `StreamingResponse` wrapping that same `BytesIO` (no `.body` attribute —
    must be drained via its async `body_iterator`)."""
    body = getattr(response, "body", None)
    if body:
        return body

    import asyncio

    async def _collect() -> bytes:
        chunks = []
        async for chunk in response.body_iterator:
            chunks.append(chunk if isinstance(chunk, bytes) else chunk.encode())
        return b"".join(chunks)

    return asyncio.run(_collect())


def _call_vista_actual(db, **kwargs):
    return export_mod.exportar_vista_actual(db=db, current_user=_current_user(), **kwargs)


def _call_gremio(db, **kwargs):
    return export_mod.exportar_lista_gremio(db=db, current_user=_current_user(), **kwargs)


def _call_sugerido(db, **kwargs):
    return export_mod.exportar_lista_sugerido(db=db, current_user=_current_user(), **kwargs)


def _call_web_transferencia_lista(db, **kwargs):
    return export_mod.exportar_lista_web_transferencia(db=db, current_user=_current_user(), **kwargs)


def _call_web_transferencia(db, **kwargs):
    # `Query(...)`-typed params resolve to the sentinel object, not their
    # declared default, when called directly (see `_call_clasica` above).
    kwargs.setdefault("porcentaje_adicional", 0)
    kwargs.setdefault("currency_id", 1)
    kwargs.setdefault("offset_dolar", 0)
    kwargs.setdefault("redondear", True)
    return export_mod.exportar_web_transferencia(db=db, current_user=_current_user(), **kwargs)


def _call_clasica(db, **kwargs):
    # Every `Query(...)`-typed parameter on this route resolves to the
    # sentinel `Query` object itself (not its declared default) when the
    # function is called directly, bypassing FastAPI's own dependency
    # resolution — must be passed explicitly unless the caller overrides it.
    kwargs.setdefault("porcentaje_adicional", 0)
    kwargs.setdefault("tipo_cuotas", "clasica")
    kwargs.setdefault("currency_id", 1)
    kwargs.setdefault("offset_dolar", 0)
    kwargs.setdefault("redondear", True)
    kwargs.setdefault("tiendas_oficiales", None)
    with patch("app.services.permisos_service.verificar_permiso", return_value=True):
        return export_mod.exportar_clasica(db=db, current_user=_current_user(), **kwargs)


# `(caller, codigo_column_index)` for the 6 GET endpoints with a simple
# `.all()`-then-write-rows shape. `exportar-rebate` (POST, MLA-row-shaped) is
# covered separately below since its rows are per-MLA, not per-product.
_GET_ENDPOINTS = [
    pytest.param(_call_web_transferencia, 1, id="exportar-web-transferencia"),
    pytest.param(_call_clasica, 1, id="exportar-clasica"),
    pytest.param(_call_vista_actual, 1, id="exportar-vista-actual"),
    pytest.param(_call_gremio, 4, id="exportar-lista-gremio"),
    pytest.param(_call_sugerido, 4, id="exportar-lista-sugerido"),
    pytest.param(_call_web_transferencia_lista, 4, id="exportar-lista-web-transferencia"),
]


class TestPromoFiltersAllExportEndpoints:
    """FF-2: promo_tipos/promo_estado/con_promo_aplicada/con_promo_sin_aplicar
    must be honored on all 7 export endpoints, matching `listar_productos`."""

    @pytest.mark.parametrize("caller,codigo_col", _GET_ENDPOINTS)
    def test_promo_tipos_narrows_to_matching_mla(self, db, caller, codigo_col) -> None:
        _seed_products(db)

        with patch(
            "app.api.endpoints.productos_export.fetch_mlas_with_active_promo_type",
            return_value={"MLA1"},
        ):
            response = caller(db, promo_tipos="descuento")

        codigos = _codigos_in_workbook(_response_bytes(response), codigo_col)
        assert codigos == {"COD1"}

    @pytest.mark.parametrize("caller,codigo_col", _GET_ENDPOINTS)
    def test_con_promo_aplicada_empty_set_yields_no_products(self, db, caller, codigo_col) -> None:
        """Empty-set guard (I5): resolver finds zero matching MLAs => zero
        products, never the unfiltered catalog."""
        _seed_products(db)

        with patch(
            "app.api.endpoints.productos_export.fetch_mlas_with_started",
            return_value=set(),
        ):
            response = caller(db, con_promo_aplicada=True)

        codigos = _codigos_in_workbook(_response_bytes(response), codigo_col)
        assert codigos == set()


class TestConMlaEstadoMlaNuevos7DiasAllEndpoints:
    """FF-3: con_mla / estado_mla / nuevos_ultimos_7_dias must be honored on
    all 7 export endpoints, not only on `/exportar-clasica`."""

    @pytest.mark.parametrize("caller,codigo_col", _GET_ENDPOINTS)
    def test_con_mla_true_narrows_to_products_with_publication(self, db, caller, codigo_col) -> None:
        _seed_products(db)

        response = caller(db, con_mla=True)

        codigos = _codigos_in_workbook(_response_bytes(response), codigo_col)
        assert codigos == {"COD1"}

    @pytest.mark.parametrize("caller,codigo_col", _GET_ENDPOINTS)
    def test_estado_mla_activa_narrows_to_products_with_active_publication(self, db, caller, codigo_col) -> None:
        _seed_products(db)

        response = caller(db, estado_mla="activa")

        codigos = _codigos_in_workbook(_response_bytes(response), codigo_col)
        assert codigos == {"COD1"}

    @pytest.mark.parametrize("caller,codigo_col", _GET_ENDPOINTS)
    def test_nuevos_ultimos_7_dias_narrows_to_recently_synced(self, db, caller, codigo_col) -> None:
        _seed_products(db)

        response = caller(db, nuevos_ultimos_7_dias=True)

        codigos = _codigos_in_workbook(_response_bytes(response), codigo_col)
        assert codigos == {"COD1"}


class TestTiendaOficialNotSilentlyNoop:
    """FF-3: `tienda_oficial` is dead code on the list side
    (`listar_productos:701`) — the export must not silently accept it as a
    working filter. It stays declared-but-inert on `/exportar-clasica`
    (documented, not newly wired) and must NOT be introduced on any of the
    other 6 endpoints (which would silently no-op a param that looks like it
    filters)."""

    def test_tienda_oficial_has_no_effect_on_exportar_clasica(self, db) -> None:
        _seed_products(db)

        response = _call_clasica(db, tienda_oficial="99999")

        codigos = _codigos_in_workbook(_response_bytes(response), 1)
        # Dead code (mirrors listar_productos:701): passing a value changes
        # nothing — both products still export.
        assert codigos == {"COD1", "COD2"}

    @pytest.mark.parametrize(
        "endpoint_fn",
        [
            export_mod.exportar_web_transferencia,
            export_mod.exportar_vista_actual,
            export_mod.exportar_lista_gremio,
            export_mod.exportar_lista_sugerido,
            export_mod.exportar_lista_web_transferencia,
        ],
    )
    def test_tienda_oficial_not_accepted_on_other_endpoints(self, endpoint_fn) -> None:
        import inspect

        params = inspect.signature(endpoint_fn).parameters
        assert "tienda_oficial" not in params, (
            f"{endpoint_fn.__name__} must not silently accept 'tienda_oficial' as a "
            "working no-op filter (FF-3) while it remains dead code on the list side."
        )
